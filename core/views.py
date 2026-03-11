import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Avg, Count, Q, Sum, OuterRef, Subquery, Prefetch, Max
from django.db import models
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.utils.dateparse import parse_date
from django.http import HttpResponse, JsonResponse
from django.template.loader import get_template
from io import BytesIO
from xhtml2pdf import pisa
from django.urls import reverse
from django.views.decorators.csrf import csrf_protect
from django.utils.http import url_has_allowed_host_and_scheme
from urllib.parse import urlencode
import pandas as pd
from .models import (
    User, Etudiant, Enseignant, Jury, Evaluation, 
    Inscription, UE, EC, Cohorte, Departement, Attribution, ParametreEvaluation,
    CommuniqueDeliberation, Deliberation, CommentaireCours, EvaluationEnseignement, BulletinNotes,
    Recours, Notification, HistoriqueAction, DocumentCours, InscriptionUE
)
from reglage.models import Classe, Categorie, Fonction, Grade, Departement as ReglDepartement, Section, TypeCharge
from .forms import UserForm, UserEditForm, EtudiantForm, EnseignantForm, UEForm, ECForm, JuryForm, EvaluationForm, CohorteForm, InscriptionForm, PhotoForm
from .decorators_permissions import require_admin, require_gestionnaire_or_admin, require_staff_or_roles


# ========== HELPERS POUR SIMULATION ADMIN ==========

def _extract_jury_classe_code(username):
    """Extrait le code classe d'un username jury.
    Supporte les formats:
      - jury_pres_L1INFO (ancien)
      - jury_pres_L1INFO_27 (nouveau, avec suffixe année)
      - jury_sec_L1INFO, jury_sec_L1INFO_27
    Retourne le code_classe (ex: 'L1INFO') ou None.
    """
    import re
    for prefix in ('jury_pres_', 'jury_sec_'):
        if username.startswith(prefix):
            rest = username[len(prefix):]  # ex: 'L1INFO' ou 'L1INFO_27'
            # Retirer un éventuel suffixe _XX (2 chiffres = année)
            cleaned = re.sub(r'_\d{2}$', '', rest)
            return cleaned
    return None

def get_simulated_etudiant(request):
    """Récupère l'étudiant simulé (pour admin) ou l'étudiant connecté"""
    if request.user.is_staff and 'simulated_etudiant' in request.session:
        return Etudiant.objects.filter(matricule_et=request.session['simulated_etudiant']).first()
    return Etudiant.objects.filter(id_lgn=request.user).first()

def get_simulated_enseignant(request):
    """Récupère l'enseignant simulé (pour admin) ou l'enseignant connecté"""
    if request.user.is_staff and 'simulated_enseignant' in request.session:
        return Enseignant.objects.select_related('grade').filter(matricule_en=request.session['simulated_enseignant']).first()
    return Enseignant.objects.select_related('grade').filter(id_lgn=request.user).first()

def get_simulated_jury(request):
    """Récupère le jury simulé (pour admin) ou le jury connecté"""
    if request.user.is_staff and 'simulated_jury' in request.session:
        return Jury.objects.select_related('code_classe').filter(code_jury=request.session['simulated_jury']).first()
    # Chercher le jury normalement
    jury = Jury.objects.filter(id_lgn=request.user).first()
    if not jury:
        username = request.user.username
        classe_code = _extract_jury_classe_code(username)
        if classe_code:
            jury = Jury.objects.filter(code_classe__code_classe=classe_code).order_by('-annee_academique').first()
    return jury


# ========== HELPERS POUR LES RÔLES ==========

def is_gestionnaire(user):
    """Vérifie si l'utilisateur est gestionnaire"""
    return user.role == 'GESTIONNAIRE' or user.is_staff and user.role == 'ADMIN'

def is_agent(user):
    """Vérifie si l'utilisateur est agent"""
    return user.role == 'AGENT'

def can_access_historique(user):
    """Vérifie si l'utilisateur peut accéder à l'historique (admin seulement)"""
    return user.is_staff and user.role == 'ADMIN'

def can_manage_enseignants(user):
    """Vérifie si l'utilisateur peut gérer les enseignants"""
    return user.is_staff or user.role in ['ADMIN', 'GESTIONNAIRE', 'AGENT']

def can_manage_etudiants(user):
    """Vérifie si l'utilisateur peut gérer les étudiants"""
    return user.is_staff or user.role in ['ADMIN', 'GESTIONNAIRE', 'AGENT']

def can_manage_ue_ec(user):
    """Vérifie si l'utilisateur peut gérer les UE et EC"""
    return user.is_staff or user.role in ['ADMIN', 'GESTIONNAIRE', 'AGENT']

def can_manage_inscriptions(user):
    """Vérifie si l'utilisateur peut gérer les inscriptions"""
    return user.is_staff or user.role in ['ADMIN', 'GESTIONNAIRE', 'AGENT']

def can_manage_all(user):
    """Vérifie si l'utilisateur peut tout gérer (admin/gestionnaire)"""
    return user.is_staff or user.role in ['ADMIN', 'GESTIONNAIRE']
def is_admin_viewing(request):
    """Vérifie si l'admin visualise un espace simulé"""
    return request.user.is_staff and (
        'simulated_etudiant' in request.session or 
        'simulated_enseignant' in request.session or 
        'simulated_jury' in request.session
    )

def clear_simulation(request):
    """Efface toutes les simulations de la session"""
    request.session.pop('simulated_etudiant', None)
    request.session.pop('simulated_enseignant', None)
    request.session.pop('simulated_jury', None)


def notifier_enseignants_communique(communique):
    """
    Notifie les enseignants qui ont eu des cours dans la classe concernée par le communiqué de délibération.
    Retourne la liste des enseignants notifiés.
    """
    enseignants_notifies = []
    
    try:
        # Récupérer toutes les attributions pour cette classe et cette année académique
        attributions = Attribution.objects.filter(
            annee_academique=communique.annee_academique
        ).select_related('matricule_en', 'matricule_en__id_lgn')
        
        # Filtrer les attributions qui concernent la classe du communiqué
        for attribution in attributions:
            cours_obj = attribution.get_cours_object()
            if cours_obj and hasattr(cours_obj, 'classe') and cours_obj.classe:
                if cours_obj.classe.code_classe == communique.code_classe.code_classe:
                    enseignant = attribution.matricule_en
                    if enseignant.id_lgn and enseignant not in enseignants_notifies:
                        enseignants_notifies.append(enseignant)
                        
                        # Créer une notification persistante pour l'enseignant
                        Notification.objects.create(
                            destinataire=enseignant.id_lgn,
                            type_notification='COMMUNIQUE',
                            titre=f'Communiqué de délibération - {communique.code_classe.code_classe}',
                            message=f'''Un nouveau communiqué de délibération a été publié pour votre classe {communique.code_classe.code_classe}.

Date de délibération: {communique.date_deliberation.strftime("%d/%m/%Y")}
Année académique: {communique.annee_academique}

{communique.contenu if communique.contenu else "Aucun message détaillé."}''',
                            lien=f"{reverse('enseignant_dashboard')}",
                            code_classe=communique.code_classe,
                            annee_academique=communique.annee_academique
                        )
                            
                        # Log de la notification (pour débogage)
                        print(f"Notification envoyée à l'enseignant: {enseignant.nom_complet} pour le communiqué de {communique.code_classe.code_classe}")
        
        return enseignants_notifies
        
    except Exception as e:
        print(f"Erreur lors de la notification des enseignants: {str(e)}")
        return []


@login_required
def admin_simulate_etudiant(request, matricule):
    """Admin : simuler la connexion d'un étudiant"""
    if not request.user.is_staff:
        return redirect('home')

    clear_simulation(request)
    request.session['simulated_etudiant'] = matricule
    next_url = request.GET.get('next')
    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        return redirect(next_url)
    return redirect('etudiant_dashboard')


@login_required
def etudiant_evaluer_enseignant(request):
    etudiant = get_simulated_etudiant(request)
    if not etudiant:
        messages.error(request, 'Profil étudiant non trouvé.')
        return redirect('home')

    if is_admin_viewing(request):
        messages.error(request, "Impossible d'évaluer un enseignant en mode visualisation Admin.")
        return redirect('etudiant_dashboard')

    from reglage.models import AnneeAcademique, Semestre

    annee_active = AnneeAcademique.get_annee_en_cours()
    annee_code = request.GET.get('annee') or (annee_active.code_anac if annee_active else None)

    inscriptions = Inscription.objects.filter(matricule_etudiant=etudiant).select_related('code_classe', 'cohorte')
    inscription_active = None
    if annee_code:
        inscription_active = inscriptions.filter(annee_academique=annee_code).first()
    if not inscription_active:
        inscription_active = inscriptions.order_by('-annee_academique').first()

    classe_active = inscription_active.code_classe if inscription_active else None
    if isinstance(classe_active, str):
        try:
            classe_active = Classe.objects.get(pk=classe_active)
        except Classe.DoesNotExist:
            classe_active = None
    if not classe_active:
        messages.error(request, "Aucune classe trouvée pour cet étudiant.")
        return redirect('etudiant_dashboard')

    # Si l'année sélectionnée n'a pas d'inscription, utiliser l'année de l'inscription active
    if annee_code and inscription_active and inscription_active.annee_academique != annee_code:
        annee_code = inscription_active.annee_academique

    semestre = request.GET.get('semestre')

    # Fonction helper pour obtenir le semestre d'un cours
    def get_semestre(cours_obj):
        sem = None
        if hasattr(cours_obj, 'semestre'):
            sem = cours_obj.semestre
        elif hasattr(cours_obj, 'code_ue') and cours_obj.code_ue:
            sem = cours_obj.code_ue.semestre
        return f"S{sem}" if sem else None

    # Calculer les semestres disponibles à partir des cours de la classe
    semestres_from_data = set()
    if classe_active:
        classe_code = classe_active.code_classe if hasattr(classe_active, 'code_classe') else classe_active
        ue_qs = UE.objects.filter(classe__code_classe=classe_code)
        ec_qs = EC.objects.filter(classe__code_classe=classe_code)
        for ue in ue_qs:
            sem = get_semestre(ue)
            if sem:
                semestres_from_data.add(sem)
        for ec in ec_qs:
            sem = get_semestre(ec)
            if sem:
                semestres_from_data.add(sem)

    attributions_annee = Attribution.objects.filter(
        annee_academique=inscription_active.annee_academique
    ).select_related('matricule_en', 'type_charge')

    attributions_cibles = []
    for attr in attributions_annee:
        cours_obj = attr.get_cours_object()
        if not cours_obj or not getattr(cours_obj, 'classe', None):
            continue
        if str(cours_obj.classe.code_classe) != str(classe_active.code_classe):
            continue
        attributions_cibles.append({
            'id': attr.pk,
            'code_cours': attr.code_cours,
            'type_cours': attr.get_type_cours(),
            'intitule': attr.get_intitule(),
            'enseignant': attr.matricule_en,
            'annee_academique': attr.annee_academique,
            'cours_obj': cours_obj,
        })

    # Fonction helper pour obtenir le semestre d'un cours
    def get_semestre(cours_obj):
        if hasattr(cours_obj, 'semestre'):
            return cours_obj.semestre
        elif hasattr(cours_obj, 'code_ue') and cours_obj.code_ue:
            return cours_obj.code_ue.semestre
        return None

    # Filtrer par semestre si sélectionné
    if semestre:
        attributions_cibles = [a for a in attributions_cibles if str(get_semestre(a['cours_obj'])) == semestre]

    # Dédoublonner par enseignant : un enseignant = une attribution (la première trouvée)
    from collections import OrderedDict
    teacher_to_attr = OrderedDict()
    for a in attributions_cibles:
        matricule = a['enseignant'].matricule_en
        if matricule not in teacher_to_attr:
            teacher_to_attr[matricule] = a
    attributions_cibles = list(teacher_to_attr.values())

    attributions_cibles.sort(key=lambda a: str(a['enseignant']))

    existing_qs = EvaluationEnseignement.objects.filter(
        etudiant=etudiant,
        annee_academique=inscription_active.annee_academique,
        attribution__in=[a['id'] for a in attributions_cibles],
    ).select_related('attribution', 'attribution__matricule_en').order_by('-date_creation')

    # Préparer les options pour les filtres
    annees = sorted(set(i.annee_academique for i in inscriptions))
    semestres = Semestre.objects.all().order_by('code_semestre')

    if request.method == 'POST':
        attribution_id = request.POST.get('attribution_id')
        try:
            attribution = Attribution.objects.select_related('matricule_en').get(pk=attribution_id)
        except (Attribution.DoesNotExist, ValueError, TypeError):
            attribution = None

        allowed_ids = {str(a['id']) for a in attributions_cibles}
        if not attribution or str(attribution.pk) not in allowed_ids:
            messages.error(request, "Cours/enseignant invalide.")
            return redirect('etudiant_evaluer_enseignant')

        def _to_int(v):
            try:
                return int(v)
            except (TypeError, ValueError):
                return None

        ponctualite = _to_int(request.POST.get('ponctualite'))
        maitrise_communication = _to_int(request.POST.get('maitrise_communication'))
        pedagogie_methodologie = _to_int(request.POST.get('pedagogie_methodologie'))
        utilisation_tic = _to_int(request.POST.get('utilisation_tic'))
        disponibilite = _to_int(request.POST.get('disponibilite'))
        commentaire = (request.POST.get('commentaire') or '').strip()

        notes = [ponctualite, maitrise_communication, pedagogie_methodologie, utilisation_tic, disponibilite]
        if any(n is None or n < 1 or n > 5 for n in notes):
            messages.error(request, "Veuillez renseigner toutes les notes (1 à 5).")
            return redirect('etudiant_evaluer_enseignant')

        EvaluationEnseignement.objects.update_or_create(
            etudiant=etudiant,
            attribution=attribution,
            defaults={
                'annee_academique': inscription_active.annee_academique,
                'ponctualite': ponctualite,
                'maitrise_communication': maitrise_communication,
                'pedagogie_methodologie': pedagogie_methodologie,
                'utilisation_tic': utilisation_tic,
                'disponibilite': disponibilite,
                'commentaire': commentaire,
            }
        )
        messages.success(request, "Évaluation enregistrée avec succès.")
        return redirect('etudiant_evaluer_enseignant')

    context = {
        'etudiant': etudiant,
        'annee_active': annee_active,
        'annee_code': annee_code,
        'inscription_active': inscription_active,
        'classe_active': classe_active,
        'attributions': attributions_cibles,
        'evaluations': existing_qs,
        'annees': annees,
        'semestres': semestres,
        'semestre': semestre,
        'semestres_from_data': semestres_from_data,
    }
    return render(request, 'etudiant/evaluer_enseignant.html', context)


@login_required
def enseignant_appreciations(request):
    enseignant = get_simulated_enseignant(request)
    if not enseignant:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')

    from reglage.models import AnneeAcademique

    annee_filter = request.GET.get('annee', '')
    cours_filter = request.GET.get('cours', '')

    annees = AnneeAcademique.objects.all().order_by('-code_anac')

    attributions = Attribution.objects.filter(matricule_en=enseignant)
    if annee_filter:
        attributions = attributions.filter(annee_academique=annee_filter)
    if cours_filter:
        attributions = attributions.filter(code_cours=cours_filter)

    cours_list = []
    for attr in Attribution.objects.filter(matricule_en=enseignant):
        cours_list.append({'code': attr.code_cours, 'intitule': attr.get_intitule()})
    seen = set()
    dedup = []
    for c in cours_list:
        if c['code'] in seen:
            continue
        seen.add(c['code'])
        dedup.append(c)
    cours_list = sorted(dedup, key=lambda x: x['code'])

    evaluations_qs = EvaluationEnseignement.objects.filter(
        attribution__in=attributions
    ).select_related('attribution', 'attribution__matricule_en').order_by('-date_creation')

    stats = evaluations_qs.aggregate(
        avg_ponctualite=Avg('ponctualite'),
        avg_maitrise_communication=Avg('maitrise_communication'),
        avg_pedagogie_methodologie=Avg('pedagogie_methodologie'),
        avg_utilisation_tic=Avg('utilisation_tic'),
        avg_disponibilite=Avg('disponibilite'),
        total=Count('id'),
    )
    overall = None
    avgs = [
        stats.get('avg_ponctualite'),
        stats.get('avg_maitrise_communication'),
        stats.get('avg_pedagogie_methodologie'),
        stats.get('avg_utilisation_tic'),
        stats.get('avg_disponibilite'),
    ]
    if all(v is not None for v in avgs) and stats.get('total'):
        overall = round(sum(float(v) for v in avgs) / 5, 2)

    chart_data = {
        'labels': json.dumps(['Ponctualité', 'Maîtrise Communication', 'Pédagogie Méthodologie', 'Utilisation TIC', 'Disponibilité']),
        'data': json.dumps([stats.get('avg_ponctualite') or 0, stats.get('avg_maitrise_communication') or 0, stats.get('avg_pedagogie_methodologie') or 0, stats.get('avg_utilisation_tic') or 0, stats.get('avg_disponibilite') or 0])
    }

    context = {
        'enseignant': enseignant,
        'annees': annees,
        'annee_filter': annee_filter,
        'cours_filter': cours_filter,
        'cours_list': cours_list,
        'evaluations': evaluations_qs,
        'stats': stats,
        'overall': overall,
        'chart_data': chart_data,
    }
    return render(request, 'enseignant/appreciations.html', context)

@login_required
def admin_simulate_enseignant(request, matricule):
    """Admin : simuler la connexion d'un enseignant"""
    if not request.user.is_staff:
        return redirect('home')
    clear_simulation(request)
    request.session['simulated_enseignant'] = matricule
    return redirect('enseignant_dashboard')

@login_required
def admin_simulate_jury(request, code_jury):
    """Admin : simuler la connexion d'un jury"""
    if not request.user.is_staff:
        return redirect('home')
    clear_simulation(request)
    request.session['simulated_jury'] = code_jury
    return redirect('jury_dashboard')

@login_required
def admin_stop_simulation(request):
    """Admin : arrêter la simulation et revenir à l'accueil"""
    clear_simulation(request)
    return redirect('home')


# Page d'accueil
def home(request):
    """Page d'accueil du système"""
    if request.user.is_authenticated:
        if request.user.role == 'ETUDIANT':
            return redirect('etudiant_dashboard')
        elif request.user.role == 'ENSEIGNANT':
            return redirect('enseignant_dashboard')
        elif request.user.role == 'JURY':
            return redirect('jury_dashboard')
        elif request.user.is_staff or request.user.role == 'GESTIONNAIRE':
            # Statistiques pour l'administrateur et gestionnaire
            from django.db.models import Sum, Count
            
            stats = {
                'total_etudiants': Etudiant.objects.count(),
                'total_enseignants': Enseignant.objects.count(),
                'total_ues': UE.objects.count(),
                'total_ecs': EC.objects.count(),
                'total_credits_ue': UE.objects.aggregate(total=Sum('credit'))['total'] or 0,
                'total_credits_ec': EC.objects.aggregate(total=Sum('credit'))['total'] or 0,
                'total_inscriptions': Inscription.objects.count(),
                'total_attributions': Attribution.objects.count(),
                'is_gestionnaire': request.user.role == 'GESTIONNAIRE'
            }
            
            return render(request, 'home.html', stats)
    
    return render(request, 'home.html')


# Authentification
def login_view(request):
    """Vue de connexion"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, f'Bienvenue {user.username}!')
            
            # Redirection selon le rôle
            if user.role == 'ETUDIANT':
                return redirect('etudiant_dashboard')
            elif user.role == 'ENSEIGNANT':
                return redirect('enseignant_dashboard')
            elif user.role == 'JURY':
                return redirect('jury_dashboard')
            else:
                return redirect('home')
        else:
            messages.error(request, 'Nom d\'utilisateur ou mot de passe incorrect.')
    
    return render(request, 'login.html')


def logout_view(request):
    """Vue de déconnexion"""
    logout(request)
    messages.info(request, 'Vous avez été déconnecté.')
    return redirect('login')


@login_required
def change_password(request):
    if is_admin_viewing(request):
        messages.error(request, "Impossible de changer le mot de passe en mode visualisation Admin.")
        return redirect('home')

    template_name = 'change_password.html'
    extra_context = {}
    if request.user.role == 'ENSEIGNANT':
        template_name = 'enseignant/change_password.html'
        enseignant = get_simulated_enseignant(request)
        if enseignant:
            extra_context['enseignant'] = enseignant

    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Mot de passe modifié avec succès!')
            if request.user.role == 'ETUDIANT':
                return redirect('etudiant_dashboard')
            if request.user.role == 'ENSEIGNANT':
                return redirect('enseignant_dashboard')
            if request.user.role == 'JURY':
                return redirect('jury_dashboard')
            return redirect('home')
    else:
        form = PasswordChangeForm(user=request.user)

    context = {'form': form}
    context.update(extra_context)
    return render(request, template_name, context)


@login_required
def modifier_ma_photo(request):
    """Permet à l'utilisateur connecté de modifier sa photo de profil"""
    user = request.user
    profil = None
    photo_actuelle = None
    
    # Récupérer le profil selon le rôle
    if user.role == 'ETUDIANT':
        try:
            profil = Etudiant.objects.get(id_lgn=user)
            photo_actuelle = profil.photo
        except Etudiant.DoesNotExist:
            messages.error(request, 'Profil étudiant non trouvé.')
            return redirect('home')
    elif user.role == 'ENSEIGNANT':
        try:
            profil = Enseignant.objects.get(id_lgn=user)
            photo_actuelle = profil.photo
        except Enseignant.DoesNotExist:
            messages.error(request, 'Profil enseignant non trouvé.')
            return redirect('home')
    else:
        messages.error(request, 'Modification de photo non disponible pour ce type de compte.')
        return redirect('home')
    
    if request.method == 'POST':
        form = PhotoForm(request.POST, request.FILES)
        if form.is_valid() and request.FILES.get('photo'):
            profil.photo = request.FILES['photo']
            profil.save()
            messages.success(request, 'Photo mise à jour avec succès!')
            
            # Rediriger vers le dashboard approprié
            if user.role == 'ETUDIANT':
                return redirect('etudiant_dashboard')
            elif user.role == 'ENSEIGNANT':
                return redirect('enseignant_dashboard')
        else:
            messages.error(request, 'Veuillez sélectionner une photo valide.')
    else:
        form = PhotoForm()
    
    context = {
        'form': form,
        'photo_actuelle': photo_actuelle,
        'profil': profil,
    }
    return render(request, 'modifier_photo.html', context)


# ========== VUE COHORTE/DETTES JURY ==========

@login_required
def jury_cohorte(request):
    """Affiche la cohorte de la classe du jury connecté et liste ses étudiants.
    Pour les classes montantes (L2, L3, M1, M2), affiche aussi les résultats
    des classes antérieures de chaque étudiant de la cohorte."""
    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        logout(request)
        return redirect('login')
    
    classe = jury.code_classe
    
    # Utiliser l'année académique du jury connecté
    annee_code = jury.annee_academique
    if not annee_code:
        from reglage.models import AnneeAcademique
        annee_active = AnneeAcademique.objects.filter(active=True).first()
        annee_code = annee_active.code_anac if annee_active else None
    
    # Récupérer les inscriptions de la classe du jury
    inscriptions = Inscription.objects.filter(
        code_classe=classe,
        annee_academique=annee_code
    ).select_related('matricule_etudiant', 'cohorte')
    
    # Identifier la cohorte de la classe
    cohorte = None
    cohortes_distinctes = inscriptions.values_list('cohorte', flat=True).distinct()
    if cohortes_distinctes:
        cohorte = Cohorte.objects.filter(code_cohorte__in=[c for c in cohortes_distinctes if c]).first()
    
    # Récupérer tous les étudiants de cette cohorte, regroupés par classe
    etudiants_par_classe = {}
    palmares_par_classe = {}

    def _mention_for_moyenne(moyenne):
        if moyenne is None:
            return 'A déterminer'
        n = float(moyenne)
        if n >= 18:
            return 'Excellent (A)'
        if n >= 16:
            return 'Très bien (B)'
        if n >= 14:
            return 'Bien (C)'
        if n >= 12:
            return 'Assez Bien (D)'
        if n >= 10:
            return 'Passable (E)'
        if n >= 8:
            return 'Insuffisant (F)'
        return 'Insatisfaisant (G)'

    if cohorte:
        all_inscriptions_cohorte = Inscription.objects.filter(
            cohorte=cohorte,
            annee_academique=annee_code
        ).select_related('matricule_etudiant', 'code_classe').order_by('code_classe__code_classe', 'matricule_etudiant__nom_complet')
        
        for insc in all_inscriptions_cohorte:
            classe_label = insc.code_classe.code_classe if insc.code_classe else 'Sans classe'
            if classe_label not in etudiants_par_classe:
                etudiants_par_classe[classe_label] = []
            etudiants_par_classe[classe_label].append({
                'etudiant': insc.matricule_etudiant,
                'inscription': insc,
            })

        # Calcul palmarès (par classe) aligné avec jury_imprimable_palmare
        for classe_label, items in etudiants_par_classe.items():
            rows = []
            for item in items:
                etudiant = item.get('etudiant')
                insc = item.get('inscription')
                classe_obj = getattr(insc, 'code_classe', None)
                if not etudiant or not classe_obj:
                    continue

                delib_data = _jury_compute_delib_ues(classe_obj, etudiant, 'annuel', None, annee_code)
                if not delib_data.get('rows'):
                    continue

                # Dettes (cours non validés) pour cet étudiant (basé sur les données de profil/délibération)
                dettes_etudiant_map = {}
                for r in delib_data.get('rows', []):
                    if r.get('statut_code') == 'NON_VALIDE':
                        code = r.get('code_ec') or r.get('code_ue')
                        lib = r.get('intitule_ec') or r.get('intitule_ue') or ''
                        if code and code not in dettes_etudiant_map:
                            dettes_etudiant_map[code] = {
                                'code': code,
                                'code_ue': r.get('code_ue') or '',
                                'code_ec': r.get('code_ec') or '',
                                'intitule_ue': r.get('intitule_ue') or '',
                                'intitule_ec': r.get('intitule_ec') or '',
                                'libelle': lib,
                                'categorie': r.get('categorie', ''),
                                'semestre': r.get('semestre'),
                                'credit': r.get('credit', 0) or 0,
                            }
                dettes_etudiant = sorted(
                    dettes_etudiant_map.values(),
                    key=lambda x: (x.get('semestre') or 0, str(x.get('code') or ''))
                )

                moyenne = delib_data.get('moyenne', 0) or 0
                pourcentage = delib_data.get('pourcentage', 0) or 0
                credits_capitalises = delib_data.get('credits_valides', 0)
                credits_total = delib_data.get('credits_total', 0)
                decision_code = delib_data.get('decision_code')

                if decision_code == 'DEF':
                    decision = 'DEF'
                elif decision_code == 'ADM':
                    decision = 'ADM'
                elif decision_code in ['ADMD', 'COMP']:
                    decision = 'COMP'
                else:
                    decision = 'AJ'

                rows.append({
                    'etudiant': etudiant,
                    'moyenne': moyenne,
                    'pourcentage': pourcentage,
                    'credits_capitalises': credits_capitalises,
                    'credits_total': credits_total,
                    'decision': decision,
                    'mention': _mention_for_moyenne(moyenne),
                    'dettes': dettes_etudiant,
                })

            rows.sort(key=lambda x: float(x.get('moyenne') or 0), reverse=True)
            for idx, r in enumerate(rows, 1):
                r['rang'] = idx

            palmares_par_classe[classe_label] = rows
    
    # Trier les classes
    etudiants_par_classe = dict(sorted(etudiants_par_classe.items()))

    classes_data = []
    for classe_label, etudiants in etudiants_par_classe.items():
        classes_data.append({
            'classe_label': classe_label,
            'etudiants': etudiants,
            'palmares_rows': palmares_par_classe.get(classe_label, []),
        })
    
    total_etudiants = sum(len(v) for v in etudiants_par_classe.values())
    
    # ==================================================================
    # RÉSULTATS CLASSES ANTÉRIEURES (pour classes montantes L2,L3,M1,M2)
    # ==================================================================
    niveau_precedent_map = {
        'L2': 'L1', 'L3': 'L2', 'M1': None, 'M2': 'M1',
    }
    # Niveaux cumulés à remonter pour chaque niveau actuel
    niveaux_anterieurs_map = {
        'L2': ['L1'],
        'L3': ['L2', 'L1'],
        'M2': ['M1'],
    }
    
    niveau_actuel = str(classe.code_niveau_id) if classe.code_niveau_id else ''
    mention_id = str(classe.code_mention_id) if classe.code_mention_id else ''
    niveaux_anterieurs = niveaux_anterieurs_map.get(niveau_actuel, [])
    
    resultats_anterieurs = []  # Liste de dicts {classe_label, annee, palmares_rows}
    
    if niveaux_anterieurs and cohorte:
        # Collecter tous les étudiants de la cohorte actuelle
        etudiants_cohorte = set()
        for items in etudiants_par_classe.values():
            for item in items:
                etudiants_cohorte.add(item['etudiant'].matricule_et)
        
        # Pour chaque niveau antérieur, trouver les inscriptions passées
        for niv_ant in niveaux_anterieurs:
            # Trouver la classe antérieure (même mention + niveau antérieur)
            try:
                classe_ant = Classe.objects.get(
                    code_niveau_id=niv_ant,
                    code_mention_id=mention_id
                )
            except Classe.DoesNotExist:
                continue
            
            # Trouver les inscriptions de ces étudiants dans cette classe (toutes années)
            inscriptions_ant = Inscription.objects.filter(
                matricule_etudiant__matricule_et__in=etudiants_cohorte,
                code_classe=classe_ant,
            ).select_related('matricule_etudiant', 'code_classe').order_by('-annee_academique')
            
            # Regrouper par année académique
            annees_vues = {}
            for insc_ant in inscriptions_ant:
                annee_ant = insc_ant.annee_academique
                if annee_ant not in annees_vues:
                    annees_vues[annee_ant] = []
                # Éviter les doublons d'étudiant dans la même année
                matricules_deja = [x['etudiant'].matricule_et for x in annees_vues[annee_ant]]
                if insc_ant.matricule_etudiant.matricule_et not in matricules_deja:
                    annees_vues[annee_ant].append({
                        'etudiant': insc_ant.matricule_etudiant,
                        'inscription': insc_ant,
                        'classe_obj': insc_ant.code_classe,
                    })
            
            # Calculer le palmarès pour chaque année trouvée
            for annee_ant, items_ant in sorted(annees_vues.items(), reverse=True):
                rows_ant = []
                for item in items_ant:
                    etudiant = item['etudiant']
                    classe_obj_ant = item['classe_obj']
                    
                    delib_data = _jury_compute_delib_ues(classe_obj_ant, etudiant, 'annuel', None, annee_ant)
                    if not delib_data.get('rows'):
                        # Même sans délibérations, inclure l'étudiant
                        rows_ant.append({
                            'etudiant': etudiant,
                            'moyenne': 0,
                            'pourcentage': 0,
                            'credits_capitalises': 0,
                            'credits_total': 0,
                            'decision': '-',
                            'mention': '-',
                            'dettes': [],
                        })
                        continue
                    
                    # Dettes de cette classe antérieure
                    dettes_map_ant = {}
                    for r in delib_data.get('rows', []):
                        if r.get('statut_code') == 'NON_VALIDE':
                            code = r.get('code_ec') or r.get('code_ue')
                            if code and code not in dettes_map_ant:
                                dettes_map_ant[code] = {
                                    'code': code,
                                    'intitule_ue': r.get('intitule_ue') or '',
                                    'intitule_ec': r.get('intitule_ec') or '',
                                    'semestre': r.get('semestre'),
                                    'credit': r.get('credit', 0) or 0,
                                }
                    dettes_ant = sorted(
                        dettes_map_ant.values(),
                        key=lambda x: (x.get('semestre') or 0, str(x.get('code') or ''))
                    )
                    
                    moyenne = delib_data.get('moyenne', 0) or 0
                    dc = delib_data.get('decision_code')
                    if dc == 'DEF':
                        decision = 'DEF'
                    elif dc == 'ADM':
                        decision = 'ADM'
                    elif dc in ['ADMD', 'COMP']:
                        decision = 'COMP'
                    else:
                        decision = 'AJ'
                    
                    rows_ant.append({
                        'etudiant': etudiant,
                        'moyenne': moyenne,
                        'pourcentage': delib_data.get('pourcentage', 0) or 0,
                        'credits_capitalises': delib_data.get('credits_valides', 0),
                        'credits_total': delib_data.get('credits_total', 0),
                        'decision': decision,
                        'mention': _mention_for_moyenne(moyenne),
                        'dettes': dettes_ant,
                    })
                
                rows_ant.sort(key=lambda x: float(x.get('moyenne') or 0), reverse=True)
                for idx, r in enumerate(rows_ant, 1):
                    r['rang'] = idx
                
                resultats_anterieurs.append({
                    'classe_label': classe_ant.code_classe,
                    'annee': annee_ant,
                    'palmares_rows': rows_ant,
                    'nb_etudiants': len(rows_ant),
                })
    
    # ---- Suivi des dettes (InscriptionUE) ----
    from .models import InscriptionUE
    suivi_dettes = []
    if cohorte:
        # Récupérer tous les étudiants de la cohorte  
        etudiants_cohorte = [
            insc.matricule_etudiant
            for insc in Inscription.objects.filter(
                cohorte=cohorte, annee_academique=annee_code
            ).select_related('matricule_etudiant')
        ]
        inscriptions_ue = InscriptionUE.objects.filter(
            matricule_etudiant__in=etudiants_cohorte,
            type_inscription__in=['DETTE_COMPENSEE', 'DETTE_LIQUIDEE']
        ).select_related(
            'matricule_etudiant', 'code_ue', 'code_ec', 'code_classe'
        ).order_by('matricule_etudiant__nom_complet', 'annee_academique')
        
        for iu in inscriptions_ue:
            if iu.code_ec and iu.code_ec.code_ue:
                sem = iu.code_ec.code_ue.semestre
            elif iu.code_ue:
                sem = iu.code_ue.semestre
            else:
                sem = None
            suivi_dettes.append({
                'etudiant': iu.matricule_etudiant,
                'code': iu.code_ec.code_ec if iu.code_ec else (iu.code_ue.code_ue if iu.code_ue else '-'),
                'intitule': iu.code_ec.intitule_ue if iu.code_ec else (iu.code_ue.intitule_ue if iu.code_ue else '-'),
                'semestre': sem,
                'classe_origine': iu.code_classe.code_classe if iu.code_classe else '-',
                'annee': iu.annee_academique,
                'type': iu.type_inscription,
                'est_liquidee': iu.type_inscription == 'DETTE_LIQUIDEE',
            })
    
    context = {
        'jury': jury,
        'classe': classe,
        'annee': annee_code,
        'cohorte': cohorte,
        'etudiants_par_classe': etudiants_par_classe,
        'palmares_par_classe': palmares_par_classe,
        'classes_data': classes_data,
        'total_etudiants': total_etudiants,
        'resultats_anterieurs': resultats_anterieurs,
        'niveau_actuel': niveau_actuel,
        'est_classe_montante': niveau_actuel in ('L2', 'L3', 'M2'),
        'suivi_dettes': suivi_dettes,
        'nb_dettes_en_cours': sum(1 for d in suivi_dettes if not d['est_liquidee']),
        'nb_dettes_liquidees': sum(1 for d in suivi_dettes if d['est_liquidee']),
        'delib_annuelle_faite': Deliberation.objects.filter(
            code_classe=classe,
            annee_academique=annee_code,
            type_deliberation='ANNEE'
        ).exists(),
    }
    return render(request, 'jury/cohorte.html', context)


# ========== VUES RECOURS JURY ==========

@login_required
def jury_recours(request):
    """Liste des recours pour le jury"""
    jury = get_simulated_jury(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')
    
    # Récupérer les recours pour la classe du jury (filtrer par année)
    annee_code = jury.annee_academique
    ins_qs = Inscription.objects.filter(code_classe=jury.code_classe)
    if annee_code:
        ins_qs = ins_qs.filter(annee_academique=annee_code)
    etudiants_classe = [ins.matricule_etudiant for ins in ins_qs]
    
    recours = Recours.objects.filter(etudiant__in=etudiants_classe).select_related('etudiant').order_by('-date_envoi')
    
    # Résoudre les intitulés UE/EC pour chaque recours
    # Le champ ue_ec_concerne peut contenir: "UE001", "EC001", "UE001-EC001", "UE001-UE001", etc.
    all_codes = set()
    for rec in recours:
        raw = rec.ue_ec_concerne.strip()
        for part in raw.split('-'):
            p = part.strip()
            if p:
                all_codes.add(p)
    
    ue_dict = {ue.code_ue: ue.intitule_ue for ue in UE.objects.filter(code_ue__in=all_codes)}
    ec_dict = {ec.code_ec: ec.intitule_ue for ec in EC.objects.filter(code_ec__in=all_codes)}
    
    recours_enrichis = []
    for rec in recours:
        raw = rec.ue_ec_concerne.strip()
        parts = [p.strip() for p in raw.split('-') if p.strip()]
        # Dédupliquer (ex: UE001-UE001 → juste UE001)
        seen = []
        for p in parts:
            if p not in seen:
                seen.append(p)
        labels = []
        for p in seen:
            if p in ue_dict:
                labels.append(f"{p} - {ue_dict[p]}")
            elif p in ec_dict:
                labels.append(f"{p} - {ec_dict[p]}")
            else:
                labels.append(p)
        rec.intitule_ue_ec = ' / '.join(labels) if labels else raw
        recours_enrichis.append(rec)
    
    # Traitement POST pour sauvegarder traitement_jury et decision_finale
    if request.method == 'POST':
        code_recours = request.POST.get('code_recours')
        traitement = request.POST.get('traitement_jury', '')
        decision = request.POST.get('decision_finale', '')
        try:
            rec = Recours.objects.get(code_recours=code_recours)
            rec.traitement_jury = traitement
            rec.decision_finale = decision
            if decision:
                rec.statut = 'TRAITE'
                rec.traite_par = request.user
                rec.date_traitement = timezone.now()
            rec.save()
            messages.success(request, f'Recours {code_recours} mis à jour.')
        except Recours.DoesNotExist:
            messages.error(request, 'Recours non trouvé.')
        return redirect('jury_recours')
    
    context = {
        'jury': jury,
        'classe': jury.code_classe,
        'recours': recours_enrichis,
        'nb_recours_en_attente': recours.filter(statut='EN_ATTENTE').count(),
        'nb_recours_en_examen': recours.filter(statut='EN_EXAMEN').count(),
        'nb_recours_traites': recours.filter(statut='TRAITE').count(),
        'nb_recours_rejetes': recours.filter(statut='REJETE').count(),
    }
    return render(request, 'jury/recours.html', context)

@login_required
def jury_detail_recours(request, code_recours):
    """Détail d'un recours pour le jury"""
    jury = get_simulated_jury(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')
    
    recours = get_object_or_404(Recours, code_recours=code_recours)
    
    # Vérifier que le recours appartient bien à la classe du jury
    inscriptions = Inscription.objects.filter(code_classe=jury.code_classe)
    etudiants_classe = [ins.matricule_etudiant for ins in inscriptions]
    
    if recours.etudiant not in etudiants_classe:
        messages.error(request, 'Ce recours ne concerne pas votre classe.')
        return redirect('jury_recours')
    
    # Résoudre les intitulés UE/EC
    raw = recours.ue_ec_concerne.strip()
    all_codes = set()
    for part in raw.split('-'):
        p = part.strip()
        if p:
            all_codes.add(p)
    ue_dict = {ue.code_ue: ue.intitule_ue for ue in UE.objects.filter(code_ue__in=all_codes)}
    ec_dict = {ec.code_ec: ec.intitule_ue for ec in EC.objects.filter(code_ec__in=all_codes)}
    parts = [p.strip() for p in raw.split('-') if p.strip()]
    seen = []
    for p in parts:
        if p not in seen:
            seen.append(p)
    labels = []
    for p in seen:
        if p in ue_dict:
            labels.append(f"{p} - {ue_dict[p]}")
        elif p in ec_dict:
            labels.append(f"{p} - {ec_dict[p]}")
        else:
            labels.append(p)
    recours.intitule_ue_ec = ' / '.join(labels) if labels else raw
    
    context = {
        'jury': jury,
        'classe': jury.code_classe,
        'recours': recours,
        'fichiers': recours.fichiers.all(),
    }
    return render(request, 'jury/detail_recours.html', context)

@login_required
def jury_traiter_recours(request, code_recours):
    """Traiter un recours"""
    jury = get_simulated_jury(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')
    
    recours = get_object_or_404(Recours, code_recours=code_recours)
    
    # Vérifier que le recours appartient bien à la classe du jury
    inscriptions = Inscription.objects.filter(code_classe=jury.code_classe)
    etudiants_classe = [ins.matricule_etudiant for ins in inscriptions]
    
    if recours.etudiant not in etudiants_classe:
        messages.error(request, 'Ce recours ne concerne pas votre classe.')
        return redirect('jury_recours')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        commentaire = request.POST.get('commentaire', '')
        
        if action == 'accepter':
            recours.statut = 'TRAITE'
            recours.traite_par = request.user
            recours.date_traitement = timezone.now()
            recours.commentaire_traitement = commentaire
            recours.save()
            messages.success(request, f'Recours {code_recours} accepté et traité.')
        elif action == 'rejeter':
            recours.statut = 'REJETE'
            recours.traite_par = request.user
            recours.date_traitement = timezone.now()
            recours.commentaire_traitement = commentaire
            recours.save()
            messages.success(request, f'Recours {code_recours} rejeté.')
        elif action == 'examiner':
            recours.statut = 'EN_EXAMEN'
            recours.traite_par = request.user
            recours.date_traitement = timezone.now()
            recours.commentaire_traitement = commentaire
            recours.save()
            messages.success(request, f'Recours {code_recours} mis en examen.')
        
        return redirect('jury_detail_recours', code_recours=code_recours)
    
    return redirect('jury_detail_recours', code_recours=code_recours)


@login_required
def jury_modifier_recours(request, code_recours):
    """Modifier un recours"""
    jury = get_simulated_jury(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')

    recours = get_object_or_404(Recours, code_recours=code_recours)

    # Vérifier que le recours appartient à la classe du jury
    inscriptions = Inscription.objects.filter(code_classe=jury.code_classe)
    etudiants_classe = [ins.matricule_etudiant for ins in inscriptions]
    if recours.etudiant not in etudiants_classe:
        messages.error(request, 'Ce recours ne concerne pas votre classe.')
        return redirect('jury_recours')

    if request.method == 'POST':
        objet = request.POST.get('objet', recours.objet)
        ue_ec = request.POST.get('ue_ec_concerne', recours.ue_ec_concerne)
        description = request.POST.get('description', recours.description)
        statut = request.POST.get('statut', recours.statut)
        traitement = request.POST.get('traitement_jury', recours.traitement_jury)
        decision = request.POST.get('decision_finale', recours.decision_finale)

        recours.objet = objet
        recours.ue_ec_concerne = ue_ec
        recours.description = description
        recours.statut = statut
        recours.traitement_jury = traitement
        recours.decision_finale = decision
        recours.save()
        messages.success(request, f'Recours {code_recours} modifié avec succès.')
        return redirect('jury_recours')

    # Construire la liste des UE/EC avec intitulés pour la classe
    from core.models import UE as UEModel, EC as ECModel
    ues = UEModel.objects.filter(classe=jury.code_classe).order_by('semestre', 'code_ue')
    ecs = ECModel.objects.filter(classe=jury.code_classe).order_by('code_ue', 'code_ec')
    ue_ec_list = []
    for ue in ues:
        ue_ec_list.append({'code': ue.code_ue, 'label': f"{ue.code_ue} - {ue.intitule_ue}"})
    for ec in ecs:
        ue_ec_list.append({'code': ec.code_ec, 'label': f"{ec.code_ec} - {ec.intitule_ue}"})

    context = {
        'jury': jury,
        'classe': jury.code_classe,
        'recours': recours,
        'ue_ec_list': ue_ec_list,
    }
    return render(request, 'jury/modifier_recours.html', context)


@login_required
def jury_supprimer_recours(request, code_recours):
    """Supprimer un recours"""
    jury = get_simulated_jury(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')

    recours = get_object_or_404(Recours, code_recours=code_recours)

    # Vérifier que le recours appartient à la classe du jury
    inscriptions = Inscription.objects.filter(code_classe=jury.code_classe)
    etudiants_classe = [ins.matricule_etudiant for ins in inscriptions]
    if recours.etudiant not in etudiants_classe:
        messages.error(request, 'Ce recours ne concerne pas votre classe.')
        return redirect('jury_recours')

    if request.method == 'POST':
        recours.delete()
        messages.success(request, f'Recours {code_recours} supprimé avec succès.')
        return redirect('jury_recours')

    return redirect('jury_recours')


@login_required
def jury_recours_pdf(request):
    """Générer le PDF des recours pour le jury"""
    from io import BytesIO
    from django.conf import settings
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    import os
    from datetime import datetime

    jury = get_simulated_jury(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')

    # Récupérer les recours
    inscriptions = Inscription.objects.filter(code_classe=jury.code_classe)
    etudiants_classe = [ins.matricule_etudiant for ins in inscriptions]
    recours_qs = Recours.objects.filter(etudiant__in=etudiants_classe).select_related('etudiant').order_by('-date_envoi')

    # Résoudre intitulés UE/EC (format composite: UE_CODE-EC_CODE)
    all_codes_pdf = set()
    for rec in recours_qs:
        raw = rec.ue_ec_concerne.strip()
        for part in raw.split('-'):
            p = part.strip()
            if p:
                all_codes_pdf.add(p)
    ue_dict = {ue.code_ue: ue.intitule_ue for ue in UE.objects.filter(code_ue__in=all_codes_pdf)}
    ec_dict = {ec.code_ec: ec.intitule_ue for ec in EC.objects.filter(code_ec__in=all_codes_pdf)}

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=0.8*cm,
        bottomMargin=0.8*cm
    )

    elements = []
    styles = getSampleStyleSheet()

    normal_style = ParagraphStyle('NormalStyle', parent=styles['Normal'], fontSize=8, leading=10, fontName='Helvetica')
    bold_style = ParagraphStyle('BoldStyle', parent=styles['Normal'], fontSize=8, leading=10, fontName='Helvetica-Bold')
    center_style = ParagraphStyle('CenterStyle', parent=styles['Normal'], fontSize=9, leading=11, alignment=TA_CENTER, fontName='Helvetica')
    right_style = ParagraphStyle('RightStyle', parent=styles['Normal'], fontSize=9, leading=11, alignment=TA_RIGHT, fontName='Helvetica')
    title_style = ParagraphStyle('TitleStyle', parent=styles['Normal'], fontSize=14, leading=16, alignment=TA_CENTER, fontName='Helvetica-Bold')
    subtitle_style = ParagraphStyle('SubtitleStyle', parent=styles['Normal'], fontSize=11, leading=13, alignment=TA_CENTER, fontName='Helvetica')
    small_style = ParagraphStyle('SmallStyle', parent=styles['Normal'], fontSize=7, leading=9, fontName='Helvetica')
    small_center = ParagraphStyle('SmallCenter', parent=styles['Normal'], fontSize=7, leading=9, fontName='Helvetica', alignment=TA_CENTER)

    # En-tête avec image
    page_width = landscape(A4)[0] - 2*cm
    from lmdmanagersystem.middleware import get_entete_path
    entete_path = get_entete_path()
    if os.path.exists(entete_path):
        from PIL import Image as PILImage
        pil_img = PILImage.open(entete_path)
        img_width, img_height = pil_img.size
        ratio = img_height / img_width
        desired_width = page_width
        desired_height = desired_width * ratio
        img = RLImage(entete_path, width=desired_width, height=desired_height)
        elements.append(img)
        elements.append(Spacer(1, 0.3*cm))

    # Titre
    classe_str = str(jury.code_classe) if jury.code_classe else ''
    elements.append(Paragraph(f"BUREAU DU JURY ({classe_str})", title_style))
    elements.append(Spacer(1, 0.2*cm))
    elements.append(Paragraph("TABLEAU DES RECOURS", subtitle_style))
    elements.append(Spacer(1, 0.4*cm))

    # Style spécifique pour les en-têtes du tableau (texte blanc)
    header_style = ParagraphStyle('HeaderStyle', parent=styles['Normal'], fontSize=7, leading=9, fontName='Helvetica-Bold', alignment=TA_CENTER, textColor=colors.white)

    # Construire le tableau
    header = [
        Paragraph('<b>N°</b>', header_style),
        Paragraph('<b>Code</b>', header_style),
        Paragraph('<b>Étudiant</b>', header_style),
        Paragraph('<b>Objet</b>', header_style),
        Paragraph('<b>UE/EC concerné(e)</b>', header_style),
        Paragraph('<b>Description</b>', header_style),
        Paragraph('<b>Date</b>', header_style),
        Paragraph('<b>Statut</b>', header_style),
        Paragraph('<b>Traitement du Jury</b>', header_style),
        Paragraph('<b>Décision finale</b>', header_style),
    ]

    data = [header]
    for i, rec in enumerate(recours_qs, 1):
        raw = rec.ue_ec_concerne.strip()
        parts = [p.strip() for p in raw.split('-') if p.strip()]
        seen = []
        for p in parts:
            if p not in seen:
                seen.append(p)
        labels = []
        for p in seen:
            if p in ue_dict:
                labels.append(f"{p} - {ue_dict[p]}")
            elif p in ec_dict:
                labels.append(f"{p} - {ec_dict[p]}")
            else:
                labels.append(p)
        intitule = ' / '.join(labels) if labels else raw

        statut_map = {'EN_ATTENTE': 'En attente', 'EN_EXAMEN': 'En examen', 'TRAITE': 'Traité', 'REJETE': 'Rejeté'}
        decision_map = {'ACCEPTE': 'Accepté', 'REJETE': 'Rejeté', 'PARTIELLEMENT_ACCEPTE': 'Part. accepté', 'SANS_SUITE': 'Sans suite'}

        data.append([
            Paragraph(str(i), small_center),
            Paragraph(str(rec.code_recours), small_style),
            Paragraph(str(rec.etudiant.nom_complet), small_style),
            Paragraph(str(rec.get_objet_display()), small_style),
            Paragraph(intitule, small_style),
            Paragraph(str(rec.description or '')[:80], small_style),
            Paragraph(rec.date_envoi.strftime('%d/%m/%Y') if rec.date_envoi else '', small_center),
            Paragraph(statut_map.get(rec.statut, rec.statut), small_center),
            Paragraph(str(rec.traitement_jury or ''), small_style),
            Paragraph(decision_map.get(rec.decision_finale, rec.decision_finale or '—'), small_center),
        ])

    col_widths = [0.7*cm, 2*cm, 3.2*cm, 2.5*cm, 3.5*cm, 4*cm, 1.8*cm, 1.8*cm, 4*cm, 2.5*cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343a40')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.6*cm))

    # Date du jour
    date_str = datetime.now().strftime('%d/%m/%Y')
    elements.append(Paragraph(f"Fait à .................................................. le {date_str}", right_style))
    elements.append(Spacer(1, 0.6*cm))

    # Signatures des membres du jury
    from core.models import Enseignant as EnseignantModel
    try:
        jury_obj = Jury.objects.filter(code_classe=jury.code_classe, annee_academique=jury.annee_academique).first()
        membre_nom = '____________________'
        membre_grade = 'Grade: _______________'
        president_nom = '____________________'
        president_grade = 'Grade: _______________'
        secretaire_nom = '____________________'
        secretaire_grade = 'Grade: _______________'

        if jury_obj:
            if jury_obj.membre:
                try:
                    ens = EnseignantModel.objects.get(matricule_en=jury_obj.membre)
                    membre_nom = ens.nom_complet
                    membre_grade = ens.grade.designation_grade if ens.grade else 'Grade: _______________'
                except:
                    membre_nom = jury_obj.membre
            if jury_obj.president:
                try:
                    ens = EnseignantModel.objects.get(matricule_en=jury_obj.president)
                    president_nom = ens.nom_complet
                    president_grade = ens.grade.designation_grade if ens.grade else 'Grade: _______________'
                except:
                    president_nom = jury_obj.president
            if jury_obj.secretaire:
                try:
                    ens = EnseignantModel.objects.get(matricule_en=jury_obj.secretaire)
                    secretaire_nom = ens.nom_complet
                    secretaire_grade = ens.grade.designation_grade if ens.grade else 'Grade: _______________'
                except:
                    secretaire_nom = jury_obj.secretaire
    except Exception:
        pass

    underline_style = ParagraphStyle('UnderlineStyle', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, fontName='Helvetica')
    italic_style = ParagraphStyle('ItalicStyle', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, fontName='Helvetica-Oblique')

    sig_table = Table([
        [Paragraph('<b>Membre du jury</b>', center_style),
         Paragraph('<b>Président du jury</b>', center_style),
         Paragraph('<b>Secrétaire du jury</b>', center_style)],
        [Paragraph('<br/><br/>', center_style),
         Paragraph('<br/><br/>', center_style),
         Paragraph('<br/><br/>', center_style)],
        [Paragraph(f'<u>{membre_nom}</u>', underline_style),
         Paragraph(f'<u>{president_nom}</u>', underline_style),
         Paragraph(f'<u>{secretaire_nom}</u>', underline_style)],
        [Paragraph(f'<i>{membre_grade}</i>', italic_style),
         Paragraph(f'<i>{president_grade}</i>', italic_style),
         Paragraph(f'<i>{secretaire_grade}</i>', italic_style)],
    ], colWidths=[page_width/3]*3)

    sig_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0, colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
    ]))
    elements.append(sig_table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    classe_name = str(jury.code_classe).replace(' ', '_')
    response['Content-Disposition'] = f'inline; filename="Recours_{classe_name}.pdf"'
    return response


# ========== VUES ÉTUDIANT ==========

@login_required
def etudiant_dashboard(request):
    """Tableau de bord étudiant"""
    # Si admin sans simulation active, afficher la liste pour sélection
    if request.user.is_staff and 'simulated_etudiant' not in request.session:
        etudiants = Etudiant.objects.all()[:50]
        return render(request, 'admin/select_etudiant.html', {'etudiants': etudiants})
    
    # Récupérer l'étudiant (simulé ou connecté)
    etudiant = get_simulated_etudiant(request)
    if not etudiant:
        messages.error(request, 'Profil étudiant non trouvé.')
        return redirect('home')
    
    from reglage.models import AnneeAcademique

    annee_active = AnneeAcademique.get_annee_en_cours()
    annee_code = annee_active.code_anac if annee_active else None
    inscriptions = Inscription.objects.filter(matricule_etudiant=etudiant).select_related('code_classe', 'cohorte')

    inscription_active = None
    if annee_code:
        inscription_active = inscriptions.filter(annee_academique=annee_code).first()
    if not inscription_active:
        inscription_active = inscriptions.order_by('-annee_academique').first()

    classe_active = inscription_active.code_classe if inscription_active else None
    classe_code = classe_active.code_classe if classe_active else None

    evaluations = Evaluation.objects.filter(matricule_etudiant=etudiant)
    if classe_code:
        evaluations = evaluations.filter(
            Q(code_ue__classe__code_classe=classe_code) |
            Q(code_ec__classe__code_classe=classe_code)
        )

    def _mention_for_note(note):
        if note is None:
            return None
        n = float(note)
        if n >= 18:
            return 'Excellent (A)'
        if n >= 16:
            return 'Très bien (B)'
        if n >= 14:
            return 'Bien (C)'
        if n >= 12:
            return 'Assez Bien (D)'
        if n >= 10:
            return 'Passable (E)'
        if n >= 8:
            return 'Insuffisant (F)'
        return 'Insatisfaisant (G)'

    def _weighted_average(values_with_weights):
        total_w = 0
        total_v = 0
        for value, weight in values_with_weights:
            if value is None or weight is None:
                continue
            total_w += float(weight)
            total_v += float(value) * float(weight)
        if total_w <= 0:
            return None
        return round(total_v / total_w, 2)

    # Calcul des moyennes par semestre + annuelle (pondérées par crédits)
    niveau_to_semestres = {
        'L1': (1, 2),
        'L2': (3, 4),
        'L3': (5, 6),
        'M1': (7, 8),
        'M2': (9, 10),
    }
    niveau_code = None
    code_classe_str = str(classe_active.code_classe) if classe_active else ''
    for prefix in ['L1', 'L2', 'L3', 'M1', 'M2']:
        if code_classe_str.startswith(prefix):
            niveau_code = prefix
            break
    semestres_niveau = niveau_to_semestres.get(niveau_code)

    def _moyenne_for_semestres(semestres_cibles):
        if not semestres_cibles:
            return None
        evals = Evaluation.objects.filter(matricule_etudiant=etudiant).select_related('code_ue', 'code_ec', 'code_ec__code_ue')
        evals = evals.filter(
            Q(code_ue__semestre__in=semestres_cibles) |
            Q(code_ec__code_ue__semestre__in=semestres_cibles)
        )
        values = []
        for ev in evals:
            note = ev.calculer_note_finale()
            if note is None:
                continue
            credit = None
            if ev.code_ec_id and ev.code_ec and ev.code_ec.credit is not None:
                credit = ev.code_ec.credit
            elif ev.code_ue_id and ev.code_ue and ev.code_ue.credit is not None:
                credit = ev.code_ue.credit
            values.append((note, credit or 0))
        return _weighted_average(values)

    moyenne_s1 = None
    moyenne_s2 = None
    moyenne_annuelle = None
    mention_s1 = None
    mention_s2 = None
    mention_annuelle = None
    if semestres_niveau:
        s1, s2 = semestres_niveau
        moyenne_s1 = _moyenne_for_semestres([s1])
        moyenne_s2 = _moyenne_for_semestres([s2])
        moyenne_annuelle = _moyenne_for_semestres([s1, s2])
        mention_s1 = _mention_for_note(moyenne_s1)
        mention_s2 = _mention_for_note(moyenne_s2)
        mention_annuelle = _mention_for_note(moyenne_annuelle)

    # Fallback: moyenne simple (non pondérée) si semestres indéterminés
    notes_finales = [n for n in (ev.calculer_note_finale() for ev in evaluations) if n is not None]
    moyenne_generale = round(sum(notes_finales) / len(notes_finales), 2) if notes_finales else 0

    # Cours de l'année (UE + EC) pour la classe active
    cours_list = []
    if classe_active:
        ue_qs = UE.objects.filter(classe=classe_active)
        ec_qs = EC.objects.filter(classe=classe_active)
        for ue in ue_qs:
            cours_list.append({
                'code': ue.code_ue,
                'intitule': ue.intitule_ue,
                'type': 'UE',
                'credit': ue.credit,
                'semestre': ue.semestre,
                'categorie': ue.get_categorie_display(),
            })
        for ec in ec_qs:
            semestre = ec.code_ue.semestre if ec.code_ue else None
            cours_list.append({
                'code': ec.code_ec,
                'intitule': ec.intitule_ue,
                'type': 'EC',
                'credit': ec.credit,
                'semestre': semestre,
                'categorie': ec.categorie,
            })
        cours_list.sort(key=lambda c: ((c['semestre'] is None), c['semestre'] or 0, c['code']))

    # Nombre de cours réellement attribués aux enseignants (distinct par code_cours)
    cours_attribues_count = 0
    if cours_list:
        cours_codes = [c['code'] for c in cours_list]
        attributions_qs = Attribution.objects.filter(code_cours__in=cours_codes)
        if inscription_active and inscription_active.annee_academique:
            attributions_qs = attributions_qs.filter(annee_academique=inscription_active.annee_academique)
        cours_attribues_codes = set(attributions_qs.values_list('code_cours', flat=True))
        cours_attribues_count = len(cours_attribues_codes)

        # Utiliser la même règle pour l'aperçu des cours: ne montrer que les cours attribués
        cours_list = [c for c in cours_list if c['code'] in cours_attribues_codes]

    # Filtrer par semestre si sélectionné
    semestre_filter = request.GET.get('semestre')
    if semestre_filter:
        cours_list = [c for c in cours_list if str(c['semestre']) == semestre_filter]

    # Semestres disponibles pour le filtre
    semestres = sorted(set(c['semestre'] for c in cours_list if c['semestre'] is not None))

    # Communiqué de délibération (dernier) pour la classe/année
    communique = None
    if classe_active and inscription_active:
        communique = CommuniqueDeliberation.objects.filter(
            code_classe=classe_active,
            annee_academique=inscription_active.annee_academique,
        ).order_by('-date_deliberation', '-date_creation').first()

    # Résultats publiés ? Si aucun jury n'existe pour cette classe, afficher normalement
    jury_publie = True
    if classe_active:
        jury_obj = Jury.objects.filter(code_classe=classe_active, annee_academique=annee_code).first()
        if jury_obj:
            jury_publie = jury_obj.resultat_publie

    # Vérifier par semestre si la délibération a été effectuée
    semestres_deliberes = {}
    if inscription_active and annee_code and semestres_niveau:
        sem1, sem2 = semestres_niveau
        delib_s1 = Deliberation.objects.filter(
            code_classe=classe_active,
            annee_academique=annee_code,
            type_deliberation='S1',
        ).exists()
        delib_s2 = Deliberation.objects.filter(
            code_classe=classe_active,
            annee_academique=annee_code,
            type_deliberation='S2',
        ).exists()
        semestres_deliberes = {sem1: delib_s1, sem2: delib_s2}

    nb_commentaires = 0
    if inscription_active:
        nb_commentaires = CommentaireCours.objects.filter(
            etudiant=etudiant,
            annee_academique=inscription_active.annee_academique,
        ).count()

    # Crédits capitalisés par semestre
    credits_par_semestre = {}
    if classe_active and jury_publie:
        toutes_ue = UE.objects.filter(classe=classe_active)
        # Récupérer les évaluations indexées par code_ue
        evals_etudiant = {}
        for ev in evaluations:
            if ev.code_ue_id:
                evals_etudiant[ev.code_ue_id] = ev
            if ev.code_ec_id and ev.code_ec and ev.code_ec.code_ue_id:
                evals_etudiant.setdefault(ev.code_ec.code_ue_id, ev)
        for ue in toutes_ue:
            semestre = ue.semestre
            if semestre not in credits_par_semestre:
                credits_par_semestre[semestre] = {
                    'total_credits': 0, 'credits_capitalises': 0,
                    'ues_validees': [], 'credits_restants': 0, 'pourcentage': 0,
                }
            credits_par_semestre[semestre]['total_credits'] += (ue.credit or 0)
            # Vérifier si l'UE est capitalisée (moyenne >= 10)
            ev = evals_etudiant.get(ue.code_ue)
            if ev:
                note = ev.calculer_note_finale()
                if note is not None and note >= 10:
                    credits_par_semestre[semestre]['credits_capitalises'] += (ue.credit or 0)
                    credits_par_semestre[semestre]['ues_validees'].append(ue)
        for data in credits_par_semestre.values():
            data['credits_restants'] = data['total_credits'] - data['credits_capitalises']
            if data['total_credits'] > 0:
                data['pourcentage'] = round((data['credits_capitalises'] / data['total_credits']) * 100, 1)

    # Nombre de recours de l'étudiant
    nb_recours = Recours.objects.filter(etudiant=etudiant).count()
    nb_recours_en_attente = Recours.objects.filter(etudiant=etudiant, statut__in=['EN_ATTENTE', 'EN_EXAMEN']).count()

    # Nombre d'évaluations enseignants effectuées
    nb_eval_enseignants = EvaluationEnseignement.objects.filter(etudiant=etudiant).count()

    context = {
        'etudiant': etudiant,
        'annee_active': annee_active,
        'inscription_active': inscription_active,
        'classe_active': classe_active,
        'cours_list': cours_list,
        'cours_count': cours_attribues_count,
        'nb_commentaires': nb_commentaires,
        'semestres_niveau': semestres_niveau,
        'moyenne_s1': moyenne_s1,
        'moyenne_s2': moyenne_s2,
        'moyenne_annuelle': moyenne_annuelle,
        'mention_s1': mention_s1,
        'mention_s2': mention_s2,
        'mention_annuelle': mention_annuelle,
        'is_admin_view': is_admin_viewing(request),
        'semestres': semestres,
        'semestre_filter': semestre_filter,
        'jury_publie': jury_publie,
        'semestres_deliberes': semestres_deliberes,
        'moyenne_generale': moyenne_generale,
        'communique': communique,
        'credits_par_semestre': credits_par_semestre,
        'nb_recours': nb_recours,
        'nb_recours_en_attente': nb_recours_en_attente,
        'nb_eval_enseignants': nb_eval_enseignants,
    }
    return render(request, 'etudiant/dashboard.html', context)


@login_required
def etudiant_profil(request):
    """Profil de l'étudiant"""
    etudiant = get_simulated_etudiant(request)
    if not etudiant:
        messages.error(request, 'Profil étudiant non trouvé.')
        return redirect('home')

    from reglage.models import AnneeAcademique

    annee_active = AnneeAcademique.get_annee_en_cours()
    annee_code = annee_active.code_anac if annee_active else None

    inscriptions = Inscription.objects.filter(matricule_etudiant=etudiant).select_related('code_classe', 'cohorte')
    inscription_active = None
    if annee_code:
        inscription_active = inscriptions.filter(annee_academique=annee_code).first()
    if not inscription_active:
        inscription_active = inscriptions.order_by('-annee_academique').first()

    context = {
        'etudiant': etudiant,
        'annee_active': annee_active,
        'inscription_active': inscription_active,
        'is_simulation': bool(request.user.is_staff and request.session.get('simulated_etudiant')),
    }
    return render(request, 'etudiant/profil.html', context)


@login_required
def etudiant_notes(request):
    """Consultation des notes par l'étudiant"""
    etudiant = get_simulated_etudiant(request)
    if not etudiant:
        messages.error(request, 'Profil étudiant non trouvé.')
        return redirect('home')
    try:
        from reglage.models import AnneeAcademique
        annee_active = AnneeAcademique.get_annee_en_cours()
        annee_code = request.GET.get('annee') or (annee_active.code_anac if annee_active else None)
        ue_filtre = request.GET.get('ue')

        inscriptions = Inscription.objects.filter(matricule_etudiant=etudiant).select_related('code_classe', 'cohorte')
        inscription_active = None
        if annee_code:
            inscription_active = inscriptions.filter(annee_academique=annee_code).first()
        if not inscription_active:
            inscription_active = inscriptions.order_by('-annee_academique').first()

        classe_code = inscription_active.code_classe.code_classe if inscription_active else None
        classe_active = inscription_active.code_classe if inscription_active else None

        # Vérifier si le jury a publié les résultats. Si aucun jury n'existe, afficher normalement
        jury_publie = True
        if classe_active:
            jury_obj = Jury.objects.filter(code_classe=classe_active, annee_academique=annee_code).first()
            if jury_obj:
                jury_publie = jury_obj.resultat_publie

        # Vérifier par semestre si la délibération a été effectuée
        # semestres_deliberes = {numéro_semestre: True/False}
        semestres_deliberes = {}
        if inscription_active and annee_code:
            classe_obj = inscription_active.code_classe
            # Déterminer les semestres du niveau (ex: L1 → S1,S2 ; L2 → S3,S4)
            niveau_code = None
            if getattr(classe_obj, 'code_niveau', None):
                niveau_code = classe_obj.code_niveau.code_niveau
            else:
                code_cl = str(classe_obj.code_classe)
                for prefix in ['L1', 'L2', 'L3', 'M1', 'M2']:
                    if code_cl.startswith(prefix):
                        niveau_code = prefix
                        break
            niveau_to_semestres = {
                'L1': (1, 2), 'L2': (3, 4), 'L3': (5, 6),
                'M1': (7, 8), 'M2': (9, 10),
            }
            semestres_niveau = niveau_to_semestres.get(niveau_code, (1, 2))
            sem1, sem2 = semestres_niveau
            delib_s1 = Deliberation.objects.filter(
                code_classe=classe_obj,
                annee_academique=annee_code,
                type_deliberation='S1',
            ).exists()
            delib_s2 = Deliberation.objects.filter(
                code_classe=classe_obj,
                annee_academique=annee_code,
                type_deliberation='S2',
            ).exists()
            semestres_deliberes = {sem1: delib_s1, sem2: delib_s2}

        # Récupérer les évaluations de l'étudiant
        evaluations = Evaluation.objects.filter(matricule_etudiant=etudiant).select_related('code_ue', 'code_ec', 'code_ec__code_ue')
        if classe_code:
            evaluations = evaluations.filter(
                Q(code_ue__classe__code_classe=classe_code) |
                Q(code_ec__classe__code_classe=classe_code)
            )
        
        # Créer des dictionnaires d'évaluations par code
        eval_by_ec = {}
        eval_by_ue = {}
        for ev in evaluations:
            if ev.code_ec_id:
                eval_by_ec[ev.code_ec_id] = ev
            if ev.code_ue_id:
                eval_by_ue.setdefault(ev.code_ue_id, ev)
        
        # Récupérer les UE de la classe (approche identique au jury)
        ues_classe = []
        if classe_code:
            ues_classe = list(UE.objects.filter(classe__code_classe=classe_code).order_by('semestre', 'code_ue'))
        
        # Récupérer les EC liés aux UE via code_ue (relation correcte)
        ecs_classe = list(EC.objects.filter(code_ue__in=ues_classe).select_related('code_ue')) if ues_classe else []
        ec_by_ue = {}
        for ec in ecs_classe:
            ec_by_ue.setdefault(ec.code_ue_id, []).append(ec)
        
        # Filtrer les UE pour le filtre dropdown
        toutes_les_ue = [ue for ue in ues_classe if ue.code_ue in eval_by_ue or ue.code_ue in ec_by_ue]
        
        # Récupérer les statuts de délibération (source officielle, avec compensation)
        delib_statut_by_ec = {}
        delib_statut_by_ue = {}
        if classe_active and annee_code:
            deliberations = Deliberation.objects.filter(
                matricule_etudiant=etudiant,
                code_classe=classe_active,
                annee_academique=annee_code,
            ).select_related('code_ec', 'code_ue')
            for d in deliberations:
                if d.code_ec_id:
                    delib_statut_by_ec[d.code_ec_id] = d.statut
                elif d.code_ue_id:
                    delib_statut_by_ue[d.code_ue_id] = d.statut

        # Appliquer la compensation pour calculer les statuts finaux
        # On utilise _jury_compute_delib_ues qui gère déjà la compensation
        delib_statut_final_by_code = {}
        if classe_active and annee_code and semestres_deliberes:
            for sem_num in (sem1, sem2):
                delib_data = _jury_compute_delib_ues(classe_active, etudiant, 'semestriel', sem_num, annee_code)
                if delib_data:
                    for row in delib_data.get('rows', []):
                        code = row.get('code_ec', '')
                        delib_statut_final_by_code[code] = {
                            'statut_code': row.get('statut_code', ''),
                            'statut': row.get('statut', ''),
                            'compense': row.get('compense', False),
                        }

        # Organiser les évaluations par UE
        evaluations_par_ue = {}
        moyennes_ue = {}
        moyennes_ue_list = []
        
        for ue in ues_classe:
            # Appliquer le filtre UE si sélectionné
            if ue_filtre and ue.code_ue != ue_filtre:
                continue
            
            # Récupérer les EC de cette UE via la relation code_ue
            ecs = ec_by_ue.get(ue.code_ue, [])
            
            if ecs:
                # L'UE a des EC, récupérer les évaluations des EC
                evals_list = []
                total_weighted = 0
                total_credits = 0
                for ec in ecs:
                    ev = eval_by_ec.get(ec.code_ec)
                    if ev:
                        note_finale = ev.calculer_note_finale()
                        # Statut depuis la délibération (avec compensation)
                        delib_info = delib_statut_final_by_code.get(ec.code_ec, {})
                        evals_list.append({
                            'evaluation': ev,
                            'note_finale': note_finale,
                            'ec': ec,
                            'delib_statut_code': delib_info.get('statut_code', ''),
                            'delib_statut': delib_info.get('statut', ''),
                            'delib_compense': delib_info.get('compense', False),
                        })
                        if note_finale is not None:
                            total_weighted += note_finale * ec.credit
                            total_credits += ec.credit
                
                if evals_list:
                    evaluations_par_ue[ue] = evals_list
                    if total_credits > 0:
                        moyenne_ue = round(total_weighted / total_credits, 2)
                        moyennes_ue[ue.code_ue] = moyenne_ue
                        moyennes_ue_list.append({'ue': ue, 'moyenne': moyenne_ue})
            else:
                # L'UE n'a pas d'EC, utiliser l'évaluation directe de l'UE
                ev_ue = eval_by_ue.get(ue.code_ue)
                if ev_ue:
                    note_finale = ev_ue.calculer_note_finale()
                    delib_info = delib_statut_final_by_code.get(ue.code_ue, {})
                    evaluations_par_ue[ue] = [{
                        'evaluation': ev_ue,
                        'note_finale': note_finale,
                        'ec': None,
                        'delib_statut_code': delib_info.get('statut_code', ''),
                        'delib_statut': delib_info.get('statut', ''),
                        'delib_compense': delib_info.get('compense', False),
                    }]
                    if note_finale is not None:
                        moyennes_ue[ue.code_ue] = note_finale
                        moyennes_ue_list.append({'ue': ue, 'moyenne': note_finale})
        
        # Calculer les crédits capitalisés par semestre via _jury_compute_delib_ues
        # (même source que la page Résultats : table Deliberation + compensation)
        credits_par_semestre = {}
        if inscription_active and annee_code and classe_active and semestres_deliberes:
            for sem_num in (sem1, sem2):
                delib_data = _jury_compute_delib_ues(classe_active, etudiant, 'semestriel', sem_num, annee_code)
                if delib_data and delib_data.get('rows'):
                    credits_par_semestre[sem_num] = {
                        'total_credits': delib_data.get('credits_total', 0),
                        'credits_capitalises': delib_data.get('credits_valides', 0),
                    }
        
        context = {
            'etudiant': etudiant,
            'annee_active': annee_active,
            'annee_code': annee_code,
            'inscriptions': inscriptions,
            'evaluations_par_ue': evaluations_par_ue,
            'moyennes_ue': moyennes_ue,
            'moyennes_ue_list': moyennes_ue_list,
            'toutes_les_ue': toutes_les_ue,
            'ue_filtre': ue_filtre,
            'credits_par_semestre': credits_par_semestre,
            'semestres_deliberes': semestres_deliberes,
            'jury_publie': jury_publie,
        }
        return render(request, 'etudiant/notes.html', context)
    except Etudiant.DoesNotExist:
        messages.error(request, 'Profil étudiant non trouvé.')
        return redirect('home')


@login_required
def jury_dashboard_grille_pdf(request):
    """Imprime en PDF la grille du dashboard jury en 2 pages (S1 puis S2)"""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    import os
    from django.conf import settings
    from PIL import Image as PILImage
    from reportlab.platypus import Image as RLImage

    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')

    classe = jury.code_classe
    if not classe:
        messages.error(request, 'Classe introuvable pour ce jury.')
        return redirect('home')

    # Année académique du jury (éviter doublons)
    annee_code = jury.annee_academique
    if not annee_code:
        from reglage.models import AnneeAcademique as AA
        aa = AA.get_annee_en_cours()
        annee_code = aa.code_anac if aa else None

    # Reprendre la même logique de cours que jury_dashboard
    ins_qs = Inscription.objects.filter(code_classe=classe).select_related('matricule_etudiant', 'cohorte')
    if annee_code:
        ins_qs = ins_qs.filter(annee_academique=annee_code)
    inscriptions = ins_qs

    ue_codes = list(UE.objects.filter(classe__code_classe=classe.code_classe).values_list('code_ue', flat=True))
    ec_codes = list(EC.objects.filter(classe__code_classe=classe.code_classe).values_list('code_ec', flat=True))

    ue_parents_avec_ec = set()
    for ec_code in ec_codes:
        ec = EC.objects.filter(code_ec=ec_code).select_related('code_ue').first()
        if ec and ec.code_ue:
            ue_parents_avec_ec.add(ec.code_ue.code_ue)

    ue_codes_filtrees = [ue_code for ue_code in ue_codes if ue_code not in ue_parents_avec_ec]
    cours_codes = ue_codes_filtrees + ec_codes
    attributions = Attribution.objects.filter(code_cours__in=cours_codes)

    cours_map = {}
    for attr in attributions:
        cours_obj = attr.get_cours_object()
        cours_type = attr.get_type_cours()
        if not (cours_obj and cours_type):
            continue
        semestre = None
        if cours_type == 'UE':
            semestre = getattr(cours_obj, 'semestre', None)
        elif cours_type == 'EC':
            semestre = getattr(getattr(cours_obj, 'code_ue', None), 'semestre', None)
        cours_map[attr.code_cours] = {
            'code': attr.code_cours,
            'intitule': getattr(cours_obj, 'intitule_ue', attr.code_cours),
            'type': cours_type,
            'semestre': semestre,
        }

    cours_list = list(cours_map.values())
    cours_list.sort(key=lambda c: ((c['semestre'] is None), c['semestre'] or 0, c['code']))

    # Récupérer les évaluations pour chaque étudiant et chaque cours
    etudiants_data = []
    for inscription in inscriptions:
        etudiant = inscription.matricule_etudiant
        notes = {}
        for cours in cours_list:
            eval_filter = {'matricule_etudiant': etudiant}
            if annee_code:
                eval_filter['annee_academique'] = annee_code
                eval_filter['code_classe'] = classe
            if cours['type'] == 'EC':
                eval_filter['code_ec__code_ec'] = cours['code']
            else:
                eval_filter['code_ue__code_ue'] = cours['code']
            eval_obj = Evaluation.objects.filter(**eval_filter).first()

            notes[cours['code']] = _format_evaluation_for_display(eval_obj) if eval_obj else None

        etudiants_data.append({
            'etudiant': etudiant,
            'notes': notes,
        })

    def _cours_for_semestre(target_semestre):
        return [c for c in cours_list if c.get('semestre') == target_semestre]

    from reportlab.platypus import Flowable
    from reportlab.pdfbase.pdfmetrics import stringWidth

    class VerticalText(Flowable):
        """Flowable qui dessine du texte pivoté à 90° (bas vers haut)."""
        def __init__(self, text, font_name='Helvetica-Bold', font_size=11, cell_w=None, cell_h=None):
            Flowable.__init__(self)
            self.text = text
            self.font_name = font_name
            self.font_size = font_size
            self._cell_w = cell_w
            self._cell_h = cell_h
            self.width = cell_w or 0
            self.height = cell_h or 0

        def draw(self):
            canvas = self.canv
            canvas.saveState()
            canvas.setFont(self.font_name, self.font_size)
            # Partir du bas de la cellule, texte monte vers le haut
            canvas.translate(self.width / 2, 0.15*cm)
            canvas.rotate(90)
            canvas.drawString(0, -self.font_size / 3, self.text)
            canvas.restoreState()

    def _build_table_for_semestre(target_semestre):
        cours_sem = _cours_for_semestre(target_semestre)
        col_w_note = 1.0*cm
        header_row_h = 6.0*cm  # hauteur suffisante pour lire les intitulés en entier

        def _abreger(texte, max_car=30):
            return texte if len(texte) <= max_car else texte[:max_car - 1] + '…'

        # Ligne 1 : intitulés verticaux (fusionnés sur 2 lignes pour la col Nom)
        header_1 = ['NOMS']  # Nom Complet — sera fusionné sur 2 lignes
        for c in cours_sem:
            header_1.append(VerticalText(_abreger(c.get('intitule', '')), cell_w=col_w_note, cell_h=header_row_h))

        # Ligne 2 : "" + "NF" pour chaque cours
        header_2 = ['']
        for _ in cours_sem:
            header_2.append('NF')

        data = [header_1, header_2]
        for row in etudiants_data:
            etu = row['etudiant']
            line = [etu.nom_complet or '']
            for c in cours_sem:
                n = row['notes'].get(c['code'])
                line.append(n.get('note_finale', '-') if n else '-')
            data.append(line)

        col_widths = [6.5*cm] + [col_w_note] * len(cours_sem)
        row_heights = [header_row_h, 0.5*cm] + [None] * len(etudiants_data)

        t = Table(data, colWidths=col_widths, rowHeights=row_heights, repeatRows=2)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.white),
            ('FONTNAME', (0, 0), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 2), (0, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, 1), 'MIDDLE'),
            ('VALIGN', (0, 2), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('SPAN', (0, 0), (0, 1)),  # Fusionner "Nom Complet" sur 2 lignes
            ('ROWBACKGROUNDS', (0, 2), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
        ]))
        return t

    # Déterminer les 2 semestres de la classe (si possible)
    semestres = list(
        UE.objects.filter(classe__code_classe=classe.code_classe)
        .values_list('semestre', flat=True)
        .distinct()
        .order_by('semestre')
    )
    semestres = [s for s in semestres if s is not None]
    sem1 = semestres[0] if len(semestres) >= 1 else None
    sem2 = semestres[1] if len(semestres) >= 2 else None

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=0.7*cm, rightMargin=0.7*cm, topMargin=0.7*cm, bottomMargin=0.7*cm)
    styles = getSampleStyleSheet()
    page_w = landscape(A4)[0] - 1.4*cm

    bureau_style = ParagraphStyle('bureau', fontSize=11, fontName='Helvetica-Bold', alignment=TA_CENTER, textColor=colors.HexColor('#2c3e50'), spaceAfter=2)
    title_style  = ParagraphStyle('title',  fontSize=16, fontName='Helvetica-Bold', alignment=TA_CENTER, textColor=colors.HexColor('#2c3e50'), spaceAfter=2)
    sub_style    = ParagraphStyle('sub',    fontSize=9,  fontName='Helvetica',      alignment=TA_CENTER, textColor=colors.HexColor('#555555'), spaceAfter=4)

    from lmdmanagersystem.middleware import get_entete_path
    entete_path = get_entete_path()

    def _build_header_elements(semestre_label):
        elems = []
        if os.path.exists(entete_path):
            pil_img = PILImage.open(entete_path)
            iw, ih = pil_img.size
            desired_w = page_w
            desired_h = desired_w * (ih / iw)
            elems.append(RLImage(entete_path, width=desired_w, height=desired_h))
            elems.append(Spacer(1, 0.3*cm))
        elems.append(Paragraph(f"BUREAU DU JURY {classe.code_classe}", bureau_style))
        elems.append(Paragraph(
            f"Classe : <b>{classe.code_classe} – {classe.designation_classe}</b>"
            f" &nbsp;&nbsp;|&nbsp;&nbsp; Semestre : <b>{semestre_label}</b>"
            f" &nbsp;&nbsp;|&nbsp;&nbsp; Année académique : <b>{annee_code}</b>",
            sub_style
        ))
        elems.append(Spacer(1, 0.3*cm))
        return elems

    from reglage.models import AnneeAcademique as _AA
    annee_active = _AA.get_annee_en_cours()
    annee_code = jury.annee_academique or (annee_active.code_anac if annee_active else '')

    elements = []

    if sem1 is not None:
        elements += _build_header_elements(sem1)
        elements.append(_build_table_for_semestre(sem1))
    if sem2 is not None:
        elements.append(PageBreak())
        elements += _build_header_elements(sem2)
        elements.append(_build_table_for_semestre(sem2))

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="grille_jury_{classe.code_classe}.pdf"'
    return response


@login_required
def jury_deliberer_resultats_pdf(request):
    """Imprime en PDF le tableau des résultats affichés dans Délibérer"""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')

    classe = jury.code_classe
    if not classe:
        messages.error(request, 'Classe non trouvée.')
        return redirect('jury_deliberer')

    from reglage.models import AnneeAcademique
    annee_active = AnneeAcademique.get_annee_en_cours()
    annee_code = jury.annee_academique or (annee_active.code_anac if annee_active else None)
    if not annee_code:
        annee_code = (
            Inscription.objects.filter(code_classe=classe)
            .values_list('annee_academique', flat=True)
            .order_by('-annee_academique')
            .first()
        )

    def _mention_for_note(note):
        if note is None:
            return None
        n = float(note)
        if n >= 18:
            return 'Excellent (A)'
        if n >= 16:
            return 'Très bien (B)'
        if n >= 14:
            return 'Bien (C)'
        if n >= 12:
            return 'Assez Bien (D)'
        if n >= 10:
            return 'Passable (E)'
        if n >= 8:
            return 'Insuffisant (F)'
        return 'Insatisfaisant (G)'

    inscriptions = Inscription.objects.filter(code_classe=classe, annee_academique=annee_code).select_related('matricule_etudiant')
    resultats = []
    for inscription in inscriptions:
        etudiant = inscription.matricule_etudiant
        stats_s1 = _jury_compute_delib_ues(classe, etudiant, 'semestriel', 1, annee_code)
        stats_s2 = _jury_compute_delib_ues(classe, etudiant, 'semestriel', 2, annee_code)
        stats_annuel = _jury_compute_delib_ues(classe, etudiant, 'annuel', None, annee_code)

        credits_s1 = stats_s1.get('credits_valides', 0)
        credits_s2 = stats_s2.get('credits_valides', 0)
        credits_annuel = stats_annuel.get('credits_valides', 0)
        moyenne = stats_annuel.get('moyenne', 0) or 0
        decision = stats_annuel.get('decision_label', 'A déterminer')
        mention = _mention_for_note(moyenne)

        resultats.append({
            'matricule': etudiant.matricule_et,
            'nom': etudiant.nom_complet,
            'credits_s1': credits_s1,
            'credits_s2': credits_s2,
            'credits_annuel': credits_annuel,
            'moyenne': round(float(moyenne), 2),
            'mention': mention or '',
            'decision': decision or '',
        })

    resultats.sort(key=lambda r: (r.get('nom') or '', r.get('matricule') or ''))

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()

    data = [[
        'Matricule', 'Nom Complet',
        'Crédits S1', 'Crédits S2', 'Crédits Annuel',
        'Moyenne', 'Mention', 'Décision'
    ]]
    for r in resultats:
        data.append([
            r.get('matricule', ''),
            r.get('nom', ''),
            str(r.get('credits_s1', 0)),
            str(r.get('credits_s2', 0)),
            str(r.get('credits_annuel', 0)),
            f"{r.get('moyenne', 0):.2f}",
            r.get('mention', ''),
            r.get('decision', ''),
        ])

    col_widths = [2.8*cm, 7.5*cm, 2.2*cm, 2.2*cm, 2.6*cm, 1.8*cm, 3.6*cm, 4.2*cm]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343a40')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (2, 1), (5, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))

    elements = [
        Paragraph(f"RESULTATS DE DELIBERATION - Classe {classe.code_classe} - {annee_code}", styles['Title']),
        Spacer(1, 0.3*cm),
        t,
    ]
    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="resultats_deliberation_{classe.code_classe}_{annee_code}.pdf"'
    return response


@login_required
def etudiant_mes_cours(request):
    etudiant = get_simulated_etudiant(request)
    if not etudiant:
        messages.error(request, 'Profil étudiant non trouvé.')
        return redirect('home')

    from reglage.models import AnneeAcademique
    annee_active = AnneeAcademique.get_annee_en_cours()
    annee_code = request.GET.get('annee') or (annee_active.code_anac if annee_active else None)

    # Récupérer les filtres
    semestre_filter = request.GET.get('semestre')

    inscriptions = Inscription.objects.filter(matricule_etudiant=etudiant).select_related('code_classe', 'cohorte')
    inscription_active = None
    if annee_code:
        inscription_active = inscriptions.filter(annee_academique=annee_code).first()
    if not inscription_active:
        inscription_active = inscriptions.order_by('-annee_academique').first()

    classe_active = inscription_active.code_classe if inscription_active else None
    classe_code = classe_active.code_classe if classe_active else None

    cours_list = []
    if classe_code:
        ue_qs = UE.objects.filter(classe__code_classe=classe_code)
        ec_qs = EC.objects.filter(classe__code_classe=classe_code)
        for ue in ue_qs:
            cours_list.append({
                'code': ue.code_ue,
                'intitule': ue.intitule_ue,
                'type': 'UE',
                'credit': ue.credit,
                'semestre': ue.semestre,
                'categorie': ue.get_categorie_display(),
            })
        for ec in ec_qs:
            semestre = ec.code_ue.semestre if ec.code_ue else None
            cours_list.append({
                'code': ec.code_ec,
                'intitule': ec.intitule_ue,
                'type': 'EC',
                'credit': ec.credit,
                'semestre': semestre,
                'categorie': ec.categorie,
            })
        cours_list.sort(key=lambda c: ((c['semestre'] is None), c['semestre'] or 0, c['code']))

    # Ajouter les enseignants pour chaque cours
    attributions = Attribution.objects.filter(
        code_cours__in=[c['code'] for c in cours_list],
        annee_academique=annee_code
    ).select_related('matricule_en')

    # Ne garder que les cours réellement attribués
    cours_attribues_codes = set(attributions.values_list('code_cours', flat=True))
    cours_list = [c for c in cours_list if c['code'] in cours_attribues_codes]
    
    attr_dict = {}
    for attr in attributions:
        if attr.code_cours not in attr_dict:
            attr_dict[attr.code_cours] = []
        attr_dict[attr.code_cours].append(attr.matricule_en.nom_complet)
    
    for c in cours_list:
        c['enseignants'] = attr_dict.get(c['code'], [])

    # Récupérer les documents disponibles pour chaque cours
    docs_qs = DocumentCours.objects.filter(
        code_cours__in=[c['code'] for c in cours_list],
        annee_academique=annee_code
    ).order_by('code_cours', '-date_ajout')
    docs_dict = {}
    for doc in docs_qs:
        if doc.code_cours not in docs_dict:
            docs_dict[doc.code_cours] = []
        docs_dict[doc.code_cours].append(doc)
    for c in cours_list:
        c['documents'] = docs_dict.get(c['code'], [])

    # Récupérer toutes les options pour les filtres (avant filtrage)
    tous_les_semestres = sorted(set(c['semestre'] for c in cours_list if c['semestre'] is not None))

    # Appliquer les filtres
    if semestre_filter:
        cours_list = [c for c in cours_list if str(c['semestre']) == semestre_filter]

    context = {
        'etudiant': etudiant,
        'annee_active': annee_active,
        'annee_code': annee_code,
        'inscription_active': inscription_active,
        'classe_active': classe_active,
        'cours_list': cours_list,
        'semestres': tous_les_semestres,
        'semestre_filter': semestre_filter,
        'inscriptions': inscriptions,
    }
    return render(request, 'etudiant/mes_cours.html', context)


@login_required
def etudiant_communique(request):
    etudiant = get_simulated_etudiant(request)
    if not etudiant:
        messages.error(request, 'Profil étudiant non trouvé.')
        return redirect('home')

    from reglage.models import AnneeAcademique
    annee_active = AnneeAcademique.get_annee_en_cours()
    annee_code = request.GET.get('annee') or (annee_active.code_anac if annee_active else None)

    inscriptions = Inscription.objects.filter(matricule_etudiant=etudiant).select_related('code_classe', 'cohorte')
    inscription_active = None
    if annee_code:
        inscription_active = inscriptions.filter(annee_academique=annee_code).first()
    if not inscription_active:
        inscription_active = inscriptions.order_by('-annee_academique').first()

    classe_active = inscription_active.code_classe if inscription_active else None

    communiques = CommuniqueDeliberation.objects.none()
    if inscription_active and classe_active:
        communiques = CommuniqueDeliberation.objects.filter(
            code_classe=classe_active,
            annee_academique=inscription_active.annee_academique,
        ).order_by('-date_deliberation', '-date_creation')

    context = {
        'etudiant': etudiant,
        'annee_active': annee_active,
        'annee_code': annee_code,
        'inscription_active': inscription_active,
        'classe_active': classe_active,
        'communiques': communiques,
    }
    return render(request, 'etudiant/communique.html', context)


@login_required
def etudiant_commentaires(request):
    etudiant = get_simulated_etudiant(request)
    if not etudiant:
        messages.error(request, 'Profil étudiant non trouvé.')
        return redirect('home')

    from reglage.models import AnneeAcademique
    annee_active = AnneeAcademique.get_annee_en_cours()
    annee_code = request.GET.get('annee') or (annee_active.code_anac if annee_active else None)

    inscriptions = Inscription.objects.filter(matricule_etudiant=etudiant).select_related('code_classe', 'cohorte')
    inscription_active = None
    if annee_code:
        inscription_active = inscriptions.filter(annee_academique=annee_code).first()
    if not inscription_active:
        inscription_active = inscriptions.order_by('-annee_academique').first()

    classe_active = inscription_active.code_classe if inscription_active else None
    classe_code = classe_active.code_classe if classe_active else None

    cours_list = []
    if classe_code:
        ue_qs = UE.objects.filter(classe__code_classe=classe_code)
        ec_qs = EC.objects.filter(classe__code_classe=classe_code)
        for ue in ue_qs:
            cours_list.append({'code': ue.code_ue, 'intitule': ue.intitule_ue, 'type': 'UE'})
        for ec in ec_qs:
            cours_list.append({'code': ec.code_ec, 'intitule': ec.intitule_ue, 'type': 'EC'})
        cours_list.sort(key=lambda c: (c['type'], c['code']))

    if request.method == 'POST':
        code_cours = request.POST.get('code_cours')
        type_cours = request.POST.get('type_cours')
        contenu = (request.POST.get('contenu') or '').strip()
        if not (code_cours and type_cours and contenu and inscription_active):
            messages.error(request, 'Veuillez remplir tous les champs.')
        else:
            CommentaireCours.objects.create(
                etudiant=etudiant,
                annee_academique=inscription_active.annee_academique,
                type_cours=type_cours,
                code_cours=code_cours,
                contenu=contenu,
            )
            messages.success(request, 'Commentaire envoyé avec succès!')
        return redirect('etudiant_commentaires')

    commentaires = CommentaireCours.objects.none()
    if inscription_active:
        commentaires = CommentaireCours.objects.filter(
            etudiant=etudiant,
            annee_academique=inscription_active.annee_academique,
        ).order_by('-date_creation')
    
    # Enrichir chaque commentaire avec l'intitulé du cours et l'enseignant
    for comm in commentaires:
        comm.intitule_cours = comm.code_cours
        comm.enseignant_nom = None
        try:
            ue = UE.objects.get(code_ue=comm.code_cours)
            comm.intitule_cours = ue.intitule_ue
        except UE.DoesNotExist:
            try:
                ec = EC.objects.get(code_ec=comm.code_cours)
                comm.intitule_cours = ec.intitule_ue
            except EC.DoesNotExist:
                pass
        attr = Attribution.objects.filter(
            code_cours=comm.code_cours,
            annee_academique=comm.annee_academique
        ).select_related('matricule_en').first()
        if attr and attr.matricule_en:
            comm.enseignant_nom = attr.matricule_en.nom_complet

    context = {
        'etudiant': etudiant,
        'annee_active': annee_active,
        'annee_code': annee_code,
        'inscription_active': inscription_active,
        'classe_active': classe_active,
        'cours_list': cours_list,
        'commentaires': commentaires,
    }
    return render(request, 'etudiant/commentaires.html', context)


@login_required
def etudiant_resultats(request):
    etudiant = get_simulated_etudiant(request)
    if not etudiant:
        messages.error(request, 'Profil étudiant non trouvé.')
        return redirect('home')

    from reglage.models import AnneeAcademique
    annee_active = AnneeAcademique.get_annee_en_cours()
    annee_code = request.GET.get('annee') or (annee_active.code_anac if annee_active else None)

    inscriptions = Inscription.objects.filter(matricule_etudiant=etudiant).select_related('code_classe', 'cohorte')
    inscription_active = None
    if annee_code:
        inscription_active = inscriptions.filter(annee_academique=annee_code).first()
    if not inscription_active:
        inscription_active = inscriptions.order_by('-annee_academique').first()

    classe_active = inscription_active.code_classe if inscription_active else None
    if not classe_active:
        messages.error(request, 'Aucune classe trouvée pour cet étudiant.')
        return redirect('etudiant_dashboard')

    jury_obj = Jury.objects.filter(code_classe=classe_active, annee_academique=annee_code).first()
    if jury_obj:
        jury_publie = jury_obj.resultat_publie
    else:
        jury_publie = True

    if not jury_publie:
        context = {
            'etudiant': etudiant,
            'annee_active': annee_active,
            'annee_code': annee_code,
            'inscription_active': inscription_active,
            'classe_active': classe_active,
            'jury_publie': False,
        }
        return render(request, 'etudiant/resultats.html', context)

    niveau_to_semestres = {
        'L1': (1, 2),
        'L2': (3, 4),
        'L3': (5, 6),
        'M1': (7, 8),
        'M2': (9, 10),
    }
    niveau_code = None
    code_classe = str(classe_active.code_classe)
    for prefix in ['L1', 'L2', 'L3', 'M1', 'M2']:
        if code_classe.startswith(prefix):
            niveau_code = prefix
            break
    semestres_niveau = niveau_to_semestres.get(niveau_code)

    # Calculer les résumés par semestre et annuel via _jury_compute_delib_ues
    def _extract_summary(delib_data):
        if not delib_data:
            return None
        moy = delib_data.get('moyenne')
        # Mention basée sur la moyenne
        mention = None
        if moy is not None:
            n = float(moy)
            if n >= 18: mention = 'Excellent (A)'
            elif n >= 16: mention = 'Très bien (B)'
            elif n >= 14: mention = 'Bien (C)'
            elif n >= 12: mention = 'Assez Bien (D)'
            elif n >= 10: mention = 'Passable (E)'
            elif n >= 8: mention = 'Insuffisant (F)'
            else: mention = 'Insatisfaisant (G)'
        has_rows = bool(delib_data.get('rows'))
        return {
            'moyenne': moy,
            'mention': mention,
            'decision_label': delib_data.get('decision_label', ''),
            'decision_code': delib_data.get('decision_code', ''),
            'moyenne_cat_a': delib_data.get('moyenne_cat_a'),
            'moyenne_cat_b': delib_data.get('moyenne_cat_b'),
            'credits_total': delib_data.get('credits_total', 0),
            'credits_valides': delib_data.get('credits_valides', 0),
            'has_data': has_rows,
        }

    resume_s1 = None
    resume_s2 = None
    resume_annuel = None

    if semestres_niveau:
        s1, s2 = semestres_niveau
        delib_s1 = _jury_compute_delib_ues(classe_active, etudiant, 'semestriel', s1, annee_code)
        delib_s2 = _jury_compute_delib_ues(classe_active, etudiant, 'semestriel', s2, annee_code)
        delib_annuel = _jury_compute_delib_ues(classe_active, etudiant, 'annuel', None, annee_code)
        resume_s1 = _extract_summary(delib_s1)
        resume_s2 = _extract_summary(delib_s2)
        resume_annuel = _extract_summary(delib_annuel)

    context = {
        'etudiant': etudiant,
        'annee_active': annee_active,
        'annee_code': annee_code,
        'inscription_active': inscription_active,
        'classe_active': classe_active,
        'jury_publie': True,
        'semestres_niveau': semestres_niveau,
        'resume_s1': resume_s1,
        'resume_s2': resume_s2,
        'resume_annuel': resume_annuel,
    }
    return render(request, 'etudiant/resultats.html', context)


@login_required
def etudiant_bulletin_pdf(request):
    """Génère et télécharge le bulletin de l'étudiant en PDF (style profil_pdf avec titre BULLETIN)"""
    etudiant = get_simulated_etudiant(request)
    if not etudiant:
        messages.error(request, 'Profil étudiant non trouvé.')
        return redirect('home')

    from reglage.models import AnneeAcademique
    annee_active = AnneeAcademique.get_annee_en_cours()
    annee_code = request.GET.get('annee') or (annee_active.code_anac if annee_active else None)

    inscriptions = Inscription.objects.filter(matricule_etudiant=etudiant).select_related('code_classe', 'cohorte')
    inscription_active = None
    if annee_code:
        inscription_active = inscriptions.filter(annee_academique=annee_code).first()
    if not inscription_active:
        inscription_active = inscriptions.order_by('-annee_academique').first()

    classe_active = inscription_active.code_classe if inscription_active else None
    if not classe_active:
        messages.error(request, 'Aucune classe trouvée.')
        return redirect('etudiant_resultats')

    # Vérifier que le jury a publié
    jury_obj = Jury.objects.filter(code_classe=classe_active, annee_academique=annee_code).first()
    jury_publie = bool(jury_obj and jury_obj.resultat_publie)
    if not jury_publie:
        messages.error(request, 'Les résultats ne sont pas encore publiés.')
        return redirect('etudiant_resultats')

    # Déterminer le type demandé
    vue = request.GET.get('vue', 'annuel')  # 's1', 's2', 'annuel'

    niveau_to_semestres = {
        'L1': (1, 2), 'L2': (3, 4), 'L3': (5, 6), 'M1': (7, 8), 'M2': (9, 10),
    }
    niveau_code = None
    code_classe_str = str(classe_active.code_classe)
    for prefix in ['L1', 'L2', 'L3', 'M1', 'M2']:
        if code_classe_str.startswith(prefix):
            niveau_code = prefix
            break
    semestres_niveau = niveau_to_semestres.get(niveau_code)
    if not semestres_niveau:
        messages.error(request, 'Impossible de déterminer les semestres.')
        return redirect('etudiant_resultats')

    s1, s2 = semestres_niveau

    if vue == 's1':
        type_delib = 'semestriel'
        semestre_int = s1
    elif vue == 's2':
        type_delib = 'semestriel'
        semestre_int = s2
    else:
        type_delib = 'annuel'
        semestre_int = None

    delib_data = _jury_compute_delib_ues(classe_active, etudiant, type_delib, semestre_int, annee_code)

    # Récupérer les dettes (InscriptionUE) pour affichage dans le bulletin
    from .views_passage_automatique import recuperer_dettes_classe_inferieure
    dettes = recuperer_dettes_classe_inferieure(etudiant, classe_active, annee_code)

    return _generer_bulletin_pdf(request, etudiant, classe_active, annee_code, semestre_int, delib_data, vue, dettes)


def _generer_bulletin_pdf(request, etudiant, classe_obj, annee, semestre, delib, vue_type, dettes=None):
    """Génère le bulletin de l'étudiant en PDF — même format que profil_pdf mais titre BULLETIN DE L'ÉTUDIANT"""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    import os
    from django.conf import settings

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=1.2*cm, leftMargin=1.2*cm,
                            topMargin=1*cm, bottomMargin=1*cm)
    elements_pdf = []
    styles = getSampleStyleSheet()

    normal_style = ParagraphStyle('NS', parent=styles['Normal'], fontSize=9, leading=11, fontName='Helvetica')
    bold_style = ParagraphStyle('BS', parent=styles['Normal'], fontSize=9, leading=11, fontName='Helvetica-Bold')
    center_style = ParagraphStyle('CS', parent=styles['Normal'], fontSize=9, leading=11, fontName='Helvetica', alignment=TA_CENTER)
    right_style = ParagraphStyle('RS', parent=styles['Normal'], fontSize=9, leading=11, fontName='Helvetica', alignment=TA_RIGHT)
    table_header_style = ParagraphStyle('THS', parent=styles['Normal'], fontSize=8, leading=9, fontName='Helvetica-Bold', alignment=TA_CENTER)
    table_cell_style = ParagraphStyle('TCS', parent=styles['Normal'], fontSize=8, leading=9, fontName='Helvetica', alignment=TA_LEFT)
    table_cell_center = ParagraphStyle('TCC', parent=styles['Normal'], fontSize=8, leading=9, fontName='Helvetica', alignment=TA_CENTER)

    # En-tête image
    from lmdmanagersystem.middleware import get_entete_path
    entete_path = get_entete_path()
    if os.path.exists(entete_path):
        from PIL import Image as PILImage
        pil_img = PILImage.open(entete_path)
        img_w, img_h = pil_img.size
        desired_w = 18*cm
        desired_h = desired_w * (img_h / img_w)
        img = RLImage(entete_path, width=desired_w, height=desired_h)
        elements_pdf.append(img)
        elements_pdf.append(Spacer(1, 0.5*cm))

    # Titre
    title_style = ParagraphStyle('TS', parent=styles['Normal'], fontSize=14, leading=16,
                                  fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=10)
    elements_pdf.append(Paragraph("<b>BULLETIN DE L'ÉTUDIANT</b>", title_style))
    elements_pdf.append(Spacer(1, 0.3*cm))

    # Infos étudiant
    nom_complet = getattr(etudiant, 'nom_complet', '') or ''
    classe_label = getattr(classe_obj, 'code_classe', '')
    if vue_type == 's1':
        periode_txt = f"Semestre {semestre}"
    elif vue_type == 's2':
        periode_txt = f"Semestre {semestre}"
    else:
        periode_txt = "Annuel"

    info_table = Table([
        [Paragraph("<b>Matricule :</b>", right_style),
         Paragraph(f"{etudiant.matricule_et}", normal_style),
         Paragraph("<b>Année académique :</b>", right_style),
         Paragraph(f"{annee}", normal_style),
         Paragraph(f"<b>Période :</b> {periode_txt}", normal_style)],
        [Paragraph("<b>Noms :</b>", right_style),
         Paragraph(f"{nom_complet}", normal_style),
         Paragraph("<b>Classe :</b>", right_style),
         Paragraph(f"{classe_label}", normal_style),
         Paragraph("", normal_style)]
    ], colWidths=[2.5*cm, 4*cm, 3.5*cm, 4*cm, 4*cm])
    info_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0, colors.white),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements_pdf.append(info_table)
    elements_pdf.append(Spacer(1, 4*mm))

    # Tableau des notes
    table_data = [[
        Paragraph('<b>Code</b>', table_header_style),
        Paragraph("<b>Intitulé de l'UE</b>", table_header_style),
        Paragraph('<b>Éléments constitutifs</b>', table_header_style),
        Paragraph('<b>Cat.</b>', table_header_style),
        Paragraph('<b>Cr</b>', table_header_style),
        Paragraph('<b>CC</b>', table_header_style),
        Paragraph('<b>Exam</b>', table_header_style),
        Paragraph('<b>Note</b>', table_header_style),
        Paragraph('<b>N.Pd</b>', table_header_style),
        Paragraph('<b>Ratt</b>', table_header_style),
        Paragraph('<b>Statut</b>', table_header_style)
    ]]

    rows = delib.get('rows', []) if delib else []
    credits_total = delib.get('credits_total', 0) if delib else 0
    credits_valides = delib.get('credits_valides', 0) if delib else 0
    moyenne = delib.get('moyenne') if delib else None
    moyenne_cat_a = delib.get('moyenne_cat_a') if delib else None
    moyenne_cat_b = delib.get('moyenne_cat_b') if delib else None
    decision_label = delib.get('decision_label', '') if delib else ''
    decision_code = delib.get('decision_code', '') if delib else ''

    def _fmt(v):
        return f"{v:.1f}" if v is not None else '-'

    for row in rows:
        table_data.append([
            Paragraph(str(row.get('code_ec', row.get('code_ue', ''))), table_cell_style),
            Paragraph(str(row.get('intitule_ue', '')), table_cell_style),
            Paragraph(str(row.get('intitule_ec', '-')), table_cell_style),
            Paragraph(str(row.get('categorie', '') or '-'), table_cell_center),
            Paragraph(str(row.get('credit', '') or '-'), table_cell_center),
            Paragraph(_fmt(row.get('cc')), table_cell_center),
            Paragraph(_fmt(row.get('examen')), table_cell_center),
            Paragraph(_fmt(row.get('note')), table_cell_center),
            Paragraph(_fmt(row.get('note_ponderee')), table_cell_center),
            Paragraph(_fmt(row.get('rattrapage')), table_cell_center),
            Paragraph(str(row.get('statut', '-')), table_cell_center)
        ])

    col_widths = [1.5*cm, 3.8*cm, 5.2*cm, 0.9*cm, 0.8*cm, 0.9*cm, 1.0*cm, 0.9*cm, 1.0*cm, 0.9*cm, 1.8*cm]
    cours_table = Table(table_data, colWidths=col_widths)
    cours_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e0e0e0')),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elements_pdf.append(cours_table)
    elements_pdf.append(Spacer(1, 4*mm))

    # Mention
    mention = ''
    if moyenne is not None:
        n = float(moyenne)
        if n >= 18: mention = 'Excellent (A)'
        elif n >= 16: mention = 'Très bien (B)'
        elif n >= 14: mention = 'Bien (C)'
        elif n >= 12: mention = 'Assez Bien (D)'
        elif n >= 10: mention = 'Passable (E)'
        elif n >= 8: mention = 'Insuffisant (F)'
        else: mention = 'Insatisfaisant (G)'

    def _fmt_dec(v):
        return f"{v:.2f}" if v is not None else '-'

    # Résumé ligne 1
    summary1 = Table([
        [Paragraph('<b>Total crédits</b>', table_cell_style),
         Paragraph(str(credits_total), table_cell_center),
         Paragraph('<b>Crédits capitalisés</b>', table_cell_style),
         Paragraph(str(credits_valides), table_cell_center),
         Paragraph('<b>Moyenne</b>', table_cell_style),
         Paragraph(_fmt_dec(moyenne), table_cell_center),
         Paragraph('<b>Mention</b>', table_cell_style),
         Paragraph(mention, table_cell_center)]
    ], colWidths=[3*cm, 1.2*cm, 3.8*cm, 1.2*cm, 2*cm, 1.2*cm, 2*cm, 3.6*cm])
    summary1.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements_pdf.append(summary1)
    elements_pdf.append(Spacer(1, 2*mm))

    # Résumé ligne 2
    moy_a_str = _fmt_dec(moyenne_cat_a) if moyenne_cat_a else '-'
    moy_b_str = _fmt_dec(moyenne_cat_b) if moyenne_cat_b else '-'
    decision_jury = f"{decision_label} ({decision_code})" if decision_label else '-'

    summary2 = Table([
        [Paragraph('<b>Catégorie A</b>', table_cell_style),
         Paragraph(moy_a_str, table_cell_center),
         Paragraph('<b>Catégorie B</b>', table_cell_style),
         Paragraph(moy_b_str, table_cell_center),
         Paragraph('<b>Décision du jury</b>', table_cell_style),
         Paragraph(decision_jury, table_cell_center)]
    ], colWidths=[3.5*cm, 3*cm, 3.5*cm, 3*cm, 3.5*cm, 2.5*cm])
    summary2.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements_pdf.append(summary2)
    elements_pdf.append(Spacer(1, 4*mm))

    # Section Dettes (si l'étudiant a des dettes)
    if dettes:
        dette_title_style = ParagraphStyle('DTS', parent=styles['Normal'], fontSize=9, leading=11,
                                            fontName='Helvetica-Bold', textColor=colors.HexColor('#856404'))
        elements_pdf.append(Paragraph('<b>⚠ Suivi des Dettes</b>', dette_title_style))
        elements_pdf.append(Spacer(1, 2*mm))

        dette_data = [[
            Paragraph('<b>Code UE/EC</b>', table_header_style),
            Paragraph('<b>Intitulé</b>', table_header_style),
            Paragraph("<b>Classe d'origine</b>", table_header_style),
            Paragraph('<b>Statut</b>', table_header_style),
        ]]
        for d in dettes:
            code = d.code_ec.code_ec if d.code_ec else (d.code_ue.code_ue if d.code_ue else '-')
            intitule = d.code_ec.intitule_ue if d.code_ec else (d.code_ue.intitule_ue if d.code_ue else '-')
            classe_orig = d.code_classe.code_classe if d.code_classe else '-'
            statut = 'Liquidée' if d.type_inscription == 'DETTE_LIQUIDEE' else 'En cours'
            dette_data.append([
                Paragraph(code, table_cell_style),
                Paragraph(intitule, table_cell_style),
                Paragraph(classe_orig, table_cell_center),
                Paragraph(statut, table_cell_center),
            ])

        dette_table = Table(dette_data, colWidths=[3*cm, 7.5*cm, 3.5*cm, 4*cm])
        dette_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fff3cd')),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#856404')),
        ]))
        elements_pdf.append(dette_table)

    elements_pdf.append(Spacer(1, 0.6*cm))

    # Date
    from datetime import datetime
    date_str = datetime.now().strftime("%d/%m/%Y")
    elements_pdf.append(Paragraph(f"Fait à .................................................. le {date_str}", right_style))
    elements_pdf.append(Spacer(1, 0.6*cm))

    # Signatures
    underline_style = ParagraphStyle('US', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, fontName='Helvetica')
    italic_style = ParagraphStyle('IS', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, fontName='Helvetica-Oblique')

    jury_info = Jury.objects.filter(code_classe=classe_obj, annee_academique=annee).first()
    membre_nom = president_nom = secretaire_nom = "____________________"
    membre_grade = president_grade = secretaire_grade = "Grade: _______________"
    if jury_info:
        for field, name_var, grade_var in [
            ('membre', 'membre_nom', 'membre_grade'),
            ('president', 'president_nom', 'president_grade'),
            ('secretaire', 'secretaire_nom', 'secretaire_grade')
        ]:
            mat = getattr(jury_info, field, None)
            if mat:
                try:
                    ens = Enseignant.objects.get(matricule_en=mat)
                    if field == 'membre':
                        membre_nom = ens.nom_complet
                        membre_grade = ens.grade.designation_grade if ens.grade else "Grade: _______________"
                    elif field == 'president':
                        president_nom = ens.nom_complet
                        president_grade = ens.grade.designation_grade if ens.grade else "Grade: _______________"
                    elif field == 'secretaire':
                        secretaire_nom = ens.nom_complet
                        secretaire_grade = ens.grade.designation_grade if ens.grade else "Grade: _______________"
                except Exception:
                    pass

    sig_table = Table([
        [Paragraph("<b>Membre du jury</b>", center_style),
         Paragraph("<b>Président du jury</b>", center_style),
         Paragraph("<b>Secrétaire du jury</b>", center_style)],
        [Paragraph("<br/><br/>", center_style)] * 3,
        [Paragraph(f"<u>{membre_nom}</u>", underline_style),
         Paragraph(f"<u>{president_nom}</u>", underline_style),
         Paragraph(f"<u>{secretaire_nom}</u>", underline_style)],
        [Paragraph(f"<i>{membre_grade}</i>", italic_style),
         Paragraph(f"<i>{president_grade}</i>", italic_style),
         Paragraph(f"<i>{secretaire_grade}</i>", italic_style)]
    ], colWidths=[6*cm, 6*cm, 6*cm])
    sig_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0, colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
    ]))
    elements_pdf.append(sig_table)

    doc.build(elements_pdf)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    nom_clean = (nom_complet or etudiant.matricule_et).replace(" ", "_").replace("/", "_")
    type_str = f"S{semestre}" if semestre else "Annuel"
    filename = f"Bulletin_{nom_clean}_{type_str}_{annee}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def etudiant_telecharger_bulletin(request):
    """Téléchargement du bulletin de notes par l'étudiant"""
    etudiant = get_simulated_etudiant(request)
    if not etudiant:
        messages.error(request, 'Profil étudiant non trouvé.')
        return redirect('home')
    
    from reglage.models import AnneeAcademique
    annee_active = AnneeAcademique.get_annee_en_cours()
    annee_code = request.GET.get('annee') or (annee_active.code_anac if annee_active else None)
    
    inscriptions = Inscription.objects.filter(matricule_etudiant=etudiant).select_related('code_classe', 'cohorte')
    inscription_active = None
    if annee_code:
        inscription_active = inscriptions.filter(annee_academique=annee_code).first()
    if not inscription_active:
        inscription_active = inscriptions.order_by('-annee_academique').first()
    
    classe_active = inscription_active.code_classe if inscription_active else None
    
    if not classe_active:
        messages.error(request, 'Aucune classe trouvée.')
        return redirect('etudiant_resultats')
    
    # Récupérer le bulletin
    bulletin = BulletinNotes.objects.filter(
        etudiant=etudiant,
        annee_academique=annee_code or inscription_active.annee_academique,
        code_classe=classe_active,
        disponible=True
    ).first()
    
    if not bulletin:
        messages.error(request, 'Aucun bulletin disponible pour le moment.')
        return redirect('etudiant_resultats')
    
    if not bulletin.fichier_pdf:
        messages.error(request, 'Le fichier du bulletin n\'est pas disponible.')
        return redirect('etudiant_resultats')
    
    # Télécharger le fichier
    try:
        response = HttpResponse(bulletin.fichier_pdf.read(), content_type='application/pdf')
        filename = f"Bulletin_{etudiant.matricule_et}_{annee_code or inscription_active.annee_academique}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        messages.error(request, f'Erreur lors du téléchargement: {str(e)}')
        return redirect('etudiant_resultats')


# ========== VUES ENSEIGNANT ==========

@login_required
def enseignant_dashboard(request):
    """Tableau de bord enseignant"""
    # Si admin sans simulation active, afficher la liste pour sélection
    if request.user.is_staff and 'simulated_enseignant' not in request.session:
        enseignants = Enseignant.objects.select_related('grade').all()[:50]
        return render(request, 'admin/select_enseignant.html', {'enseignants': enseignants})
    
    # Récupérer l'enseignant (simulé ou connecté)
    enseignant = get_simulated_enseignant(request)
    if not enseignant:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')
    
    from reglage.models import AnneeAcademique
    
    # Année académique active
    try:
        annee_active = AnneeAcademique.objects.get(active=True)
    except AnneeAcademique.DoesNotExist:
        annee_active = None
    
    # Récupérer les attributions de l'enseignant
    attributions = Attribution.objects.filter(matricule_en=enseignant)
    total_cours = attributions.count()
    
    # Récupérer les classes des cours
    classes_codes = set()
    for attr in attributions:
        try:
            ue = UE.objects.get(code_ue=attr.code_cours)
            if ue.classe:
                classes_codes.add(ue.classe)
        except UE.DoesNotExist:
            try:
                ec = EC.objects.get(code_ec=attr.code_cours)
                if ec.classe:
                    classes_codes.add(ec.classe)
            except EC.DoesNotExist:
                pass
    
    # Nombre d'étudiants concernés
    total_etudiants = Etudiant.objects.filter(
        inscription__code_classe__code_classe__in=classes_codes
    ).distinct().count()
    
    # Récupérer les évaluations pour les cours de l'enseignant
    total_evaluations = 0
    total_validees = 0
    total_non_validees = 0
    total_en_cours = 0
    somme_notes = 0
    count_notes = 0
    
    for attr in attributions:
        try:
            ue = UE.objects.get(code_ue=attr.code_cours)
            evals = Evaluation.objects.filter(code_ue=ue)
        except UE.DoesNotExist:
            try:
                ec = EC.objects.get(code_ec=attr.code_cours)
                evals = Evaluation.objects.filter(code_ec=ec)
            except EC.DoesNotExist:
                evals = Evaluation.objects.none()
        
        for ev in evals:
            total_evaluations += 1
            if ev.cc is not None and ev.examen is not None:
                note_finale = ev.cc + ev.examen
                somme_notes += note_finale
                count_notes += 1
                if note_finale >= 10:
                    total_validees += 1
                else:
                    total_non_validees += 1
            else:
                total_en_cours += 1
    
    moyenne_generale = round(somme_notes / count_notes, 2) if count_notes > 0 else 0
    taux_reussite = round((total_validees / count_notes) * 100, 1) if count_notes > 0 else 0
    
    # Derniers cours (5 derniers)
    derniers_cours = []
    for attr in attributions[:5]:
        cours_info = {'code': attr.code_cours, 'intitule': attr.code_cours, 'classe': None}
        try:
            ue = UE.objects.get(code_ue=attr.code_cours)
            cours_info['intitule'] = ue.intitule_ue
            cours_info['classe'] = ue.classe
        except UE.DoesNotExist:
            try:
                ec = EC.objects.get(code_ec=attr.code_cours)
                cours_info['intitule'] = ec.intitule_ue
                cours_info['classe'] = ec.classe
            except EC.DoesNotExist:
                pass
        derniers_cours.append(cours_info)
    
    # Récupérer les notifications non lues pour l'enseignant
    notifications_non_lues = Notification.objects.filter(
        destinataire=request.user,
        lue=False
    ).order_by('-date_creation')[:5]
    
    total_notifications_non_lues = Notification.objects.filter(
        destinataire=request.user,
        lue=False
    ).count()
    
    context = {
        'enseignant': enseignant,
        'annee_active': annee_active,
        'total_cours': total_cours,
        'total_etudiants': total_etudiants,
        'total_evaluations': total_evaluations,
        'total_validees': total_validees,
        'total_non_validees': total_non_validees,
        'total_en_cours': total_en_cours,
        'moyenne_generale': moyenne_generale,
        'taux_reussite': taux_reussite,
        'derniers_cours': derniers_cours,
        'notifications_non_lues': notifications_non_lues,
        'total_notifications_non_lues': total_notifications_non_lues,
        'is_admin_view': is_admin_viewing(request),
    }
    return render(request, 'enseignant/dashboard.html', context)


@login_required
def enseignant_notifications(request):
    """Afficher toutes les notifications de l'enseignant"""
    enseignant = get_simulated_enseignant(request)
    if not enseignant:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')
    
    # Récupérer toutes les notifications de l'enseignant
    notifications = Notification.objects.filter(
        destinataire=request.user
    ).order_by('-date_creation')
    
    # Marquer toutes les notifications comme lues si demandé
    if request.method == 'POST' and request.POST.get('marquer_toutes_lues'):
        notifications.update(lue=True, date_lecture=timezone.now())
        messages.success(request, 'Toutes les notifications ont été marquées comme lues.')
        return redirect('enseignant_notifications')
    
    # Marquer une notification spécifique comme lue
    if request.method == 'POST' and request.POST.get('marquer_lue'):
        notification_id = request.POST.get('notification_id')
        try:
            notification = Notification.objects.get(id_notification=notification_id, destinataire=request.user)
            notification.marquer_comme_lue()
            messages.success(request, 'Notification marquée comme lue.')
        except Notification.DoesNotExist:
            messages.error(request, 'Notification non trouvée.')
        return redirect('enseignant_notifications')
    
    context = {
        'enseignant': enseignant,
        'notifications': notifications,
        'total_notifications': notifications.count(),
        'notifications_non_lues': notifications.filter(lue=False).count(),
        'is_admin_view': is_admin_viewing(request),
    }
    return render(request, 'enseignant/notifications.html', context)


@login_required
def enseignant_encoder_notes(request):
    """Encodage des notes par l'enseignant"""
    enseignant = get_simulated_enseignant(request)
    if not enseignant:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')
    try:
        pass  # enseignant déjà récupéré
        
        if request.method == 'POST':
            eval_id = request.POST.get('eval_id')
            cc = float(request.POST.get('cc', 0))
            examen = float(request.POST.get('examen', 0))
            rattrapage = request.POST.get('rattrapage')
            rachat = request.POST.get('rachat')
            
            evaluation = get_object_or_404(Evaluation, id_ev=eval_id)
            evaluation.cc = cc
            evaluation.examen = examen
            
            if rattrapage:
                evaluation.rattrapage = float(rattrapage)
            if rachat:
                evaluation.rachat = float(rachat)
            
            evaluation.save()
            messages.success(request, 'Notes enregistrées avec succès!')
            return redirect('enseignant_encoder_notes')
        
        # Afficher toutes les évaluations
        evaluations = Evaluation.objects.all().select_related('matricule_etudiant', 'code_ue', 'code_ec')
        
        context = {
            'enseignant': enseignant,
            'evaluations': evaluations,
        }
        return render(request, 'enseignant/encoder_notes.html', context)
    except Enseignant.DoesNotExist:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')


@login_required
def enseignant_evaluer_cours(request, code_cours, annee):
    """Évaluer les étudiants d'un cours spécifique"""
    enseignant = get_simulated_enseignant(request)
    if not enseignant:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')
    try:
        pass  # enseignant déjà récupéré
        
        # Vérifier que l'enseignant a bien ce cours attribué
        attribution = Attribution.objects.filter(
            matricule_en=enseignant,
            code_cours=code_cours,
            annee_academique=annee
        ).first()
        
        if not attribution:
            messages.error(request, "Vous n'êtes pas autorisé à évaluer ce cours.")
            return redirect('enseignant_mes_cours')
        
        # Récupérer les infos du cours (UE ou EC)
        cours_info = {'code': code_cours, 'intitule': code_cours, 'type': None, 'classe': None}
        try:
            ue = UE.objects.get(code_ue=code_cours)
            cours_info['intitule'] = ue.intitule_ue
            cours_info['type'] = 'UE'
            cours_info['classe'] = ue.classe
        except UE.DoesNotExist:
            try:
                ec = EC.objects.get(code_ec=code_cours)
                cours_info['intitule'] = ec.intitule_ue
                cours_info['type'] = 'EC'
                cours_info['classe'] = ec.classe
            except EC.DoesNotExist:
                pass
        
        # Vérifier si le rattrapage/rachat est activé par le jury
        rattrapage_actif = False
        rachat_actif = False
        if cours_info.get('classe'):
            param_eval = ParametreEvaluation.objects.filter(
                code_classe__code_classe=cours_info['classe'],
                annee_academique=annee
            ).first()
            if param_eval:
                rattrapage_actif = param_eval.rattrapage_actif
                rachat_actif = param_eval.rachat_actif
        
        # Récupérer les étudiants inscrits dans la classe pour cette année
        etudiants = []
        matricules_deja_ajoutes = set()
        if cours_info.get('classe'):
            inscriptions = Inscription.objects.filter(
                code_classe__code_classe=cours_info['classe'],
                annee_academique=annee
            ).select_related('matricule_etudiant')
            
            for insc in inscriptions:
                # Chercher l'évaluation existante (avec filtre annee_academique)
                eval_existante = Evaluation.objects.filter(
                    matricule_etudiant=insc.matricule_etudiant,
                    code_ue__code_ue=code_cours if cours_info['type'] == 'UE' else None,
                    code_ec__code_ec=code_cours if cours_info['type'] == 'EC' else None,
                    annee_academique=annee,
                ).first()
                
                etudiants.append({
                    'matricule': insc.matricule_etudiant.matricule_et,
                    'nom_complet': insc.matricule_etudiant.nom_complet,
                    'evaluation': eval_existante,
                    'cc': eval_existante.cc if eval_existante else '',
                    'examen': eval_existante.examen if eval_existante else '',
                    'rattrapage': eval_existante.rattrapage if eval_existante else '',
                    'rachat': eval_existante.rachat if eval_existante else '',
                    'est_dette': False,
                })
                matricules_deja_ajoutes.add(insc.matricule_etudiant.matricule_et)
            
            # === NOUVEAU: Récupérer les étudiants avec dettes (InscriptionUE) ===
            # Ces étudiants sont inscrits en classe supérieure mais doivent
            # repasser ce cours dans cette classe
            from .models import InscriptionUE
            
            inscriptions_ue_filter = {}
            if cours_info['type'] == 'UE':
                inscriptions_ue_filter['code_ue__code_ue'] = code_cours
            elif cours_info['type'] == 'EC':
                inscriptions_ue_filter['code_ec__code_ec'] = code_cours
            
            inscriptions_ue = InscriptionUE.objects.filter(
                code_classe__code_classe=cours_info['classe'],
                annee_academique=annee,
                type_inscription='DETTE_COMPENSEE',
                **inscriptions_ue_filter
            ).select_related('matricule_etudiant')
            
            for insc_ue in inscriptions_ue:
                # Éviter les doublons
                if insc_ue.matricule_etudiant.matricule_et in matricules_deja_ajoutes:
                    continue
                
                # Chercher l'évaluation existante pour la dette
                eval_existante = Evaluation.objects.filter(
                    matricule_etudiant=insc_ue.matricule_etudiant,
                    code_ue__code_ue=code_cours if cours_info['type'] == 'UE' else None,
                    code_ec__code_ec=code_cours if cours_info['type'] == 'EC' else None,
                    annee_academique=annee,
                ).first()
                
                etudiants.append({
                    'matricule': insc_ue.matricule_etudiant.matricule_et,
                    'nom_complet': f"{insc_ue.matricule_etudiant.nom_complet} (Dette)",
                    'evaluation': eval_existante,
                    'cc': eval_existante.cc if eval_existante else '',
                    'examen': eval_existante.examen if eval_existante else '',
                    'rattrapage': eval_existante.rattrapage if eval_existante else '',
                    'rachat': eval_existante.rachat if eval_existante else '',
                    'est_dette': True,
                })
                matricules_deja_ajoutes.add(insc_ue.matricule_etudiant.matricule_et)
        
        # Vérifier si les notes ont été envoyées au jury
        envoye_au_jury = attribution.envoye_au_jury
        
        if request.method == 'POST':
            # Bloquer toute modification si déjà envoyé au jury
            if envoye_au_jury:
                messages.error(request, 'Les notes ont déjà été envoyées au jury. Modification impossible.')
                return redirect('enseignant_evaluer_cours', code_cours=code_cours, annee=annee)
            
            # Sauvegarder les notes
            matricules = request.POST.getlist('matricule')
            ccs = request.POST.getlist('cc')
            examens = request.POST.getlist('examen')
            rattrapages = request.POST.getlist('rattrapage') if rattrapage_actif else []
            rachats = request.POST.getlist('rachat') if rachat_actif else []
            
            def parse_note(val):
                """Convertir une note en float, en gérant la virgule comme séparateur décimal."""
                if not val:
                    return None
                return float(str(val).replace(',', '.'))
            
            for i, matricule in enumerate(matricules):
                etudiant = Etudiant.objects.get(matricule_et=matricule)
                
                # Créer ou mettre à jour l'évaluation
                eval_data = {
                    'matricule_etudiant': etudiant,
                    'annee_academique': annee,
                    'cc': parse_note(ccs[i]) if i < len(ccs) else None,
                    'examen': parse_note(examens[i]) if i < len(examens) else None,
                }
                
                # Rattrapage et rachat seulement si activés par le jury
                if rattrapage_actif and i < len(rattrapages):
                    eval_data['rattrapage'] = parse_note(rattrapages[i])
                if rachat_actif and i < len(rachats):
                    eval_data['rachat'] = parse_note(rachats[i])
                
                if cours_info['type'] == 'UE':
                    eval_data['code_ue'] = UE.objects.get(code_ue=code_cours)
                    eval_data['code_classe'] = cours_info['classe']
                    Evaluation.objects.update_or_create(
                        matricule_etudiant=etudiant,
                        code_ue=eval_data['code_ue'],
                        annee_academique=annee,
                        code_classe=cours_info['classe'],
                        defaults=eval_data
                    )
                else:
                    eval_data['code_ec'] = EC.objects.get(code_ec=code_cours)
                    eval_data['code_classe'] = cours_info['classe']
                    Evaluation.objects.update_or_create(
                        matricule_etudiant=etudiant,
                        code_ec=eval_data['code_ec'],
                        annee_academique=annee,
                        code_classe=cours_info['classe'],
                        defaults=eval_data
                    )
            
            messages.success(request, 'Notes enregistrées avec succès!')
            return redirect('enseignant_evaluer_cours', code_cours=code_cours, annee=annee)
        
        context = {
            'enseignant': enseignant,
            'cours': cours_info,
            'annee': annee,
            'etudiants': etudiants,
            'rattrapage_actif': rattrapage_actif,
            'rachat_actif': rachat_actif,
            'envoye_au_jury': envoye_au_jury,
            'date_envoi_jury': attribution.date_envoi_jury,
        }
        return render(request, 'enseignant/evaluer_cours.html', context)
    except Enseignant.DoesNotExist:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')


@login_required
def envoyer_au_jury(request, code_cours, annee):
    """Envoyer les notes au jury - bloque toute modification ultérieure par l'enseignant"""
    enseignant = get_simulated_enseignant(request)
    if not enseignant:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')
    
    attribution = Attribution.objects.filter(
        matricule_en=enseignant,
        code_cours=code_cours,
        annee_academique=annee
    ).first()
    
    if not attribution:
        messages.error(request, "Vous n'êtes pas autorisé à accéder à ce cours.")
        return redirect('enseignant_mes_cours')
    
    if attribution.envoye_au_jury:
        messages.warning(request, 'Les notes ont déjà été envoyées au jury.')
        return redirect('enseignant_evaluer_cours', code_cours=code_cours, annee=annee)
    
    if request.method == 'POST':
        attribution.envoye_au_jury = True
        attribution.date_envoi_jury = timezone.now()
        attribution.save()
        messages.success(request, 'Les notes ont été envoyées au jury avec succès ! Aucune modification ne sera possible.')
    
    return redirect('enseignant_evaluer_cours', code_cours=code_cours, annee=annee)


@login_required
def telecharger_grille_evaluation(request, code_cours, annee):
    """Télécharger le modèle Excel de la grille d'évaluation"""
    try:
        enseignant = Enseignant.objects.get(id_lgn=request.user)
        
        # Vérifier que l'enseignant a bien ce cours attribué
        attribution = Attribution.objects.filter(
            matricule_en=enseignant,
            code_cours=code_cours,
            annee_academique=annee
        ).first()
        
        if not attribution:
            messages.error(request, "Vous n'êtes pas autorisé à accéder à ce cours.")
            return redirect('enseignant_mes_cours')
        
        # Récupérer les infos du cours
        cours_info = {'code': code_cours, 'intitule': code_cours, 'type': None, 'classe': None}
        try:
            ue = UE.objects.get(code_ue=code_cours)
            cours_info['intitule'] = ue.intitule_ue
            cours_info['type'] = 'UE'
            cours_info['classe'] = ue.classe
        except UE.DoesNotExist:
            try:
                ec = EC.objects.get(code_ec=code_cours)
                cours_info['intitule'] = ec.intitule_ue
                cours_info['type'] = 'EC'
                cours_info['classe'] = ec.classe
            except EC.DoesNotExist:
                pass
        
        # Vérifier si rattrapage/rachat activé
        rattrapage_actif = False
        rachat_actif = False
        if cours_info.get('classe'):
            param_eval = ParametreEvaluation.objects.filter(
                code_classe__code_classe=cours_info['classe'],
                annee_academique=annee
            ).first()
            if param_eval:
                rattrapage_actif = param_eval.rattrapage_actif
                rachat_actif = param_eval.rachat_actif
        
        # Récupérer les étudiants
        data = []
        if cours_info.get('classe'):
            inscriptions = Inscription.objects.filter(
                code_classe__code_classe=cours_info['classe'],
                annee_academique=annee
            ).select_related('matricule_etudiant')
            
            for insc in inscriptions:
                # Chercher l'évaluation existante
                eval_existante = Evaluation.objects.filter(
                    matricule_etudiant=insc.matricule_etudiant,
                    code_ue__code_ue=code_cours if cours_info['type'] == 'UE' else None,
                    code_ec__code_ec=code_cours if cours_info['type'] == 'EC' else None,
                ).first()
                
                row = {
                    'Matricule': insc.matricule_etudiant.matricule_et,
                    'Nom Complet': insc.matricule_etudiant.nom_complet,
                    'CC (0-10)': eval_existante.cc if eval_existante and eval_existante.cc else '',
                    'Examen (0-10)': eval_existante.examen if eval_existante and eval_existante.examen else '',
                }
                if rattrapage_actif:
                    row['Rattrapage (0-20)'] = eval_existante.rattrapage if eval_existante and eval_existante.rattrapage else ''
                if rachat_actif:
                    row['Rachat (0-20)'] = eval_existante.rachat if eval_existante and eval_existante.rachat else ''
                data.append(row)
        
        # Créer le fichier Excel avec formatage
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
        from io import BytesIO
        
        df = pd.DataFrame(data)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Grille d'évaluation"
        
        # En-tête du document
        ws.merge_cells('A1:F1')
        ws['A1'] = f"GRILLE D'ÉVALUATION - {cours_info['intitule']}"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Infos du cours
        ws['A2'] = f"Code: {code_cours}"
        ws['B2'] = f"Année: {annee}"
        ws['C2'] = f"Classe: {cours_info.get('classe', '-')}"
        ws.row_dimensions[2].height = 20
        
        # Ligne vide
        ws.row_dimensions[3].height = 10
        
        # En-têtes des colonnes
        headers = list(df.columns) if not df.empty else ['Matricule', 'Nom Complet', 'CC (0-10)', 'Examen (0-10)']
        header_row = 4
        
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws.row_dimensions[header_row].height = 25
        
        # Données
        data_alignment = Alignment(horizontal='center', vertical='center')
        data_font = Font(size=11)
        alt_fill = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
        
        for row_idx, row_data in enumerate(df.values, header_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.font = data_font
                if col_idx <= 2:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = data_alignment
                # Alternance de couleurs
                if (row_idx - header_row) % 2 == 0:
                    cell.fill = alt_fill
        
        # Largeurs des colonnes
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 14
        
        # Instructions en bas
        last_row = header_row + len(df) + 2
        ws.cell(row=last_row, column=1, value="Instructions:").font = Font(bold=True, italic=True)
        ws.cell(row=last_row + 1, column=1, value="• CC et Examen : notes sur 10").font = Font(italic=True, size=10)
        ws.cell(row=last_row + 2, column=1, value="• Rattrapage et Rachat : notes sur 20").font = Font(italic=True, size=10)
        ws.cell(row=last_row + 3, column=1, value="• Ne modifiez pas les colonnes Matricule et Nom").font = Font(italic=True, size=10, color="FF0000")
        
        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=grille_{code_cours}_{annee}.xlsx'
        
        return response
    except Enseignant.DoesNotExist:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')


@login_required
def importer_grille_evaluation(request, code_cours, annee):
    """Importer les notes depuis un fichier Excel"""
    try:
        enseignant = get_simulated_enseignant(request)
        if not enseignant:
            messages.error(request, 'Profil enseignant non trouvé.')
            return redirect('home')
        
        # Vérifier que l'enseignant a bien ce cours attribué
        attribution = Attribution.objects.filter(
            matricule_en=enseignant,
            code_cours=code_cours,
            annee_academique=annee
        ).first()
        
        if not attribution:
            messages.error(request, "Vous n'êtes pas autorisé à accéder à ce cours.")
            return redirect('enseignant_mes_cours')
        
        # Bloquer l'import si déjà envoyé au jury
        if attribution.envoye_au_jury:
            messages.error(request, 'Les notes ont déjà été envoyées au jury. Import impossible.')
            return redirect('enseignant_evaluer_cours', code_cours=code_cours, annee=annee)
        
        if request.method == 'POST' and request.FILES.get('fichier_excel'):
            fichier = request.FILES['fichier_excel']
            
            try:
                # Les en-têtes sont à la ligne 4 (index 3) dans le fichier exporté
                # Essayer d'abord avec header=3, sinon header=0
                df = pd.read_excel(fichier, header=3)
                # Vérifier si les colonnes attendues existent
                if 'Matricule' not in df.columns:
                    # Remonter au début et essayer header=0
                    fichier.seek(0)
                    df = pd.read_excel(fichier, header=0)
                if 'Matricule' not in df.columns:
                    raise ValueError("Le fichier Excel ne contient pas la colonne 'Matricule'. Utilisez le modèle téléchargé.")
                
                # Récupérer le type de cours
                cours_type = None
                cours_classe = None
                try:
                    ue_obj = UE.objects.get(code_ue=code_cours)
                    cours_type = 'UE'
                    cours_classe = ue_obj.classe
                except UE.DoesNotExist:
                    try:
                        ec_obj = EC.objects.get(code_ec=code_cours)
                        cours_type = 'EC'
                        cours_classe = ec_obj.classe
                    except EC.DoesNotExist:
                        pass
                
                count = 0
                for _, row in df.iterrows():
                    matricule = str(row.get('Matricule', '')).strip()
                    if not matricule or matricule == 'nan':
                        continue
                    
                    try:
                        etudiant = Etudiant.objects.get(matricule_et=matricule)
                        
                        eval_data = {
                            'matricule_etudiant': etudiant,
                            'annee_academique': annee,
                        }
                        
                        # CC
                        cc_val = row.get('CC (0-10)', '')
                        if pd.notna(cc_val) and cc_val != '':
                            eval_data['cc'] = float(cc_val)
                        
                        # Examen
                        examen_val = row.get('Examen (0-10)', '')
                        if pd.notna(examen_val) and examen_val != '':
                            eval_data['examen'] = float(examen_val)
                        
                        # Rattrapage
                        rattrapage_val = row.get('Rattrapage (0-20)', '')
                        if pd.notna(rattrapage_val) and rattrapage_val != '':
                            eval_data['rattrapage'] = float(rattrapage_val)
                        
                        # Rachat
                        rachat_val = row.get('Rachat (0-20)', '')
                        if pd.notna(rachat_val) and rachat_val != '':
                            eval_data['rachat'] = float(rachat_val)
                        
                        if cours_type == 'UE':
                            eval_data['code_ue'] = UE.objects.get(code_ue=code_cours)
                            eval_data['code_classe'] = cours_classe
                            Evaluation.objects.update_or_create(
                                matricule_etudiant=etudiant,
                                code_ue=eval_data['code_ue'],
                                annee_academique=annee,
                                code_classe=cours_classe,
                                defaults=eval_data
                            )
                        else:
                            eval_data['code_ec'] = EC.objects.get(code_ec=code_cours)
                            eval_data['code_classe'] = cours_classe
                            Evaluation.objects.update_or_create(
                                matricule_etudiant=etudiant,
                                code_ec=eval_data['code_ec'],
                                annee_academique=annee,
                                code_classe=cours_classe,
                                defaults=eval_data
                            )
                        count += 1
                    except Etudiant.DoesNotExist:
                        continue
                
                messages.success(request, f'{count} notes importées avec succès!')
            except Exception as e:
                messages.error(request, f'Erreur lors de l\'importation: {str(e)}')
        
        return redirect('enseignant_evaluer_cours', code_cours=code_cours, annee=annee)
    except Enseignant.DoesNotExist:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')


@login_required
def enseignant_profil(request):
    """Profil de l'enseignant"""
    enseignant = get_simulated_enseignant(request)
    if not enseignant:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')
    try:
        # Recharger avec les relations
        enseignant = Enseignant.objects.select_related('grade', 'fonction', 'categorie', 'code_dpt', 'code_section').get(matricule_en=enseignant.matricule_en)
        context = {'enseignant': enseignant}
        return render(request, 'enseignant/profil.html', context)
    except Enseignant.DoesNotExist:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')


@login_required
def enseignant_mes_cours(request):
    """Liste des cours attribués à l'enseignant"""
    enseignant = get_simulated_enseignant(request)
    if not enseignant:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')
    try:
        from .models import UE, EC
        from reglage.models import AnneeAcademique
        
        # Récupérer les attributions de l'enseignant
        attributions_raw = Attribution.objects.filter(matricule_en=enseignant).select_related('type_charge')
        
        # Liste des années académiques pour le filtre
        annees = AnneeAcademique.objects.all().order_by('-code_anac')
        
        # Filtre par année académique (par défaut : la dernière année)
        annee_filter = request.GET.get('annee', '')
        if not annee_filter and 'annee' not in request.GET and annees.exists():
            annee_filter = annees.first().code_anac
        if annee_filter:
            attributions_raw = attributions_raw.filter(annee_academique=annee_filter)
        
        # Enrichir avec les infos des cours
        attributions = []
        for attr in attributions_raw:
            cours_info = {
                'code_cours': attr.code_cours,
                'type_charge': attr.type_charge,
                'annee_academique': attr.annee_academique,
                'intitule': None,
                'credit': None,
                'classe': None,
                'semestre': None,
                'type_cours': None,
            }
            # Chercher dans UE
            try:
                ue = UE.objects.get(code_ue=attr.code_cours)
                cours_info['intitule'] = ue.intitule_ue
                cours_info['credit'] = ue.credit
                cours_info['classe'] = ue.classe
                cours_info['semestre'] = ue.semestre
                cours_info['type_cours'] = 'UE'
            except UE.DoesNotExist:
                # Chercher dans EC
                try:
                    ec = EC.objects.select_related('code_ue').get(code_ec=attr.code_cours)
                    cours_info['intitule'] = ec.intitule_ue
                    cours_info['credit'] = ec.credit
                    cours_info['classe'] = ec.classe
                    cours_info['semestre'] = ec.code_ue.semestre if ec.code_ue else None
                    cours_info['type_cours'] = 'EC'
                except EC.DoesNotExist:
                    pass
            attributions.append(cours_info)
        
        # Calculer le total de crédits
        total_credits = sum(attr['credit'] or 0 for attr in attributions)
        
        context = {
            'enseignant': enseignant,
            'attributions': attributions,
            'annees': annees,
            'annee_filter': annee_filter,
            'total_credits': total_credits,
        }
        return render(request, 'enseignant/mes_cours.html', context)
    except Enseignant.DoesNotExist:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')


@login_required
def enseignant_documents_cours(request, code_cours, annee):
    """Gestion des documents PDF d'un cours par l'enseignant"""
    enseignant = get_simulated_enseignant(request)
    if not enseignant:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')

    # Vérifier que l'enseignant est bien attributaire de ce cours
    attribution = Attribution.objects.filter(
        matricule_en=enseignant, code_cours=code_cours, annee_academique=annee
    ).first()
    if not attribution:
        messages.error(request, "Vous n'êtes pas attributaire de ce cours.")
        return redirect('enseignant_mes_cours')

    # Récupérer intitulé du cours
    from .models import UE, EC
    intitule = code_cours
    try:
        ue = UE.objects.get(code_ue=code_cours)
        intitule = ue.intitule_ue
    except UE.DoesNotExist:
        try:
            ec = EC.objects.get(code_ec=code_cours)
            intitule = ec.intitule_ue
        except EC.DoesNotExist:
            pass

    # Upload de document
    if request.method == 'POST':
        action = request.POST.get('action', 'upload')

        if action == 'upload':
            titre = request.POST.get('titre', '').strip()
            type_document = request.POST.get('type_document', 'COURS')
            fichier = request.FILES.get('fichier')

            if not titre or not fichier:
                messages.error(request, 'Veuillez remplir tous les champs.')
            elif not fichier.name.lower().endswith('.pdf'):
                messages.error(request, 'Seuls les fichiers PDF sont acceptés.')
            elif fichier.size > 10 * 1024 * 1024:  # 10 Mo max
                messages.error(request, 'Le fichier ne doit pas dépasser 10 Mo.')
            else:
                DocumentCours.objects.create(
                    enseignant=enseignant,
                    code_cours=code_cours,
                    annee_academique=annee,
                    titre=titre,
                    type_document=type_document,
                    fichier=fichier,
                )
                messages.success(request, f'Document « {titre} » ajouté avec succès.')
                return redirect('enseignant_documents_cours', code_cours=code_cours, annee=annee)

        elif action == 'delete':
            doc_id = request.POST.get('doc_id')
            doc = DocumentCours.objects.filter(id=doc_id, enseignant=enseignant, code_cours=code_cours).first()
            if doc:
                doc.fichier.delete(save=False)
                doc.delete()
                messages.success(request, 'Document supprimé.')
            else:
                messages.error(request, 'Document introuvable.')
            return redirect('enseignant_documents_cours', code_cours=code_cours, annee=annee)

    documents = DocumentCours.objects.filter(
        enseignant=enseignant, code_cours=code_cours, annee_academique=annee
    )

    context = {
        'enseignant': enseignant,
        'code_cours': code_cours,
        'annee': annee,
        'intitule': intitule,
        'documents': documents,
        'type_choices': DocumentCours.TYPE_DOCUMENT_CHOICES,
    }
    return render(request, 'enseignant/documents_cours.html', context)


@login_required
def etudiant_telecharger_document(request, doc_id):
    """Télécharger un document cours (étudiant)"""
    etudiant = get_simulated_etudiant(request)
    if not etudiant:
        messages.error(request, 'Profil étudiant non trouvé.')
        return redirect('home')

    doc = get_object_or_404(DocumentCours, id=doc_id)

    # Vérifier que l'étudiant est inscrit dans une classe qui a ce cours
    inscriptions = Inscription.objects.filter(matricule_etudiant=etudiant).values_list('code_classe__code_classe', flat=True)
    cours_classes = set()
    from .models import UE, EC
    for ue in UE.objects.filter(code_ue=doc.code_cours):
        if ue.classe:
            cours_classes.add(ue.classe.code_classe if hasattr(ue.classe, 'code_classe') else str(ue.classe))
    for ec in EC.objects.filter(code_ec=doc.code_cours):
        if ec.classe:
            cours_classes.add(ec.classe.code_classe if hasattr(ec.classe, 'code_classe') else str(ec.classe))

    if not cours_classes.intersection(set(inscriptions)):
        messages.error(request, "Vous n'avez pas accès à ce document.")
        return redirect('etudiant_mes_cours')

    import mimetypes
    content_type = mimetypes.guess_type(doc.fichier.name)[0] or 'application/pdf'
    response = HttpResponse(doc.fichier.read(), content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{doc.fichier.name.split("/")[-1]}"'
    return response


@login_required
def enseignant_commentaires(request):
    """Gestion des commentaires de l'enseignant"""
    enseignant = get_simulated_enseignant(request)
    if not enseignant:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')
    try:
        from reglage.models import AnneeAcademique
        
        # Traitement POST (ajouter / modifier / supprimer)
        if request.method == 'POST':
            action = request.POST.get('action', 'add')
            
            if action == 'add':
                etudiant_mat = request.POST.get('etudiant')
                cours_code = request.POST.get('cours')
                contenu = request.POST.get('commentaire', '').strip()
                annee_com = request.POST.get('annee_commentaire', '')
                
                if etudiant_mat and cours_code and contenu and annee_com:
                    try:
                        etudiant = Etudiant.objects.get(matricule_et=etudiant_mat)
                        type_cours = 'UE' if UE.objects.filter(code_ue=cours_code).exists() else 'EC'
                        CommentaireCours.objects.create(
                            etudiant=etudiant,
                            annee_academique=annee_com,
                            type_cours=type_cours,
                            code_cours=cours_code,
                            contenu=contenu,
                        )
                        messages.success(request, 'Commentaire envoyé avec succès !')
                    except Etudiant.DoesNotExist:
                        messages.error(request, 'Étudiant non trouvé.')
                else:
                    messages.error(request, 'Veuillez remplir tous les champs.')
            
            elif action == 'edit':
                comm_id = request.POST.get('comm_id')
                contenu = request.POST.get('commentaire', '').strip()
                if comm_id and contenu:
                    try:
                        comm = CommentaireCours.objects.get(id=comm_id)
                        comm.contenu = contenu
                        comm.save()
                        messages.success(request, 'Commentaire modifié avec succès !')
                    except CommentaireCours.DoesNotExist:
                        messages.error(request, 'Commentaire non trouvé.')
            
            elif action == 'delete':
                comm_id = request.POST.get('comm_id')
                if comm_id:
                    try:
                        comm = CommentaireCours.objects.get(id=comm_id)
                        comm.delete()
                        messages.success(request, 'Commentaire supprimé avec succès !')
                    except CommentaireCours.DoesNotExist:
                        messages.error(request, 'Commentaire non trouvé.')
            
            return redirect('enseignant_commentaires')
        
        # Filtres
        cours_filter = request.GET.get('cours', '')
        annee_filter = request.GET.get('annee', '')
        
        # Liste des années académiques
        annees = AnneeAcademique.objects.all().order_by('-code_anac')
        
        # Récupérer les cours attribués à l'enseignant
        attributions = Attribution.objects.filter(matricule_en=enseignant)
        
        # Construire la liste des cours pour le filtre
        cours_list = []
        for attr in attributions:
            cours_info = {'code': attr.code_cours, 'intitule': attr.code_cours, 'classe': None}
            try:
                ue = UE.objects.get(code_ue=attr.code_cours)
                cours_info['intitule'] = ue.intitule_ue
                cours_info['classe'] = ue.classe
            except UE.DoesNotExist:
                try:
                    ec = EC.objects.get(code_ec=attr.code_cours)
                    cours_info['intitule'] = ec.intitule_ue
                    cours_info['classe'] = ec.classe
                except EC.DoesNotExist:
                    pass
            cours_list.append(cours_info)
        
        # Filtrer par année si spécifié
        if annee_filter:
            attributions = attributions.filter(annee_academique=annee_filter)
        
        # Récupérer les classes des cours de l'enseignant
        classes_codes = set()
        for attr in attributions:
            # Si filtre par cours, ne prendre que ce cours
            if cours_filter and attr.code_cours != cours_filter:
                continue
            try:
                ue = UE.objects.get(code_ue=attr.code_cours)
                if ue.classe:
                    classes_codes.add(ue.classe)
            except UE.DoesNotExist:
                try:
                    ec = EC.objects.get(code_ec=attr.code_cours)
                    if ec.classe:
                        classes_codes.add(ec.classe)
                except EC.DoesNotExist:
                    pass
        
        # Récupérer les étudiants inscrits dans ces classes
        etudiants = Etudiant.objects.filter(
            inscription__code_classe__code_classe__in=classes_codes
        ).distinct().order_by('nom_complet')
        
        # Filtrer par année académique d'inscription
        if annee_filter:
            etudiants = etudiants.filter(inscription__annee_academique=annee_filter).distinct()
        
        cours_codes = [c['code'] for c in cours_list]

        commentaires_qs = CommentaireCours.objects.filter(code_cours__in=cours_codes)
        if annee_filter:
            commentaires_qs = commentaires_qs.filter(annee_academique=annee_filter)
        if cours_filter:
            commentaires_qs = commentaires_qs.filter(code_cours=cours_filter)
        commentaires = commentaires_qs.select_related('etudiant').order_by('-date_creation')
        
        context = {
            'enseignant': enseignant,
            'etudiants': etudiants,
            'commentaires': commentaires,
            'cours_list': cours_list,
            'annees': annees,
            'cours_filter': cours_filter,
            'annee_filter': annee_filter,
        }
        return render(request, 'enseignant/commentaires.html', context)
    except Enseignant.DoesNotExist:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')


@login_required
def enseignant_evaluations(request):
    """Liste des évaluations de l'enseignant classifiées par cours"""
    enseignant = get_simulated_enseignant(request)
    if not enseignant:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')
    try:
        pass  # enseignant déjà récupéré
        from reglage.models import AnneeAcademique
        
        # Traitement des actions POST (modifier / supprimer)
        if request.method == 'POST':
            action = request.POST.get('action')
            eval_id = request.POST.get('eval_id')
            if eval_id:
                try:
                    evaluation = Evaluation.objects.get(id_ev=eval_id)
                    # Vérifier si le cours est verrouillé (envoyé au jury)
                    code_cours_eval = evaluation.code_ue.code_ue if evaluation.code_ue else (evaluation.code_ec.code_ec if evaluation.code_ec else None)
                    attr_locked = Attribution.objects.filter(
                        matricule_en=enseignant,
                        code_cours=code_cours_eval,
                        annee_academique=evaluation.annee_academique,
                        envoye_au_jury=True
                    ).exists() if code_cours_eval else False
                    if attr_locked:
                        messages.error(request, 'Les notes de ce cours ont été envoyées au jury. Modification impossible.')
                    elif action == 'edit':
                        def parse_note(val):
                            if not val:
                                return None
                            return float(str(val).replace(',', '.'))
                        evaluation.cc = parse_note(request.POST.get('cc'))
                        evaluation.examen = parse_note(request.POST.get('examen'))
                        evaluation.save()
                        messages.success(request, 'Évaluation modifiée avec succès !')
                    elif action == 'delete':
                        evaluation.delete()
                        messages.success(request, 'Évaluation supprimée avec succès !')
                except Evaluation.DoesNotExist:
                    messages.error(request, 'Évaluation non trouvée.')
            return redirect('enseignant_evaluations')
        
        # Filtres
        annee_filter = request.GET.get('annee', '')
        cours_filter = request.GET.get('cours', '')
        
        # Liste des années académiques pour le filtre
        annees = AnneeAcademique.objects.all().order_by('-code_anac')
        
        # Récupérer les cours attribués à l'enseignant
        attributions = Attribution.objects.filter(matricule_en=enseignant).select_related('type_charge')
        
        # Construire la liste des cours pour le filtre
        cours_list = []
        for attr in attributions:
            cours_info = {'code': attr.code_cours, 'intitule': attr.code_cours}
            try:
                ue = UE.objects.get(code_ue=attr.code_cours)
                cours_info['intitule'] = ue.intitule_ue
            except UE.DoesNotExist:
                try:
                    ec = EC.objects.get(code_ec=attr.code_cours)
                    cours_info['intitule'] = ec.intitule_ue
                except EC.DoesNotExist:
                    pass
            if cours_info not in cours_list:
                cours_list.append(cours_info)
        
        # Filtrer par année si spécifié
        if annee_filter:
            attributions = attributions.filter(annee_academique=annee_filter)
        
        # Filtrer par cours si spécifié
        if cours_filter:
            attributions = attributions.filter(code_cours=cours_filter)
        
        # Regrouper les évaluations par cours
        cours_evaluations = []
        total_evaluations = 0
        total_validees = 0
        
        for attr in attributions:
            code_cours = attr.code_cours
            
            # Récupérer les infos du cours
            cours_info = {'code': code_cours, 'intitule': code_cours, 'type': None, 'classe': None}
            try:
                ue = UE.objects.get(code_ue=code_cours)
                cours_info['intitule'] = ue.intitule_ue
                cours_info['type'] = 'UE'
                cours_info['classe'] = ue.classe
            except UE.DoesNotExist:
                try:
                    ec = EC.objects.get(code_ec=code_cours)
                    cours_info['intitule'] = ec.intitule_ue
                    cours_info['type'] = 'EC'
                    cours_info['classe'] = ec.classe
                except EC.DoesNotExist:
                    pass
            
            # Récupérer les évaluations pour ce cours
            if cours_info['type'] == 'UE':
                evals = Evaluation.objects.filter(code_ue__code_ue=code_cours).select_related('matricule_etudiant')
            elif cours_info['type'] == 'EC':
                evals = Evaluation.objects.filter(code_ec__code_ec=code_cours).select_related('matricule_etudiant')
            else:
                evals = Evaluation.objects.none()
            
            nb_evals = evals.count()
            # Compter les validés où CC + Examen >= 10 et calculer la moyenne
            nb_validees = 0
            total_notes = 0
            count_notes = 0
            for ev in evals:
                if ev.cc is not None and ev.examen is not None:
                    note_finale = ev.cc + ev.examen
                    total_notes += note_finale
                    count_notes += 1
                    if note_finale >= 10:
                        nb_validees += 1
            moyenne = round(total_notes / count_notes, 2) if count_notes > 0 else 0
            
            total_evaluations += nb_evals
            total_validees += nb_validees
            
            if nb_evals > 0:
                cours_evaluations.append({
                    'cours': cours_info,
                    'annee': attr.annee_academique,
                    'evaluations': evals,
                    'nb_evals': nb_evals,
                    'nb_validees': nb_validees,
                    'moyenne': round(moyenne, 2) if moyenne else 0,
                    'envoye_au_jury': attr.envoye_au_jury,
                    'date_envoi_jury': attr.date_envoi_jury,
                })
        
        context = {
            'enseignant': enseignant,
            'cours_evaluations': cours_evaluations,
            'total_evaluations': total_evaluations,
            'total_validees': total_validees,
            'annees': annees,
            'annee_filter': annee_filter,
            'cours_list': cours_list,
            'cours_filter': cours_filter,
        }
        return render(request, 'enseignant/evaluations.html', context)
    except Enseignant.DoesNotExist:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')


# ========== VUES JURY ==========

def _format_evaluation_for_display(eval_obj):
    cc_val = getattr(eval_obj, 'cc', None)
    examen_val = getattr(eval_obj, 'examen', None)
    ratt_val = getattr(eval_obj, 'rattrapage', None)

    cc_val = float(cc_val or 0)
    examen_val = float(examen_val or 0)

    nf = None
    try:
        nf = eval_obj.calculer_note_finale()
    except Exception:
        nf = None
    note_finale_num = float(nf) if nf is not None else (cc_val + examen_val)

    return {
        'cc': f"{cc_val:.1f}",
        'examen': f"{examen_val:.1f}",
        'rattrapage': f"{float(ratt_val):.1f}" if ratt_val is not None else '-',
        'note_finale': f"{note_finale_num:.1f}",
        'note_finale_num': float(note_finale_num),
    }

@login_required
def jury_dashboard(request):
    """Tableau de bord jury"""
    # Si admin sans simulation active, afficher la liste pour sélection
    if request.user.is_staff and 'simulated_jury' not in request.session:
        jurys = Jury.objects.select_related('code_classe').all()[:50]
        return render(request, 'admin/select_jury.html', {'jurys': jurys})
    
    # Récupérer le jury (simulé ou connecté)
    jury = get_simulated_jury(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')

    classe = jury.code_classe
    if not classe:
        messages.error(request, 'Classe introuvable pour ce jury.')
        return redirect('home')

    # Année académique du jury (éviter doublons si des doublants existent)
    annee_code = jury.annee_academique
    if not annee_code:
        from reglage.models import AnneeAcademique
        annee_active = AnneeAcademique.get_annee_en_cours()
        annee_code = annee_active.code_anac if annee_active else None

    # Récupérer les étudiants de la classe pour l'année du jury
    ins_qs = Inscription.objects.filter(code_classe=classe).select_related('matricule_etudiant', 'cohorte')
    if annee_code:
        ins_qs = ins_qs.filter(annee_academique=annee_code)
    inscriptions = ins_qs
    
    # Récupérer les cours évalués pour cette classe (via Attribution)
    ue_codes = list(UE.objects.filter(classe__code_classe=classe.code_classe).values_list('code_ue', flat=True))
    ec_codes = list(EC.objects.filter(classe__code_classe=classe.code_classe).values_list('code_ec', flat=True))
    
    # Identifier les UE qui ont des EC (UE parentes) et les exclure
    ue_parents_avec_ec = set()
    for ec_code in ec_codes:
        try:
            ec = EC.objects.get(code_ec=ec_code)
            if ec.code_ue:
                ue_parents_avec_ec.add(ec.code_ue.code_ue)
        except EC.DoesNotExist:
            continue
    
    # Filtrer les UE pour exclure celles qui ont des EC
    ue_codes_filtrees = [ue_code for ue_code in ue_codes if ue_code not in ue_parents_avec_ec]
    
    cours_codes = ue_codes_filtrees + ec_codes
    attributions = Attribution.objects.filter(code_cours__in=cours_codes)
    
    # Liste des cours (UE/EC) pour cette classe
    cours_map = {}
    for attr in attributions:
        cours_obj = attr.get_cours_object()
        cours_type = attr.get_type_cours()
        if not (cours_obj and cours_type):
            continue

        semestre = None
        if cours_type == 'UE':
            semestre = getattr(cours_obj, 'semestre', None)
        elif cours_type == 'EC':
            semestre = getattr(getattr(cours_obj, 'code_ue', None), 'semestre', None)

        cours_map[attr.code_cours] = {
            'code': attr.code_cours,
            'intitule': getattr(cours_obj, 'intitule_ue', attr.code_cours),
            'type': cours_type,
            'semestre': semestre,
        }

    cours_list = list(cours_map.values())
    cours_list.sort(key=lambda c: ((c['semestre'] is None), c['semestre'] or 0, c['code']))

    # Marquer le début de chaque semestre (pour afficher une séparation visuelle dans le tableau)
    _prev_semestre = object()
    for c in cours_list:
        c['semester_start'] = (c.get('semestre') != _prev_semestre)
        _prev_semestre = c.get('semestre')

    # Total des crédits par semestre (basé uniquement sur les UE pour éviter le double comptage avec les EC)
    credits_par_semestre = {
        row['semestre']: (row['total'] or 0)
        for row in UE.objects.filter(classe__code_classe=classe.code_classe)
        .values('semestre')
        .annotate(total=Sum('credit'))
    }

    # Calculer les groupes par semestre pour l'en-tête (4 colonnes par cours: CC/Exam/Ratt/Total)
    semestres = []
    current_sem = None
    current_count = 0
    for cours in cours_list:
        sem = cours.get('semestre')
        if current_sem is None:
            current_sem = sem
            current_count = 1
            continue
        if sem == current_sem:
            current_count += 1
            continue
        semestres.append({
            'semestre': current_sem,
            'colspan': current_count * 4,
            'credits': credits_par_semestre.get(current_sem, 0),
        })
        current_sem = sem
        current_count = 1
    if current_sem is not None:
        semestres.append({
            'semestre': current_sem,
            'colspan': current_count * 4,
            'credits': credits_par_semestre.get(current_sem, 0),
        })

    total_credits_semestres = sum((s.get('credits') or 0) for s in semestres)
    
    # Récupérer les évaluations pour chaque étudiant
    etudiants_data = []
    for inscription in inscriptions:
        etudiant = inscription.matricule_etudiant
        notes = {}
        total = 0
        count = 0
        
        for cours in cours_list:
            eval_filter = {'matricule_etudiant': etudiant}
            if annee_code:
                eval_filter['annee_academique'] = annee_code
                eval_filter['code_classe'] = classe
            if cours['type'] == 'EC':
                eval_filter['code_ec__code_ec'] = cours['code']
            else:
                eval_filter['code_ue__code_ue'] = cours['code']
            eval_obj = Evaluation.objects.filter(**eval_filter).first()
            
            if eval_obj:
                formatted = _format_evaluation_for_display(eval_obj)
                notes[cours['code']] = formatted
                total += float(formatted['note_finale_num'])
                count += 1
            else:
                notes[cours['code']] = None
        
        moyenne = round(total / count, 2) if count > 0 else 0
        
        # Construire la liste des notes dans l'ordre des cours
        notes_list = []
        for cours in cours_list:
            note_data = notes.get(cours['code'])
            notes_list.append({
                'note': note_data,
                'semester_start': cours.get('semester_start', False),
            })
        
        etudiants_data.append({
            'etudiant': etudiant,
            'notes_list': notes_list,
            'moyenne': moyenne,
            'decision': 'Admis' if moyenne >= 10 else 'Ajourné'
        })
    
    # Calculer les statistiques
    nb_admis = sum(1 for e in etudiants_data if e['moyenne'] >= 10)
    nb_ajournes = len(etudiants_data) - nb_admis
    
    # Calculer le pourcentage de réussite
    total_etudiants = len(etudiants_data)
    pourcentage_admis = round((nb_admis / total_etudiants * 100), 1) if total_etudiants > 0 else 0
    
    # Calculer les recours pour la classe du jury (même année)
    ins_recours = Inscription.objects.filter(code_classe=classe)
    if annee_code:
        ins_recours = ins_recours.filter(annee_academique=annee_code)
    etudiants_classe = [ins.matricule_etudiant for ins in ins_recours]
    recours_classe = Recours.objects.filter(etudiant__in=etudiants_classe)
    nb_recours_en_attente = recours_classe.filter(statut='EN_ATTENTE').count()
    
    # Récupérer les informations sur les délibérations
    # annee_code déjà définie en haut de la vue
    
    # Récupérer la dernière délibération pour cette classe
    dernieres_deliberations = Deliberation.objects.filter(
        code_classe=classe,
        annee_academique=annee_code
    ).order_by('-date_mise_a_jour')
    
    derniere_deliberation = dernieres_deliberations.first()
    
    # Vérifier si une délibération est nécessaire
    deliberations_necessaires = []
    if annee_code:
        # Pour L1: S1 et S2
        # Pour L2: S3 et S4
        # Pour L3: S5 et S6
        niveau_code = None
        if classe and getattr(classe, 'code_niveau', None):
            niveau_code = classe.code_niveau.code_niveau
        elif classe and getattr(classe, 'code_classe', None):
            code_classe = str(classe.code_classe)
            for prefix in ['L1', 'L2', 'L3', 'M1', 'M2']:
                if code_classe.startswith(prefix):
                    niveau_code = prefix
                    break
        
        if niveau_code == 'L1':
            deliberations_necessaires = ['S1', 'S2']
        elif niveau_code == 'L2':
            deliberations_necessaires = ['S3', 'S4']
        elif niveau_code == 'L3':
            deliberations_necessaires = ['S5', 'S6']
        elif niveau_code == 'M1':
            deliberations_necessaires = ['S7', 'S8']
        elif niveau_code == 'M2':
            deliberations_necessaires = ['S9', 'S10']
        
        # Vérifier quelles délibérations sont déjà faites
        deliberations_faites = list(dernieres_deliberations.values_list('type_deliberation', flat=True))
        deliberations_manquantes = [d for d in deliberations_necessaires if d not in deliberations_faites]
    else:
        deliberations_manquantes = []
    
    # Récupérer les noms complets avec grades
    pres = Enseignant.objects.filter(matricule_en=jury.president).select_related('grade').first()
    pres_display = f"{pres.grade.code_grade} {pres.nom_complet}" if pres and pres.grade else (pres.nom_complet if pres else jury.president)
    
    sec = Enseignant.objects.filter(matricule_en=jury.secretaire).select_related('grade').first()
    sec_display = f"{sec.grade.code_grade} {sec.nom_complet}" if sec and sec.grade else (sec.nom_complet if sec else jury.secretaire)
    
    membre_display = '-'
    if jury.membre:
        membre = Enseignant.objects.filter(matricule_en=jury.membre).select_related('grade').first()
        membre_display = f"{membre.grade.code_grade} {membre.nom_complet}" if membre and membre.grade else (membre.nom_complet if membre else jury.membre)
    
    context = {
        'jury': jury,
        'classe': classe,
        'inscriptions': inscriptions,
        'is_president': request.user.username.startswith('jury_pres_') if not request.user.is_staff else False,
        'is_secretaire': request.user.username.startswith('jury_sec_') if not request.user.is_staff else False,
        'president_display': pres_display,
        'secretaire_display': sec_display,
        'membre_display': membre_display,
        'cours_list': cours_list,
        'semestres': semestres,
        'etudiants_data': etudiants_data,
        'nb_admis': nb_admis,
        'nb_ajournes': nb_ajournes,
        'nb_recours_en_attente': nb_recours_en_attente,
        'pourcentage_admis': pourcentage_admis,
        'total_credits_semestres': total_credits_semestres,
        'derniere_deliberation': derniere_deliberation,
        'deliberations_manquantes': deliberations_manquantes,
        'annee_academique': annee_code,
    }
    
    # État actuel du rattrapage/rachat pour cette classe/année
    param_eval = ParametreEvaluation.objects.filter(
        code_classe=classe,
        annee_academique=annee_code
    ).first()
    context['rattrapage_actif'] = param_eval.rattrapage_actif if param_eval else False
    context['rachat_actif'] = param_eval.rachat_actif if param_eval else False
    
    return render(request, 'jury/dashboard.html', context)


@login_required
def jury_toggle_parametre(request):
    """Activer/désactiver le rattrapage ou le rachat pour la classe du jury"""
    if request.method != 'POST':
        return redirect('jury_dashboard')
    
    jury = get_simulated_jury(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')
    
    classe = jury.code_classe
    if not classe:
        messages.error(request, 'Classe introuvable.')
        return redirect('jury_dashboard')
    
    from reglage.models import AnneeAcademique
    annee_active = AnneeAcademique.get_annee_en_cours()
    annee_code = annee_active.code_anac if annee_active else None
    if not annee_code:
        annee_code = Inscription.objects.filter(code_classe=classe).order_by('-annee_academique').values_list('annee_academique', flat=True).first()
    
    if not annee_code:
        messages.error(request, 'Année académique introuvable.')
        return redirect('jury_dashboard')
    
    param_type = request.POST.get('param_type')  # 'rattrapage' ou 'rachat'
    
    param_eval, created = ParametreEvaluation.objects.get_or_create(
        code_classe=classe,
        annee_academique=annee_code,
        defaults={'active_par': request.user}
    )
    
    from django.utils import timezone as tz
    if param_type == 'rattrapage':
        param_eval.rattrapage_actif = not param_eval.rattrapage_actif
        if param_eval.rattrapage_actif:
            param_eval.date_activation_rattrapage = tz.now()
        etat = 'activé' if param_eval.rattrapage_actif else 'désactivé'
        messages.success(request, f'Rattrapage {etat} pour {classe}.')
    elif param_type == 'rachat':
        param_eval.rachat_actif = not param_eval.rachat_actif
        if param_eval.rachat_actif:
            param_eval.date_activation_rachat = tz.now()
        etat = 'activé' if param_eval.rachat_actif else 'désactivé'
        messages.success(request, f'Rachat {etat} pour {classe}.')
    else:
        messages.error(request, 'Paramètre inconnu.')
        return redirect('jury_dashboard')
    
    param_eval.active_par = request.user
    param_eval.save()
    return redirect('jury_dashboard')


def get_jury_for_user(request):
    """Helper pour trouver le jury associé à l'utilisateur ou simulé par admin"""
    # Vérifier simulation admin
    if request.user.is_staff and 'simulated_jury' in request.session:
        return Jury.objects.select_related('code_classe').filter(code_jury=request.session['simulated_jury']).first()
    
    # Sinon, chercher le jury normal (même logique que get_simulated_jury)
    user = request.user
    jury = Jury.objects.filter(id_lgn=user).first()
    if not jury:
        username = user.username
        classe_code = _extract_jury_classe_code(username)
        if classe_code:
            jury = Jury.objects.filter(code_classe__code_classe=classe_code).order_by('-annee_academique').first()
    return jury


def _calculer_resultats_semestre(etudiant, evaluations, semestre, classe, _weighted_average, _apply_compensation_pairs_1_to_1):
    """Calculer les résultats pour un semestre spécifique avec compensations intra-semestre"""
    from django.db.models import Q, Sum
    
    eval_by_ec = {}
    eval_by_ue = {}
    for ev in evaluations:
        if ev.code_ec_id:
            eval_by_ec[ev.code_ec_id] = ev
        if ev.code_ue_id:
            eval_by_ue.setdefault(ev.code_ue_id, ev)

    # Récupérer les UE du semestre
    code_classe_str = classe.code_classe if classe else None
    ues_sem = list(
        UE.objects.filter(
            classe__code_classe=code_classe_str,
            semestre=semestre
        ).order_by('code_ue')
    )

    ecs_sem = []
    if ues_sem:
        ecs_sem = list(EC.objects.filter(code_ue__in=ues_sem).select_related('code_ue'))
    ec_by_ue = {}
    for ec in ecs_sem:
        ec_by_ue.setdefault(ec.code_ue_id, []).append(ec)

    # Calcul des notes UE et compensations intra-UE au niveau EC
    ue_notes = {}
    ec_compensated_ids = set()
    ec_comp_pairs = []

    for ue in ues_sem:
        ecs = ec_by_ue.get(ue.code_ue, [])
        if ecs:
            ec_notes = []
            all_present = True
            for ec in ecs:
                ev = eval_by_ec.get(ec.code_ec)
                note = ev.calculer_note_finale() if ev else None
                if note is None:
                    all_present = False
                    break
                ec_notes.append((float(note), float(ec.credit)))

            if all_present and ec_notes:
                ue_notes[ue.code_ue] = _weighted_average(ec_notes)
                # Compensation intra-UE au niveau EC (1-to-1)
                fails = []
                donors = []
                for (note, _), ec in zip(ec_notes, ecs):
                    if 8 <= note <= 9:
                        fails.append((ec.code_ec, note))
                    elif note > 10:
                        donors.append((ec.code_ec, note))
                comp_ids, comp_pairs = _apply_compensation_pairs_1_to_1(fails, donors)
                ec_compensated_ids |= comp_ids
                for p in comp_pairs:
                    ec_comp_pairs.append({'ue': ue.code_ue, 'from': p['from'], 'to': p['to']})
            else:
                ue_notes[ue.code_ue] = None
        else:
            ev_ue = eval_by_ue.get(ue.code_ue)
            ue_notes[ue.code_ue] = ev_ue.calculer_note_finale() if ev_ue else None

    # Moyennes par catégorie (pondérées par crédit UE)
    cat_values = {'A': [], 'B': []}
    for ue in ues_sem:
        note_ue = ue_notes.get(ue.code_ue)
        if note_ue is None:
            continue
        if ue.categorie in cat_values:
            cat_values[ue.categorie].append((note_ue, ue.credit))

    moyenne_cat_a = _weighted_average(cat_values['A'])
    moyenne_cat_b = _weighted_average(cat_values['B'])

    # Compensation intra-catégorie au niveau UE (condition: moyenne catégorie >= 10)
    ue_compensated_ids = set()
    ue_comp_pairs = []
    for cat, moyenne_cat in [('A', moyenne_cat_a), ('B', moyenne_cat_b)]:
        if moyenne_cat is None or moyenne_cat < 10:
            continue
        fails_ue = []
        donors_ue = []
        for ue in ues_sem:
            if ue.categorie != cat:
                continue
            note_ue = ue_notes.get(ue.code_ue)
            if note_ue is None:
                continue
            if 8 <= float(note_ue) <= 9:
                fails_ue.append((ue.code_ue, float(note_ue)))
            elif float(note_ue) > 10:
                donors_ue.append((ue.code_ue, float(note_ue)))

        comp_ids, comp_pairs = _apply_compensation_pairs_1_to_1(fails_ue, donors_ue)
        ue_compensated_ids |= comp_ids
        for p in comp_pairs:
            ue_comp_pairs.append({'categorie': cat, 'from': p['from'], 'to': p['to']})

    # Moyenne globale du semestre (pondérée par crédits UE)
    moyenne_semestre = _weighted_average(
        [
            (ue_notes.get(ue.code_ue), ue.credit)
            for ue in ues_sem
            if ue_notes.get(ue.code_ue) is not None
        ]
    )
    
    if moyenne_semestre is None:
        return None
        
    return {
        'etudiant': etudiant,  # Ajouter l'objet étudiant pour éviter l'erreur
        'semestre': semestre,
        'ue_notes': ue_notes,
        'ues_sem': ues_sem,
        'moyenne_cat_a': moyenne_cat_a,
        'moyenne_cat_b': moyenne_cat_b,
        'moyenne_semestre': moyenne_semestre,
        'ec_compensated_ids': ec_compensated_ids,
        'ue_compensated_ids': ue_compensated_ids,
        'ec_comp_pairs': ec_comp_pairs,
        'ue_comp_pairs': ue_comp_pairs,
    }


def _appliquer_compensations_annuelles(resultats_s1, resultats_s2, _weighted_average, _apply_compensation_pairs_1_to_1):
    """Appliquer les compensations entre semestres pour la délibération annuelle
    IMPORTANT: La délibération annuelle peut seulement améliorer les statuts, jamais les détériorer
    """
    # Fusionner les UE des deux semestres
    toutes_ues = resultats_s1['ues_sem'] + resultats_s2['ues_sem']
    
    # Fusionner les notes UE
    ue_notes_annuel = {}
    ue_notes_annuel.update(resultats_s1['ue_notes'])
    ue_notes_annuel.update(resultats_s2['ue_notes'])
    
    # Calculer les moyennes annuelles par catégorie
    cat_values_annuel = {'A': [], 'B': []}
    for ue in toutes_ues:
        note_ue = ue_notes_annuel.get(ue.code_ue)
        if note_ue is None:
            continue
        if ue.categorie in cat_values_annuel:
            cat_values_annuel[ue.categorie].append((note_ue, ue.credit))
    
    moyenne_cat_a_annuel = _weighted_average(cat_values_annuel['A'])
    moyenne_cat_b_annuel = _weighted_average(cat_values_annuel['B'])
    
    # Compensations entre semestres par catégorie
    # RÈGLE: Une UE/EC déjà validé(e) ne peut pas changer de statut (ne peut pas être "désutilisée")
    # Mais une UE/EC non validé(e) peut devenir validé(e) par compensation annuelle
    fails_annuel = {'A': [], 'B': []}
    donors_annuel = {'A': [], 'B': []}
    
    # Récupérer les UE déjà validées (>=10) ou déjà compensées intra-semestre
    ue_valides_s1 = set()
    ue_valides_s2 = set()
    
    for ue in resultats_s1['ues_sem']:
        note_ue = resultats_s1['ue_notes'].get(ue.code_ue)
        if note_ue is not None and float(note_ue) >= 10:
            ue_valides_s1.add(ue.code_ue)
        elif ue.code_ue in resultats_s1['ue_compensated_ids']:
            ue_valides_s1.add(ue.code_ue)  # Déjà validée par compensation
    
    for ue in resultats_s2['ues_sem']:
        note_ue = resultats_s2['ue_notes'].get(ue.code_ue)
        if note_ue is not None and float(note_ue) >= 10:
            ue_valides_s2.add(ue.code_ue)
        elif ue.code_ue in resultats_s2['ue_compensated_ids']:
            ue_valides_s2.add(ue.code_ue)  # Déjà validée par compensation
    
    for cat in ['A', 'B']:
        # UE en échec dans S1 mais pouvant être compensées par S2
        for ue in resultats_s1['ues_sem']:
            if ue.categorie != cat:
                continue
            note_ue = resultats_s1['ue_notes'].get(ue.code_ue)
            # Seulement les UE non validées (échec ou <10) et non déjà compensées
            if (note_ue is not None and float(note_ue) < 10 and 
                ue.code_ue not in ue_valides_s1 and 
                ue.code_ue not in resultats_s1['ue_compensated_ids']):
                fails_annuel[cat].append((f"S1_{ue.code_ue}", float(note_ue)))
        
        # UE donneuses dans S2 (uniquement celles déjà validées naturellement)
        # IMPORTANT: Ne pas utiliser les UE déjà compensées intra-semestre pour préserver leurs bénéficiaires
        for ue in resultats_s2['ues_sem']:
            if ue.categorie != cat:
                continue
            note_ue = resultats_s2['ue_notes'].get(ue.code_ue)
            # Seulement les UE naturellement validées (>=10) et non déjà utilisées en compensation intra-semestre
            if (note_ue is not None and float(note_ue) >= 10 and 
                ue.code_ue not in resultats_s2['ue_compensated_ids']):
                donors_annuel[cat].append((f"S2_{ue.code_ue}", float(note_ue)))
        
        # UE en échec dans S2 mais pouvant être compensées par S1
        for ue in resultats_s2['ues_sem']:
            if ue.categorie != cat:
                continue
            note_ue = resultats_s2['ue_notes'].get(ue.code_ue)
            # Seulement les UE non validées (échec ou <10) et non déjà compensées
            if (note_ue is not None and float(note_ue) < 10 and 
                ue.code_ue not in ue_valides_s2 and 
                ue.code_ue not in resultats_s2['ue_compensated_ids']):
                fails_annuel[cat].append((f"S2_{ue.code_ue}", float(note_ue)))
        
        # UE donneuses dans S1 (uniquement celles déjà validées naturellement)
        for ue in resultats_s1['ues_sem']:
            if ue.categorie != cat:
                continue
            note_ue = resultats_s1['ue_notes'].get(ue.code_ue)
            # Seulement les UE naturellement validées (>=10) et non déjà utilisées en compensation intra-semestre
            if (note_ue is not None and float(note_ue) >= 10 and 
                ue.code_ue not in resultats_s1['ue_compensated_ids']):
                donors_annuel[cat].append((f"S1_{ue.code_ue}", float(note_ue)))
    
    # Appliquer les compensations entre semestres
    ue_compensated_annuel_ids = set()
    ue_comp_annuel_pairs = []
    
    for cat in ['A', 'B']:
        # Condition: moyenne annuelle de la catégorie >= 10
        moyenne_cat_annuel = moyenne_cat_a_annuel if cat == 'A' else moyenne_cat_b_annuel
        if moyenne_cat_annuel is None or moyenne_cat_annuel < 10:
            continue
            
        comp_ids, comp_pairs = _apply_compensation_pairs_1_to_1(fails_annuel[cat], donors_annuel[cat])
        ue_compensated_annuel_ids |= comp_ids
        for p in comp_pairs:
            ue_comp_annuel_pairs.append({'categorie': cat, 'from': p['from'], 'to': p['to']})
    
    # Fusionner tous les IDs compensés
    tous_ec_compensated = resultats_s1['ec_compensated_ids'] | resultats_s2['ec_compensated_ids']
    
    # IMPORTANT: Conserver TOUTES les UE déjà validées (naturellement ou par compensation)
    # et ajouter seulement celles nouvellement validées par compensation annuelle
    tous_ue_compensated = resultats_s1['ue_compensated_ids'] | resultats_s2['ue_compensated_ids']
    
    # Ajouter les UE nouvellement compensées entre semestres
    for comp_id in ue_compensated_annuel_ids:
        # Extraire le code UE réel (enlever le préfixe S1_ ou S2_)
        if comp_id.startswith('S1_') or comp_id.startswith('S2_'):
            ue_code = comp_id.split('_', 1)[1]
            tous_ue_compensated.add(ue_code)
    
    # Calculer la moyenne annuelle finale
    moyenne_annuelle = _weighted_average(
        [
            (ue_notes_annuel.get(ue.code_ue), ue.credit)
            for ue in toutes_ues
            if ue_notes_annuel.get(ue.code_ue) is not None
        ]
    )
    
    # Déterminer la mention annuelle
    def _mention_for_note(note):
        if note is None:
            return None
        n = float(note)
        if n >= 18:
            return 'Excellent (A)'
        if n >= 16:
            return 'Très bien (B)'
        if n >= 14:
            return 'Bien (C)'
        if n >= 12:
            return 'Assez Bien (D)'
        if n >= 10:
            return 'Passable (E)'
        if n >= 8:
            return 'Insuffisant (F)'
        return 'Insatisfaisant (G) : un travail considérable est nécessaire pour réussir'
    
    decision_annuelle = _mention_for_note(moyenne_annuelle)
    
    return {
        'etudiant': resultats_s1.get('etudiant'),  # Les deux ont le même étudiant
        'moyenne': round(moyenne_annuelle, 2) if moyenne_annuelle is not None else None,
        'decision': decision_annuelle,
        'moyenne_cat_a': moyenne_cat_a_annuel,
        'moyenne_cat_b': moyenne_cat_b_annuel,
        'ec_compensated_ids': list(tous_ec_compensated),
        'ue_compensated_ids': list(tous_ue_compensated),
        'ec_comp_pairs': resultats_s1['ec_comp_pairs'] + resultats_s2['ec_comp_pairs'],
        'ue_comp_pairs': resultats_s1['ue_comp_pairs'] + resultats_s2['ue_comp_pairs'] + ue_comp_annuel_pairs,
        'compensations_annuelles': ue_comp_annuel_pairs,  # Pour suivi des compensations inter-semestres
        'ue_valides_s1': list(ue_valides_s1),  # Pour suivi/transparence
        'ue_valides_s2': list(ue_valides_s2),  # Pour suivi/transparence
    }


@login_required
def jury_annuler_deliberation(request):
    """Annuler/supprimer les délibérations"""
    if request.method != 'POST':
        return redirect('jury_deliberer')
    
    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')
    
    classe = jury.code_classe
    if not classe:
        messages.error(request, 'Classe non trouvée.')
        return redirect('jury_deliberer')
    
    # Récupérer l'année académique
    from reglage.models import AnneeAcademique
    annee_active = AnneeAcademique.get_annee_en_cours()
    annee_code = annee_active.code_anac if annee_active else None
    if not annee_code:
        annee_code = (
            Inscription.objects.filter(code_classe=classe)
            .values_list('annee_academique', flat=True)
            .order_by('-annee_academique')
            .first()
        )
    
    type_suppression = request.POST.get('type', '')
    
    # Construire le filtre de base
    base_filter = {
        'code_classe': classe,
        'annee_academique': annee_code
    }
    
    # Déterminer quelles délibérations supprimer
    if type_suppression == 'S1':
        count, _ = Deliberation.objects.filter(**base_filter, type_deliberation='S1').delete()
        messages.success(request, f'Délibérations du Semestre 1 supprimées ({count} enregistrements).')
    elif type_suppression == 'S2':
        count, _ = Deliberation.objects.filter(**base_filter, type_deliberation='S2').delete()
        messages.success(request, f'Délibérations du Semestre 2 supprimées ({count} enregistrements).')
    elif type_suppression == 'ANNEE':
        count, _ = Deliberation.objects.filter(**base_filter, type_deliberation='ANNEE').delete()
        messages.success(request, f'Délibérations Annuelles supprimées ({count} enregistrements).')
    elif type_suppression == 'TOUT':
        count, _ = Deliberation.objects.filter(**base_filter).delete()
        messages.success(request, f'Toutes les délibérations supprimées ({count} enregistrements).')
    else:
        messages.error(request, 'Type de suppression non reconnu.')
    
    return redirect('jury_deliberer')


@login_required
def jury_deliberer(request):
    """Délibération par le jury"""
    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')
    try:
        classe = jury.code_classe

        # Utiliser l'année académique du jury connecté par défaut
        annee_code = jury.annee_academique
        if not annee_code:
            from reglage.models import AnneeAcademique
            annee_active = AnneeAcademique.get_annee_en_cours()
            annee_code = annee_active.code_anac if annee_active else None
        if not annee_code and classe:
            annee_code = (
                Inscription.objects.filter(code_classe=classe)
                .values_list('annee_academique', flat=True)
                .order_by('-annee_academique')
                .first()
            )

        niveau_code = None
        if classe and getattr(classe, 'code_niveau', None):
            niveau_code = classe.code_niveau.code_niveau
        elif classe and getattr(classe, 'code_classe', None):
            code_classe = str(classe.code_classe)
            for prefix in ['L1', 'L2', 'L3', 'M1', 'M2']:
                if code_classe.startswith(prefix):
                    niveau_code = prefix
                    break

        niveau_to_semestres = {
            'L1': (1, 2),
            'L2': (3, 4),
            'L3': (5, 6),
            'M1': (7, 8),
            'M2': (9, 10),
        }
        semestres_niveau = niveau_to_semestres.get(niveau_code, None)

        selected_type = request.GET.get('type', 'ANNEE')
        if request.method == 'POST':
            selected_type = request.POST.get('type', 'ANNEE')

        # Semestres cibles pour la délibération actuelle
        semestres_cibles = None
        if semestres_niveau:
            sem1, sem2 = semestres_niveau
            if selected_type == 'S1':
                semestres_cibles = [sem1]
            elif selected_type == 'S2':
                semestres_cibles = [sem2]
            else:
                semestres_cibles = [sem1, sem2]

        # Stats: nombre de cours réellement évalués (présents dans Evaluation avec cc+examen) et crédits correspondants
        # Remarque: on calcule au niveau de la classe (tous étudiants), sur la sélection S1/S2/ANNEE.
        stats_semestres = []
        nb_cours_selection = 0
        credits_selection = 0

        if classe and semestres_niveau:
            evals_classe = Evaluation.objects.filter(
                matricule_etudiant__inscription__code_classe=classe,
                cc__isnull=False,
                examen__isnull=False,
            )

            if semestres_cibles:
                evals_classe = evals_classe.filter(
                    Q(code_ue__semestre__in=semestres_cibles) |
                    Q(code_ec__code_ue__semestre__in=semestres_cibles)
                )

            # Stats par semestre (toujours affichées pour les 2 semestres du niveau)
            for sem in semestres_niveau:
                evals_sem = Evaluation.objects.filter(
                    matricule_etudiant__inscription__code_classe=classe,
                    cc__isnull=False,
                    examen__isnull=False,
                ).filter(
                    Q(code_ue__semestre=sem) |
                    Q(code_ec__code_ue__semestre=sem)
                )

                ue_ids = list(evals_sem.filter(code_ue__isnull=False).values_list('code_ue', flat=True).distinct())
                ec_ids = list(evals_sem.filter(code_ec__isnull=False).values_list('code_ec', flat=True).distinct())

                credits_ue = UE.objects.filter(code_ue__in=ue_ids).aggregate(total=Sum('credit'))['total'] or 0
                credits_ec = EC.objects.filter(code_ec__in=ec_ids).aggregate(total=Sum('credit'))['total'] or 0

                stats_semestres.append({
                    'semestre': sem,
                    'nb_cours': len(ue_ids) + len(ec_ids),
                    'credits': (credits_ue + credits_ec),
                })

            ue_ids_sel = list(evals_classe.filter(code_ue__isnull=False).values_list('code_ue', flat=True).distinct())
            ec_ids_sel = list(evals_classe.filter(code_ec__isnull=False).values_list('code_ec', flat=True).distinct())
            nb_cours_selection = len(ue_ids_sel) + len(ec_ids_sel)

            credits_ue_sel = UE.objects.filter(code_ue__in=ue_ids_sel).aggregate(total=Sum('credit'))['total'] or 0
            credits_ec_sel = EC.objects.filter(code_ec__in=ec_ids_sel).aggregate(total=Sum('credit'))['total'] or 0
            credits_selection = credits_ue_sel + credits_ec_sel

        def _weighted_average(values_with_weights):
            total_w = 0
            total_v = 0
            for value, weight in values_with_weights:
                if value is None or weight is None:
                    continue
                total_w += float(weight)
                total_v += float(value) * float(weight)
            if total_w <= 0:
                return None
            return round(total_v / total_w, 2)

        def _apply_compensation_1_to_1(fails, donors):
            """
            fails: list of tuples (key, note) where note in [8,9]
            donors: list of tuples (key, note) where note > 10
            returns set of keys compensated in fails
            """
            fail_items = []
            for key, note in fails:
                deficit = 10 - float(note)
                fail_items.append((key, float(note), deficit))

            donor_items = []
            for key, note in donors:
                excess = float(note) - 10
                donor_items.append((key, float(note), excess))

            # Stratégie: traiter les plus gros déficits d'abord, et utiliser le plus petit excédent suffisant (1 pour 1)
            fail_items.sort(key=lambda x: x[2], reverse=True)
            donor_items.sort(key=lambda x: x[2])

            compensated = set()
            used_donors = set()
            for f_key, f_note, deficit in fail_items:
                donor_index = None
                for idx, (d_key, d_note, excess) in enumerate(donor_items):
                    if d_key in used_donors:
                        continue
                    if excess >= deficit:
                        donor_index = idx
                        used_donors.add(d_key)
                        compensated.add(f_key)
                        break
                if donor_index is None:
                    continue
            return compensated

        def _apply_compensation_pairs_1_to_1(fails, donors):
            """Retourne (set_compenses, pairs)

            pairs: list of dicts {"from": donor_key, "to": fail_key}
            """
            fail_items = []
            for key, note in fails:
                deficit = 10 - float(note)
                fail_items.append((key, float(note), deficit))

            donor_items = []
            for key, note in donors:
                excess = float(note) - 10
                donor_items.append((key, float(note), excess))

            fail_items.sort(key=lambda x: x[2], reverse=True)
            donor_items.sort(key=lambda x: x[2])

            compensated = set()
            pairs = []
            used_donors = set()
            for f_key, f_note, deficit in fail_items:
                for d_key, d_note, excess in donor_items:
                    if d_key in used_donors:
                        continue
                    if excess >= deficit:
                        used_donors.add(d_key)
                        compensated.add(f_key)
                        pairs.append({'from': d_key, 'to': f_key})
                        break
            return compensated, pairs

        def _mention_for_note(note):
            if note is None:
                return None
            n = float(note)
            if n >= 18:
                return 'Excellent (A)'
            if n >= 16:
                return 'Très bien (B)'
            if n >= 14:
                return 'Bien (C)'
            if n >= 12:
                return 'Assez Bien (D)'
            if n >= 10:
                return 'Passable (E)'
            if n >= 8:
                return 'Insuffisant (F)'
            return 'Insatisfaisant (G) : un travail considérable est nécessaire pour réussir'

        def _calculer_credits_capitalises(etudiant, classe, semestre, annee_academique):
            """Calcule les crédits capitalisés pour un étudiant selon le semestre"""
            from .models import Deliberation
            
            credits_capitalises = 0
            ues_comptees = set()  # Pour éviter de compter plusieurs fois la même UE
            
            # Si semestre est None, on calcule pour l'année entière (S1 + S2)
            if semestre is None:
                # Calculer pour S1 et S2
                for type_delib in ['S1', 'S2']:
                    delib_sem = Deliberation.objects.filter(
                        matricule_etudiant=etudiant,
                        code_classe=classe,
                        annee_academique=annee_academique,
                        type_deliberation=type_delib,
                        statut='VALIDE'
                    ).select_related('code_ue', 'code_ec')
                    
                    for delib in delib_sem:
                        # Priorité: si c'est une UE directe, on compte le crédit de l'UE
                        if delib.code_ue and not delib.code_ec:
                            ue_key = f"{type_delib}_{delib.code_ue.code_ue}"
                            if ue_key not in ues_comptees:
                                credits_capitalises += delib.code_ue.credit
                                ues_comptees.add(ue_key)
                        # Si c'est un EC, on compte le crédit de l'EC
                        elif delib.code_ec:
                            credits_capitalises += delib.code_ec.credit
            else:
                # Calculer pour un semestre spécifique
                delib_sem = Deliberation.objects.filter(
                    matricule_etudiant=etudiant,
                    code_classe=classe,
                    annee_academique=annee_academique,
                    semestre=semestre,
                    statut='VALIDE'
                ).select_related('code_ue', 'code_ec')
                
                for delib in delib_sem:
                    # Priorité: si c'est une UE directe, on compte le crédit de l'UE
                    if delib.code_ue and not delib.code_ec:
                        ue_key = delib.code_ue.code_ue
                        if ue_key not in ues_comptees:
                            credits_capitalises += delib.code_ue.credit
                            ues_comptees.add(ue_key)
                    # Si c'est un EC, on compte le crédit de l'EC
                    elif delib.code_ec:
                        credits_capitalises += delib.code_ec.credit
            
            return credits_capitalises

        # Récupérer les étudiants avec leurs moyennes
        inscriptions = Inscription.objects.filter(code_classe=classe, annee_academique=annee_code).select_related('matricule_etudiant', 'cohorte')
        
        # Construire les résultats pour le template en utilisant _jury_compute_delib_ues
        resultats = []
        for inscription in inscriptions:
            etudiant = inscription.matricule_etudiant
            
            # Calculer les stats pour S1, S2 et Annuel en utilisant la même fonction que jury_deliberations
            stats_s1 = _jury_compute_delib_ues(classe, etudiant, 'semestriel', 1, annee_code) if semestres_niveau and len(semestres_niveau) >= 1 else {'credits_valides': 0, 'moyenne': 0}
            stats_s2 = _jury_compute_delib_ues(classe, etudiant, 'semestriel', 2, annee_code) if semestres_niveau and len(semestres_niveau) >= 2 else {'credits_valides': 0, 'moyenne': 0}
            stats_annuel = _jury_compute_delib_ues(classe, etudiant, 'annuel', None, annee_code)
            
            # Extraire les données
            credits_s1 = stats_s1.get('credits_valides', 0)
            credits_s2 = stats_s2.get('credits_valides', 0)
            credits_annuel = stats_annuel.get('credits_valides', 0)
            
            # Décisions par type
            decision_s1 = stats_s1.get('decision_label', 'A déterminer')
            decision_s2 = stats_s2.get('decision_label', 'A déterminer')
            decision_annuel = stats_annuel.get('decision_label', 'A déterminer')
            
            # Moyenne et décision selon le type sélectionné
            # Détecter quels semestres ont réellement des données
            s1_has_data = stats_s1.get('credits_total', 0) > 0
            s2_has_data = stats_s2.get('credits_total', 0) > 0
            
            if selected_type == 'S1':
                moyenne = stats_s1.get('moyenne', 0)
                decision = decision_s1
            elif selected_type == 'S2':
                moyenne = stats_s2.get('moyenne', 0)
                decision = decision_s2
            else:
                # ANNEE: si un seul semestre a des données, utiliser sa décision
                if s1_has_data and not s2_has_data:
                    moyenne = stats_s1.get('moyenne', 0)
                    decision = decision_s1
                elif s2_has_data and not s1_has_data:
                    moyenne = stats_s2.get('moyenne', 0)
                    decision = decision_s2
                else:
                    # Les deux semestres ont des données (ou aucun) → décision annuelle
                    moyenne = stats_annuel.get('moyenne', 0)
                    decision = decision_annuel
            
            # Déterminer la mention
            mention = _mention_for_note(moyenne)
            
            resultats.append({
                'etudiant': etudiant,
                'moyenne': round(moyenne, 2),
                'decision': decision,
                'decision_s1': decision_s1,
                'decision_s2': decision_s2,
                'decision_annuel': decision_annuel,
                'credits_s1_capitalises': credits_s1,
                'credits_s2_capitalises': credits_s2,
                'credits_annuel_capitalises': credits_annuel,
                'mention': mention,
            })

        # Calculer les statistiques par décision pour S1, S2 et ANNEE
        stats_s1 = {'admis': 0, 'compensable': 0, 'ajournes': 0, 'defaillants': 0}
        stats_s2 = {'admis': 0, 'compensable': 0, 'ajournes': 0, 'defaillants': 0}
        stats_annuel = {'admis': 0, 'compensable': 0, 'ajournes': 0, 'defaillants': 0}
        
        # Calculer les stats basées sur les vraies décisions
        def _classify_decision(decision_str, stats_dict):
            """Classe une décision dans le bon compteur de stats."""
            if 'Admis' in decision_str and 'dette' not in decision_str:
                stats_dict['admis'] += 1
            elif 'Admis avec dette' in decision_str:
                stats_dict['compensable'] += 1
            elif 'Compensable' in decision_str:
                stats_dict['compensable'] += 1
            elif 'Ajourné' in decision_str:
                stats_dict['ajournes'] += 1
            else:
                stats_dict['defaillants'] += 1
        
        for r in resultats:
            _classify_decision(r.get('decision_s1', ''), stats_s1)
            _classify_decision(r.get('decision_s2', ''), stats_s2)
            _classify_decision(r.get('decision_annuel', ''), stats_annuel)

        if request.method == 'POST':
            semestre_val = None
            if selected_type in ('S1', 'S2') and semestres_cibles:
                semestre_val = int(semestres_cibles[0])

            # Créer les délibérations individuelles pour chaque étudiant
            if annee_code and classe:
                # Supprimer les anciennes délibérations pour ce type
                Deliberation.objects.filter(
                    code_classe=classe,
                    annee_academique=annee_code,
                    type_deliberation=selected_type,
                    semestre=semestre_val
                ).delete()
                
                # Créer les nouvelles délibérations avec les statuts calculés
                for r in resultats:
                    etudiant = r['etudiant']
                    
                    # Récupérer les évaluations de l'étudiant pour cette délibération
                    # IMPORTANT: filtrer par année et classe pour éviter les doublons
                    # après passage automatique (notes transférées aux doublants)
                    evaluations = Evaluation.objects.filter(
                        matricule_etudiant=etudiant,
                        annee_academique=annee_code,
                        code_classe=classe
                    )
                    if semestres_cibles:
                        evaluations = evaluations.filter(
                            Q(code_ue__semestre__in=semestres_cibles) |
                            Q(code_ec__code_ue__semestre__in=semestres_cibles)
                        )
                    
                    # Créer une délibération pour chaque évaluation
                    for eval in evaluations:
                        # Déterminer le statut en fonction des compensations
                        statut = 'NON_VALIDE'
                        note_finale = eval.calculer_note_finale()
                        
                        if note_finale is not None:
                            if note_finale >= 10:
                                statut = 'VALIDE'
                            elif note_finale >= 8:
                                # Vérifier si l'EC est compensé
                                if eval.code_ec_id and eval.code_ec_id in r.get('ec_compensated_ids', []):
                                    statut = 'VALIDE'
                                elif eval.code_ue_id and eval.code_ue_id in r.get('ue_compensated_ids', []):
                                    statut = 'VALIDE'
                        
                        # Créer la délibération
                        Deliberation.objects.create(
                            cc=eval.cc,
                            examen=eval.examen,
                            rattrapage=eval.rattrapage,
                            rachat=eval.rachat,
                            statut=statut,
                            code_ue=eval.code_ue,
                            matricule_etudiant=eval.matricule_etudiant,
                            code_ec=eval.code_ec,
                            type_deliberation=selected_type,
                            annee_academique=annee_code,
                            code_classe=classe,
                            semestre=semestre_val,
                            cree_par=request.user
                        )
                
                # Appliquer la compensation annuelle si c'est une délibération annuelle
                if selected_type == 'ANNEE':
                    compensations = Deliberation.appliquer_compensation_annuelle(classe, annee_code)
                    if compensations:
                        messages.info(request, f'{len(compensations)} cours compensés par la moyenne annuelle.')
                
                messages.success(request, f'Délibération effectuée avec succès! {len(resultats)} étudiants traités.')

            # Construire l'URL imprimable pour le bouton d'impression
            params_print = {
                'annee': annee_code,
            }
            if classe:
                params_print['classe'] = classe.code_classe
            if selected_type == 'ANNEE':
                params_print['type'] = 'annuel'
            else:
                params_print['type'] = 'semestriel'
                if semestre_val is not None:
                    params_print['semestre'] = semestre_val
            
            # Rester sur la même page avec le type sélectionné
            return redirect(f"{reverse('jury_deliberer')}?type={selected_type}")
        
        # Construire l'URL imprimable pour le bouton d'impression
        params_print = {'annee': annee_code}
        if classe:
            params_print['classe'] = classe.code_classe
        if selected_type == 'ANNEE':
            params_print['type'] = 'annuel'
        else:
            params_print['type'] = 'semestriel'
            if semestres_niveau:
                sem1, sem2 = semestres_niveau
                if selected_type == 'S1':
                    params_print['semestre'] = sem1
                elif selected_type == 'S2':
                    params_print['semestre'] = sem2
        url_imprimables = f"{reverse('jury_imprimables')}?{urlencode(params_print)}"

        # Vérifier les délibérations déjà effectuées
        delib_s1_faite = False
        delib_s2_faite = False
        delib_annee_faite = False
        if classe and annee_code and semestres_niveau:
            sem1, sem2 = semestres_niveau
            delib_s1_faite = Deliberation.objects.filter(
                code_classe=classe, annee_academique=annee_code, type_deliberation='S1'
            ).exists()
            delib_s2_faite = Deliberation.objects.filter(
                code_classe=classe, annee_academique=annee_code, type_deliberation='S2'
            ).exists()
            delib_annee_faite = Deliberation.objects.filter(
                code_classe=classe, annee_academique=annee_code, type_deliberation='ANNEE'
            ).exists()

        context = {
            'classe': classe,
            'annee_code': annee_code,
            'semestres_niveau': semestres_niveau,
            'selected_type': selected_type,
            'stats_semestres': stats_semestres,
            'nb_cours_selection': nb_cours_selection,
            'credits_selection': credits_selection,
            'resultats': resultats,
            'stats_s1': stats_s1,
            'stats_s2': stats_s2,
            'stats_annuel': stats_annuel,
            'url_imprimables': url_imprimables,
            'delib_s1_faite': delib_s1_faite,
            'delib_s2_faite': delib_s2_faite,
            'delib_annee_faite': delib_annee_faite,
        }
        return render(request, 'jury/deliberer.html', context)
    except Exception as e:
        messages.error(request, f'Erreur: {str(e)}')
        return redirect('home')


def _deliberation_exists_for_selection(classe_obj, annee, selected_type, selected_semestre, semestres_niveau):
    if not (classe_obj and annee):
        return False

    if selected_type == 'annuel':
        return Deliberation.objects.filter(
            code_classe=classe_obj,
            annee_academique=annee,
            type_deliberation='ANNEE',
            semestre__isnull=True,
        ).exists()

    if selected_type == 'semestriel':
        if not (selected_semestre and semestres_niveau):
            return False
        sem1, sem2 = semestres_niveau
        type_delib = 'S1' if str(selected_semestre) == str(sem1) else 'S2'
        return Deliberation.objects.filter(
            code_classe=classe_obj,
            annee_academique=annee,
            type_deliberation=type_delib,
            semestre=int(selected_semestre),
        ).exists()

    return False


def _require_deliberation_for_imprimable(request, jury, classe_obj, annee, selected_type, selected_semestre):
    niveau_code = None
    if classe_obj and getattr(classe_obj, 'code_niveau', None):
        niveau_code = classe_obj.code_niveau.code_niveau
    elif classe_obj and getattr(classe_obj, 'code_classe', None):
        code_classe = str(classe_obj.code_classe)
        for prefix in ['L1', 'L2', 'L3', 'M1', 'M2']:
            if code_classe.startswith(prefix):
                niveau_code = prefix
                break

    niveau_to_semestres = {
        'L1': (1, 2),
        'L2': (3, 4),
        'L3': (5, 6),
        'M1': (7, 8),
        'M2': (9, 10),
    }
    semestres_niveau = niveau_to_semestres.get(niveau_code, None)

    ok = _deliberation_exists_for_selection(
        classe_obj,
        annee,
        selected_type,
        selected_semestre,
        semestres_niveau,
    )
    if not ok:
        messages.error(request, "Aucune délibération n'a encore été effectuée pour cette sélection.")
        return False
    return True


@login_required
def jury_communique(request):
    """Publication des communiqués (date de délibération) par le jury"""
    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')

    try:
        from reglage.models import AnneeAcademique

        classe = jury.code_classe
        annees = AnneeAcademique.objects.all().order_by('-code_anac')

        # Année sélectionnée (par défaut: année en cours si définie, sinon dernière inscription de la classe)
        annee_active = AnneeAcademique.get_annee_en_cours()
        annee_code = request.GET.get('annee') or (annee_active.code_anac if annee_active else None)
        if not annee_code:
            annee_code = Inscription.objects.filter(code_classe=classe).values_list('annee_academique', flat=True).order_by('-annee_academique').first()

        if request.method == 'POST':
            annee_code = request.POST.get('annee_academique') or annee_code
            date_delib_str = request.POST.get('date_deliberation')
            contenu = (request.POST.get('contenu') or '').strip()
            date_delib = parse_date(date_delib_str) if date_delib_str else None

            if not (annee_code and date_delib):
                messages.error(request, 'Veuillez renseigner l\'année académique et la date de délibération.')
            else:
                # Créer le communiqué
                communique = CommuniqueDeliberation.objects.create(
                    code_classe=classe,
                    annee_academique=annee_code,
                    date_deliberation=date_delib,
                    contenu=contenu,
                    cree_par=request.user,
                )
                
                # Notifier les enseignants concernés
                enseignants_notifies = notifier_enseignants_communique(communique)
                
                if enseignants_notifies:
                    messages.success(request, f'Communiqué publié avec succès! {len(enseignants_notifies)} enseignant(s) notifié(s).')
                else:
                    messages.success(request, 'Communiqué publié avec succès!')
                    
                return redirect(f"{reverse('jury_communique')}?{urlencode({'annee': annee_code})}")

        communiques = CommuniqueDeliberation.objects.filter(code_classe=classe)
        if annee_code:
            communiques = communiques.filter(annee_academique=annee_code)
        communiques = communiques.order_by('-date_deliberation', '-date_creation')

        context = {
            'jury': jury,
            'classe': classe,
            'annees': annees,
            'annee_code': annee_code,
            'communiques': communiques,
        }
        return render(request, 'jury/communique.html', context)
    except Exception as e:
        messages.error(request, f'Erreur: {str(e)}')
        return redirect('jury_dashboard')


@login_required
def jury_publier(request):
    """Publication des résultats par le jury"""
    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')
    try:
        if request.method == 'POST':
            from reglage.models import AnneeAcademique
            from io import BytesIO
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            from django.core.files.base import ContentFile
            
            # Publier les résultats
            jury.resultat_publie = True
            jury.date_publication = timezone.now()
            jury.save()
            
            # Générer les bulletins pour tous les étudiants de la classe
            classe = jury.code_classe
            annee_active = AnneeAcademique.get_annee_en_cours()
            annee_code = annee_active.code_anac if annee_active else None
            
            if not annee_code:
                annee_code = Inscription.objects.filter(code_classe=classe).values_list('annee_academique', flat=True).order_by('-annee_academique').first()
            
            inscriptions = Inscription.objects.filter(
                code_classe=classe,
                annee_academique=annee_code
            ).select_related('matricule_etudiant')
            
            bulletins_generes = 0
            for inscription in inscriptions:
                etudiant = inscription.matricule_etudiant
                
                # Vérifier si le bulletin existe déjà
                bulletin, created = BulletinNotes.objects.get_or_create(
                    etudiant=etudiant,
                    annee_academique=annee_code,
                    code_classe=classe,
                    defaults={'genere_par': request.user, 'disponible': True}
                )
                
                if not created and bulletin.fichier_pdf:
                    continue
                
                # Générer le PDF du bulletin
                buffer = BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
                elements = []
                styles = getSampleStyleSheet()
                
                # Style personnalisé pour le titre
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=18,
                    textColor=colors.HexColor('#667eea'),
                    spaceAfter=30,
                    alignment=TA_CENTER,
                    fontName='Helvetica-Bold'
                )
                
                # En-tête
                elements.append(Paragraph("BULLETIN DE NOTES", title_style))
                elements.append(Spacer(1, 0.5*cm))
                
                # Informations étudiant
                info_data = [
                    ['Matricule:', etudiant.matricule_et, 'Année:', annee_code],
                    ['Nom:', etudiant.nom_complet, 'Classe:', classe.code_classe],
                ]
                info_table = Table(info_data, colWidths=[3*cm, 6*cm, 2.5*cm, 4.5*cm])
                info_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ]))
                elements.append(info_table)
                elements.append(Spacer(1, 1*cm))
                
                # Récupérer les notes
                evaluations = Evaluation.objects.filter(
                    matricule_etudiant=etudiant
                ).select_related('code_ue', 'code_ec')
                
                if evaluations.exists():
                    notes_data = [['UE/EC', 'CC', 'Examen', 'Note Finale', 'Statut']]
                    
                    for ev in evaluations:
                        cours = str(ev.code_ue) if ev.code_ue else str(ev.code_ec)
                        cc = f"{ev.cc:.2f}" if ev.cc is not None else '-'
                        examen = f"{ev.examen:.2f}" if ev.examen is not None else '-'
                        note_finale = ev.calculer_note_finale()
                        note_str = f"{note_finale:.2f}" if note_finale is not None else '-'
                        statut = 'Validé' if ev.statut == 'VALIDE' else 'Non validé'
                        
                        notes_data.append([cours, cc, examen, note_str, statut])
                    
                    notes_table = Table(notes_data, colWidths=[6*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm])
                    notes_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 11),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                        ('FONTSIZE', (0, 1), (-1, -1), 9),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ]))
                    elements.append(notes_table)
                else:
                    elements.append(Paragraph("Aucune note disponible", styles['Normal']))
                
                elements.append(Spacer(1, 2*cm))
                elements.append(Paragraph(f"Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}", styles['Normal']))
                
                doc.build(elements)
                
                # Sauvegarder le PDF
                pdf_content = buffer.getvalue()
                buffer.close()
                
                filename = f"bulletin_{etudiant.matricule_et}_{annee_code}.pdf"
                bulletin.fichier_pdf.save(filename, ContentFile(pdf_content), save=False)
                bulletin.disponible = True
                bulletin.genere_par = request.user
                bulletin.save()
                bulletins_generes += 1
            
            messages.success(request, f'Résultats publiés avec succès! {bulletins_generes} bulletin(s) généré(s).')
            return redirect('jury_dashboard')
        
        context = {
            'jury': jury,
        }
        return render(request, 'jury/publier.html', context)
    except Jury.DoesNotExist:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')
    except Exception as e:
        messages.error(request, f'Erreur lors de la publication: {str(e)}')
        return redirect('jury_dashboard')


@login_required
def jury_depublier(request):
    """Dépublication des résultats par le jury"""
    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')
    
    try:
        if request.method == 'POST':
            # Dépublier les résultats
            jury.resultat_publie = False
            jury.date_publication = None
            jury.save()
            
            messages.success(request, 'Les résultats ont été dépubliés avec succès. Les étudiants ne peuvent plus y accéder.')
            return redirect('jury_dashboard')
        
        # Afficher la page de confirmation
        context = {
            'jury': jury,
        }
        return render(request, 'jury/depublier.html', context)
    except Exception as e:
        messages.error(request, f'Erreur lors de la dépublication: {str(e)}')
        return redirect('jury_dashboard')


@login_required
def jury_grille_cours(request):
    """Liste des cours de la classe pour le jury"""
    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')
    
    try:
        classe = jury.code_classe
        code_classe_str = classe.code_classe  # Récupérer le code string
        
        # Récupérer les cours (UE/EC) de la classe (filtrer par code_classe string)
        ue_list = UE.objects.filter(classe__code_classe=code_classe_str).order_by('semestre', 'code_ue')
        ec_list = EC.objects.filter(classe__code_classe=code_classe_str).order_by('code_ue__semestre', 'code_ec')
        
        # Identifier les UE qui ont des EC
        ues_avec_ec = set(ec_list.values_list('code_ue__code_ue', flat=True).distinct())
        
        # Année académique actuelle (depuis les paramètres ou inscriptions)
        annee = Inscription.objects.filter(code_classe=classe).values_list('annee_academique', flat=True).first() or '2024-2025'
        
        # Nombre d'étudiants inscrits dans la classe
        nb_etudiants = Inscription.objects.filter(code_classe=classe, annee_academique=annee).count()
        
        # Pré-charger les attributions pour tous les cours de cette classe/année
        attributions = Attribution.objects.filter(
            annee_academique=annee
        ).select_related('matricule_en')
        attribution_map = {}
        for attr in attributions:
            ens = attr.matricule_en
            if ens:
                attribution_map[attr.code_cours] = f"{ens.matricule_en} - {ens.nom_complet}"
            else:
                attribution_map[attr.code_cours] = '-'
        
        # Grouper les cours par semestre
        cours_par_semestre = {}
        
        # Ajouter seulement les UE qui n'ont PAS d'EC
        for ue in ue_list:
            if ue.code_ue in ues_avec_ec:
                continue  # Cette UE a des EC, on ne l'affiche pas
            
            semestre = ue.semestre or 0
            nb_evals = Evaluation.objects.filter(code_ue=ue, annee_academique=annee).count()
            is_evalué = nb_evals > 0
            is_complet = nb_evals >= nb_etudiants if nb_etudiants > 0 else False
            
            if semestre not in cours_par_semestre:
                cours_par_semestre[semestre] = []
            
            cours_par_semestre[semestre].append({
                'code': ue.code_ue,
                'intitule': ue.intitule_ue,
                'type': 'UE',
                'credit': ue.credit,
                'nb_evals': nb_evals,
                'nb_etudiants': nb_etudiants,
                'is_evalue': is_evalué,
                'is_complet': is_complet,
                'enseignant': attribution_map.get(ue.code_ue, '-'),
            })
        
        for ec in ec_list:
            semestre = ec.code_ue.semestre if ec.code_ue else 0
            nb_evals = Evaluation.objects.filter(code_ec=ec, annee_academique=annee).count()
            is_evalué = nb_evals > 0
            is_complet = nb_evals >= nb_etudiants if nb_etudiants > 0 else False
            
            if semestre not in cours_par_semestre:
                cours_par_semestre[semestre] = []
            
            cours_par_semestre[semestre].append({
                'code': ec.code_ec,
                'intitule': ec.intitule_ue,
                'type': 'EC',
                'credit': ec.credit,
                'nb_evals': nb_evals,
                'nb_etudiants': nb_etudiants,
                'is_evalue': is_evalué,
                'is_complet': is_complet,
                'enseignant': attribution_map.get(ec.code_ec, '-'),
            })
        
        # Trier les semestres
        semestres_tries = sorted(cours_par_semestre.keys())
        cours_groupes = [(sem, cours_par_semestre[sem]) for sem in semestres_tries]
        
        context = {
            'jury': jury,
            'classe': classe,
            'cours_groupes': cours_groupes,
            'annee': annee,
            'nb_etudiants': nb_etudiants,
        }
        return render(request, 'jury/grille_cours.html', context)
    except Exception as e:
        messages.error(request, f'Erreur: {str(e)}')
        return redirect('jury_dashboard')


@login_required
def jury_evaluer_cours(request, code_cours, annee):
    """Évaluer les étudiants d'un cours spécifique - Vue Jury"""
    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')
    
    try:
        classe = jury.code_classe
        
        # Récupérer les infos du cours (UE ou EC)
        cours_info = {
            'code': code_cours,
            'intitule': code_cours,
            'type': None,
            'classe': None,
            'categorie': None,
            'categorie_label': None,
        }
        try:
            ue = UE.objects.get(code_ue=code_cours)
            cours_info['intitule'] = ue.intitule_ue
            cours_info['type'] = 'UE'
            cours_info['classe'] = ue.classe
            cours_info['categorie'] = ue.categorie
            cours_info['categorie_label'] = ue.get_categorie_display()
        except UE.DoesNotExist:
            try:
                ec = EC.objects.get(code_ec=code_cours)
                cours_info['intitule'] = ec.intitule_ue
                cours_info['type'] = 'EC'
                cours_info['classe'] = ec.classe
                cours_info['categorie'] = ec.categorie
                cours_info['categorie_label'] = ec.categorie
            except EC.DoesNotExist:
                messages.error(request, 'Cours non trouvé.')
                return redirect('jury_grille_cours')
        
        # Vérifier que le cours appartient à la classe du jury (comparer les code_classe strings)
        cours_classe_code = cours_info['classe'].code_classe if cours_info['classe'] else None
        if cours_classe_code != classe.code_classe:
            messages.error(request, "Ce cours n'appartient pas à votre classe.")
            return redirect('jury_grille_cours')
        
        # Récupérer l'enseignant du cours pour cette année via Attribution
        enseignant_nom = None
        attribution = Attribution.objects.filter(
            code_cours=code_cours,
            annee_academique=annee
        ).select_related('matricule_en', 'matricule_en__grade').first()
        if attribution and attribution.matricule_en:
            ens = attribution.matricule_en
            if ens.grade:
                enseignant_nom = f"{ens.grade.code_grade} {ens.nom_complet} ({ens.matricule_en})"
            else:
                enseignant_nom = f"{ens.nom_complet} ({ens.matricule_en})"
        cours_info['enseignant'] = enseignant_nom
        
        # Le jury a toujours accès au rattrapage et rachat
        rattrapage_actif = True
        rachat_actif = True
        
        # Récupérer les étudiants inscrits dans la classe pour cette année
        etudiants = []
        inscriptions = Inscription.objects.filter(
            code_classe=classe,
            annee_academique=annee
        ).select_related('matricule_etudiant')
        
        for insc in inscriptions:
            # Chercher l'évaluation existante
            if cours_info['type'] == 'UE':
                eval_existante = Evaluation.objects.filter(
                    matricule_etudiant=insc.matricule_etudiant,
                    code_ue__code_ue=code_cours,
                    annee_academique=annee,
                ).first()
            else:
                eval_existante = Evaluation.objects.filter(
                    matricule_etudiant=insc.matricule_etudiant,
                    code_ec__code_ec=code_cours,
                    annee_academique=annee,
                ).first()
            
            # Formater en STRING pour éviter les problèmes de précision float
            cc_str = f"{eval_existante.cc:.1f}" if eval_existante and eval_existante.cc else ''
            examen_str = f"{eval_existante.examen:.1f}" if eval_existante and eval_existante.examen else ''
            note_finale_val = None
            note_finale_str = ''
            if eval_existante and eval_existante.cc and eval_existante.examen:
                note_finale_val = float(eval_existante.cc) + float(eval_existante.examen)
                note_finale_str = f"{note_finale_val:.1f}"
            
            etudiants.append({
                'matricule': insc.matricule_etudiant.matricule_et,
                'nom_complet': insc.matricule_etudiant.nom_complet,
                'evaluation': eval_existante,
                'cc': cc_str,
                'examen': examen_str,
                'note_finale': note_finale_str,
                'note_finale_num': float(note_finale_str) if note_finale_str else None,
                'rattrapage': f"{eval_existante.rattrapage:.1f}" if eval_existante and eval_existante.rattrapage else '',
                'rachat': f"{eval_existante.rachat:.1f}" if eval_existante and eval_existante.rachat else '',
            })
        
        if request.method == 'POST':
            action = request.POST.get('action', 'save')
            
            if action == 'delete':
                # Supprimer une évaluation
                eval_id = request.POST.get('eval_id')
                if eval_id:
                    Evaluation.objects.filter(id_ev=eval_id).delete()
                    messages.success(request, 'Évaluation supprimée avec succès!')
                return redirect('jury_evaluer_cours', code_cours=code_cours, annee=annee)
            
            # Sauvegarder les notes (jury peut modifier toutes les notes)
            # Le template poste uniquement les champs effectivement édités (inputs désactivés => non envoyés)
            matricules = request.POST.getlist('matricule')
            
            # Récupérer le nom du membre du jury connecté
            jury_nom = ''
            if hasattr(jury, 'president'):
                try:
                    ens_jury = Enseignant.objects.get(matricule_en=jury.president)
                    jury_nom = ens_jury.nom_complet
                except Enseignant.DoesNotExist:
                    jury_nom = jury.president
            if not jury_nom:
                jury_nom = request.user.get_full_name() or request.user.username
            
            # Trouver l'enseignant titulaire du cours pour la notification
            attribution_cours = Attribution.objects.filter(
                code_cours=code_cours,
                annee_academique=annee
            ).select_related('matricule_en').first()
            enseignant_titulaire = attribution_cours.matricule_en if attribution_cours else None
            
            etudiants_modifies = []
            
            for matricule in matricules:
                etudiant = Etudiant.objects.get(matricule_et=matricule)
                
                # Chercher l'évaluation existante
                if cours_info['type'] == 'UE':
                    ue_obj = UE.objects.get(code_ue=code_cours)
                    eval_obj, created = Evaluation.objects.get_or_create(
                        matricule_etudiant=etudiant,
                        code_ue=ue_obj,
                        code_classe=cours_info['classe'],
                        defaults={'code_ec': None, 'annee_academique': annee}
                    )
                else:
                    ec_obj = EC.objects.get(code_ec=code_cours)
                    eval_obj, created = Evaluation.objects.get_or_create(
                        matricule_etudiant=etudiant,
                        code_ec=ec_obj,
                        code_classe=cours_info['classe'],
                        defaults={'code_ue': None, 'annee_academique': annee}
                    )
                
                # Sauvegarder les anciennes valeurs pour détecter les modifications
                ancien_cc = eval_obj.cc
                ancien_examen = eval_obj.examen
                ancien_rattrapage = eval_obj.rattrapage
                ancien_rachat = eval_obj.rachat
                
                # Mettre à jour uniquement les champs réellement envoyés
                cc_key = f'cc_{matricule}'
                examen_key = f'examen_{matricule}'
                rattrapage_key = f'rattrapage_{matricule}'
                rachat_key = f'rachat_{matricule}'

                if cc_key in request.POST:
                    cc_val = request.POST.get(cc_key)
                    eval_obj.cc = float(cc_val) if cc_val else None
                if examen_key in request.POST:
                    examen_val = request.POST.get(examen_key)
                    eval_obj.examen = float(examen_val) if examen_val else None
                if rattrapage_key in request.POST:
                    rattrapage_val = request.POST.get(rattrapage_key)
                    eval_obj.rattrapage = float(rattrapage_val) if rattrapage_val else None
                if rachat_key in request.POST:
                    rachat_val = request.POST.get(rachat_key)
                    eval_obj.rachat = float(rachat_val) if rachat_val else None
                
                # Détecter si le jury a modifié une note existante (pas une création)
                note_modifiee = not created and (
                    eval_obj.cc != ancien_cc or
                    eval_obj.examen != ancien_examen or
                    eval_obj.rattrapage != ancien_rattrapage or
                    eval_obj.rachat != ancien_rachat
                )
                
                if note_modifiee:
                    from django.utils import timezone as tz
                    eval_obj.modifie_par_jury = True
                    eval_obj.jury_modificateur = jury_nom
                    eval_obj.date_modification_jury = tz.now()
                    eval_obj.ancien_cc = ancien_cc
                    eval_obj.ancien_examen = ancien_examen
                    eval_obj.ancien_rattrapage = ancien_rattrapage
                    eval_obj.ancien_rachat = ancien_rachat
                    etudiants_modifies.append(etudiant.nom_complet)

                eval_obj.save()
            
            # Envoyer une notification à l'enseignant titulaire si des notes ont été modifiées
            if etudiants_modifies and enseignant_titulaire and enseignant_titulaire.id_lgn:
                liste_noms = ', '.join(etudiants_modifies)
                Notification.objects.create(
                    destinataire=enseignant_titulaire.id_lgn,
                    type_notification='EVALUATION',
                    titre=f'Notes modifiées par le jury - {cours_info["intitule"]}',
                    message=f'Le membre du jury "{jury_nom}" a modifié les notes du cours '
                            f'{code_cours} ({cours_info["intitule"]}) pour les étudiants suivants : '
                            f'{liste_noms}.',
                    code_classe=cours_info['classe'],
                    annee_academique=annee,
                )
            
            messages.success(request, 'Notes enregistrées avec succès!')
            return redirect('jury_evaluer_cours', code_cours=code_cours, annee=annee)
        
        context = {
            'jury': jury,
            'classe': classe,
            'cours': cours_info,
            'annee': annee,
            'etudiants': etudiants,
            'rattrapage_actif': rattrapage_actif,
            'rachat_actif': rachat_actif,
        }
        return render(request, 'jury/evaluer_cours.html', context)
    except Exception as e:
        messages.error(request, f'Erreur: {str(e)}')
        return redirect('jury_grille_cours')


@login_required
def jury_imprimables(request):
    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        logout(request)
        return redirect('login')

    # Sélections (GET)
    selected_classe = request.GET.get('classe', '')
    selected_annee = request.GET.get('annee', '')
    selected_type = request.GET.get('type', 'annuel')
    selected_semestre = request.GET.get('semestre', '')
    selected_date_delib = request.GET.get('date_delib', '')

    # Classe: par défaut celle du jury
    classe_obj = jury.code_classe
    classes = [classe_obj]
    if request.user.is_staff:
        classes = list(Classe.objects.all().order_by('code_classe'))
        if selected_classe:
            classe_tmp = Classe.objects.filter(code_classe=selected_classe).first()
            if classe_tmp:
                classe_obj = classe_tmp

    # Années disponibles pour la classe
    annees = list(
        Inscription.objects.filter(code_classe=classe_obj)
        .values_list('annee_academique', flat=True)
        .distinct()
        .order_by('-annee_academique')
    )
    # Par défaut, utiliser l'année académique du jury connecté
    if not selected_annee:
        if jury.annee_academique:
            selected_annee = jury.annee_academique
        elif annees:
            selected_annee = annees[0]

    niveau_code = None
    if classe_obj and getattr(classe_obj, 'code_niveau', None):
        niveau_code = classe_obj.code_niveau.code_niveau
    elif classe_obj and getattr(classe_obj, 'code_classe', None):
        code_classe = str(classe_obj.code_classe)
        for prefix in ['L1', 'L2', 'L3', 'M1', 'M2']:
            if code_classe.startswith(prefix):
                niveau_code = prefix
                break

    niveau_to_semestres = {
        'L1': (1, 2),
        'L2': (3, 4),
        'L3': (5, 6),
        'M1': (7, 8),
        'M2': (9, 10),
    }
    semestres_niveau = niveau_to_semestres.get(niveau_code, None)
    semestres_choices = list(semestres_niveau) if semestres_niveau else []

    # Si l'utilisateur arrive sans paramètres explicites (ex: /jury/imprimables/),
    # on tente de se caler automatiquement sur la dernière délibération existante
    # afin que les imprimables s'activent immédiatement.
    if selected_annee and (('type' not in request.GET) and ('semestre' not in request.GET)):
        last_delib = (
            Deliberation.objects.filter(code_classe=classe_obj, annee_academique=selected_annee)
            .order_by('-date_mise_a_jour', '-date_creation')
            .first()
        )
        if last_delib:
            params = {
                'annee': selected_annee,
            }
            if request.user.is_staff:
                params['classe'] = classe_obj.code_classe
            if last_delib.type_deliberation == 'ANNEE':
                params['type'] = 'annuel'
            else:
                params['type'] = 'semestriel'
                if last_delib.semestre is not None:
                    params['semestre'] = last_delib.semestre
            return redirect(f"{reverse('jury_imprimables')}?{urlencode(params)}")

    if selected_type != 'semestriel':
        selected_semestre = ''
    elif not selected_semestre and semestres_choices:
        selected_semestre = str(semestres_choices[0])

    dates_deliberation = []
    if selected_annee:
        dates_deliberation = list(
            CommuniqueDeliberation.objects.filter(code_classe=classe_obj, annee_academique=selected_annee)
            .order_by('-date_deliberation')
            .values_list('date_deliberation', flat=True)
            .distinct()
        )
        dates_deliberation = [d for d in dates_deliberation if d]
    if not selected_date_delib and dates_deliberation:
        selected_date_delib = dates_deliberation[0].isoformat()

    inscriptions_qs = Inscription.objects.filter(code_classe=classe_obj).select_related('matricule_etudiant')
    if selected_annee:
        inscriptions_qs = inscriptions_qs.filter(annee_academique=selected_annee)

    query_params = {
        'annee': selected_annee,
        'type': selected_type,
        'semestre': selected_semestre,
        'date_delib': selected_date_delib,
    }
    if request.user.is_staff:
        query_params['classe'] = classe_obj.code_classe
    query_params = {k: v for k, v in query_params.items() if v}
    qs = urlencode(query_params)
    qs_suffix = f"?{qs}" if qs else ''

    links = {
        'palmare': reverse('jury_imprimable_palmare') + qs_suffix,
        'pv': reverse('jury_imprimable_pv') + qs_suffix,
        'releves': reverse('jury_imprimable_releves') + qs_suffix,
    }

    can_imprimer = _deliberation_exists_for_selection(
        classe_obj,
        selected_annee,
        selected_type,
        selected_semestre,
        semestres_niveau,
    )

    etudiants = []
    for ins in inscriptions_qs:
        matricule = ins.matricule_etudiant.matricule_et
        etudiants.append({
            'matricule': matricule,
            'nom': ins.matricule_etudiant.nom_complet,
            'releve_url': reverse('jury_imprimable_releve', args=[matricule]) + qs_suffix,
            'profil_url': reverse('jury_imprimable_profil', args=[matricule]) + qs_suffix,
            'profil_pdf_url': reverse('jury_imprimable_profil_pdf', args=[matricule]) + qs_suffix,
        })
    etudiants.sort(key=lambda e: (e['nom'] or '', e['matricule'] or ''))

    # Vérifier les délibérations disponibles pour chaque option
    deliberations_disponibles = {
        'annuel': False,
        'semestres': {}
    }
    
    if selected_annee:
        # Vérifier la délibération annuelle
        deliberation_annuelle = Deliberation.objects.filter(
            code_classe=classe_obj,
            annee_academique=selected_annee,
            type_deliberation='ANNEE'
        ).first()
        deliberations_disponibles['annuel'] = deliberation_annuelle is not None
        
        # Vérifier chaque délibération semestrielle
        # semestres_choices contient les n° absolus (ex: [3,4] pour L2)
        # mais type_deliberation est toujours 'S1' ou 'S2'
        for i, sem in enumerate(semestres_choices):
            type_delib_code = f'S{i + 1}'  # S1 pour le 1er semestre du niveau, S2 pour le 2ème
            deliberation_semestrielle = Deliberation.objects.filter(
                code_classe=classe_obj,
                annee_academique=selected_annee,
                type_deliberation=type_delib_code
            ).first()
            deliberations_disponibles['semestres'][sem] = deliberation_semestrielle is not None
    
    context = {
        'jury': jury,
        'classes': classes,
        'annees': annees,
        'semestres_choices': semestres_choices,
        'dates_deliberation': dates_deliberation,
        'etudiants': etudiants,
        'links': links,
        'can_imprimer': can_imprimer,
        'deliberations_disponibles': deliberations_disponibles,
        'selected': {
            'classe': classe_obj.code_classe,
            'annee': selected_annee,
            'type': selected_type,
            'semestre': selected_semestre,
            'date_delib': selected_date_delib,
        }
    }
    return render(request, 'jury/imprimables/index.html', context)


@login_required
def jury_evaluations(request):
    """Affichage des évaluations pour le jury"""
    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')

    try:
        classe = jury.code_classe

        # Utiliser l'année académique du jury connecté par défaut
        annee_code = jury.annee_academique
        if not annee_code:
            from reglage.models import AnneeAcademique
            annee_active = AnneeAcademique.get_annee_en_cours()
            annee_code = annee_active.code_anac if annee_active else None
        if not annee_code and classe:
            annee_code = (
                Inscription.objects.filter(code_classe=classe)
                .values_list('annee_academique', flat=True)
                .order_by('-annee_academique')
                .first()
            )

        # Récupérer les filtres
        selected_annee = request.GET.get('annee', annee_code)
        selected_semestre = request.GET.get('semestre', '')
        selected_categorie = request.GET.get('categorie', '')
        selected_etudiant = request.GET.get('etudiant', '')

        # Debug pour voir tous les paramètres GET
        print(f"DEBUG: Paramètres GET complets: {dict(request.GET)}")
        print(f"DEBUG: selected_annee={selected_annee}, selected_semestre={selected_semestre}, selected_categorie={selected_categorie}, selected_etudiant={selected_etudiant}")

        # Debug temporaire pour voir les catégories disponibles
        categories_ue = list(UE.objects.filter(classe=classe).values_list('categorie', flat=True).distinct())
        categories_ec = list(EC.objects.filter(classe=classe).values_list('categorie', flat=True).distinct())
        print(f"DEBUG: Catégories UE: {categories_ue}")
        print(f"DEBUG: Catégories EC: {categories_ec}")
        print(f"DEBUG: Catégorie sélectionnée: {selected_categorie}")
        
        # Debug pour voir les valeurs "vides" dans les catégories
        ue_empty = UE.objects.filter(classe=classe).filter(
            Q(categorie='') | Q(categorie__isnull=True)
        ).count()
        ec_empty = EC.objects.filter(classe=classe).filter(
            Q(categorie='') | Q(categorie__isnull=True)
        ).count()
        print(f"DEBUG: UE sans catégorie: {ue_empty}, EC sans catégorie: {ec_empty}")

        # Récupérer les évaluations - filtrer par UE/EC de la classe du jury
        evaluations = Evaluation.objects.filter(
            Q(code_ue__classe=classe) | Q(code_ec__classe=classe)
        ).select_related(
            'matricule_etudiant',
            'code_ue',
            'code_ec',
            'code_ec__code_ue'
        )

        # Appliquer les filtres
        if selected_annee:
            evaluations = evaluations.filter(annee_academique=selected_annee)

        if selected_semestre:
            evaluations = evaluations.filter(
                Q(code_ue__semestre=selected_semestre) |
                Q(code_ec__code_ue__semestre=selected_semestre)
            )

        if selected_categorie:
            if selected_categorie == 'AUCUNE':
                evaluations_before = evaluations.count()
                # Filtrer les UE/EC sans catégorie (vide, null, ou chaîne vide)
                evaluations = evaluations.filter(
                    Q(code_ue__categorie='') | 
                    Q(code_ue__categorie__isnull=True) |
                    Q(code_ec__categorie='') | 
                    Q(code_ec__categorie__isnull=True)
                )
                evaluations_after = evaluations.count()
                print(f"DEBUG: Filtre AUCUNE appliqué: {evaluations_before} -> {evaluations_after} résultats")
            else:
                # Filtrer par catégorie en gérant différents formats (A, 'A', etc.)
                evaluations_before = evaluations.count()
                evaluations = evaluations.filter(
                    Q(code_ue__categorie=selected_categorie) |
                    Q(code_ec__categorie=selected_categorie)
                )
                evaluations_after = evaluations.count()
                print(f"DEBUG: Filtre {selected_categorie} appliqué: {evaluations_before} -> {evaluations_after} résultats")

        if selected_etudiant:
            evaluations = evaluations.filter(
                matricule_etudiant__matricule_et=selected_etudiant
            )

        # Dédupliquer pour éviter les doublons (même étudiant + même UE/EC)
        evaluations = evaluations.distinct()

        # Ordonner les résultats: d'abord S1 pour tous les étudiants, puis S2
        evaluations = evaluations.order_by(
            'code_ue__semestre',
            'matricule_etudiant__nom_complet',
            'code_ue__code_ue',
            'code_ec__code_ec'
        )

        # Préparer les données pour l'affichage
        evaluations_data = []
        total_credits = 0
        total_credits_capitalises = 0
        notes_cat_a = []
        notes_cat_b = []
        notes_ponderees = []
        debug_counter = 0
        
        for eval in evaluations:
            # Déterminer l'UE et l'EC principaux
            ue = eval.code_ue if eval.code_ue else (eval.code_ec.code_ue if eval.code_ec else None)
            ec = eval.code_ec
            
            # Calculer les notes
            note_finale = eval.calculer_note_finale()
            
            # Calculer la note pondérée (cc+examen) × crédit
            # Priorité: crédit de l'EC, sinon crédit de l'UE
            credit = ec.credit if ec else (ue.credit if ue else 0)
            note_finale_calc = (eval.cc or 0) + (eval.examen or 0)
            note_ponderee = note_finale_calc * credit if credit > 0 else 0
            
            # Récupérer les informations avec des valeurs par défaut
            intitule_ue = ue.intitule_ue if ue else (ec.intitule_ue if ec else 'Non spécifié')
            intitule_ec = ec.intitule_ue if ec else ''
            categorie = ue.categorie if ue else (ec.categorie if ec else '')
            semestre = ue.semestre if ue else (ec.code_ue.semestre if ec and ec.code_ue else '')
            
            # Debug: Afficher quelques catégories pour vérification
            if debug_counter < 5:  # Seulement les 5 premiers
                print(f"DEBUG Evaluation {debug_counter}: UE={ue.code_ue if ue else 'None'}, UE_cat={ue.categorie if ue else 'None'}, EC={ec.code_ec if ec else 'None'}, EC_cat={ec.categorie if ec else 'None'}, categorie_finale={categorie}")
            debug_counter += 1
            
            # Calculer les statistiques
            if credit > 0:
                total_credits += credit
                
                # Crédits capitalisés (validés)
                if eval.statut == 'VALIDE' or (note_finale is not None and note_finale >= 10):
                    total_credits_capitalises += credit
                
                # Moyennes par catégorie
                if note_finale is not None:
                    if categorie == 'A':
                        notes_cat_a.append(note_finale)
                    elif categorie == 'B':
                        notes_cat_b.append(note_finale)
                    
                    # Note pondérée pour la moyenne globale
                    notes_ponderees.append((note_finale, credit))
            
            evaluations_data.append({
                'matricule_etudiant': eval.matricule_etudiant.matricule_et,
                'nom_etudiant': eval.matricule_etudiant.nom_complet,
                'intitule_ue': intitule_ue,
                'intitule_ec': intitule_ec,
                'categorie': categorie,
                'credit': credit,
                'cc': eval.cc,
                'examen': eval.examen,
                'note_finale': note_finale,
                'note_ponderee': round(note_ponderee, 2),
                'rattrapage': eval.rattrapage,
                'statut': eval.statut,
                'semestre': semestre,
                'id': eval.id_ev,  # Ajouter l'ID pour les actions
            })
        
        # Calculer les moyennes
        moyenne_cat_a = round(sum(notes_cat_a) / len(notes_cat_a), 2) if notes_cat_a else 0
        moyenne_cat_b = round(sum(notes_cat_b) / len(notes_cat_b), 2) if notes_cat_b else 0
        
        # Moyenne pondérée globale
        if notes_ponderees:
            total_points = sum(note * credit for note, credit in notes_ponderees)
            total_credits_pondere = sum(credit for note, credit in notes_ponderees)
            moyenne_ponderee_globale = round(total_points / total_credits_pondere, 2) if total_credits_pondere > 0 else 0
        else:
            moyenne_ponderee_globale = 0

        # Récupérer les données pour les filtres
        annees = list(
            Inscription.objects.filter(code_classe=classe)
            .values_list('annee_academique', flat=True)
            .distinct()
            .order_by('-annee_academique')
        )

        ues = UE.objects.filter(classe=classe).order_by('semestre', 'code_ue')
        semestres = sorted(list(set(ue.semestre for ue in ues if ue.semestre)))
        
        # Récupérer la liste des étudiants de la classe pour l'année sélectionnée
        etudiants_filter = {'inscription__code_classe': classe}
        if selected_annee:
            etudiants_filter['inscription__annee_academique'] = selected_annee
        etudiants = Etudiant.objects.filter(
            **etudiants_filter
        ).order_by('nom_complet').distinct()

        context = {
            'jury': jury,
            'classe': classe,
            'evaluations': evaluations_data,
            'annees': annees,
            'semestres': semestres,
            'etudiants': etudiants,
            'selected_annee': selected_annee,
            'selected_semestre': selected_semestre,
            'selected_categorie': selected_categorie,
            'selected_etudiant': selected_etudiant,
            'total_evaluations': len(evaluations_data),
            'stats': {
                'total_credits': total_credits,
                'total_credits_capitalises': total_credits_capitalises,
                'moyenne_cat_a': moyenne_cat_a,
                'moyenne_cat_b': moyenne_cat_b,
                'moyenne_ponderee_globale': moyenne_ponderee_globale,
                'taux_capitalisation': round((total_credits_capitalises / total_credits * 100), 1) if total_credits > 0 else 0
            }
        }

        return render(request, 'jury/evaluations.html', context)

    except Exception as e:
        messages.error(request, f'Erreur: {str(e)}')
        return redirect('home')


@login_required
def jury_evaluations_action(request):
    """Actions groupées sur les évaluations"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    action = request.POST.get('action')
    evaluation_ids = request.POST.getlist('evaluation_ids[]')
    
    if not action or not evaluation_ids:
        return JsonResponse({'error': 'Paramètres manquants'}, status=400)
    
    try:
        evaluations = Evaluation.objects.filter(id_ev__in=evaluation_ids)
        count = 0
        
        if action == 'validate':
            evaluations.update(statut='VALIDE')
            count = evaluations.count()
            message = f'{count} évaluation(s) validée(s) avec succès'
            
        elif action == 'invalidate':
            evaluations.update(statut='NON_VALIDE')
            count = evaluations.count()
            message = f'{count} évaluation(s) invalidée(s) avec succès'
            
        elif action == 'delete':
            count = evaluations.count()
            evaluations.delete()
            message = f'{count} évaluation(s) supprimée(s) avec succès'
            
        elif action == 'export':
            # Préparer les données pour l'export
            data = []
            for eval in evaluations.select_related('matricule_etudiant', 'code_ue', 'code_ec'):
                data.append({
                    'matricule': eval.matricule_etudiant.matricule_et,
                    'nom': eval.matricule_etudiant.nom_complet,
                    'ue': eval.code_ue.intitule_ue if eval.code_ue else '',
                    'ec': eval.code_ec.intitule_ue if eval.code_ec else '',
                    'cc': eval.cc,
                    'examen': eval.examen,
                    'note_finale': eval.calculer_note_finale(),
                    'statut': eval.statut,
                })
            return JsonResponse({'data': data, 'message': f'{len(data)} évaluation(s) exportée(s)'})
        
        else:
            return JsonResponse({'error': 'Action non reconnue'}, status=400)
        
        return JsonResponse({'success': True, 'message': message, 'count': count})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def jury_evaluation_edit(request, eval_id):
    """Modifier une évaluation"""
    try:
        evaluation = Evaluation.objects.get(id_ev=eval_id)
        
        if request.method == 'POST':
            # Mettre à jour les champs
            evaluation.cc = request.POST.get('cc')
            evaluation.examen = request.POST.get('examen')
            evaluation.rattrapage = request.POST.get('rattrapage')
            evaluation.statut = request.POST.get('statut')
            evaluation.save()
            
            messages.success(request, 'Évaluation modifiée avec succès')
            return redirect('jury_evaluations')
        
        # Afficher le formulaire de modification
        # Calculer les crédits pour l'affichage
        credits = evaluation.code_ec.credit if evaluation.code_ec else (evaluation.code_ue.credit if evaluation.code_ue else 0)
        
        # Calculer la note finale pour l'affichage
        note_finale = evaluation.calculer_note_finale()
        
        context = {
            'evaluation': evaluation,
            'statut_choices': Evaluation.STATUT_CHOICES,
            'credits': credits,
            'note_finale': note_finale,
        }
        return render(request, 'jury/evaluation_edit.html', context)
        
    except Evaluation.DoesNotExist:
        messages.error(request, 'Évaluation non trouvée')
        return redirect('jury_evaluations')
    except Exception as e:
        messages.error(request, f'Erreur: {str(e)}')
        return redirect('jury_evaluations')


@login_required
def jury_evaluation_delete(request, eval_id):
    """Supprimer une évaluation"""
    try:
        evaluation = Evaluation.objects.get(id_ev=eval_id)
        evaluation.delete()
        messages.success(request, 'Évaluation supprimée avec succès')
        
    except Evaluation.DoesNotExist:
        messages.error(request, 'Évaluation non trouvée')
    except Exception as e:
        messages.error(request, f'Erreur: {str(e)}')
    
    return redirect('jury_evaluations')


def jury_deliberations(request):
    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        logout(request)
        return redirect('login')

    selected_classe = request.GET.get('classe', '')
    selected_annee = request.GET.get('annee', '')
    selected_type = request.GET.get('type', 'ANNEE')

    classe_obj = jury.code_classe
    classes = [classe_obj]
    if request.user.is_staff:
        classes = list(Classe.objects.all().order_by('code_classe'))
        if selected_classe:
            classe_tmp = Classe.objects.filter(code_classe=selected_classe).first()
            if classe_tmp:
                classe_obj = classe_tmp

    annees = list(
        Inscription.objects.filter(code_classe=classe_obj)
        .values_list('annee_academique', flat=True)
        .distinct()
        .order_by('-annee_academique')
    )
    # Par défaut, utiliser l'année académique du jury connecté
    if not selected_annee:
        if jury.annee_academique:
            selected_annee = jury.annee_academique
        elif annees:
            selected_annee = annees[0]

    # Récupérer les filtres pour les évaluations
    selected_semestre = request.GET.get('semestre', '')
    selected_categorie = request.GET.get('categorie', '')
    selected_etudiant = request.GET.get('etudiant', '')

    # Récupérer les délibérations (source principale)
    deliberations = Deliberation.objects.filter(
        code_classe=classe_obj
    ).select_related(
        'matricule_etudiant',
        'code_ue',
        'code_ec',
        'code_ec__code_ue',
        'cree_par'
    )
    
    # Appliquer les filtres
    if selected_annee:
        deliberations = deliberations.filter(
            annee_academique=selected_annee
        )

    if selected_semestre:
        if selected_semestre == 'annuel':
            # Pour l'annuel, afficher S1 + S2 combinés (avec compensation annuelle appliquée)
            deliberations = deliberations.filter(
                type_deliberation__in=['S1', 'S2']
            )
        else:
            # Filtrer par semestre numérique
            deliberations = deliberations.filter(
                semestre=selected_semestre
            )

    if selected_categorie:
        if selected_categorie == 'AUCUNE':
            # Filtrer les UE/EC sans catégorie (vide, null, ou chaîne vide)
            deliberations = deliberations.filter(
                Q(code_ue__categorie='') | 
                Q(code_ue__categorie__isnull=True) |
                Q(code_ec__categorie='') | 
                Q(code_ec__categorie__isnull=True)
            )
        else:
            # Filtrer par catégorie
            deliberations = deliberations.filter(
                Q(code_ue__categorie=selected_categorie) |
                Q(code_ec__categorie=selected_categorie)
            )

    if selected_etudiant:
        deliberations = deliberations.filter(
            matricule_etudiant__matricule_et=selected_etudiant
        )

    # Ordonner les résultats
    deliberations = deliberations.order_by(
        'semestre',
        'matricule_etudiant__nom_complet',
        'code_ue__code_ue',
        'code_ec__code_ec'
    )

    # Préparer les données pour l'affichage
    deliberations_data = []
    
    for delib in deliberations:
        # Déterminer l'UE et l'EC principaux
        ue = delib.code_ue if delib.code_ue else (delib.code_ec.code_ue if delib.code_ec else None)
        ec = delib.code_ec
        
        # Calculer les notes
        note_finale = delib.calculer_note_finale()
        
        # Calculer la note pondérée (cc+examen) × crédit
        # Priorité: crédit de l'EC, sinon crédit de l'UE
        credit = ec.credit if ec else (ue.credit if ue else 0)
        note_finale_calc = (delib.cc or 0) + (delib.examen or 0)
        note_ponderee = note_finale_calc * credit if credit > 0 else 0
        
        # Récupérer les informations avec des valeurs par défaut
        intitule_ue = ue.intitule_ue if ue else (ec.intitule_ue if ec else 'Non spécifié')
        intitule_ec = ec.intitule_ue if ec else ''
        categorie = ue.categorie if ue else (ec.categorie if ec else '')
        semestre = delib.semestre or (ue.semestre if ue else (ec.code_ue.semestre if ec and ec.code_ue else ''))
        
        deliberations_data.append({
            'matricule_etudiant': delib.matricule_etudiant.matricule_et,
            'nom_etudiant': delib.matricule_etudiant.nom_complet,
            'intitule_ue': intitule_ue,
            'intitule_ec': intitule_ec,
            'categorie': categorie,
            'credit': credit,
            'cc': delib.cc,
            'examen': delib.examen,
            'note_finale': note_finale,
            'note_ponderee': round(note_ponderee, 2),
            'rattrapage': delib.rattrapage,
            'statut': delib.statut,
            'code_ec': ec.code_ec if ec else None,
            'code_ue': ue.code_ue if ue else None,
            'semestre': semestre,
            'id': delib.id_delib,
            'type_deliberation': delib.type_deliberation,
            'annee_academique': delib.annee_academique,
            'date_creation': delib.date_creation,
            'cree_par': delib.cree_par,
        })
    
    # Calculer les stats correctement en utilisant _jury_compute_delib_ues pour chaque étudiant
    # Déterminer le type de délibération
    if selected_semestre == 'annuel':
        type_delib = 'annuel'
        sem = None
    elif selected_semestre and str(selected_semestre).isdigit():
        type_delib = 'semestriel'
        sem = int(selected_semestre)
    else:
        type_delib = 'annuel'
        sem = None
    
    # Map compensé: (matricule, code_ec_ou_ue) -> statut_code (avec compensation)
    compensated_status_map = {}

    def _collect_compensated_statuses(stats_result, matricule):
        """Collecter les statuts compensés depuis les rows de _jury_compute_delib_ues"""
        for row in stats_result.get('rows', []):
            code = row.get('code_ec') or row.get('code_ue')
            if code:
                compensated_status_map[(matricule, code)] = row.get('statut_code', 'NON_VALIDE')
    
    # Si un étudiant spécifique est sélectionné, calculer ses stats
    if selected_etudiant:
        etudiant_obj = Etudiant.objects.filter(matricule_et=selected_etudiant).first()
        if etudiant_obj:
            stats = _jury_compute_delib_ues(classe_obj, etudiant_obj, type_delib, sem, selected_annee)
            _collect_compensated_statuses(stats, selected_etudiant)
            total_credits = stats.get('credits_total', 0)
            total_credits_capitalises = stats.get('credits_valides', 0)
            moyenne_cat_a = stats.get('moyenne_cat_a', 0)
            moyenne_cat_b = stats.get('moyenne_cat_b', 0)
            moyenne_ponderee_globale = stats.get('moyenne', 0)
            
            decision_code = stats.get('decision_code')
            nb_admis = 1 if decision_code == 'ADM' else 0
            nb_compensable = 1 if decision_code in ['ADMD', 'COMP'] else 0
            nb_ajournes = 1 if decision_code == 'AJ' else 0
            nb_defaillants = 1 if decision_code == 'DEF' else 0
        else:
            total_credits = 0
            total_credits_capitalises = 0
            moyenne_cat_a = 0
            moyenne_cat_b = 0
            moyenne_ponderee_globale = 0
            nb_admis = 0
            nb_compensable = 0
            nb_ajournes = 0
            nb_defaillants = 0
    else:
        # Calculer les stats agrégées pour TOUS les étudiants uniques
        etudiants_uniques = deliberations.values_list('matricule_etudiant', flat=True).distinct()
        
        total_credits = 0
        total_credits_capitalises = 0
        somme_moyenne_cat_a = 0
        somme_moyenne_cat_b = 0
        somme_moyenne_globale = 0
        nb_etudiants = 0
        
        # Stats pour les décisions
        nb_admis = 0
        nb_compensable = 0
        nb_ajournes = 0
        nb_defaillants = 0
        
        for etudiant_id in etudiants_uniques:
            etudiant_obj = Etudiant.objects.filter(pk=etudiant_id).first()
            if etudiant_obj:
                stats = _jury_compute_delib_ues(classe_obj, etudiant_obj, type_delib, sem, selected_annee)
                _collect_compensated_statuses(stats, etudiant_obj.matricule_et)
                total_credits += stats.get('credits_total', 0)
                total_credits_capitalises += stats.get('credits_valides', 0)
                somme_moyenne_cat_a += stats.get('moyenne_cat_a', 0) or 0
                somme_moyenne_cat_b += stats.get('moyenne_cat_b', 0) or 0
                somme_moyenne_globale += stats.get('moyenne', 0) or 0
                nb_etudiants += 1
                
                decision_code = stats.get('decision_code')
                if decision_code == 'DEF':
                    nb_defaillants += 1
                elif decision_code == 'ADM':
                    nb_admis += 1
                elif decision_code in ['ADMD', 'COMP']:
                    nb_compensable += 1
                elif decision_code == 'AJ':
                    nb_ajournes += 1
                else:
                    nb_ajournes += 1
        
        # Moyennes agrégées
        moyenne_cat_a = round(somme_moyenne_cat_a / nb_etudiants, 2) if nb_etudiants > 0 else 0
        moyenne_cat_b = round(somme_moyenne_cat_b / nb_etudiants, 2) if nb_etudiants > 0 else 0
        moyenne_ponderee_globale = round(somme_moyenne_globale / nb_etudiants, 2) if nb_etudiants > 0 else 0

    # Appliquer les statuts compensés aux lignes de délibérations
    for d in deliberations_data:
        mat = d['matricule_etudiant']
        lookup_code = d.get('code_ec') or d.get('code_ue')
        if lookup_code and (mat, lookup_code) in compensated_status_map:
            d['statut'] = compensated_status_map[(mat, lookup_code)]

    # Récupérer les données pour les filtres
    ues = UE.objects.filter(classe=classe_obj).order_by('semestre', 'code_ue')
    semestres = sorted(list(set(ue.semestre for ue in ues if ue.semestre)))
    
    # Récupérer la liste des étudiants de la classe
    etudiants = Etudiant.objects.filter(
        inscription__code_classe=classe_obj
    ).order_by('nom_complet').distinct()

    # Obtenir les types de délibération disponibles
    types_deliberation = Deliberation.objects.filter(
        code_classe=classe_obj
    ).values_list('type_deliberation', flat=True).distinct()
    if not types_deliberation:
        types_deliberation = ['S1', 'S2', 'ANNEE']

    context = {
        'jury': jury,
        'classes': classes,
        'annees': annees,
        # Ajouter les données pour les délibérations
        'evaluations': deliberations_data,  # Garder le nom pour le template
        'deliberations': deliberations_data,  # Ajouter aussi avec le nom correct
        'semestres': semestres,
        'etudiants': etudiants,
        'selected_semestre': selected_semestre,
        'selected_categorie': selected_categorie,
        'selected_etudiant': selected_etudiant,
        'total_evaluations': len(deliberations_data),
        'stats': {
            'total_credits': total_credits,
            'total_credits_capitalises': total_credits_capitalises,
            'moyenne_cat_a': moyenne_cat_a,
            'moyenne_cat_b': moyenne_cat_b,
            'moyenne_ponderee_globale': moyenne_ponderee_globale,
            'taux_capitalisation': round((total_credits_capitalises / total_credits * 100), 1) if total_credits > 0 else 0,
            'nb_admis': nb_admis,
            'nb_compensable': nb_compensable,
            'nb_ajournes': nb_ajournes,
            'nb_defaillants': nb_defaillants,
        },
        'types_deliberation': types_deliberation,
    }
    return render(request, 'jury/deliberations/index.html', context)


@login_required
def jury_deliberations_action(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

    jury = get_jury_for_user(request)
    if not jury and not request.user.is_staff:
        return JsonResponse({'error': 'Profil jury non trouvé'}, status=403)

    action = request.POST.get('action')
    deliberation_ids = request.POST.getlist('deliberation_ids[]')

    if not action or not deliberation_ids:
        return JsonResponse({'error': 'Paramètres manquants'}, status=400)

    try:
        deliberations = Deliberation.objects.filter(id_delib__in=deliberation_ids)
        if not request.user.is_staff and jury:
            deliberations = deliberations.filter(code_classe=jury.code_classe)

        if action == 'delete':
            count = deliberations.count()
            deliberations.delete()
            message = f'{count} délibération(s) supprimée(s) avec succès'
            return JsonResponse({'success': True, 'message': message, 'count': count})

        return JsonResponse({'error': 'Action non reconnue'}, status=400)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def jury_deliberation_edit(request, delib_id):
    """Modifier une délibération"""
    try:
        deliberation = Deliberation.objects.get(id_delib=delib_id)
        
        if request.method == 'POST':
            # Mettre à jour les champs
            deliberation.cc = request.POST.get('cc')
            deliberation.examen = request.POST.get('examen')
            deliberation.rattrapage = request.POST.get('rattrapage')
            deliberation.rachat = request.POST.get('rachat')
            deliberation.statut = request.POST.get('statut')
            deliberation.save()
            
            messages.success(request, 'Délibération modifiée avec succès')
            return redirect('jury_deliberations')
        
        # Afficher le formulaire de modification
        # Calculer les crédits pour l'affichage
        credits = deliberation.code_ec.credit if deliberation.code_ec else (deliberation.code_ue.credit if deliberation.code_ue else 0)
        
        # Calculer la note finale pour l'affichage
        note_finale = deliberation.calculer_note_finale()
        
        context = {
            'deliberation': deliberation,
            'evaluation': deliberation,  # Pour réutiliser le même template
            'statut_choices': Deliberation.STATUT_CHOICES,
            'credits': credits,
            'note_finale': note_finale,
        }
        return render(request, 'jury/evaluation_edit.html', context)
        
    except Deliberation.DoesNotExist:
        messages.error(request, 'Délibération non trouvée')
        return redirect('jury_deliberations')
    except Exception as e:
        messages.error(request, f'Erreur: {str(e)}')
        return redirect('jury_deliberations')


@login_required
def jury_deliberation_delete(request, delib_id):
    """Supprimer une délibération"""
    try:
        deliberation = Deliberation.objects.get(id_delib=delib_id)
        deliberation.delete()
        messages.success(request, 'Délibération supprimée avec succès')
        
    except Deliberation.DoesNotExist:
        messages.error(request, 'Délibération non trouvée')
    except Exception as e:
        messages.error(request, f'Erreur: {str(e)}')
    
    return redirect('jury_deliberations')


def get_profil_etudiant_data(etudiant, classe_obj, annee, palmares_type, palmares_semestre=None):
    """Récupère les données complètes du profil d'un étudiant"""
    try:
        from .models import UE, EC, Evaluation, Inscription
        
        # Récupérer les UE selon le type de palmarès
        if palmares_type == 'semestriel' and palmares_semestre:
            ues_classe = UE.objects.filter(
                classe=classe_obj,
                semestre=palmares_semestre
            )
        else:
            ues_classe = UE.objects.filter(classe=classe_obj)
        
        # Données du profil
        profil_data = {
            'matricule': etudiant.matricule_et,
            'nom': etudiant.nom_complet,
            'sexe': etudiant.sexe,
            'nationalite': etudiant.nationalite,
            'ues_details': [],
            'total_credits_ue': 0,
            'credits_capitalises': 0,
            'moyenne_generale': 0,
            'moyenne_ponderee': 0,
            'decision': '',
            'statistiques': {
                'total_ues': 0,
                'ues_validees': 0,
                'ues_en_ajournement': 0,
                'notes_count': 0
            }
        }
        
        total_credits_ue = 0
        total_credits_capitalises = 0
        notes_ponderees = []
        total_pondération = 0
        
        for ue in ues_classe:
            total_credits_ue += ue.credit
            profil_data['statistiques']['total_ues'] += 1
            
            ue_data = {
                'code_ue': ue.code_ue,
                'intitule': ue.intitule_ue,
                'credit': ue.credit,
                'semestre': ue.semestre,
                'evaluations': [],
                'moyenne_ue': 0,
                'decision_ue': 'AJ',
                'notes': []
            }
            
            notes_ue = []
            notes_ponderees_ue = []
            total_credits_ec = 0
            
            # 1. Récupérer les évaluations liées directement à l'UE (sans EC)
            evaluations_ue_directes = Evaluation.objects.filter(
                matricule_etudiant=etudiant,
                code_ue=ue,
                code_ec__isnull=True
            )
            
            for eval in evaluations_ue_directes:
                eval_data = {
                    'code_ec': ue.code_ue,
                    'intitule_ec': ue.intitule_ue,
                    'credit_ec': ue.credit,
                    'cc': eval.cc,
                    'examen': eval.examen,
                    'rattrapage': eval.rattrapage,
                    'rachat': eval.rachat,
                    'statut': eval.statut
                }
                ue_data['evaluations'].append(eval_data)
                
                # Calculer la note finale
                note_ec = None
                if eval.rattrapage is not None and eval.rattrapage > 0:
                    note_ec = eval.rattrapage
                elif eval.rachat is not None and eval.rachat > 0:
                    note_ec = eval.rachat
                elif eval.cc is not None and eval.examen is not None:
                    note_ec = eval.cc + eval.examen
                
                if note_ec is not None:
                    notes_ue.append(note_ec)
                    notes_ponderees_ue.append(note_ec * ue.credit)
                    total_credits_ec += ue.credit
            
            # 2. Récupérer les évaluations liées via les EC
            ecs_ue = EC.objects.filter(code_ue=ue)
            evaluations_ue_ec = Evaluation.objects.filter(
                matricule_etudiant=etudiant,
                code_ec__in=ecs_ue
            )
            
            for eval in evaluations_ue_ec:
                ec = eval.code_ec
                eval_data = {
                    'code_ec': ec.code_ec if ec else '',
                    'intitule_ec': ec.intitule_ue if ec else '',
                    'credit_ec': ec.credit if ec else 0,
                    'cc': eval.cc,
                    'examen': eval.examen,
                    'rattrapage': eval.rattrapage,
                    'rachat': eval.rachat,
                    'statut': eval.statut
                }
                ue_data['evaluations'].append(eval_data)
                
                # Calculer la note finale de l'EC
                note_ec = None
                if eval.rattrapage is not None and eval.rattrapage > 0:
                    note_ec = eval.rattrapage
                elif eval.rachat is not None and eval.rachat > 0:
                    note_ec = eval.rachat
                elif eval.cc is not None and eval.examen is not None:
                    note_ec = eval.cc + eval.examen
                
                if note_ec is not None and ec:
                    notes_ue.append(note_ec)
                    credit_ec = ec.credit if ec.credit else 1
                    notes_ponderees_ue.append(note_ec * credit_ec)
                    total_credits_ec += credit_ec
            
            if notes_ponderees_ue and total_credits_ec > 0:
                # Moyenne pondérée de l'UE
                moyenne_ue = sum(notes_ponderees_ue) / total_credits_ec
                ue_data['moyenne_ue'] = moyenne_ue
                
                # Décision pour l'UE
                if moyenne_ue >= 10:
                    ue_data['decision_ue'] = 'ADM'
                    total_credits_capitalises += ue.credit
                    profil_data['statistiques']['ues_validees'] += 1
                elif moyenne_ue >= 8:
                    ue_data['decision_ue'] = 'COMP'
                    profil_data['statistiques']['ues_en_ajournement'] += 1
                else:
                    ue_data['decision_ue'] = 'AJ'
                    profil_data['statistiques']['ues_en_ajournement'] += 1
                
                # Ajouter à la moyenne pondérée
                notes_ponderees.append(moyenne_ue * ue.credit)
                total_pondération += ue.credit
                profil_data['statistiques']['notes_count'] += len(notes_ue)
            
            profil_data['ues_details'].append(ue_data)
        
        # Calculer les moyennes générales
        profil_data['total_credits_ue'] = total_credits_ue
        profil_data['credits_capitalises'] = total_credits_capitalises
        
        if total_pondération > 0:
            profil_data['moyenne_ponderee'] = sum(notes_ponderees) / total_pondération
            profil_data['moyenne_generale'] = profil_data['moyenne_ponderee']
        
        # Décision générale
        if profil_data['moyenne_generale'] >= 10:
            profil_data['decision'] = 'ADMIS'
        elif profil_data['moyenne_generale'] >= 8:
            profil_data['decision'] = 'COMPENSABLE'
        else:
            profil_data['decision'] = 'AJOURNÉ'
        
        return profil_data
        
    except Exception as e:
        return None


@login_required
def jury_imprimable_palmare(request):
    """Génère le palmarès en PDF avec ReportLab"""
    from .utils_palmares_pdf import generer_palmares_pdf
    
    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        logout(request)
        return redirect('login')

    classe_obj = jury.code_classe
    classe_code = request.GET.get('classe', '')
    if request.user.is_staff and classe_code:
        classe_tmp = Classe.objects.filter(code_classe=classe_code).first()
        if classe_tmp:
            classe_obj = classe_tmp

    annee = request.GET.get('annee', '')

    selected_type = request.GET.get('type', 'annuel')
    selected_semestre = request.GET.get('semestre', '')
    if not _require_deliberation_for_imprimable(request, jury, classe_obj, annee, selected_type, selected_semestre):
        return redirect('jury_imprimables')
    
    # Récupérer les inscriptions
    inscriptions = Inscription.objects.filter(code_classe=classe_obj).select_related('matricule_etudiant')
    if annee:
        inscriptions = inscriptions.filter(annee_academique=annee)
    
    # Déterminer le niveau de la classe
    niveau_code = None
    if classe_obj and getattr(classe_obj, 'code_niveau', None):
        niveau_code = classe_obj.code_niveau.code_niveau
    elif classe_obj and getattr(classe_obj, 'code_classe', None):
        code_classe = str(classe_obj.code_classe)
        for prefix in ['L1', 'L2', 'L3', 'M1', 'M2']:
            if code_classe.startswith(prefix):
                niveau_code = prefix
                break
    
    # Déterminer si le palmarès doit être semestriel ou annuel selon le choix de l'utilisateur
    if selected_type == 'semestriel':
        palmares_type = 'semestriel'
        palmares_semestre = int(selected_semestre) if selected_semestre else 1
    else:
        # Annuel - respecter le choix de l'utilisateur
        palmares_type = 'annuel'
        palmares_semestre = None
    
    # Récupérer la délibération appropriée
    deliberation = None
    if palmares_type == 'semestriel' and palmares_semestre:
        deliberation = Deliberation.objects.filter(
            code_classe=classe_obj,
            annee_academique=annee,
            type_deliberation=f'S{palmares_semestre}'
        ).first()
    else:
        deliberation = Deliberation.objects.filter(
            code_classe=classe_obj,
            annee_academique=annee,
            type_deliberation='ANNEE'
        ).first()
    
    def _mention_for_moyenne(moyenne):
        """Calculer la mention selon la moyenne générale"""
        if moyenne is None:
            return 'A déterminer'
        n = float(moyenne)
        if n >= 18:
            return 'Excellent (A)'
        if n >= 16:
            return 'Très bien (B)'
        if n >= 14:
            return 'Bien (C)'
        if n >= 12:
            return 'Assez Bien (D)'
        if n >= 10:
            return 'Passable (E)'
        if n >= 8:
            return 'Insuffisant (F)'
        return 'Insatisfaisant (G)'
    
    # Préparer les données des étudiants en utilisant _jury_compute_delib_ues (aligné avec profil/relevé)
    etudiants_data = []
    stats = {'admis': 0, 'comp': 0, 'aj': 0, 'def': 0, 'total': 0}
    
    # Déterminer le type de délibération pour _jury_compute_delib_ues
    type_delib = 'annuel' if palmares_type == 'annuel' else 'semestriel'
    semestre_int = int(palmares_semestre) if palmares_semestre else None
    
    # Récupérer tous les étudiants inscrits
    for inscription in inscriptions:
        etudiant = inscription.matricule_etudiant
        
        # Utiliser _jury_compute_delib_ues pour obtenir les données (même logique que profil/relevé)
        delib_data = _jury_compute_delib_ues(classe_obj, etudiant, type_delib, semestre_int, annee)
        
        if not delib_data.get('rows'):
            continue
        
        # Extraire les données calculées
        moyenne = delib_data.get('moyenne', 0) or 0
        credits_capitalises = delib_data.get('credits_valides', 0)
        total_credits = delib_data.get('credits_total', 0)
        pourcentage = delib_data.get('pourcentage', 0) or 0
        
        decision_code = delib_data.get('decision_code')
        if decision_code == 'DEF':
            decision = 'DEF'
            stats['def'] += 1
        elif decision_code == 'ADM':
            decision = 'ADM'
            stats['admis'] += 1
        elif decision_code in ['ADMD', 'COMP']:
            decision = 'COMP'
            stats['comp'] += 1
        else:
            decision = 'AJ'
            stats['aj'] += 1
        
        stats['total'] += 1
        
        mention = _mention_for_moyenne(moyenne)
        
        etudiant_info = {
            'nom': etudiant.nom_complet or '',
            'sexe': etudiant.sexe or '',
            'nationalite': etudiant.nationalite or '',
            'matricule': etudiant.matricule_et or '',
            'moyenne': f"{moyenne:.2f}".replace('.', ','),
            'pourcentage': f"{pourcentage:.1f}".replace('.', ',') + '%',
            'credits_capitalises': str(credits_capitalises),
            'decision': decision,
            'mention': mention
        }
        
        etudiants_data.append(etudiant_info)
    
    # Trier par moyenne décroissante
    etudiants_data.sort(key=lambda x: float(str(x['moyenne']).replace(',', '.')), reverse=True)
    
    # Préparer le titre selon le type
    if palmares_type == 'semestriel':
        titre_type = f"PALMARES DES RESULTATS SEMESTRE {palmares_semestre} {annee}"
    else:
        titre_type = f"PALMARES DES RESULTATS ANNUEL {annee}"
    
    # Ajouter la classe en bas du titre
    classe_nom = classe_obj.code_classe if classe_obj else ''
    
    # Générer le PDF
    return generer_palmares_pdf(request, classe_obj, annee, etudiants_data, stats, titre_type, classe_nom)


@login_required
def jury_imprimable_releves(request):
    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        logout(request)
        return redirect('login')

    classe_obj = jury.code_classe
    classe_code = request.GET.get('classe', '')
    if request.user.is_staff and classe_code:
        classe_tmp = Classe.objects.filter(code_classe=classe_code).first()
        if classe_tmp:
            classe_obj = classe_tmp

    annee = request.GET.get('annee', '')
    inscriptions = Inscription.objects.filter(code_classe=classe_obj).select_related('matricule_etudiant')
    if annee:
        inscriptions = inscriptions.filter(annee_academique=annee)

    cohortes = list(
        inscriptions.filter(cohorte__isnull=False)
        .values_list('cohorte__code_cohorte', 'cohorte__lib_cohorte')
        .distinct()
    )
    cohorte_label = ''
    if len(cohortes) == 1:
        c_code, c_lib = cohortes[0]
        cohorte_label = f"{c_code} - {c_lib}" if c_lib else str(c_code)
    elif len(cohortes) > 1:
        cohorte_label = ", ".join([c[0] for c in cohortes if c and c[0]])

    context = {
        'jury': jury,
        'classe': classe_obj,
        'cohorte_label': cohorte_label,
        'inscriptions': inscriptions,
        'date': timezone.now().date(),
        'annee': annee,
    }
    return render(request, 'jury/imprimables/releves.html', context)


def _jury_compute_delib_ues(classe_obj, etudiant, type_delib, semestre, annee_academique=None):
    """Calculer les données de délibération pour un étudiant depuis la table Deliberation.
    Utilise directement les statuts du modèle Deliberation et regroupe par UE pour éviter le double comptage.
    Pour la délibération annuelle, calcule S1 et S2 séparément (chacun avec sa compensation) puis additionne."""
    
    # Déterminer le(s) type(s) de délibération à récupérer
    is_annuel = type_delib == 'annuel'
    
    # Pour l'annuel: calculer S1 et S2 séparément, puis combiner les résultats
    if is_annuel:
        s1 = _jury_compute_delib_ues(classe_obj, etudiant, 'semestriel', 1, annee_academique)
        s2 = _jury_compute_delib_ues(classe_obj, etudiant, 'semestriel', 2, annee_academique)
        
        # Combiner les rows
        combined_rows = s1.get('rows', []) + s2.get('rows', [])
        
        # Additionner les crédits (somme simple S1 + S2)
        credits_total = s1.get('credits_total', 0) + s2.get('credits_total', 0)
        credits_valides = s1.get('credits_valides', 0) + s2.get('credits_valides', 0)
        credits_total_attendus = s1.get('credits_total_attendus', 0) + s2.get('credits_total_attendus', 0)
        
        # Recalculer les moyennes pondérées depuis les rows combinées
        # (évite les erreurs de moyenne arithmétique simple entre semestres)
        total_points = 0
        total_credits_pondere = 0
        points_cat_a = 0
        credits_cat_a = 0
        points_cat_b = 0
        credits_cat_b = 0
        
        for row in combined_rows:
            note = row.get('note')
            credit = row.get('credit', 0) or 0
            categorie = (row.get('categorie', '') or '').upper()
            
            if note is not None and credit > 0:
                total_points += note * credit
                total_credits_pondere += credit
                
                if categorie == 'A':
                    points_cat_a += note * credit
                    credits_cat_a += credit
                elif categorie == 'B':
                    points_cat_b += note * credit
                    credits_cat_b += credit
        
        moyenne_generale = total_points / total_credits_pondere if total_credits_pondere > 0 else 0
        moyenne_cat_a = points_cat_a / credits_cat_a if credits_cat_a > 0 else 0
        moyenne_cat_b = points_cat_b / credits_cat_b if credits_cat_b > 0 else 0
        
        # Combiner les EC compensés
        ec_compensated_ids = s1.get('ec_compensated_ids', []) + s2.get('ec_compensated_ids', [])
        ue_compensated_ids = s1.get('ue_compensated_ids', []) + s2.get('ue_compensated_ids', [])
        
        # === COMPENSATION ANNUELLE ===
        # En MASTER (M1/M2): PAS de compensation annuelle
        # Compenser les cours NON_VALIDE (note 8-9.99) si moyenne annuelle catégorie >= 10
        # Ne jamais dégrader un cours déjà validé
        is_master = classe_obj and hasattr(classe_obj, 'code_niveau_id') and str(classe_obj.code_niveau_id) in ('M1', 'M2')
        
        for row in combined_rows:
            if row.get('statut_code') == 'NON_VALIDE':
                note = row.get('note')
                categorie = row.get('categorie', '').upper()
                credit = row.get('credit', 0) or 0
                
                # Déterminer la moyenne annuelle de la catégorie
                if categorie == 'A':
                    moyenne_cat = moyenne_cat_a
                elif categorie == 'B':
                    moyenne_cat = moyenne_cat_b
                else:
                    moyenne_cat = 0
                
                # Critères: note entre 8 et 9.99 ET moyenne catégorie >= 10
                # En Master: compensation désactivée
                if not is_master and note is not None and 8 <= note < 10 and moyenne_cat >= 10:
                    row['statut'] = 'Validé'
                    row['statut_code'] = 'VALIDE_COMP'
                    row['est_valide'] = True
                    row['compense'] = True
                    credits_valides += credit
        
        # Recalculer crédits non validés après compensation
        credits_non_valides = credits_total - credits_valides
        
        # Pourcentage et décision
        pourcentage = (moyenne_generale / 20) * 100 if moyenne_generale > 0 else 0
        
        decision_label = 'A déterminer'
        decision_code = 'ATT'
        
        # Seuil de crédits attendus : 60 pour une année (S1 + S2)
        seuil_credits = 60
        
        # Indicateur : l'étudiant n'a pas atteint le seuil de crédits évalués
        has_missing_evaluations = credits_total < seuil_credits
        
        # Étape 1 : Si credits_total < seuil → Défaillant (évaluations manquantes)
        # Étape 2 : Sinon → décision basée sur les crédits validés
        if has_missing_evaluations:
            decision_label, decision_code = 'Défaillant', 'DEF'
        elif credits_total > 0:
            if credits_valides >= 60:
                decision_label, decision_code = 'Admis', 'ADM'
            elif not is_master and credits_valides >= 45:
                # Licence uniquement: admis avec dette (compensation autorisée)
                decision_label, decision_code = 'Admis avec dette', 'ADMD'
            else:
                decision_label, decision_code = 'Ajourné', 'AJ'
        
        return {
            'rows': combined_rows,
            'moyenne': round(moyenne_generale, 2),
            'decision_label': decision_label,
            'decision_code': decision_code,
            'moyenne_cat_a': round(moyenne_cat_a, 2),
            'moyenne_cat_b': round(moyenne_cat_b, 2),
            'credits_total': credits_total,
            'credits_total_attendus': credits_total_attendus,
            'credits_valides': credits_valides,
            'credits_non_valides': credits_non_valides,
            'pourcentage': round(pourcentage, 2),
            'has_missing_evaluations': has_missing_evaluations,
            'ec_compensated_ids': ec_compensated_ids,
            'ue_compensated_ids': ue_compensated_ids
        }
    
    # Pour semestriel: logique normale
    # Semestre impair (1,3,5,...) → S1, pair (2,4,6,...) → S2
    type_delib_codes = ['S1' if semestre % 2 == 1 else 'S2']
    
    # Filtrer les délibérations
    deliberations_query = Deliberation.objects.filter(
        matricule_etudiant=etudiant,
        type_deliberation__in=type_delib_codes
    )
    
    # Ajouter le filtre code_classe seulement si spécifié
    if classe_obj:
        deliberations_query = deliberations_query.filter(code_classe=classe_obj)
    
    if annee_academique:
        deliberations_query = deliberations_query.filter(annee_academique=annee_academique)
    
    deliberations = deliberations_query.select_related(
        'code_ue', 'code_ec', 'code_ec__code_ue'
    ).order_by('code_ue__semestre', 'code_ue__code_ue', 'code_ec__code_ec')
    
    # Si pas de délibérations, retourner des données vides
    if not deliberations.exists():
        return {
            'rows': [],
            'moyenne': 0,
            'decision_label': 'Aucune délibération',
            'decision_code': 'NONE',
            'moyenne_cat_a': 0,
            'moyenne_cat_b': 0,
            'credits_total': 0,
            'credits_valides': 0,
            'credits_non_valides': 0,
            'pourcentage': 0,
            'ec_compensated_ids': [],
            'ue_compensated_ids': []
        }
    
    # Regrouper par UE pour l'affichage, mais utiliser les crédits EC pour les calculs
    ue_data = {}
    
    for delib in deliberations:
        # Déterminer l'UE et l'EC
        ec = delib.code_ec
        if ec and ec.code_ue:
            ue = ec.code_ue
            ue_code = ue.code_ue
            # Crédit de l'EC (ou 0 si non défini)
            credit_ec = ec.credit if ec.credit else 0
            is_direct_ue = False
        elif delib.code_ue:
            ue = delib.code_ue
            ue_code = ue.code_ue
            # Pas d'EC, utiliser le crédit de l'UE
            credit_ec = ue.credit if ue.credit else 0
            is_direct_ue = True
        else:
            continue
        
        # Initialiser l'UE si pas encore faite
        if ue_code not in ue_data:
            ue_data[ue_code] = {
                'ue': ue,
                'code_ue': ue_code,
                'intitule_ue': ue.intitule_ue,
                'categorie': getattr(ue, 'categorie', ''),
                'credit_ue': ue.credit or 0,
                'semestre': ue.semestre if ue else None,
                'deliberations': []
            }
        
        # Ajouter la délibération à cette UE avec le crédit de l'EC
        note_finale = delib.calculer_note_finale()
        ue_data[ue_code]['deliberations'].append({
            'delib': delib,
            'code_ec': ec.code_ec if ec else ue_code,
            'intitule_ec': ec.intitule_ue if ec else '-',
            'credit_ec': credit_ec,  # Crédit de l'EC
            'cc': delib.cc,
            'examen': delib.examen,
            'note_finale': note_finale,
            'rattrapage': delib.rattrapage,
            'statut': delib.statut,
            'is_direct_ue': is_direct_ue
        })
    
    # Construire les lignes et calculer les statistiques
    rows = []
    credits_total = 0
    credits_valides = 0
    credits_total_attendus = 0
    
    # Pour les moyennes par catégorie (pondérées par crédits EC)
    points_cat_a = 0
    credits_cat_a = 0
    points_cat_b = 0
    credits_cat_b = 0
    
    # Points totaux pour moyenne générale (pondérés par crédits EC)
    total_points = 0
    total_credits_pondere = 0
    
    for ue_code, data in ue_data.items():
        ue = data['ue']
        categorie = data['categorie']
        
        # Créer une ligne par EC pour l'affichage
        for ec_data in data['deliberations']:
            note_finale = ec_data['note_finale']
            credit_ec = ec_data['credit_ec']
            note_ponderee = note_finale * credit_ec if note_finale is not None else 0
            
            # Utiliser directement le statut du modèle Deliberation (pas de recalcul)
            statut_delib = ec_data['statut']  # VALIDE, NON_VALIDE, EN_COURS
            
            # Mapping du statut pour l'affichage PDF
            statut_mapping = {
                'VALIDE': 'Validé',
                'NON_VALIDE': 'Non validé',
                'EN_COURS': 'En cours'
            }
            statut_affiche = statut_mapping.get(statut_delib, statut_delib)
            
            row = {
                'code_ue': data['code_ue'],
                'intitule_ue': data['intitule_ue'],
                'code_ec': ec_data['code_ec'],
                'intitule_ec': ec_data['intitule_ec'],
                'categorie': categorie,
                'credit': credit_ec,  # Crédit de l'EC (ou UE si pas d'EC)
                'cc': ec_data['cc'],
                'examen': ec_data['examen'],
                'note': note_finale,
                'note_finale': note_finale,
                'note_ponderee': round(note_ponderee, 2),
                'rattrapage': ec_data['rattrapage'],
                'statut': statut_affiche,
                'statut_code': statut_delib,
                'est_valide': statut_delib == 'VALIDE',
                'semestre': data['semestre'],
                'compense': False
            }
            rows.append(row)
            
            # Calculer les statistiques par EC
            credits_total += credit_ec
            
            # Crédits validés: basé sur le statut de l'EC dans Deliberation
            if statut_delib == 'VALIDE':
                credits_valides += credit_ec
            
            # Ajouter aux moyennes (pondérées par crédits EC)
            if note_finale is not None and credit_ec > 0:
                total_points += note_finale * credit_ec
                total_credits_pondere += credit_ec
                
                if categorie == 'A':
                    points_cat_a += note_finale * credit_ec
                    credits_cat_a += credit_ec
                elif categorie == 'B':
                    points_cat_b += note_finale * credit_ec
                    credits_cat_b += credit_ec

    if semestre:
        expected_ec_qs = EC.objects.filter(code_ue__semestre=semestre)
        if classe_obj:
            expected_ec_qs = expected_ec_qs.filter(Q(classe=classe_obj) | Q(code_ue__classe=classe_obj))
        expected_ue_direct_qs = UE.objects.filter(semestre=semestre)
        if classe_obj:
            expected_ue_direct_qs = expected_ue_direct_qs.filter(classe=classe_obj)
        expected_ue_direct_qs = expected_ue_direct_qs.filter(ec__isnull=True)

        expected_ec_credits = expected_ec_qs.aggregate(total=Sum('credit'))['total'] or 0
        expected_ue_direct_credits = expected_ue_direct_qs.aggregate(total=Sum('credit'))['total'] or 0
        credits_total_attendus = expected_ec_credits + expected_ue_direct_credits
    
    # Calculer les moyennes
    moyenne_generale = total_points / total_credits_pondere if total_credits_pondere > 0 else 0
    moyenne_cat_a = points_cat_a / credits_cat_a if credits_cat_a > 0 else 0
    moyenne_cat_b = points_cat_b / credits_cat_b if credits_cat_b > 0 else 0
    
    # Pourcentage
    pourcentage = (moyenne_generale / 20) * 100 if moyenne_generale > 0 else 0
    
    # === COMPENSATION (Licence uniquement, PAS en Master M1/M2) ===
    # En MASTER (M1/M2): aucune compensation semestrielle ni annuelle
    # Critères de compensation (Licence uniquement):
    # 1. Moyenne de la catégorie >= 10
    # 2. Note du cours entre 8 et 9.99
    # Ne jamais changer 'Validé' en 'Non validé', seulement l'inverse
    ec_compensated_ids = []
    ue_compensated_ids = []
    is_master = classe_obj and hasattr(classe_obj, 'code_niveau_id') and str(classe_obj.code_niveau_id) in ('M1', 'M2')
    
    # Parcourir les rows et appliquer/détecter la compensation
    for row in rows:
        note = row['note']
        categorie = row['categorie']
        statut_code = row['statut_code']
        
        # Déterminer la moyenne de la catégorie
        if categorie == 'A':
            moyenne_cat = moyenne_cat_a
        elif categorie == 'B':
            moyenne_cat = moyenne_cat_b
        else:
            moyenne_cat = 0
        
        # Critères de compensation: moyenne_cat >= 10 ET note entre 8 et 9.99
        # En Master: compensation désactivée
        est_compensable = not is_master and moyenne_cat >= 10 and note is not None and 8 <= note < 10
        
        # Pour l'annuel: PAS de nouvelle compensation, juste utiliser les statuts existants
        # L'annuel = somme des crédits capitalisés S1 + S2 (déjà compensés dans chaque semestre)
        if not is_annuel and statut_code == 'NON_VALIDE' and est_compensable:
            # Compenser cet EC (seulement pour semestriel)
            row['statut'] = 'Validé'
            row['statut_code'] = 'VALIDE_COMP'
            row['est_valide'] = True
            row['compense'] = True
            
            # Ajouter aux crédits validés
            credits_valides += row['credit']
            ec_compensated_ids.append(row['code_ec'])
            
        elif statut_code == 'VALIDE' and est_compensable:
            # Déjà compensé dans la base (note < 10 mais statut VALIDE)
            row['compense'] = True
            ec_compensated_ids.append(row['code_ec'])
    
    # Calculer les crédits non validés après compensation
    credits_non_valides = credits_total - credits_valides
    
    # Décision basée sur le total de crédits évalués et les crédits validés
    decision_label = 'A déterminer'
    decision_code = 'ATT'
    
    # Seuil de crédits attendus : 30 pour un semestre
    seuil_credits = 30
    
    # Indicateur : l'étudiant n'a pas atteint le seuil de crédits évalués
    has_missing_evaluations = credits_total < seuil_credits
    
    # Étape 1 : Si credits_total < seuil → Défaillant (évaluations manquantes)
    # Étape 2 : Sinon → décision basée sur les crédits validés
    if has_missing_evaluations:
        decision_label = 'Défaillant'
        decision_code = 'DEF'
    elif credits_total > 0:
        taux_validation = (credits_valides / credits_total) * 100
        if taux_validation >= 100:
            decision_label = 'Admis'
            decision_code = 'ADM'
        elif not is_master and taux_validation >= 70:
            # Licence uniquement: compensation possible
            decision_label = 'Compensable'
            decision_code = 'COMP'
        else:
            decision_label = 'Ajourné'
            decision_code = 'AJ'
    
    return {
        'rows': rows,
        'moyenne': round(moyenne_generale, 2),
        'decision_label': decision_label,
        'decision_code': decision_code,
        'moyenne_cat_a': round(moyenne_cat_a, 2),
        'moyenne_cat_b': round(moyenne_cat_b, 2),
        'credits_total': credits_total,
        'credits_total_attendus': credits_total_attendus,
        'credits_valides': credits_valides,
        'credits_non_valides': credits_non_valides,
        'pourcentage': round(pourcentage, 1),
        'has_missing_evaluations': has_missing_evaluations,
        'ec_compensated_ids': ec_compensated_ids,
        'ue_compensated_ids': ue_compensated_ids
    }


@login_required
def jury_imprimable_profil_pdf(request, matricule):
    """Génère le profil de l'étudiant en PDF avec reportlab"""
    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        logout(request)
        return redirect('login')

    classe_obj = jury.code_classe
    classe_code = request.GET.get('classe', '')
    if request.user.is_staff and classe_code:
        classe_tmp = Classe.objects.filter(code_classe=classe_code).first()
        if classe_tmp:
            classe_obj = classe_tmp

    annee = request.GET.get('annee', '')
    selected_type = request.GET.get('type', 'annuel')
    selected_semestre = request.GET.get('semestre', '')
    if not _require_deliberation_for_imprimable(request, jury, classe_obj, annee, selected_type, selected_semestre):
        return redirect('jury_imprimables')

    type_delib = request.GET.get('type', 'annuel')
    semestre_str = request.GET.get('semestre', '')
    semestre = int(semestre_str) if semestre_str and semestre_str.isdigit() else None

    etudiant = get_object_or_404(Etudiant, matricule_et=matricule)
    
    # Calculer les détails des UE/EC depuis les délibérations
    delib_ues = _jury_compute_delib_ues(classe_obj, etudiant, type_delib, semestre, annee)

    # Récupérer les dettes (InscriptionUE) pour affichage informatif
    from .views_passage_automatique import recuperer_dettes_classe_inferieure
    dettes = recuperer_dettes_classe_inferieure(etudiant, classe_obj, annee)

    # En mode semestriel, filtrer les dettes par parité de semestre
    if type_delib == 'semestriel' and semestre:
        parite = semestre % 2
        dettes_filtrees = []
        for d in dettes:
            if d.code_ec and d.code_ec.code_ue:
                sem_dette = d.code_ec.code_ue.semestre
            elif d.code_ue:
                sem_dette = d.code_ue.semestre
            else:
                sem_dette = None
            if sem_dette is not None and sem_dette % 2 == parite:
                dettes_filtrees.append(d)
        dettes = dettes_filtrees

    from core.utils_profil_pdf import generer_profil_pdf
    return generer_profil_pdf(request, etudiant, classe_obj, annee, semestre, delib_ues, dettes=dettes)


@login_required
def jury_imprimable_pv(request):
    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        logout(request)
        return redirect('login')

    classe_obj = jury.code_classe
    classe_code = request.GET.get('classe', '')
    if request.user.is_staff and classe_code:
        classe_tmp = Classe.objects.filter(code_classe=classe_code).first()
        if classe_tmp:
            classe_obj = classe_tmp

    annee = request.GET.get('annee', '')
    type_delib = request.GET.get('type', 'annuel')
    semestre_str = request.GET.get('semestre', '')
    semestre = int(semestre_str) if semestre_str and semestre_str.isdigit() else None
    date_delib_str = request.GET.get('date_delib', '')
    date_delib = parse_date(date_delib_str) if date_delib_str else None

    selected_type = request.GET.get('type', 'annuel')
    selected_semestre = request.GET.get('semestre', '')
    if not _require_deliberation_for_imprimable(request, jury, classe_obj, annee, selected_type, selected_semestre):
        return redirect('jury_imprimables')

    # Récupérer les inscriptions
    inscriptions = Inscription.objects.filter(code_classe=classe_obj).select_related('matricule_etudiant')
    if annee:
        inscriptions = inscriptions.filter(annee_academique=annee)

    cohortes = list(
        inscriptions.filter(cohorte__isnull=False)
        .values_list('cohorte__code_cohorte', 'cohorte__lib_cohorte')
        .distinct()
    )
    cohorte_label = ''
    if len(cohortes) == 1:
        c_code, c_lib = cohortes[0]
        cohorte_label = f"{c_code} - {c_lib}"
    elif len(cohortes) > 1:
        cohorte_label = f"{len(cohortes)} cohortes"

    # Préparer les données des étudiants depuis les délibérations individuelles
    etudiants_data = []
    # Semestre impair (1,3,5,...) → S1, pair (2,4,6,...) → S2
    type_delib_code = 'ANNEE' if type_delib == 'annuel' else ('S1' if semestre % 2 == 1 else 'S2')
    
    for inscription in inscriptions:
        etudiant = inscription.matricule_etudiant
        
        # Récupérer les délibérations de cet étudiant
        deliberations_etudiant = Deliberation.objects.filter(
            matricule_etudiant=etudiant,
            code_classe=classe_obj,
            annee_academique=annee,
            type_deliberation=type_delib_code
        ).select_related('code_ue', 'code_ec')
        
        if not deliberations_etudiant.exists():
            continue
        
        # Appliquer la logique de compensation
        all_deliberations = list(deliberations_etudiant)
        fails = []
        donors = []
        
        for delib in all_deliberations:
            note_finale = delib.calculer_note_finale()
            if note_finale is not None:
                key = delib.code_ec.code_ec if delib.code_ec else delib.code_ue.code_ue
                if 8 <= note_finale < 10:
                    fails.append((key, note_finale))
                elif note_finale > 10:
                    donors.append((key, note_finale))
        
        # Appliquer la compensation 1-à-1
        compensated_ids = set()
        for f_key, f_note, deficit in sorted([(k, n, 10-n) for k, n in fails], key=lambda x: x[2], reverse=True):
            for d_key, d_note, excess in sorted([(k, n, n-10) for k, n in donors], key=lambda x: x[2]):
                if excess >= deficit and d_key not in compensated_ids:
                    compensated_ids.add(f_key)
                    break
        
        # Calculer les statistiques pour cet étudiant
        total_credits = 0
        credits_valides = 0
        notes = []
        
        for delib in deliberations_etudiant:
            # Crédits
            credit = delib.code_ec.credit if delib.code_ec else (delib.code_ue.credit if delib.code_ue else 0)
            total_credits += credit
            
            # Note finale et validation (utiliser le statut de la table Deliberation + compensation)
            note_finale = delib.calculer_note_finale()
            if note_finale is not None:
                notes.append(note_finale)
                key = delib.code_ec.code_ec if delib.code_ec else delib.code_ue.code_ue
                # Utiliser le statut de la table Deliberation
                est_valide = delib.statut == 'VALIDE' or (delib.statut in ['EN_COURS', 'NON_VALIDE'] and (note_finale >= 10 or (key in compensated_ids)))
                if est_valide:
                    credits_valides += credit
        
        # Calculer la moyenne
        moyenne = sum(notes) / len(notes) if notes else 0
        
        # Déterminer la décision
        decision = 'AJOURNÉ'
        if total_credits > 0:
            taux_validation = (credits_valides / total_credits) * 100
            if taux_validation >= 100:
                decision = 'ADMIS'
            elif taux_validation >= 75:
                decision = 'COMPENSABLE'
        
        etudiant_info = {
            'matricule': etudiant.matricule_et,
            'nom': etudiant.nom_complet,
            'sexe': etudiant.sexe,
            'nationalite': etudiant.nationalite,
            'date_naissance': etudiant.date_naissance,
            'lieu_naissance': etudiant.lieu_naissance,
            'moyenne': round(moyenne, 2),
            'decision': decision,
            'moyenne_cat_a': 0,  # À calculer si nécessaire
            'moyenne_cat_b': 0,  # À calculer si nécessaire
            'credits_capitalises': credits_valides,
            'credits_total': total_credits,
        }
        etudiants_data.append(etudiant_info)

    context = {
        'jury': jury,
        'classe': classe_obj,
        'annee': annee,
        'type_delib': type_delib,
        'semestre': semestre,
        'date_delib': date_delib or timezone.now().date(),
        'cohorte_label': cohorte_label,
        'etudiants': etudiants_data,
        'total_etudiants': len(etudiants_data),
    }
    
    return render(request, 'jury/imprimables/pv.html', context)


@login_required
def jury_imprimable_releve(request, matricule):
    """Génère le relevé de notes de l'étudiant depuis les délibérations"""
    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        logout(request)
        return redirect('login')

    classe_obj = jury.code_classe
    classe_code = request.GET.get('classe', '')
    if request.user.is_staff and classe_code:
        classe_tmp = Classe.objects.filter(code_classe=classe_code).first()
        if classe_tmp:
            classe_obj = classe_tmp

    annee = request.GET.get('annee', '')
    type_delib = request.GET.get('type', 'annuel')
    semestre_str = request.GET.get('semestre', '')
    semestre = int(semestre_str) if semestre_str and semestre_str.isdigit() else None

    selected_type = request.GET.get('type', 'annuel')
    selected_semestre = request.GET.get('semestre', '')
    if not _require_deliberation_for_imprimable(request, jury, classe_obj, annee, selected_type, selected_semestre):
        return redirect('jury_imprimables')

    ins_qs = Inscription.objects.select_related('matricule_etudiant').filter(
        code_classe=classe_obj,
        matricule_etudiant__matricule_et=matricule,
    )
    if annee:
        ins_qs = ins_qs.filter(annee_academique=annee)
    inscription = get_object_or_404(ins_qs)
    etudiant = inscription.matricule_etudiant
    
    # Utiliser la fonction modifiée qui récupère les données depuis Deliberation
    delib = _jury_compute_delib_ues(classe_obj, etudiant, type_delib, semestre, annee)
    
    # Récupérer les dettes (InscriptionUE) pour affichage informatif
    from .views_passage_automatique import recuperer_dettes_classe_inferieure
    dettes = recuperer_dettes_classe_inferieure(etudiant, classe_obj, annee)
    
    cohorte_label = ''
    if inscription.cohorte:
        cohorte_label = str(inscription.cohorte)

    context = {
        'jury': jury,
        'classe': classe_obj,
        'cohorte_label': cohorte_label,
        'inscription': inscription,
        'etudiant': etudiant,
        'lignes': delib['rows'],
        'moyenne': delib['moyenne'],
        'decision': delib['decision_label'],
        'decision_code': delib['decision_code'],
        'moyenne_cat_a': delib['moyenne_cat_a'],
        'moyenne_cat_b': delib['moyenne_cat_b'],
        'credits_total': delib['credits_total'],
        'credits_valides': delib['credits_valides'],
        'credits_non_valides': delib['credits_non_valides'],
        'pourcentage': delib['pourcentage'],
        'dettes': dettes,
        'date': timezone.now().date(),
        'annee': annee,
        'type_delib': type_delib,
        'semestre': semestre,
    }
    return render(request, 'jury/imprimables/releve_etudiant.html', context)


@login_required
def jury_imprimable_profil(request, matricule):
    """Profil détaillé de l'étudiant - relevé complet avec informations personnelles"""
    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        logout(request)
        return redirect('login')

    classe_obj = jury.code_classe
    classe_code = request.GET.get('classe', '')
    if request.user.is_staff and classe_code:
        classe_tmp = Classe.objects.filter(code_classe=classe_code).first()
        if classe_tmp:
            classe_obj = classe_tmp

    annee = request.GET.get('annee', '')
    type_delib = request.GET.get('type', 'annuel')
    semestre_str = request.GET.get('semestre', '')
    semestre = int(semestre_str) if semestre_str and semestre_str.isdigit() else None

    selected_type = request.GET.get('type', 'annuel')
    selected_semestre = request.GET.get('semestre', '')
    if not _require_deliberation_for_imprimable(request, jury, classe_obj, annee, selected_type, selected_semestre):
        return redirect('jury_imprimables')

    # Récupérer l'étudiant
    etudiant = get_object_or_404(Etudiant, matricule_et=matricule)
    
    # Toutes les inscriptions de l'étudiant
    inscriptions = Inscription.objects.filter(
        matricule_etudiant=etudiant
    ).select_related('code_classe', 'cohorte').order_by('-annee_academique')
    
    # Inscription courante
    inscription_courante = inscriptions.filter(code_classe=classe_obj)
    if annee:
        inscription_courante = inscription_courante.filter(annee_academique=annee)
    inscription_courante = inscription_courante.first()

    # Récupérer tous les cours de la classe
    ue_qs = UE.objects.filter(classe__code_classe=classe_obj.code_classe).order_by('semestre', 'code_ue')
    ec_qs = EC.objects.filter(classe__code_classe=classe_obj.code_classe).order_by('code_ue__semestre', 'code_ec')

    if type_delib == 'semestriel' and semestre:
        ue_qs = ue_qs.filter(semestre=semestre)
        ec_qs = ec_qs.filter(code_ue__semestre=semestre)

    delib = _jury_compute_delib_ues(classe_obj, etudiant, type_delib, semestre, annee)

    # Récupérer les dettes (InscriptionUE) pour affichage informatif
    from .views_passage_automatique import recuperer_dettes_classe_inferieure
    dettes = recuperer_dettes_classe_inferieure(etudiant, classe_obj, annee)

    # En mode semestriel, filtrer les dettes par parité de semestre
    # Ex: profil S3 → dettes S1/S3/S5 (impair), profil S4 → dettes S2/S4/S6 (pair)
    if type_delib == 'semestriel' and semestre:
        parite = semestre % 2  # 1 pour impair, 0 pour pair
        dettes_filtrees = []
        for d in dettes:
            if d.code_ec and d.code_ec.code_ue:
                sem_dette = d.code_ec.code_ue.semestre
            elif d.code_ue:
                sem_dette = d.code_ue.semestre
            else:
                sem_dette = None
            if sem_dette is not None and sem_dette % 2 == parite:
                dettes_filtrees.append(d)
        dettes = dettes_filtrees

    context = {
        'jury': jury,
        'classe': classe_obj,
        'etudiant': etudiant,
        'inscription_courante': inscription_courante,
        'inscriptions': inscriptions,
        'lignes': delib['rows'],
        'semestres': delib['semestres'],
        'moyenne_generale': delib['moyenne'],
        'decision': delib['decision_label'],
        'decision_code': delib['decision_code'],
        'moyenne_cat_a': delib['moyenne_cat_a'],
        'moyenne_cat_b': delib['moyenne_cat_b'],
        'total_credits': delib['credits_total'],
        'total_credits_valides': delib['credits_valides'],
        'total_credits_non_valides': delib['credits_non_valides'],
        'taux_reussite': delib['pourcentage'],
        'dettes': dettes,
        'date': timezone.now().date(),
        'annee': annee,
        'type_delib': type_delib,
        'semestre': semestre,
    }
    return render(request, 'jury/imprimables/profil_etudiant.html', context)


def _jury_imprimable_profil_like_pdf(request, matricule, titre):
    from io import BytesIO

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from reportlab.lib import colors

    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        logout(request)
        return redirect('login')

    classe_obj = jury.code_classe
    classe_code = request.GET.get('classe', '')
    if request.user.is_staff and classe_code:
        classe_tmp = Classe.objects.filter(code_classe=classe_code).first()
        if classe_tmp:
            classe_obj = classe_tmp

    annee = request.GET.get('annee', '')
    type_delib = request.GET.get('type', 'annuel')
    semestre_str = request.GET.get('semestre', '')
    semestre = int(semestre_str) if semestre_str and semestre_str.isdigit() else None

    etudiant = get_object_or_404(Etudiant, matricule_et=matricule)
    inscription = Inscription.objects.filter(
        code_classe=classe_obj,
        matricule_etudiant=etudiant,
    )
    if annee:
        inscription = inscription.filter(annee_academique=annee)
    inscription = inscription.select_related('cohorte').first()

    delib = _jury_compute_delib_ues(classe_obj, etudiant, type_delib, semestre, annee)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
        title=titre,
    )
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(titre, styles['Title']))
    story.append(Spacer(1, 0.4 * cm))

    header_rows = [
        ['Nom', etudiant.nom_complet],
        ['Matricule', etudiant.matricule_et],
        ['Classe', classe_obj.code_classe if classe_obj else ''],
        ['Année', annee or ''],
        ['Type', type_delib],
        ['Semestre', str(semestre) if semestre else '-'],
        ['Cohorte', str(inscription.cohorte) if inscription and inscription.cohorte else ''],
    ]
    header_tbl = Table(header_rows, colWidths=[4 * cm, 12.5 * cm])
    header_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
    ]))
    story.append(header_tbl)
    story.append(Spacer(1, 0.5 * cm))

    recap_rows = [
        ['Moyenne', delib.get('moyenne')],
        ['Décision', delib.get('decision_label')],
        ['Moyenne Cat A', delib.get('moyenne_cat_a')],
        ['Moyenne Cat B', delib.get('moyenne_cat_b')],
        ['Crédits total', delib.get('credits_total')],
        ['Crédits validés', delib.get('credits_valides')],
        ['Crédits non validés', delib.get('credits_non_valides')],
        ['Pourcentage', delib.get('pourcentage')],
    ]
    recap_tbl = Table(recap_rows, colWidths=[4 * cm, 12.5 * cm])
    recap_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
    ]))
    story.append(recap_tbl)
    story.append(Spacer(1, 0.5 * cm))

    rows = delib.get('rows') or []
    table_data = [['Code', 'Intitulé', 'Cat', 'Cr', 'Note', 'État']]
    for r in rows:
        table_data.append([
            r.get('code') or '',
            r.get('intitule') or '',
            r.get('categorie') or '',
            r.get('credit') or '',
            r.get('note_finale') if r.get('note_finale') is not None else '',
            r.get('etat') or '',
        ])
    cours_tbl = Table(table_data, repeatRows=1, colWidths=[2.2 * cm, 8.6 * cm, 1.1 * cm, 1.0 * cm, 1.5 * cm, 2.1 * cm])
    cours_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.black),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(cours_tbl)

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="profil_{matricule}.pdf"'
    response.write(pdf)
    return response


@login_required
def jury_imprimable_pv(request):
    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        logout(request)
        return redirect('login')

    classe_obj = jury.code_classe
    classe_code = request.GET.get('classe', '')
    if request.user.is_staff and classe_code:
        classe_tmp = Classe.objects.filter(code_classe=classe_code).first()
        if classe_tmp:
            classe_obj = classe_tmp

    annee = request.GET.get('annee', '')
    type_delib = request.GET.get('type', 'annuel')
    semestre_str = request.GET.get('semestre', '')
    semestre = int(semestre_str) if semestre_str and semestre_str.isdigit() else None
    date_delib_str = request.GET.get('date_delib', '')
    date_delib = parse_date(date_delib_str) if date_delib_str else None

    selected_type = request.GET.get('type', 'annuel')
    selected_semestre = request.GET.get('semestre', '')
    if not _require_deliberation_for_imprimable(request, jury, classe_obj, annee, selected_type, selected_semestre):
        return redirect('jury_imprimables')

    inscriptions = Inscription.objects.filter(code_classe=classe_obj).select_related('matricule_etudiant')
    if annee:
        inscriptions = inscriptions.filter(annee_academique=annee)

    cohortes = list(
        inscriptions.filter(cohorte__isnull=False)
        .values_list('cohorte__code_cohorte', 'cohorte__lib_cohorte')
        .distinct()
    )
    cohorte_label = ''
    if len(cohortes) == 1:
        c_code, c_lib = cohortes[0]
        cohorte_label = f"{c_code} - {c_lib}" if c_lib else str(c_code)
    elif len(cohortes) > 1:
        cohorte_label = ", ".join([c[0] for c in cohortes if c and c[0]])

    resultats = []
    for inscription in inscriptions:
        etudiant = inscription.matricule_etudiant
        evaluations = Evaluation.objects.filter(
            matricule_etudiant=etudiant
        ).filter(
            Q(code_ue__classe=classe_obj) | Q(code_ec__classe=classe_obj)
        )

        if type_delib == 'semestriel' and semestre:
            evaluations = evaluations.filter(
                Q(code_ue__classe=classe_obj, code_ue__semestre=semestre) |
                Q(code_ec__classe=classe_obj, code_ec__code_ue__semestre=semestre)
            )

        notes = []
        for ev in evaluations:
            nf = ev.calculer_note_finale()
            if nf is not None:
                notes.append(nf)

        moyenne = round(sum(notes) / len(notes), 2) if notes else 0
        decision = 'Admis' if moyenne >= 10 else 'Ajourné'
        resultats.append({
            'etudiant': etudiant,
            'moyenne': moyenne,
            'decision': decision,
        })

    resultats.sort(key=lambda e: (-e['moyenne'], e['etudiant'].nom_complet))
    nb_admis = sum(1 for r in resultats if r['decision'] == 'Admis')
    nb_ajournes = len(resultats) - nb_admis

    context = {
        'jury': jury,
        'classe': classe_obj,
        'cohorte_label': cohorte_label,
        'resultats': resultats,
        'nb_admis': nb_admis,
        'nb_ajournes': nb_ajournes,
        'date': date_delib or timezone.now().date(),
        'annee': annee,
        'type_delib': type_delib,
        'semestre': semestre,
    }
    return render(request, 'jury/imprimables/pv.html', context)


# ========== VUES DE GESTION ADMIN ==========

@login_required
@login_required
@require_gestionnaire_or_admin
def gestion_utilisateurs(request):
    """Vue de gestion des utilisateurs avec CRUD complet"""
    
    latest_annee_subquery = Inscription.objects.filter(
        code_classe=OuterRef('code_classe')
    ).order_by('-annee_academique').values('annee_academique')[:1]

    jury_qs = Jury.objects.select_related('code_classe').annotate(
        annee_academique_jury=Subquery(latest_annee_subquery)
    )

    utilisateurs_qs = User.objects.select_related('enseignant', 'etudiant').prefetch_related(
        Prefetch('jury_set', queryset=jury_qs)
    ).order_by('-date_joined')

    utilisateurs = list(utilisateurs_qs)
    utilisateurs_count = utilisateurs_qs.count()

    # Enrichir les comptes JURY existants (anciens comptes sans lien Jury.id_lgn)
    jury_users = [u for u in utilisateurs if getattr(u, 'role', None) == 'JURY']
    for u in utilisateurs:
        u.jury_classe = None
        u.jury_annee = None

    # D'abord, si la relation existe, on la prend
    juries_from_relation = {}
    for u in jury_users:
        jury = None
        try:
            jury = u.jury_set.all()[0]
        except Exception:
            jury = None
        if jury:
            juries_from_relation[u.id] = jury

    # Fallback : associer via username (mêmes conventions que get_simulated_jury)
    # Usernames: jury_pres_<CLASSE>[_YY], jury_sec_<CLASSE>[_YY], ou code_jury directement
    pres_classes = []
    sec_classes = []
    code_jury_usernames = []
    for u in jury_users:
        if u.id in juries_from_relation:
            continue
        username = (u.username or '').strip()
        classe_code = _extract_jury_classe_code(username)
        if classe_code and username.startswith('jury_pres_'):
            pres_classes.append(classe_code)
        elif classe_code and username.startswith('jury_sec_'):
            sec_classes.append(classe_code)
        else:
            code_jury_usernames.append(username)

    juries_by_user_id = {
        j.id_lgn_id: j
        for j in Jury.objects.filter(id_lgn__in=jury_users).select_related('code_classe')
        if j.id_lgn_id
    }
    juries_by_code = {
        j.code_jury: j
        for j in Jury.objects.filter(code_jury__in=code_jury_usernames).select_related('code_classe')
    }
    # Chercher par classe (le suffixe du username est le code_classe)
    all_classe_codes = list(set(pres_classes + sec_classes))
    juries_by_classe = {}
    if all_classe_codes:
        for j in Jury.objects.filter(code_classe__code_classe__in=all_classe_codes).select_related('code_classe').order_by('-annee_academique'):
            # Garder le premier (plus récent) par classe
            if j.code_classe_id not in juries_by_classe:
                juries_by_classe[j.code_classe_id] = j

    jury_by_user = {}
    for u in jury_users:
        jury = juries_from_relation.get(u.id)
        if not jury:
            jury = juries_by_user_id.get(u.id)
        if not jury:
            username = (u.username or '').strip()
            classe_code = _extract_jury_classe_code(username)
            if classe_code:
                jury = juries_by_classe.get(classe_code)
            else:
                jury = juries_by_code.get(username)
        if jury:
            jury_by_user[u.id] = jury

    classes_ids = [j.code_classe_id for j in jury_by_user.values() if getattr(j, 'code_classe_id', None)]
    latest_year_by_class = {}
    if classes_ids:
        latest_year_by_class = {
            row['code_classe']: row['latest']
            for row in Inscription.objects.filter(code_classe__in=classes_ids)
            .values('code_classe')
            .annotate(latest=Max('annee_academique'))
        }

    for u in jury_users:
        jury = jury_by_user.get(u.id)
        if not jury:
            continue
        u.jury_classe = getattr(jury, 'code_classe', None)
        u.jury_annee = jury.annee_academique if getattr(jury, 'annee_academique', None) else latest_year_by_class.get(getattr(jury, 'code_classe_id', None))

    # Enrichir les comptes ETUDIANT avec classe et année (dernière inscription)
    etudiant_users = [u for u in utilisateurs if getattr(u, 'role', None) == 'ETUDIANT' and hasattr(u, 'etudiant') and u.etudiant]
    etudiant_ids = [u.etudiant.matricule_et for u in etudiant_users]
    if etudiant_ids:
        # Récupérer la dernière inscription de chaque étudiant
        latest_inscriptions = {}
        for insc in Inscription.objects.filter(
            matricule_etudiant__in=etudiant_ids
        ).select_related('code_classe').order_by('-annee_academique'):
            if insc.matricule_etudiant_id not in latest_inscriptions:
                latest_inscriptions[insc.matricule_etudiant_id] = insc
        for u in etudiant_users:
            insc = latest_inscriptions.get(u.etudiant.matricule_et)
            if insc:
                u.etu_classe = insc.code_classe
                u.etu_annee = insc.annee_academique
            else:
                u.etu_classe = None
                u.etu_annee = None
    else:
        for u in etudiant_users:
            u.etu_classe = None
            u.etu_annee = None

    form = UserForm()
    
    # Calculer les statistiques des utilisateurs
    stats_utilisateurs = {
        'total_enseignants': User.objects.filter(role='ENSEIGNANT').count(),
        'total_etudiants': User.objects.filter(role='ETUDIANT').count(),
        'total_jury': User.objects.filter(role='JURY').count(),
        'total_admins': User.objects.filter(is_staff=True).count(),
        'total_utilisateurs': utilisateurs_count,
    }
    
    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Utilisateur créé avec succès!')
            return redirect('gestion_utilisateurs')
    
    context = {
        'utilisateurs': utilisateurs,
        'utilisateurs_count': utilisateurs_count,
        'form': form,
        'stats_utilisateurs': stats_utilisateurs,
    }
    return render(request, 'gestion/utilisateurs.html', context)


@login_required
def modifier_utilisateur(request, user_id):
    """Modifier un utilisateur existant"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    utilisateur = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=utilisateur)
        if form.is_valid():
            form.save()
            messages.success(request, 'Utilisateur modifié avec succès!')
            return redirect('gestion_utilisateurs')
    else:
        form = UserEditForm(instance=utilisateur)
    
    context = {
        'form': form,
        'utilisateur': utilisateur,
    }
    return render(request, 'gestion/modifier_utilisateur.html', context)


@login_required
def supprimer_utilisateur(request, user_id):
    """Supprimer un utilisateur"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    utilisateur = get_object_or_404(User, id=user_id)
    
    # Empêcher la suppression de son propre compte
    if utilisateur == request.user:
        messages.error(request, 'Vous ne pouvez pas supprimer votre propre compte!')
        return redirect('gestion_utilisateurs')
    
    if request.method == 'POST':
        username = utilisateur.username
        utilisateur.delete()
        messages.success(request, f'Utilisateur {username} supprimé avec succès!')
        return redirect('gestion_utilisateurs')
    
    context = {
        'utilisateur': utilisateur,
    }
    return render(request, 'gestion/supprimer_utilisateur.html', context)


@login_required
def reinitialiser_mot_de_passe(request, user_id):
    """Réinitialiser le mot de passe d'un utilisateur"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    utilisateur = get_object_or_404(User, id=user_id)
    nouveau_mdp = 'changeme123'
    if getattr(utilisateur, 'role', None) == 'ETUDIANT':
        # Username étudiant = matricule
        nouveau_mdp = f"{utilisateur.username}2025"
    utilisateur.set_password(nouveau_mdp)
    utilisateur.save()
    messages.success(request, f'Mot de passe de {utilisateur.username} réinitialisé à "{nouveau_mdp}"')
    return redirect('gestion_utilisateurs')


@login_required
def toggle_utilisateur_actif(request, user_id):
    """Activer ou désactiver un compte utilisateur"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')

    utilisateur = get_object_or_404(User, id=user_id)

    if utilisateur == request.user:
        messages.error(request, 'Vous ne pouvez pas désactiver votre propre compte.')
        return redirect('gestion_utilisateurs')

    utilisateur.is_active = not utilisateur.is_active
    utilisateur.save()

    etat = 'activé' if utilisateur.is_active else 'désactivé'
    messages.success(request, f'Compte de {utilisateur.username} {etat} avec succès.')
    return redirect('gestion_utilisateurs')


@login_required
def supprimer_utilisateurs_selection(request):
    """Supprimer en lot des utilisateurs sélectionnés dans la liste."""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')

    if request.method == 'POST':
        ids = request.POST.getlist('selected_users')
        if not ids:
            messages.warning(request, 'Aucun utilisateur sélectionné.')
            return redirect('gestion_utilisateurs')

        # Empêcher la suppression de son propre compte
        queryset = User.objects.filter(id__in=ids).exclude(id=request.user.id)
        count = queryset.count()
        queryset.delete()

        if str(request.user.id) in ids:
            messages.warning(request, 'Votre propre compte n\'a pas été supprimé.')

        if count > 0:
            messages.success(request, f'{count} utilisateur(s) supprimé(s) avec succès!')
        else:
            messages.info(request, 'Aucun utilisateur n\'a été supprimé.')

    return redirect('gestion_utilisateurs')


@login_required
def exporter_credentials_utilisateurs(request):
    """Exporter les identifiants des utilisateurs dans un fichier texte (sans réinitialiser les mots de passe)"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    import re
    from django.http import HttpResponse
    
    # Filtre par rôle (paramètre GET optionnel)
    role_filter = request.GET.get('role', '')
    
    utilisateurs = User.objects.select_related('enseignant', 'etudiant').exclude(id=request.user.id)
    
    if role_filter:
        utilisateurs = utilisateurs.filter(role=role_filter)
    
    lines = []
    lines.append("=" * 60)
    if role_filter:
        lines.append(f"LISTE DES IDENTIFIANTS - {dict(User.ROLE_CHOICES).get(role_filter, role_filter).upper()}")
    else:
        lines.append("LISTE DES IDENTIFIANTS UTILISATEURS")
    lines.append(f"Généré le: {timezone.now().strftime('%d/%m/%Y %H:%M')}")
    lines.append("=" * 60)
    lines.append("")
    
    # Ajouter une section d'explication pour les jurys si inclus
    if not role_filter or role_filter == 'JURY':
        lines.append("FORMAT DES COMPTES JURY:")
        lines.append("- Username: jury_[fonction]_[CODE_CLASSE]_[ANNEE]")
        lines.append("- Fonctions: pres (président), sec (secrétaire)")
        lines.append("- Exemple: jury_pres_L1INFO_27 (Président pour classe L1INFO, année 2026-2027)")
        lines.append("- Exemple: jury_sec_L1MEC_27 (Secrétaire pour classe L1MEC, année 2026-2027)")
        lines.append("- Mot de passe: généré aléatoirement à la création du jury")
        lines.append("")
        lines.append("=" * 60)
        lines.append("")
    
    # Pré-charger les mots de passe jury stockés (indexés par username)
    jury_passwords = {}
    for j in Jury.objects.all():
        annee_suffix = j.annee_academique[-2:] if j.annee_academique else ''
        cc = j.code_classe.code_classe if j.code_classe_id else ''
        if annee_suffix and cc:
            if j.password_pres:
                jury_passwords[f'jury_pres_{cc}_{annee_suffix}'] = j.password_pres
            if j.password_sec:
                jury_passwords[f'jury_sec_{cc}_{annee_suffix}'] = j.password_sec
        elif cc:
            if j.password_pres:
                jury_passwords[f'jury_pres_{cc}'] = j.password_pres
            if j.password_sec:
                jury_passwords[f'jury_sec_{cc}'] = j.password_sec
    
    for user in utilisateurs:
        # Récupérer le nom complet
        if hasattr(user, 'enseignant') and user.enseignant:
            nom_complet = user.enseignant.nom_complet
        elif hasattr(user, 'etudiant') and user.etudiant:
            nom_complet = user.etudiant.nom_complet
        else:
            nom_complet = f"{user.first_name} {user.last_name}".strip() or "-"

        # Récupérer le mot de passe: depuis le stockage Jury si disponible, sinon format par défaut
        if user.username in jury_passwords:
            default_password = jury_passwords[user.username]
        else:
            default_password = f"{user.username}2025"
        
        lines.append(f"Nom complet: {nom_complet}")
        lines.append(f"Username: {user.username}")
        lines.append(f"Mot de passe par défaut: {default_password}")
        lines.append(f"Rôle: {user.get_role_display()}")
        
        # Ajouter des informations spécifiques pour les jurys
        if getattr(user, 'role', None) == 'JURY':
            lines.append(f"Type de compte: Jury")
            if user.username.startswith('jury_pres_'):
                lines.append(f"Fonction: Président de Jury")
                classe_code = user.username.replace('jury_pres_', '')
                lines.append(f"Classe assignée: {classe_code}")
            elif user.username.startswith('jury_sec_'):
                lines.append(f"Fonction: Secrétaire de Jury")
                classe_code = user.username.replace('jury_sec_', '')
                lines.append(f"Classe assignée: {classe_code}")
            elif user.username.startswith('jury_membre_'):
                lines.append(f"Fonction: Membre de Jury")
                classe_code = user.username.replace('jury_membre_', '')
                lines.append(f"Classe assignée: {classe_code}")
            else:
                lines.append(f"Fonction: Non spécifiée")
        
        lines.append("-" * 40)
        lines.append("")
    
    lines.append(f"Total: {utilisateurs.count()} utilisateur(s)")
    
    content = "\n".join(lines)
    
    filename = f"identifiants_{role_filter.lower() if role_filter else 'tous'}.txt"
    response = HttpResponse(content, content_type='text/plain; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response


@login_required
@require_staff_or_roles(['GESTIONNAIRE', 'AGENT'])
def gestion_etudiants(request):
    """Vue de gestion des étudiants avec formulaire et grille"""
    
    etudiants = Etudiant.objects.all().select_related('id_lgn')
    form = EtudiantForm()
    
    if request.method == 'POST':
        form = EtudiantForm(request.POST, request.FILES)
        if form.is_valid():
            etudiant = form.save(commit=False)
            # Créer automatiquement un compte utilisateur
            matricule = form.cleaned_data['matricule_et']
            user, created = User.objects.get_or_create(
                username=matricule,
                defaults={'role': 'ETUDIANT', 'is_active': True}
            )
            if created:
                user.set_password(f"{matricule}2025")
                user.save()
            etudiant.id_lgn = user
            etudiant.save()
            messages.success(request, f'Étudiant ajouté avec succès! Compte créé: {matricule} / {matricule}2025')
            return redirect('gestion_etudiants')
    
    context = {
        'etudiants': etudiants,
        'form': form,
    }
    return render(request, 'gestion/etudiants.html', context)


@login_required
@login_required
@require_staff_or_roles(['GESTIONNAIRE', 'AGENT'])
def gestion_enseignants(request):
    """Vue de gestion des enseignants avec formulaire et grille"""
    
    enseignants = Enseignant.objects.all().select_related('code_dpt', 'code_section', 'fonction', 'grade', 'categorie', 'id_lgn')
    departements = Departement.objects.all().select_related('code_section')
    if request.method == 'POST':
        form = EnseignantForm(request.POST, request.FILES)
        
        # D'abord, créer les entrées manquantes dans les tables de réglage AVANT la validation du formulaire
        fonction_code = request.POST.get('fonction')
        if fonction_code:
            Fonction.objects.get_or_create(
                code_fonction=fonction_code,
                defaults={'designation_fonction': fonction_code}
            )
        
        grade_code = request.POST.get('grade')
        if grade_code:
            Grade.objects.get_or_create(
                code_grade=grade_code,
                defaults={'designation_grade': grade_code}
            )
        
        categorie_code = request.POST.get('categorie')
        if categorie_code:
            Categorie.objects.get_or_create(
                code_categorie=categorie_code,
                defaults={'designation_categorie': categorie_code}
            )
        
        dpt_code = request.POST.get('code_dpt')
        if dpt_code:
            # Créer une section par défaut si nécessaire
            default_section, _ = Section.objects.get_or_create(
                code_section='DEFAULT',
                defaults={'designation_section': 'Section par défaut'}
            )
            ReglDepartement.objects.get_or_create(
                code_departement=dpt_code,
                defaults={
                    'designation_departement': dpt_code,
                    'code_section': default_section
                }
            )

        if form.is_valid():
            enseignant = form.save(commit=False)

            # Créer automatiquement un compte utilisateur avec username basé sur le matricule
            matricule = str(form.cleaned_data.get('matricule_en', '')).strip()
            
            if not matricule:
                matricule = 'enseignant'

            # Username : matricule
            base_username = matricule
            
            # S'assurer que le username est unique en ajoutant un suffixe numérique si nécessaire
            username = base_username
            suffix = 1
            while User.objects.filter(username=username).exists():
                username = f"{base_username}{suffix}"
                suffix += 1

            # Mot de passe : username + 2025
            password = f"{username}2025"

            user = User.objects.create_user(
                username=username,
                password=password,
                role='ENSEIGNANT'
            )

            enseignant.id_lgn = user
            enseignant.save()

            # Stocker temporairement les identifiants dans la session pour les afficher après la redirection
            request.session['last_enseignant_credentials'] = {
                'username': username,
                'password': password,
            }

            messages.success(request, 'Enseignant ajouté avec succès ! Les identifiants ont été générés.')
            return redirect('gestion_enseignants')
    else:
        form = EnseignantForm()

    # Récupérer les éventuels identifiants générés lors de la dernière création
    last_credentials = request.session.pop('last_enseignant_credentials', None)
    # Récupérer les éventuels identifiants générés en masse pour les enseignants existants
    bulk_credentials = request.session.pop('bulk_enseignants_credentials', None)

    # Compteurs pour les statistiques
    avec_compte = enseignants.filter(id_lgn__isnull=False).count()
    sans_compte = enseignants.filter(id_lgn__isnull=True).count()

    context = {
        'enseignants': enseignants,
        'departements': departements,
        'form': form,
        'last_credentials': last_credentials,
        'bulk_credentials': bulk_credentials,
        'avec_compte': avec_compte,
        'sans_compte': sans_compte,
    }
    return render(request, 'gestion/enseignants.html', context)


@login_required
def generer_comptes_enseignants_existants(request):
    """Générer des comptes utilisateurs pour les enseignants existants sans compte."""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')

    # On ne traite que les enseignants qui n'ont pas encore de compte lié
    enseignants_sans_compte = Enseignant.objects.filter(id_lgn__isnull=True)

    credentials_list = []

    for enseignant in enseignants_sans_compte:
        # Utiliser le matricule comme username
        matricule = enseignant.matricule_en.strip() if enseignant.matricule_en else 'enseignant'
        
        base_username = matricule

        username = base_username
        suffix = 1
        while User.objects.filter(username=username).exists():
            username = f"{base_username}{suffix}"
            suffix += 1

        # Mot de passe : username + 2025
        password = f"{username}2025"

        user = User.objects.create_user(
            username=username,
            password=password,
            role='ENSEIGNANT'
        )

        enseignant.id_lgn = user
        enseignant.save()

        credentials_list.append({
            'matricule': enseignant.matricule_en,
            'nom_complet': enseignant.nom_complet,
            'username': username,
            'password': password,
        })

    if credentials_list:
        request.session['bulk_enseignants_credentials'] = credentials_list
        messages.success(request, f"Comptes générés pour {len(credentials_list)} enseignant(s) sans compte.")
    else:
        messages.info(request, "Aucun enseignant sans compte utilisateur n'a été trouvé.")

    return redirect('gestion_enseignants')


@login_required
@login_required
@require_staff_or_roles(['GESTIONNAIRE', 'AGENT'])
def gestion_ue(request):
    """Vue de gestion des UE avec formulaire et grille"""
    
    ues = UE.objects.all().select_related('classe')
    
    # Recherche et filtres
    search_query = request.GET.get('q', '')
    semestre_filter = request.GET.get('semestre', '')
    classe_filter = request.GET.get('classe', '')
    
    if search_query:
        ues = ues.filter(
            Q(code_ue__icontains=search_query) | Q(intitule_ue__icontains=search_query)
        )
    if semestre_filter:
        ues = ues.filter(semestre=semestre_filter)
    if classe_filter:
        ues = ues.filter(classe__code_classe=classe_filter)
    
    # Liste des classes pour le filtre
    classes_list = Classe.objects.all().order_by('code_classe')
    
    # Calculer le total des crédits
    total_credits = ues.aggregate(total=models.Sum('credit'))['total'] or 0
    
    form = UEForm()
    
    if request.method == 'POST':
        # Vérifier si c'est une demande d'association automatique
        if 'associer_classes' in request.POST:
            return associer_classes_automatiquement(request)
        
        form = UEForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'UE ajoutée avec succès!')
            return redirect('gestion_ue')
    
    context = {
        'ues': ues,
        'form': form,
        'search_query': search_query,
        'semestre_filter': semestre_filter,
        'classe_filter': classe_filter,
        'classes_list': classes_list,
        'total_credits': total_credits,
    }
    return render(request, 'gestion/ue.html', context)


@login_required
def associer_classes_automatiquement(request):
    """Associer automatiquement les UE aux classes en se basant sur une logique intelligente"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('gestion_ue')
    
    if request.method == 'POST':
        from reglage.models import Classe
        
        classes = Classe.objects.all()
        ues_mises_a_jour = 0
        
        # Logique améliorée: associer les UE en se basant sur les mots-clés dans l'intitulé
        for ue in UE.objects.filter(classe__isnull=True):
            classe_trouvee = None
            intitule_lower = ue.intitule_ue.lower()
            
            # Logique d'association basée sur les mots-clés et le semestre
            if 'informatique' in intitule_lower or 'info' in intitule_lower or 'programmation' in intitule_lower:
                mention = 'INFO'
            elif 'mécanique' in intitule_lower or 'mecanique' in intitule_lower or 'mec' in intitule_lower:
                mention = 'MEC'
            elif 'construction' in intitule_lower or 'cons' in intitule_lower or 'bâtiment' in intitule_lower or 'batiment' in intitule_lower:
                mention = 'CONS'
            else:
                # Par défaut, utiliser INFO si aucun mot-clé spécifique
                mention = 'INFO'
            
            # Déterminer le niveau basé sur le semestre
            if ue.semestre in [1, 2]:
                niveau = 'L1'
            elif ue.semestre in [3, 4]:
                niveau = 'L2'
            elif ue.semestre in [5, 6]:
                niveau = 'L3'
            elif ue.semestre in [7, 8]:
                niveau = 'M1'
            elif ue.semestre in [9, 10]:
                niveau = 'M2'
            else:
                niveau = 'L1'  # Par défaut
            
            # Chercher la classe correspondante
            code_classe_recherche = f"{niveau}{mention}"
            classe_trouvee = classes.filter(code_classe=code_classe_recherche).first()
            
            # Si pas trouvé, prendre la première classe du niveau
            if not classe_trouvee:
                classe_trouvee = classes.filter(code_niveau__code_niveau=niveau).first()
            
            # Associer la classe si trouvée
            if classe_trouvee:
                ue.classe = classe_trouvee
                ue.save()
                ues_mises_a_jour += 1
                print(f"UE {ue.code_ue} associée à {classe_trouvee.code_classe}")
        
        if ues_mises_a_jour > 0:
            messages.success(request, f'{ues_mises_a_jour} UE(s) ont été associées à des classes automatiquement!')
        else:
            messages.info(request, 'Aucune UE n\'a pu être associée (toutes ont déjà une classe ou aucune classe correspondante trouvée).')
        
        return redirect('gestion_ue')


@login_required
@login_required
@require_staff_or_roles(['GESTIONNAIRE', 'AGENT'])
def gestion_ec(request):
    """Vue de gestion des EC avec formulaire et grille"""
    
    ecs = EC.objects.all().select_related('code_ue')
    
    # Recherche et filtres
    search_query = request.GET.get('q', '')
    ue_filter = request.GET.get('ue', '')
    classe_filter = request.GET.get('classe', '')
    
    if search_query:
        ecs = ecs.filter(
            Q(code_ec__icontains=search_query) | Q(intitule_ue__icontains=search_query)
        )
    if ue_filter:
        ecs = ecs.filter(code_ue__code_ue=ue_filter)
    if classe_filter:
        ecs = ecs.filter(classe__code_classe=classe_filter)
    
    # Liste des UE et classes pour les filtres
    ues_list = UE.objects.all().order_by('code_ue')
    classes_list = Classe.objects.all().order_by('code_classe')
    
    form = ECForm()
    
    if request.method == 'POST':
        form = ECForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'EC ajouté avec succès!')
            return redirect('gestion_ec')
    
    context = {
        'ecs': ecs,
        'form': form,
        'search_query': search_query,
        'ue_filter': ue_filter,
        'classe_filter': classe_filter,
        'ues_list': ues_list,
        'classes_list': classes_list,
    }
    return render(request, 'gestion/ec.html', context)


@login_required
@require_gestionnaire_or_admin
def gestion_jurys(request):
    """Vue de gestion des jurys avec formulaire et grille"""
    
    jurys_raw = Jury.objects.all().select_related('code_classe', 'id_lgn')
    
    # Pré-charger tous les enseignants référencés (élimine les requêtes N+1)
    all_matricules = set()
    for jury in jurys_raw:
        all_matricules.add(jury.president)
        all_matricules.add(jury.secretaire)
        if jury.membre:
            all_matricules.add(jury.membre)
    all_matricules.discard('')
    all_matricules.discard(None)
    enseignants_map = {
        ens.matricule_en: ens
        for ens in Enseignant.objects.filter(matricule_en__in=all_matricules).select_related('grade')
    }
    
    # Pré-charger les délibérations par classe (élimine les requêtes N+1)
    delib_set = set(
        Deliberation.objects.filter(
            code_classe__in=[j.code_classe_id for j in jurys_raw],
        ).values_list('code_classe', 'annee_academique', 'type_deliberation')
    )
    
    def _ens_display(matricule):
        """Formater le nom d'affichage d'un enseignant."""
        ens = enseignants_map.get(matricule)
        if ens and ens.grade:
            return f"{ens.grade.code_grade} {ens.nom_complet}"
        return ens.nom_complet if ens else (matricule or '')
    
    # Enrichir les jurys avec les noms complets des enseignants
    jurys = []
    for jury in jurys_raw:
        # Utiliser l'année du jury (pas la dernière année d'inscription)
        annee_for_delib = jury.annee_academique
        
        jurys.append({
            'jury': jury,
            'president_display': _ens_display(jury.president),
            'secretaire_display': _ens_display(jury.secretaire),
            'membre_display': _ens_display(jury.membre) if jury.membre else '',
            'delib_s1': (jury.code_classe_id, annee_for_delib, 'S1') in delib_set,
            'delib_s2': (jury.code_classe_id, annee_for_delib, 'S2') in delib_set,
            'delib_annee': (jury.code_classe_id, annee_for_delib, 'ANNEE') in delib_set,
        })
    
    form = JuryForm()
    
    if request.method == 'POST':
        form = JuryForm(request.POST)
        if form.is_valid():
            jury = form.save(commit=False)
            president = form.cleaned_data.get('president')
            secretaire = form.cleaned_data.get('secretaire')
            
            # Créer utilisateur Président avec rôle JURY
            annee_suffix = jury.annee_academique[-2:] if jury.annee_academique else ''
            username_pres = f"jury_pres_{jury.code_classe.code_classe}_{annee_suffix}" if annee_suffix else f"jury_pres_{jury.code_classe.code_classe}"
            password_pres = get_random_string(12)
            user_pres, created_pres = User.objects.get_or_create(
                username=username_pres,
                defaults={
                    'first_name': president.nom_complet.split()[0] if president.nom_complet else '',
                    'last_name': ' '.join(president.nom_complet.split()[1:]) if president.nom_complet and len(president.nom_complet.split()) > 1 else '',
                    'role': 'JURY',
                    'is_active': True
                }
            )
            user_pres.set_password(password_pres)
            user_pres.is_active = True
            user_pres.role = 'JURY'
            user_pres.save()
            
            # Créer utilisateur Secrétaire avec rôle JURY
            username_sec = f"jury_sec_{jury.code_classe.code_classe}_{annee_suffix}" if annee_suffix else f"jury_sec_{jury.code_classe.code_classe}"
            password_sec = get_random_string(12)
            user_sec, created_sec = User.objects.get_or_create(
                username=username_sec,
                defaults={
                    'first_name': secretaire.nom_complet.split()[0] if secretaire.nom_complet else '',
                    'last_name': ' '.join(secretaire.nom_complet.split()[1:]) if secretaire.nom_complet and len(secretaire.nom_complet.split()) > 1 else '',
                    'role': 'JURY',
                    'is_active': True
                }
            )
            user_sec.set_password(password_sec)
            user_sec.is_active = True
            user_sec.role = 'JURY'
            user_sec.save()
            
            # Sauvegarder avec les matricules
            jury.president = president.matricule_en if president else ''
            jury.secretaire = secretaire.matricule_en if secretaire else ''
            membre = form.cleaned_data.get('membre')
            jury.membre = membre.matricule_en if membre else ''
            # Lier le compte président au jury et stocker les mots de passe
            jury.id_lgn = user_pres
            jury.password_pres = password_pres
            jury.password_sec = password_sec
            jury.save()
            
            messages.success(request, f'Jury créé! Comptes: {username_pres} (mdp: {password_pres}) / {username_sec} (mdp: {password_sec})')
            return redirect('gestion_jurys')
    
    context = {
        'jurys': jurys,
        'form': form,
    }
    return render(request, 'gestion/jurys.html', context)


@login_required
def gestion_cohortes(request):
    """Vue de gestion des cohortes avec formulaire et grille"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    cohortes = Cohorte.objects.all()
    form = CohorteForm()
    
    if request.method == 'POST':
        form = CohorteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cohorte ajoutée avec succès!')
            return redirect('gestion_cohortes')
    
    context = {
        'cohortes': cohortes,
        'form': form,
    }
    return render(request, 'gestion/cohortes.html', context)


@login_required
@login_required
@require_staff_or_roles(['GESTIONNAIRE', 'AGENT'])
def gestion_inscriptions(request):
    """Vue de gestion des inscriptions avec formulaire et grille"""
    
    from reglage.models import AnneeAcademique, Classe as ReglageClasse
    
    inscriptions = Inscription.objects.all().select_related('matricule_etudiant', 'code_classe', 'cohorte')
    
    # Filtres
    classe_filter = request.GET.get('classe', '')
    annee_filter = request.GET.get('annee', '')
    date_filter = request.GET.get('date_filter', '')

    if classe_filter:
        inscriptions = inscriptions.filter(code_classe__code_classe=classe_filter)
    if annee_filter:
        inscriptions = inscriptions.filter(annee_academique=annee_filter)
    if date_filter:
        from datetime import datetime, timedelta
        try:
            date_obj = datetime.strptime(date_filter, '%Y-%m-%d').date()
            date_debut = datetime.combine(date_obj, datetime.min.time())
            date_fin = date_debut + timedelta(days=1)
            inscriptions = inscriptions.filter(
                date_inscription__gte=date_debut,
                date_inscription__lt=date_fin,
            )
        except ValueError:
            pass

    inscriptions = inscriptions.order_by('-date_inscription', 'code_inscription')

    # Dates distinctes pour peupler le combo (Cast compatible MySQL)
    from django.db.models import DateField
    from django.db.models.functions import Cast
    dates_distinctes = (
        Inscription.objects.annotate(date_only=Cast('date_inscription', DateField()))
        .values_list('date_only', flat=True)
        .distinct()
        .order_by('-date_only')
    )
    dates_distinctes = [d for d in dates_distinctes if d is not None]
    
    form = InscriptionForm()
    
    if request.method == 'POST':
        form = InscriptionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Inscription ajoutée avec succès!')
            return redirect('gestion_inscriptions')
    
    # Données pour les filtres
    classes = ReglageClasse.objects.all().order_by('code_classe')
    annees = AnneeAcademique.objects.all().order_by('-code_anac')
    
    context = {
        'inscriptions': inscriptions,
        'form': form,
        'classes': classes,
        'annees': annees,
        'classe_filter': classe_filter,
        'annee_filter': annee_filter,
        'date_filter': date_filter,
        'dates_distinctes': dates_distinctes,
    }
    return render(request, 'gestion/inscriptions.html', context)


@login_required
def gestion_inscriptions_pdf(request):
    """Export PDF de la liste des inscriptions avec filtres actifs"""
    from io import BytesIO
    from datetime import datetime, timedelta
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    inscriptions = Inscription.objects.all().select_related('matricule_etudiant', 'code_classe', 'cohorte')

    classe_filter = request.GET.get('classe', '')
    annee_filter = request.GET.get('annee', '')
    date_filter = request.GET.get('date_filter', '')

    if classe_filter:
        inscriptions = inscriptions.filter(code_classe__code_classe=classe_filter)
    if annee_filter:
        inscriptions = inscriptions.filter(annee_academique=annee_filter)
    if date_filter:
        try:
            date_obj = datetime.strptime(date_filter, '%Y-%m-%d').date()
            date_debut = datetime.combine(date_obj, datetime.min.time())
            date_fin = date_debut + timedelta(days=1)
            inscriptions = inscriptions.filter(date_inscription__gte=date_debut, date_inscription__lt=date_fin)
        except ValueError:
            pass

    inscriptions = inscriptions.order_by('-date_inscription', 'code_inscription')

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm, leftMargin=1.5*cm, rightMargin=1.5*cm)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER, spaceAfter=6, fontName='Helvetica-Bold')
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, spaceAfter=12, textColor=colors.grey)

    elements.append(Paragraph("LISTE DES INSCRIPTIONS", title_style))

    filtres_parts = []
    if classe_filter:
        filtres_parts.append(f"Classe : {classe_filter}")
    if annee_filter:
        filtres_parts.append(f"Année : {annee_filter}")
    if date_filter:
        try:
            d = datetime.strptime(date_filter, '%Y-%m-%d')
            filtres_parts.append(f"Date : {d.strftime('%d/%m/%Y')}")
        except ValueError:
            pass
    if filtres_parts:
        elements.append(Paragraph("Filtres actifs : " + " | ".join(filtres_parts), sub_style))
    else:
        elements.append(Paragraph(f"Toutes les inscriptions — Imprimé le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", sub_style))

    elements.append(Spacer(1, 0.3*cm))

    header = ['N°', 'Code Inscription', 'Date', 'Année Acad.', 'Étudiant', 'Classe', 'Cohorte']
    data = [header]
    for i, ins in enumerate(inscriptions, 1):
        cohorte = ins.cohorte.code_cohorte if ins.cohorte else '-'
        date_str = ins.date_inscription.strftime('%d/%m/%Y') if ins.date_inscription else '-'
        data.append([
            str(i),
            ins.code_inscription,
            date_str,
            ins.annee_academique,
            ins.matricule_etudiant.nom_complet if ins.matricule_etudiant else '-',
            str(ins.code_classe),
            cohorte,
        ])

    col_widths = [1*cm, 3.5*cm, 2.5*cm, 3*cm, 7*cm, 3*cm, 3*cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343a40')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (4, 1), (4, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#dee2e6')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="inscriptions.pdf"'
    return response


# ========== VUES D'ACTIONS (Modifier, Supprimer, Voir) ==========

@login_required
def modifier_etudiant(request, matricule):
    """Modifier un étudiant"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    etudiant = get_object_or_404(Etudiant, matricule_et=matricule)
    
    if request.method == 'POST':
        form = EtudiantForm(request.POST, request.FILES, instance=etudiant)
        if form.is_valid():
            form.save()
            messages.success(request, 'Étudiant modifié avec succès!')
            return redirect('gestion_etudiants')
    else:
        form = EtudiantForm(instance=etudiant)
    
    context = {
        'etudiant': etudiant,
        'form': form,
    }
    return render(request, 'gestion/modifier_etudiant.html', context)


@login_required
def supprimer_etudiant(request, matricule):
    """Supprimer un étudiant"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    etudiant = get_object_or_404(Etudiant, matricule_et=matricule)
    
    if request.method == 'POST':
        etudiant.delete()
        messages.success(request, 'Étudiant supprimé avec succès!')
        return redirect('gestion_etudiants')
    
    context = {'etudiant': etudiant}
    return render(request, 'gestion/supprimer_etudiant.html', context)


@login_required
def voir_etudiant(request, matricule):
    """Voir les détails d'un étudiant"""
    if not request.user.is_staff and request.user.role not in ('JURY', 'ENSEIGNANT'):
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    etudiant = get_object_or_404(Etudiant, matricule_et=matricule)
    inscriptions = Inscription.objects.filter(matricule_etudiant=etudiant).select_related('code_classe')
    evaluations = Evaluation.objects.filter(matricule_etudiant=etudiant).select_related('code_ue', 'code_ec', 'code_ec__code_ue')
    
    # Récupérer les UEs par inscription (via la classe)
    ues_par_inscription = {}
    for insc in inscriptions:
        ues = UE.objects.filter(classe=insc.code_classe).order_by('semestre', 'code_ue')
        if ues.exists():
            ues_par_inscription[insc.code_inscription] = {
                'inscription': insc,
                'ues': ues,
            }
    
    # Récupérer résumé des dettes
    dettes_compensees = InscriptionUE.objects.filter(
        matricule_etudiant=etudiant,
        type_inscription='DETTE_COMPENSEE',
    ).select_related('code_ue', 'code_ec', 'code_classe')
    dettes_liquidees = InscriptionUE.objects.filter(
        matricule_etudiant=etudiant,
        type_inscription='DETTE_LIQUIDEE',
    ).select_related('code_ue', 'code_ec', 'code_classe')

    # Auto-liquider les dettes si Evaluation VALIDE existe
    nb_auto_liquidees = 0
    for dette in dettes_compensees:
        filtre_eval = {'matricule_etudiant': etudiant, 'statut': 'VALIDE',
                       'annee_academique': dette.annee_academique}
        if dette.code_ec:
            filtre_eval['code_ec'] = dette.code_ec
        elif dette.code_ue:
            filtre_eval['code_ue'] = dette.code_ue
        else:
            continue

        # Vérifier Deliberation ou Evaluation
        filtre_delib = {'matricule_etudiant': etudiant, 'statut__in': ['VALIDE', 'VALIDE_COMP']}
        if dette.code_ec:
            filtre_delib['code_ec'] = dette.code_ec
        elif dette.code_ue:
            filtre_delib['code_ue'] = dette.code_ue

        if Deliberation.objects.filter(**filtre_delib).exists() or Evaluation.objects.filter(**filtre_eval).exists():
            dette.type_inscription = 'DETTE_LIQUIDEE'
            dette.save(update_fields=['type_inscription'])
            nb_auto_liquidees += 1

    # Re-compter après auto-liquidation
    nb_compensees = InscriptionUE.objects.filter(
        matricule_etudiant=etudiant, type_inscription='DETTE_COMPENSEE').count()
    nb_liquidees_count = InscriptionUE.objects.filter(
        matricule_etudiant=etudiant, type_inscription='DETTE_LIQUIDEE').count()
    nb_total_dettes = nb_compensees + nb_liquidees_count

    context = {
        'etudiant': etudiant,
        'inscriptions': inscriptions,
        'evaluations': evaluations,
        'ues_par_inscription': ues_par_inscription,
        'nb_dettes_compensees': nb_compensees,
        'nb_dettes_liquidees': nb_liquidees_count,
        'nb_total_dettes': nb_total_dettes,
    }
    return render(request, 'gestion/voir_etudiant.html', context)


@login_required
def historique_academique_etudiant(request, matricule):
    """Historique académique complet d'un étudiant : année → classe → décision → crédits"""
    if not request.user.is_staff and request.user.role not in ('JURY', 'ENSEIGNANT'):
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')

    etudiant = get_object_or_404(Etudiant, matricule_et=matricule)
    inscriptions = Inscription.objects.filter(
        matricule_etudiant=etudiant
    ).select_related('code_classe', 'cohorte').order_by('annee_academique')

    parcours = []
    for ins in inscriptions:
        classe = ins.code_classe
        annee = ins.annee_academique
        niveau = classe.code_niveau_id if classe else '?'
        mention = classe.code_mention_id if classe else '?'

        # Calculer crédits validés pour cette année via Deliberation
        delibs_valides = Deliberation.objects.filter(
            matricule_etudiant=etudiant,
            annee_academique=annee,
            statut__in=['VALIDE', 'VALIDE_COMP'],
            code_classe=classe,
        ).select_related('code_ue', 'code_ec')

        credits_annee = 0
        codes_comptes = set()
        for d in delibs_valides:
            if d.code_ec:
                key = f"EC-{d.code_ec.code_ec}"
                if key not in codes_comptes:
                    credits_annee += d.code_ec.credit or 0
                    codes_comptes.add(key)
            elif d.code_ue:
                key = f"UE-{d.code_ue.code_ue}"
                if key not in codes_comptes:
                    credits_annee += d.code_ue.credit or 0
                    codes_comptes.add(key)

        # Dettes en cours pour cette année
        dettes = InscriptionUE.objects.filter(
            matricule_etudiant=etudiant,
            annee_academique=annee,
            type_inscription='DETTE_COMPENSEE',
        ).count()

        parcours.append({
            'annee': annee,
            'classe': classe.code_classe if classe else '-',
            'niveau': niveau,
            'mention': mention,
            'cohorte': ins.cohorte.code_cohorte if ins.cohorte else '-',
            'decision': ins.get_decision_annuelle_display() if ins.decision_annuelle else None,
            'decision_code': ins.decision_annuelle,
            'credits_valides': credits_annee,
            'dettes': dettes,
        })

    # Crédits cumulés globaux (tous niveaux confondus pour la mention)
    if parcours:
        mention = parcours[-1]['mention']
        all_delibs = Deliberation.objects.filter(
            matricule_etudiant=etudiant,
            statut__in=['VALIDE', 'VALIDE_COMP'],
            code_classe__code_mention_id=mention,
        ).select_related('code_ue', 'code_ec')
        credits_cumules = 0
        codes_comptes = set()
        for d in all_delibs:
            if d.code_ec:
                key = f"EC-{d.code_ec.code_ec}"
                if key not in codes_comptes:
                    credits_cumules += d.code_ec.credit or 0
                    codes_comptes.add(key)
            elif d.code_ue:
                key = f"UE-{d.code_ue.code_ue}"
                if key not in codes_comptes:
                    credits_cumules += d.code_ue.credit or 0
                    codes_comptes.add(key)
    else:
        credits_cumules = 0

    context = {
        'etudiant': etudiant,
        'parcours': parcours,
        'credits_cumules': credits_cumules,
        'nb_annees': len(parcours),
    }
    return render(request, 'gestion/historique_academique.html', context)


@login_required
def suivi_dettes_etudiant(request, matricule):
    """Suivi des dettes (InscriptionUE) d'un étudiant avec liquidation automatique"""
    if not request.user.is_staff and request.user.role not in ('JURY', 'ENSEIGNANT'):
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')

    etudiant = get_object_or_404(Etudiant, matricule_et=matricule)

    # Liquidation automatique : marquer DETTE_LIQUIDEE si une Deliberation ou Evaluation VALIDE existe
    dettes_compensees = InscriptionUE.objects.filter(
        matricule_etudiant=etudiant,
        type_inscription='DETTE_COMPENSEE',
    ).select_related('code_ue', 'code_ec', 'code_classe')

    nb_liquidees = 0
    for dette in dettes_compensees:
        liquidee = False
        # Vérifier d'abord dans Deliberation
        filtre_delib = {'matricule_etudiant': etudiant, 'statut__in': ['VALIDE', 'VALIDE_COMP']}
        if dette.code_ec:
            filtre_delib['code_ec'] = dette.code_ec
        elif dette.code_ue:
            filtre_delib['code_ue'] = dette.code_ue
        else:
            continue
        if Deliberation.objects.filter(**filtre_delib).exists():
            liquidee = True
        else:
            # Vérifier aussi dans Evaluation (notes saisies mais pas encore délibérées)
            filtre_eval = {'matricule_etudiant': etudiant, 'statut': 'VALIDE',
                           'annee_academique': dette.annee_academique}
            if dette.code_ec:
                filtre_eval['code_ec'] = dette.code_ec
            elif dette.code_ue:
                filtre_eval['code_ue'] = dette.code_ue
            if Evaluation.objects.filter(**filtre_eval).exists():
                liquidee = True
        if liquidee:
            dette.type_inscription = 'DETTE_LIQUIDEE'
            dette.save(update_fields=['type_inscription'])
            nb_liquidees += 1

    if nb_liquidees > 0:
        messages.success(request, f'{nb_liquidees} dette(s) automatiquement liquidée(s).')

    # Récupérer toutes les dettes (compensées et liquidées)
    all_dettes = InscriptionUE.objects.filter(
        matricule_etudiant=etudiant,
    ).select_related('code_ue', 'code_ec', 'code_classe').order_by('annee_academique', 'type_inscription')

    dettes_list = []
    for d in all_dettes:
        code = d.code_ec.code_ec if d.code_ec else (d.code_ue.code_ue if d.code_ue else '-')
        intitule = '-'
        credit = 0
        if d.code_ec:
            intitule = d.code_ec.intitule_ue or '-'
            credit = d.code_ec.credit or 0
        elif d.code_ue:
            intitule = d.code_ue.intitule_ue or '-'
            credit = d.code_ue.credit or 0

        dettes_list.append({
            'code': code,
            'intitule': intitule,
            'credit': credit,
            'classe_origine': d.code_classe.code_classe if d.code_classe else '-',
            'annee': d.annee_academique,
            'type': d.get_type_inscription_display(),
            'type_code': d.type_inscription,
            'date': d.date_creation,
        })

    nb_compensees = sum(1 for d in dettes_list if d['type_code'] == 'DETTE_COMPENSEE')
    nb_total_liquidees = sum(1 for d in dettes_list if d['type_code'] == 'DETTE_LIQUIDEE')

    context = {
        'etudiant': etudiant,
        'dettes': dettes_list,
        'nb_compensees': nb_compensees,
        'nb_liquidees': nb_total_liquidees,
        'nb_total': len(dettes_list),
    }
    return render(request, 'gestion/suivi_dettes.html', context)


@login_required
def modifier_enseignant(request, matricule):
    """Modifier un enseignant"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    enseignant = get_object_or_404(Enseignant, matricule_en=matricule)
    
    if request.method == 'POST':
        form = EnseignantForm(request.POST, request.FILES, instance=enseignant)
        if form.is_valid():
            form.save()
            messages.success(request, 'Enseignant modifié avec succès!')
            return redirect('gestion_enseignants')
    else:
        form = EnseignantForm(instance=enseignant)
    
    context = {
        'enseignant': enseignant,
        'form': form,
    }
    return render(request, 'gestion/modifier_enseignant.html', context)


@login_required
def voir_enseignant(request, matricule):
    """Voir les détails d'un enseignant"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    enseignant = get_object_or_404(Enseignant, matricule_en=matricule)
    # Récupérer les évaluations où cet enseignant pourrait être impliqué
    # (à adapter selon votre logique métier)
    
    context = {
        'enseignant': enseignant,
    }
    return render(request, 'gestion/voir_enseignant.html', context)


@login_required
def supprimer_enseignant(request, matricule):
    """Supprimer un enseignant"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    enseignant = get_object_or_404(Enseignant, matricule_en=matricule)
    
    if request.method == 'POST':
        enseignant.delete()
        messages.success(request, 'Enseignant supprimé avec succès!')
        return redirect('gestion_enseignants')
    
    context = {'enseignant': enseignant}
    return render(request, 'gestion/supprimer_enseignant.html', context)


@login_required
def modifier_ue(request, code):
    """Modifier une UE"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    ue = get_object_or_404(UE, code_ue=code)
    
    if request.method == 'POST':
        form = UEForm(request.POST, instance=ue)
        if form.is_valid():
            form.save()
            messages.success(request, 'UE modifiée avec succès!')
            return redirect('gestion_ue')
    else:
        form = UEForm(instance=ue)
    
    context = {
        'ue': ue,
        'form': form,
    }
    return render(request, 'gestion/modifier_ue.html', context)


@login_required
def supprimer_ue(request, code):
    """Supprimer une UE"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    ue = get_object_or_404(UE, code_ue=code)
    
    if request.method == 'POST':
        ue.delete()
        messages.success(request, 'UE supprimée avec succès!')
        return redirect('gestion_ue')
    
    context = {'ue': ue}
    return render(request, 'gestion/supprimer_ue.html', context)


@login_required
def modifier_ec(request, code):
    """Modifier un EC"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    ec = get_object_or_404(EC, code_ec=code)
    
    if request.method == 'POST':
        form = ECForm(request.POST, instance=ec)
        if form.is_valid():
            form.save()
            messages.success(request, 'EC modifié avec succès!')
            return redirect('gestion_ec')
    else:
        form = ECForm(instance=ec)
    
    context = {
        'ec': ec,
        'form': form,
    }
    return render(request, 'gestion/modifier_ec.html', context)


@login_required
def supprimer_ec(request, code):
    """Supprimer un EC"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    ec = get_object_or_404(EC, code_ec=code)
    
    if request.method == 'POST':
        ec.delete()
        messages.success(request, 'EC supprimé avec succès!')
        return redirect('gestion_ec')
    
    context = {'ec': ec}
    return render(request, 'gestion/supprimer_ec.html', context)


# ========== VUES AJAX ========== 

def get_section_for_departement(request):
    """Vue AJAX pour récupérer la section d'un département."""
    departement_id = request.GET.get('departement_id')
    if not departement_id:
        return JsonResponse({'error': 'ID du département manquant.'}, status=400)

    try:
        # Utiliser le modèle de l'application 'reglage' pour la requête
        departement = ReglDepartement.objects.select_related('code_section').get(pk=departement_id)
        section = departement.code_section
        if section:
            data = {
                'code_section': section.pk,
                'designation_section': section.designation_section
            }
            return JsonResponse(data)
        else:
            return JsonResponse({'error': 'Aucune section trouvée pour ce département.'}, status=404)
    except ReglDepartement.DoesNotExist:
        return JsonResponse({'error': 'Département non trouvé.'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def modifier_jury(request, code):
    """Modifier un jury"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    jury = get_object_or_404(Jury, code_jury=code)
    old_president = jury.president
    old_secretaire = jury.secretaire
    
    if request.method == 'POST':
        form = JuryForm(request.POST, instance=jury)
        if form.is_valid():
            new_jury = form.save(commit=False)
            new_president = form.cleaned_data.get('president')
            new_secretaire = form.cleaned_data.get('secretaire')
            new_pres_mat = new_president.matricule_en if new_president else ''
            new_sec_mat = new_secretaire.matricule_en if new_secretaire else ''
            
            annee_suffix = new_jury.annee_academique[-2:] if new_jury.annee_academique else ''
            
            # Si le président a changé, mettre à jour le compte User
            if new_pres_mat != old_president:
                username_pres = f"jury_pres_{new_jury.code_classe.code_classe}_{annee_suffix}" if annee_suffix else f"jury_pres_{new_jury.code_classe.code_classe}"
                password_pres = get_random_string(12)
                user_pres, _ = User.objects.get_or_create(
                    username=username_pres,
                    defaults={
                        'first_name': new_president.nom_complet.split()[0] if new_president and new_president.nom_complet else '',
                        'last_name': ' '.join(new_president.nom_complet.split()[1:]) if new_president and new_president.nom_complet and len(new_president.nom_complet.split()) > 1 else '',
                        'role': 'JURY',
                        'is_active': True
                    }
                )
                user_pres.set_password(password_pres)
                user_pres.first_name = new_president.nom_complet.split()[0] if new_president and new_president.nom_complet else ''
                user_pres.last_name = ' '.join(new_president.nom_complet.split()[1:]) if new_president and new_president.nom_complet and len(new_president.nom_complet.split()) > 1 else ''
                user_pres.is_active = True
                user_pres.role = 'JURY'
                user_pres.save()
                new_jury.id_lgn = user_pres
                new_jury.password_pres = password_pres
                messages.info(request, f'Nouveau compte président: {username_pres} (mdp: {password_pres})')
            
            # Si le secrétaire a changé, mettre à jour le compte User
            if new_sec_mat != old_secretaire:
                username_sec = f"jury_sec_{new_jury.code_classe.code_classe}_{annee_suffix}" if annee_suffix else f"jury_sec_{new_jury.code_classe.code_classe}"
                password_sec = get_random_string(12)
                user_sec, _ = User.objects.get_or_create(
                    username=username_sec,
                    defaults={
                        'first_name': new_secretaire.nom_complet.split()[0] if new_secretaire and new_secretaire.nom_complet else '',
                        'last_name': ' '.join(new_secretaire.nom_complet.split()[1:]) if new_secretaire and new_secretaire.nom_complet and len(new_secretaire.nom_complet.split()) > 1 else '',
                        'role': 'JURY',
                        'is_active': True
                    }
                )
                user_sec.set_password(password_sec)
                user_sec.first_name = new_secretaire.nom_complet.split()[0] if new_secretaire and new_secretaire.nom_complet else ''
                user_sec.last_name = ' '.join(new_secretaire.nom_complet.split()[1:]) if new_secretaire and new_secretaire.nom_complet and len(new_secretaire.nom_complet.split()) > 1 else ''
                user_sec.is_active = True
                user_sec.role = 'JURY'
                user_sec.save()
                new_jury.password_sec = password_sec
                messages.info(request, f'Nouveau compte secrétaire: {username_sec} (mdp: {password_sec})')
            
            new_jury.president = new_pres_mat
            new_jury.secretaire = new_sec_mat
            membre = form.cleaned_data.get('membre')
            new_jury.membre = membre.matricule_en if membre else ''
            new_jury.save()
            
            messages.success(request, 'Jury modifié avec succès!')
            return redirect('gestion_jurys')
    else:
        form = JuryForm(instance=jury)
    
    context = {
        'jury': jury,
        'form': form,
    }
    return render(request, 'gestion/modifier_jury.html', context)


@login_required
def supprimer_jury(request, code):
    """Supprimer un jury et désactiver les comptes User associés"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    jury = get_object_or_404(Jury, code_jury=code)
    
    if request.method == 'POST':
        # Désactiver les comptes User liés au jury
        annee_suffix = jury.annee_academique[-2:] if jury.annee_academique else ''
        classe_code = jury.code_classe.code_classe if jury.code_classe else ''
        usernames_to_deactivate = []
        if annee_suffix and classe_code:
            usernames_to_deactivate = [
                f"jury_pres_{classe_code}_{annee_suffix}",
                f"jury_sec_{classe_code}_{annee_suffix}",
            ]
        elif classe_code:
            usernames_to_deactivate = [
                f"jury_pres_{classe_code}",
                f"jury_sec_{classe_code}",
            ]
        
        deactivated = User.objects.filter(username__in=usernames_to_deactivate, role='JURY').update(is_active=False)
        
        jury.delete()
        messages.success(request, f'Jury supprimé avec succès! ({deactivated} compte(s) désactivé(s))')
        return redirect('gestion_jurys')
    
    context = {'jury': jury}
    return render(request, 'gestion/supprimer_jury.html', context)


@login_required
def modifier_cohorte(request, code):
    """Modifier une cohorte"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    cohorte = get_object_or_404(Cohorte, code_cohorte=code)
    
    if request.method == 'POST':
        form = CohorteForm(request.POST, instance=cohorte)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cohorte modifiée avec succès!')
            return redirect('gestion_cohortes')
    else:
        form = CohorteForm(instance=cohorte)
    
    context = {
        'cohorte': cohorte,
        'form': form,
    }
    return render(request, 'gestion/modifier_cohorte.html', context)


@login_required
def supprimer_cohorte(request, code):
    """Supprimer une cohorte"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    cohorte = get_object_or_404(Cohorte, code_cohorte=code)
    
    if request.method == 'POST':
        cohorte.delete()
        messages.success(request, 'Cohorte supprimée avec succès!')
        return redirect('gestion_cohortes')
    
    context = {'cohorte': cohorte}
    return render(request, 'gestion/supprimer_cohorte.html', context)


@login_required
def voir_cohorte(request, code):
    """Voir les détails d'une cohorte et ses étudiants"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    cohorte = get_object_or_404(Cohorte, code_cohorte=code)
    etudiants = Etudiant.objects.filter(code_cohorte=cohorte)
    
    context = {
        'cohorte': cohorte,
        'etudiants': etudiants,
    }
    return render(request, 'gestion/voir_cohorte.html', context)


@login_required
def modifier_inscription(request, code):
    """Modifier une inscription"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    inscription = get_object_or_404(Inscription, code_inscription=code)
    
    if request.method == 'POST':
        form = InscriptionForm(request.POST, instance=inscription)
        if form.is_valid():
            form.save()
            messages.success(request, 'Inscription modifiée avec succès!')
            return redirect('gestion_inscriptions')
    else:
        form = InscriptionForm(instance=inscription)
    
    context = {
        'inscription': inscription,
        'form': form,
    }
    return render(request, 'gestion/modifier_inscription.html', context)


@login_required
def supprimer_inscription(request, code):
    """Supprimer une inscription"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    inscription = get_object_or_404(Inscription, code_inscription=code)
    
    if request.method == 'POST':
        inscription.delete()
        messages.success(request, 'Inscription supprimée avec succès!')
        return redirect('gestion_inscriptions')
    
    context = {'inscription': inscription}
    return render(request, 'gestion/supprimer_inscription.html', context)


@login_required
def voir_inscription(request, code):
    """Voir les détails d'une inscription"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    inscription = get_object_or_404(Inscription, code_inscription=code)
    etudiant = inscription.matricule_etudiant
    evaluations = Evaluation.objects.filter(matricule_etudiant=etudiant)
    
    context = {
        'inscription': inscription,
        'etudiant': etudiant,
        'evaluations': evaluations,
    }
    return render(request, 'gestion/voir_inscription.html', context)


# ============================================
# VUES POUR ATTRIBUTIONS
# ============================================

@login_required
@login_required
@require_gestionnaire_or_admin
def gestion_attributions(request):
    """Vue pour gérer les attributions UE/EC aux enseignants"""
    
    from .models import Attribution, Enseignant, UE, EC, CoursAttribution
    from .forms import AttributionForm
    from reglage.models import AnneeAcademique, TypeCharge
    
    # Filtres
    enseignant_filter = request.GET.get('enseignant', '')
    annee_filter = request.GET.get('annee', '')
    cours_filter = request.GET.get('cours', '')
    type_charge_filter = request.GET.get('type_charge', '')
    
    # Récupérer l'enseignant sélectionné
    enseignant_selectionne = None
    if enseignant_filter:
        enseignant_selectionne = Enseignant.objects.filter(matricule_en=enseignant_filter).select_related('grade').first()
    
    # Récupérer toutes les attributions
    attributions = Attribution.objects.all().select_related('matricule_en', 'type_charge')
    
    # Appliquer les filtres
    if enseignant_filter:
        attributions = attributions.filter(matricule_en__matricule_en=enseignant_filter)
    if annee_filter:
        attributions = attributions.filter(annee_academique=annee_filter)
    if type_charge_filter:
        attributions = attributions.filter(type_charge__code_type=type_charge_filter)
    
    # Calculer les statistiques
    total_cmi = 0
    total_tdtp = 0
    total_cours = attributions.count()
    
    # Traitement du formulaire d'ajout
    form = AttributionForm()
    if request.method == 'POST':
        form = AttributionForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Attribution ajoutée avec succès!')
                return redirect('gestion_attributions')
            except Exception as e:
                messages.error(request, f'Erreur lors de l\'ajout: {str(e)}')
    
    # Récupérer les cours pour attribution (EC en premier, puis UE)
    cours_attribution = CoursAttribution.objects.all().select_related('classe').order_by('type_cours', 'code_cours')
    
    # Calculer le total des crédits
    from django.db.models import Sum
    total_credits = cours_attribution.aggregate(Sum('credit'))['credit__sum'] or 0
    
    context = {
        'attributions': attributions,
        'cours_attribution': cours_attribution,
        'total_credits': total_credits,
        'form': form,
        'enseignants': Enseignant.objects.all(),
        'ues': UE.objects.all(),
        'ecs': EC.objects.all(),
        'annees': AnneeAcademique.objects.all(),
        'annee_en_cours': AnneeAcademique.get_annee_en_cours(),
        'types_charge': TypeCharge.objects.all(),
        'total_cmi': total_cmi,
        'total_tdtp': total_tdtp,
        'total_cmi_tdtp': total_cmi + total_tdtp,
        'total_cours': cours_attribution.count(),
        'enseignant_filter': enseignant_filter,
        'enseignant_selectionne': enseignant_selectionne,
        'annee_filter': annee_filter,
        'cours_filter': cours_filter,
        'type_charge_filter': type_charge_filter,
    }
    return render(request, 'gestion/attributions.html', context)


@login_required
def supprimer_attribution(request, code):
    """Supprimer une attribution"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    from .models import Attribution
    attribution = get_object_or_404(Attribution, code_attribution=code)
    attribution.delete()
    messages.success(request, 'Attribution supprimée avec succès!')
    return redirect('gestion_attributions')


@login_required
def modifier_attribution(request, code):
    """Modifier une attribution"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    from .models import Attribution
    attribution = get_object_or_404(Attribution, code_attribution=code)
    
    if request.method == 'POST':
        matricule_en = request.POST.get('matricule_en')
        type_charge = request.POST.get('type_charge')
        annee_academique = request.POST.get('annee_academique')
        
        if matricule_en and annee_academique:
            from reglage.models import Enseignant, TypeCharge
            enseignant = get_object_or_404(Enseignant, matricule_en=matricule_en)
            
            # Mettre à jour l'attribution
            attribution.matricule_en = enseignant
            attribution.annee_academique = annee_academique
            
            # Mettre à jour le type de charge si fourni
            if type_charge:
                type_charge_obj = TypeCharge.objects.filter(code_type=type_charge).first()
                attribution.type_charge = type_charge_obj
            
            attribution.save()
            messages.success(request, 'Attribution modifiée avec succès!')
            return redirect('liste_attributions')
        else:
            messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
    
    # Pour les requêtes GET ou en cas d'erreur, rediriger vers la liste
    return redirect('liste_attributions')


@login_required
@csrf_protect
def ajouter_cours_attribution(request):
    """Ajouter un UE ou EC à la table cours_attribution"""
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Accès non autorisé'})
    
    if request.method == 'POST':
        from .models import UE, EC, CoursAttribution
        import json
        
        try:
            data = json.loads(request.body)
            code = data.get('code')
            type_cours = data.get('type')
            
            if not code or not type_cours:
                return JsonResponse({'success': False, 'error': 'Données manquantes'})
            
            # Vérifier si le cours existe déjà
            if CoursAttribution.objects.filter(code_cours=code).exists():
                return JsonResponse({'success': False, 'error': 'Ce cours est déjà dans la liste'})
            
            if type_cours == 'UE':
                ue = UE.objects.get(code_ue=code)
                CoursAttribution.objects.create(
                    code_cours=ue.code_ue,
                    intitule=ue.intitule_ue,
                    type_cours='UE',
                    code_ue_parent=None,
                    credit=ue.credit,
                    semestre=ue.semestre,
                    classe=ue.classe
                )
            else:  # EC
                ec = EC.objects.get(code_ec=code)
                CoursAttribution.objects.create(
                    code_cours=ec.code_ec,
                    intitule=ec.intitule_ue,
                    type_cours='EC',
                    code_ue_parent=ec.code_ue_id,
                    credit=ec.credit,
                    semestre=ec.code_ue.semestre if ec.code_ue else 0,
                    classe=ec.classe
                )
            
            return JsonResponse({'success': True, 'message': f'{type_cours} ajouté avec succès'})
        except UE.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'UE non trouvée'})
        except EC.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'EC non trouvé'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})


@login_required
@csrf_protect
def supprimer_cours_attribution(request):
    """Supprimer un cours individuel de la table cours_attribution"""
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Accès non autorisé'})
    
    if request.method == 'POST':
        from .models import CoursAttribution
        import json
        
        try:
            data = json.loads(request.body)
            code = data.get('code')
            
            if not code:
                return JsonResponse({'success': False, 'error': 'Code manquant'})
            
            # Rechercher et supprimer le cours
            cours = CoursAttribution.objects.filter(code_cours=code).first()
            if not cours:
                return JsonResponse({'success': False, 'error': 'Cours non trouvé'})
            
            type_cours = cours.type_cours
            cours.delete()
            
            return JsonResponse({'success': True, 'message': f'{type_cours} supprimé avec succès'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})


@login_required
def vider_cours_attribution(request):
    """Vider la table cours_attribution"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    if request.method == 'POST':
        from .models import CoursAttribution
        
        try:
            count = CoursAttribution.objects.count()
            CoursAttribution.objects.all().delete()
            messages.success(request, f'Table vidée avec succès ! {count} cours supprimés.')
        except Exception as e:
            messages.error(request, f'Erreur lors de la suppression : {str(e)}')
        
        return redirect('gestion_attributions')
    
    return redirect('gestion_attributions')


@login_required
def liste_attributions(request):
    """Afficher la liste des cours attribués"""
    if not (request.user.is_staff or request.user.role == 'GESTIONNAIRE'):
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    from .models import Attribution, Enseignant, UE, EC
    from reglage.models import AnneeAcademique, Classe
    
    # Filtres
    enseignant_filter = request.GET.get('enseignant', '')
    annee_filter = request.GET.get('annee', '')
    classe_filter = request.GET.get('classe', '')
    
    # Récupérer l'enseignant sélectionné
    enseignant_selectionne = None
    if enseignant_filter:
        enseignant_selectionne = Enseignant.objects.filter(matricule_en=enseignant_filter).select_related('grade').first()
    
    # Récupérer les attributions
    attributions = Attribution.objects.all().select_related('matricule_en', 'type_charge').order_by('-date_attribution')
    
    if enseignant_filter:
        attributions = attributions.filter(matricule_en__matricule_en=enseignant_filter)
    if annee_filter:
        attributions = attributions.filter(annee_academique=annee_filter)
    if classe_filter:
        # Filtrer par classe (chercher dans UE et EC)
        ue_codes = list(UE.objects.filter(classe__code_classe=classe_filter).values_list('code_ue', flat=True))
        ec_codes = list(EC.objects.filter(classe__code_classe=classe_filter).values_list('code_ec', flat=True))
        cours_codes = ue_codes + ec_codes
        attributions = attributions.filter(code_cours__in=cours_codes)
    
    # Calculer le total des crédits
    total_credits = 0
    for attr in attributions:
        # Chercher dans UE
        try:
            ue = UE.objects.get(code_ue=attr.code_cours)
            total_credits += ue.credit
        except UE.DoesNotExist:
            # Chercher dans EC
            try:
                ec = EC.objects.get(code_ec=attr.code_cours)
                total_credits += ec.credit
            except EC.DoesNotExist:
                pass
    
    # Données pour les filtres
    enseignants = Enseignant.objects.all().order_by('nom_complet')
    annees = AnneeAcademique.objects.all().order_by('-code_anac')
    classes = Classe.objects.all().order_by('code_classe')
    from reglage.models import TypeCharge
    types_charge = TypeCharge.objects.all().order_by('designation_typecharge')
    
    # Récupérer toutes les UEs et ECs pour afficher les crédits
    ues = UE.objects.all()
    ecs = EC.objects.all()
    
    context = {
        'attributions': attributions,
        'enseignants': enseignants,
        'annees': annees,
        'classes': classes,
        'types_charge': types_charge,
        'ues': ues,
        'ecs': ecs,
        'enseignant_filter': enseignant_filter,
        'enseignant_selectionne': enseignant_selectionne,
        'annee_filter': annee_filter,
        'classe_filter': classe_filter,
        'total_attributions': attributions.count(),
        'total_credits': total_credits,
    }
    
    return render(request, 'gestion/liste_attributions.html', context)


@login_required
def import_attributions(request):
    """Importer des attributions depuis un fichier Excel"""
    # Vérifier l'accès : gestionnaire ou admin
    if not (request.user.role in ['ADMIN', 'GESTIONNAIRE'] or request.user.is_staff):
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        from .models import Attribution, Enseignant
        from reglage.models import TypeCharge
        import uuid
        
        excel_file = request.FILES['excel_file']
        
        try:
            df = pd.read_excel(excel_file)
            
            # Mapping flexible des colonnes possibles
            column_mapping = {
                # Matricule enseignant
                'matricule_en': ['matricule_en', 'matricule enseignant', 'matricule', 'enseignant', 'professeur', 'prof'],
                # Code cours
                'code_cours': ['code_cours', 'code cours', 'code_ue', 'code ec', 'code_ec', 'cours', 'ue', 'ec', 'module'],
                # Type charge
                'type_charge': ['type_charge', 'type charge', 'type', 'charge', 'nature'],
                # Année académique
                'annee_academique': ['annee_academique', 'année académique', 'annee', 'année', 'academique', 'académique', '2024-2025']
            }
            
            # Trouver les colonnes correspondantes dans le fichier
            found_columns = {}
            for standard_name, possible_names in column_mapping.items():
                found_col = None
                for col in df.columns:
                    if str(col).strip().lower() in [name.lower() for name in possible_names]:
                        found_col = col
                        break
                if found_col:
                    found_columns[standard_name] = found_col
                elif standard_name in ['matricule_en', 'code_cours', 'annee_academique']:
                    messages.error(request, f'Colonne requise manquante: {standard_name}. Colonnes possibles: {", ".join(possible_names)}')
                    return redirect('liste_attributions')
            
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    ligne = index + 2  # +2 car Excel commence à 1 et il y a l'en-tête
                    matricule = str(row[found_columns['matricule_en']]).strip()
                    code_cours = str(row[found_columns['code_cours']]).strip()
                    annee = str(row[found_columns['annee_academique']]).strip()
                    
                    # Récupérer l'enseignant
                    enseignant = Enseignant.objects.filter(matricule_en=matricule).first()
                    if not enseignant:
                        errors.append(f"Ligne {ligne}: Enseignant '{matricule}' non trouvé")
                        error_count += 1
                        continue
                    
                    # Récupérer le type de charge si fourni (par code ou désignation)
                    type_charge = None
                    if 'type_charge' in found_columns and pd.notna(row[found_columns['type_charge']]):
                        val = str(row[found_columns['type_charge']]).strip()
                        type_charge = TypeCharge.objects.filter(code_type=val).first()
                        if not type_charge:
                            type_charge = TypeCharge.objects.filter(designation_typecharge__iexact=val).first()
                        if not type_charge:
                            errors.append(f"Ligne {ligne}: Type charge '{val}' non trouvé (ignoré)")
                    
                    # Générer un code unique
                    code_attribution = f"ATT{uuid.uuid4().hex[:8].upper()}"
                    
                    # Créer l'attribution
                    Attribution.objects.create(
                        code_attribution=code_attribution,
                        matricule_en=enseignant,
                        code_cours=code_cours,
                        type_charge=type_charge,
                        annee_academique=annee
                    )
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Ligne {ligne}: Erreur - {str(e)}")
                    error_count += 1
            
            # Afficher le résumé
            if success_count > 0:
                messages.success(request, f'{success_count} attribution(s) importée(s) avec succès.')
            
            # Afficher les erreurs détaillées (max 10)
            if errors:
                for err in errors[:10]:
                    messages.warning(request, err)
                if len(errors) > 10:
                    messages.warning(request, f'... et {len(errors) - 10} autre(s) erreur(s)')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'importation: {str(e)}')
        
        return redirect('liste_attributions')
    
    return render(request, 'gestion/import_attributions_progress.html')


@login_required
def import_attributions_ajax(request):
    """API AJAX pour importer des attributions avec progression"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Accès non autorisé'}, status=403)
    
    if request.method != 'POST' or not request.FILES.get('excel_file'):
        return JsonResponse({'error': 'Fichier manquant'}, status=400)
    
    excel_file = request.FILES['excel_file']
    
    try:
        df = pd.read_excel(excel_file)
        total = len(df)
        success_count = 0
        error_count = 0
        errors = []
        
        # Mapping flexible des colonnes possibles
        column_mapping = {
            # Matricule enseignant
            'matricule_en': ['matricule_en', 'matricule enseignant', 'matricule', 'enseignant', 'professeur', 'prof'],
            # Code cours
            'code_cours': ['code_cours', 'code cours', 'code_ue', 'code ec', 'code_ec', 'cours', 'ue', 'ec', 'module'],
            # Type charge
            'type_charge': ['type_charge', 'type charge', 'type', 'charge', 'nature'],
            # Année académique
            'annee_academique': ['annee_academique', 'année académique', 'annee', 'année', 'academique', 'académique', '2024-2025']
        }
        
        # Trouver les colonnes correspondantes dans le fichier
        found_columns = {}
        for standard_name, possible_names in column_mapping.items():
            found_col = None
            for col in df.columns:
                if str(col).strip().lower() in [name.lower() for name in possible_names]:
                    found_col = col
                    break
            if found_col:
                found_columns[standard_name] = found_col
            elif standard_name in ['matricule_en', 'code_cours', 'annee_academique']:
                errors.append(f'Colonne requise manquante: {standard_name}. Colonnes possibles: {", ".join(possible_names)}')
                error_count += 1
        
        for index, row in df.iterrows():
            ligne = index + 2
            try:
                # Récupérer les valeurs obligatoires
                matricule_en = str(row[found_columns['matricule_en']]).strip()
                code_cours = str(row[found_columns['code_cours']]).strip()
                annee = str(row[found_columns['annee_academique']]).strip()
                
                # Vérifier l'enseignant
                enseignant = Enseignant.objects.filter(matricule_en=matricule_en).first()
                if not enseignant:
                    errors.append(f"Ligne {ligne}: Enseignant '{matricule_en}' non trouvé")
                    error_count += 1
                    continue
                
                # Vérifier si c'est un UE ou EC
                cours = None
                try:
                    cours = UE.objects.get(code_ue=code_cours)
                except UE.DoesNotExist:
                    try:
                        cours = EC.objects.get(code_ec=code_cours)
                    except EC.DoesNotExist:
                        errors.append(f"Ligne {ligne}: Cours '{code_cours}' non trouvé (ni UE ni EC)")
                        error_count += 1
                        continue
                
                # Récupérer le type de charge (optionnel)
                type_charge = None
                if 'type_charge' in found_columns and pd.notna(row[found_columns['type_charge']]):
                    val = str(row[found_columns['type_charge']]).strip()
                    from reglage.models import TypeCharge
                    type_charge = TypeCharge.objects.filter(code_type=val).first()
                    if not type_charge:
                        type_charge = TypeCharge.objects.filter(designation_typecharge__iexact=val).first()
                    if not type_charge:
                        errors.append(f"Ligne {ligne}: Type charge '{val}' non trouvé (ignoré)")
                
                # Générer un code unique
                import uuid
                code_attribution = f"ATT{uuid.uuid4().hex[:8].upper()}"
                
                # Créer l'attribution
                Attribution.objects.create(
                    code_attribution=code_attribution,
                    matricule_en=enseignant,
                    code_cours=code_cours,
                    type_charge=type_charge,
                    annee_academique=annee
                )
                success_count += 1
                
            except Exception as e:
                errors.append(f"Ligne {ligne}: Erreur - {str(e)}")
                error_count += 1
        
        return JsonResponse({
            'success': True,
            'total': total,
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors[:10]
        })
                
    except Exception as e:
        return JsonResponse({'error': f'Erreur lors de la lecture du fichier: {str(e)}'}, status=500)


@login_required
def supprimer_tout_attributions(request):
    """Supprimer toutes les attributions"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    from .models import Attribution
    
    count = Attribution.objects.count()
    Attribution.objects.all().delete()
    messages.success(request, f'{count} attribution(s) supprimée(s).')
    
    return redirect('liste_attributions')


@login_required
def attribuer_cours(request):
    """Attribuer plusieurs cours à un enseignant"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    if request.method == 'POST':
        from .models import Attribution, CoursAttribution
        import uuid
        
        matricule_en = request.POST.get('matricule_en')
        type_charge_code = request.POST.get('type_charge')
        annee_academique = request.POST.get('annee_academique')
        cours_list = request.POST.getlist('cours[]')
        
        if not matricule_en:
            messages.error(request, 'Veuillez sélectionner un enseignant.')
            return redirect('gestion_attributions')
        
        if not cours_list:
            messages.error(request, 'Veuillez cocher au moins un cours.')
            return redirect('gestion_attributions')
        
        try:
            from .models import Enseignant
            from reglage.models import TypeCharge
            
            enseignant = Enseignant.objects.get(matricule_en=matricule_en)
            type_charge = TypeCharge.objects.get(code_type=type_charge_code) if type_charge_code else None
            
            count = 0
            for code_cours in cours_list:
                # Générer un code attribution unique
                code_attribution = f"ATT-{uuid.uuid4().hex[:8].upper()}"
                
                Attribution.objects.create(
                    code_attribution=code_attribution,
                    matricule_en=enseignant,
                    code_cours=code_cours,
                    type_charge=type_charge,
                    annee_academique=annee_academique
                )
                count += 1
            
            # Supprimer les cours attribués de la table cours_attribution
            CoursAttribution.objects.filter(code_cours__in=cours_list).delete()
            
            messages.success(request, f'{count} cours attribués avec succès à {enseignant.nom_complet}!')
        except Enseignant.DoesNotExist:
            messages.error(request, 'Enseignant non trouvé.')
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'attribution : {str(e)}')
    
    return redirect('gestion_attributions')


@login_required
def migrer_ue_ec(request):
    """Migrer les UE (sans EC) et tous les EC vers la table cours_attribution"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    if request.method == 'POST':
        from .models import UE, EC, CoursAttribution
        
        try:
            # Vider la table cours_attribution avant migration
            CoursAttribution.objects.all().delete()
            
            count_ue = 0
            count_ec = 0
            
            # Récupérer les codes UE qui ont des EC
            ues_avec_ec = EC.objects.values_list('code_ue_id', flat=True).distinct()
            
            # Migrer les UE qui n'ont PAS d'EC
            ues_sans_ec = UE.objects.exclude(code_ue__in=ues_avec_ec)
            for ue in ues_sans_ec:
                CoursAttribution.objects.create(
                    code_cours=ue.code_ue,
                    intitule=ue.intitule_ue,
                    type_cours='UE',
                    code_ue_parent=None,
                    credit=ue.credit,
                    semestre=ue.semestre,
                    classe=ue.classe
                )
                count_ue += 1
            
            # Migrer TOUS les EC
            for ec in EC.objects.all():
                # Récupérer le semestre depuis l'UE parent
                semestre = ec.code_ue.semestre if ec.code_ue else 0
                
                CoursAttribution.objects.create(
                    code_cours=ec.code_ec,
                    intitule=ec.intitule_ue,
                    type_cours='EC',
                    code_ue_parent=ec.code_ue_id,
                    credit=ec.credit,
                    semestre=semestre,
                    classe=ec.classe
                )
                count_ec += 1
            
            messages.success(request, f'Migration réussie ! {count_ue} UE et {count_ec} EC migrés vers cours_attribution.')
        except Exception as e:
            messages.error(request, f'Erreur lors de la migration : {str(e)}')
        
        return redirect('gestion_attributions')
    
    return redirect('gestion_attributions')


# ============================================
# VUES POUR RÉGLAGE
# ============================================

@login_required
def gestion_reglage(request):
    """Vue pour gérer les réglages du système"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    from reglage.models import (
        Section, Departement, Mention, Niveau, Semestre, Classe, AnneeAcademique,
        Grade, Fonction, TypeCharge, Categorie
    )
    
    context = {
        'sections': Section.objects.all(),
        'departements': Departement.objects.all(),
        'mentions': Mention.objects.all(),
        'niveaux': Niveau.objects.all(),
        'semestres': Semestre.objects.all(),
        'classes': Classe.objects.all(),
        'annees': AnneeAcademique.objects.all(),
        'grades': Grade.objects.all(),
        'fonctions': Fonction.objects.all(),
        'types_charge': TypeCharge.objects.all(),
        'categories': Categorie.objects.all(),
    }
    return render(request, 'gestion/reglage.html', context)


# ============================================
# VUES POUR IMPORTATION EXCEL
# ============================================

@login_required
def import_etudiants(request):
    """Importer des étudiants depuis un fichier Excel"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            df = pd.read_excel(excel_file)
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Créer un compte utilisateur
                    username = str(row['matricule_etudiant']).strip()
                    user, created = User.objects.get_or_create(
                        username=username,
                        defaults={'role': 'ETUDIANT', 'is_active': True}
                    )
                    if created:
                        user.set_password(f"{username}2025")
                        user.save()
                    
                    # Créer ou mettre à jour l'étudiant
                    Etudiant.objects.update_or_create(
                        matricule_etudiant=username,
                        defaults={
                            'nom_complet': str(row['nom_complet']).strip(),
                            'sexe': str(row.get('sexe', 'M')).strip().upper()[:1],
                            'date_naiss': pd.to_datetime(row['date_naiss']).date() if pd.notna(row.get('date_naiss')) else None,
                            'nationalite': str(row.get('nationalite', 'Congolaise (RDC)')).strip(),
                            'telephone': str(row.get('telephone', '')).strip(),
                            'id_lgn': user,
                        }
                    )
                    success_count += 1
                except Exception as e:
                    error_count += 1
                    errors.append(f"Ligne {index + 2}: {str(e)}")
            
            if success_count > 0:
                messages.success(request, f'{success_count} étudiant(s) importé(s) avec succès!')
            if error_count > 0:
                messages.warning(request, f'{error_count} erreur(s) lors de l\'importation.')
                for error in errors[:5]:
                    messages.error(request, error)
                    
        except Exception as e:
            messages.error(request, f'Erreur lors de la lecture du fichier: {str(e)}')
        
        return redirect('gestion_etudiants')
    
    return render(request, 'gestion/import_excel_progress.html', {
        'titre': 'Étudiants',
        'colonnes': ['matricule_etudiant', 'nom_complet', 'sexe', 'date_naiss', 'nationalite', 'telephone'],
        'retour': 'gestion_etudiants'
    })


@login_required
def import_enseignants(request):
    """Importer des enseignants depuis un fichier Excel"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    return render(request, 'gestion/import_enseignants_progress.html', {
        'titre': 'Enseignants',
        'colonnes': ['matricule_en', 'nom_complet', 'code_fonction', 'code_grade', 'code_categorie', 'telephone', 'code_dpt', 'code_section'],
        'retour': 'gestion_enseignants'
    })


@login_required
def import_enseignants_ajax(request):
    """API AJAX pour importer des enseignants avec progression"""
    from django.http import JsonResponse
    from reglage.models import Grade, Fonction, Categorie, Section, Departement as ReglDepartement
    import json
    
    if not request.user.is_staff:
        return JsonResponse({'error': 'Accès non autorisé'}, status=403)
    
    if request.method != 'POST' or not request.FILES.get('excel_file'):
        return JsonResponse({'error': 'Fichier manquant'}, status=400)
    
    excel_file = request.FILES['excel_file']
    
    try:
        df = pd.read_excel(excel_file)
        total = len(df)
        success_count = 0
        error_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                matricule = str(row['matricule_en']).strip()
                nom_complet = str(row['nom_complet']).strip()
                
                # Username : matricule (unique)
                base_username = matricule if matricule else 'enseignant'
                username = base_username
                suffix = 1
                while User.objects.filter(username=username).exists():
                    username = f"{base_username}{suffix}"
                    suffix += 1
                
                # Mot de passe : matricule + 2025
                password = f"{matricule}2025"
                
                # Créer un compte utilisateur
                user, created = User.objects.get_or_create(
                    username=username,
                    defaults={'role': 'ENSEIGNANT', 'is_active': True}
                )
                if created:
                    user.set_password(password)
                    user.save()
                
                # Récupérer les FK (cherche par code ou par désignation)
                fonction = None
                if 'code_fonction' in row and pd.notna(row['code_fonction']):
                    val = str(row['code_fonction']).strip()
                    fonction = Fonction.objects.filter(code_fonction=val).first()
                    if not fonction:
                        fonction = Fonction.objects.filter(designation_fonction__iexact=val).first()
                
                grade = None
                if 'code_grade' in row and pd.notna(row['code_grade']):
                    val = str(row['code_grade']).strip()
                    grade = Grade.objects.filter(code_grade=val).first()
                    if not grade:
                        grade = Grade.objects.filter(designation_grade__iexact=val).first()
                
                categorie = None
                if 'code_categorie' in row and pd.notna(row['code_categorie']):
                    val = str(row['code_categorie']).strip()
                    categorie = Categorie.objects.filter(code_categorie=val).first()
                    if not categorie:
                        categorie = Categorie.objects.filter(designation_categorie__iexact=val).first()
                
                departement = None
                if 'code_dpt' in row and pd.notna(row['code_dpt']):
                    val = str(row['code_dpt']).strip()
                    departement = ReglDepartement.objects.filter(code_departement=val).first()
                    if not departement:
                        departement = ReglDepartement.objects.filter(designation_departement__icontains=val).first()
                
                section = None
                if 'code_section' in row and pd.notna(row['code_section']):
                    val = str(row['code_section']).strip()
                    section = Section.objects.filter(code_section=val).first()
                    if not section:
                        section = Section.objects.filter(designation_section__iexact=val).first()
                
                # Récupérer le téléphone
                telephone = None
                if 'telephone' in row and pd.notna(row['telephone']):
                    telephone = str(row['telephone']).strip()
                
                # Créer ou mettre à jour l'enseignant (avec le matricule original)
                Enseignant.objects.update_or_create(
                    matricule_en=matricule,
                    defaults={
                        'nom_complet': str(row['nom_complet']).strip(),
                        'telephone': telephone,
                        'fonction': fonction,
                        'grade': grade,
                        'categorie': categorie,
                        'code_dpt': departement,
                        'code_section': section,
                        'id_lgn': user,
                    }
                )
                success_count += 1
            except Exception as e:
                error_count += 1
                errors.append(f"Ligne {index + 2}: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'total': total,
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors[:10]
        })
                
    except Exception as e:
        return JsonResponse({'error': f'Erreur lors de la lecture du fichier: {str(e)}'}, status=500)


@login_required
def import_etudiants_ajax(request):
    """API AJAX pour importer des étudiants avec progression"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Accès non autorisé'}, status=403)
    
    if request.method != 'POST' or not request.FILES.get('excel_file'):
        return JsonResponse({'error': 'Fichier manquant'}, status=400)
    
    excel_file = request.FILES['excel_file']
    
    try:
        df = pd.read_excel(excel_file)
        total = len(df)
        success_count = 0
        error_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Créer un compte utilisateur
                username = str(row['matricule_etudiant']).strip()
                user, created = User.objects.get_or_create(
                    username=username,
                    defaults={'role': 'ETUDIANT', 'is_active': True}
                )
                if created:
                    user.set_password(f"{username}2025")
                    user.save()
                
                # Créer ou mettre à jour l'étudiant
                Etudiant.objects.update_or_create(
                    matricule_etudiant=username,
                    defaults={
                        'nom_complet': str(row['nom_complet']).strip(),
                        'sexe': str(row.get('sexe', 'M')).strip().upper()[:1],
                        'date_naiss': pd.to_datetime(row['date_naiss']).date() if pd.notna(row.get('date_naiss')) else None,
                        'nationalite': str(row.get('nationalite', 'Congolaise (RDC)')).strip(),
                        'telephone': str(row.get('telephone', '')).strip(),
                        'id_lgn': user,
                    }
                )
                success_count += 1
            except Exception as e:
                error_count += 1
                errors.append(f"Ligne {index + 2}: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'total': total,
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors[:10]
        })
                
    except Exception as e:
        return JsonResponse({'error': f'Erreur lors de la lecture du fichier: {str(e)}'}, status=500)


@login_required
def import_ue(request):
    """Importer des UE depuis un fichier Excel"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            df = pd.read_excel(excel_file)
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Récupérer la classe si spécifiée
                    classe = None
                    if 'code_classe' in row and pd.notna(row['code_classe']):
                        classe = Classe.objects.filter(code_classe=str(row['code_classe']).strip()).first()
                    
                    # Récupérer la catégorie (A ou B)
                    categorie = 'A'
                    if 'categorie' in row and pd.notna(row['categorie']):
                        cat_value = str(row['categorie']).strip().upper()
                        if cat_value in ['A', 'B']:
                            categorie = cat_value
                    
                    UE.objects.update_or_create(
                        code_ue=str(row['code_ue']).strip(),
                        defaults={
                            'intitule_ue': str(row['intitule_ue']).strip(),
                            'credit': int(row['credit']),
                            'semestre': int(row['semestre']),
                            'seuil': int(row.get('seuil', 50)),
                            'categorie': categorie,
                            'classe': classe,
                        }
                    )
                    success_count += 1
                except Exception as e:
                    error_count += 1
                    errors.append(f"Ligne {index + 2}: {str(e)}")
            
            if success_count > 0:
                messages.success(request, f'{success_count} UE importée(s) avec succès!')
            if error_count > 0:
                messages.warning(request, f'{error_count} erreur(s) lors de l\'importation.')
                for error in errors[:5]:
                    messages.error(request, error)
                    
        except Exception as e:
            messages.error(request, f'Erreur lors de la lecture du fichier: {str(e)}')
        
        return redirect('gestion_ue')
    
    return render(request, 'gestion/import_excel_progress.html', {
        'titre': 'Unités d\'Enseignement (UE)',
        'colonnes': ['code_ue', 'intitule_ue', 'credit', 'semestre', 'seuil', 'categorie', 'code_classe'],
        'retour': 'gestion_ue'
    })


@login_required
def import_ue_ajax(request):
    """API AJAX pour importer des UE avec progression"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Accès non autorisé'}, status=403)
    
    if request.method != 'POST' or not request.FILES.get('excel_file'):
        return JsonResponse({'error': 'Fichier manquant'}, status=400)
    
    excel_file = request.FILES['excel_file']
    
    try:
        df = pd.read_excel(excel_file)
        total = len(df)
        success_count = 0
        error_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Récupérer la classe si spécifiée
                classe = None
                if 'code_classe' in row and pd.notna(row['code_classe']):
                    classe = Classe.objects.filter(code_classe=str(row['code_classe']).strip()).first()
                
                # Récupérer la catégorie (A ou B)
                categorie = 'A'
                if 'categorie' in row and pd.notna(row['categorie']):
                    cat_value = str(row['categorie']).strip().upper()
                    if cat_value in ['A', 'B']:
                        categorie = cat_value
                
                UE.objects.update_or_create(
                    code_ue=str(row['code_ue']).strip(),
                    defaults={
                        'intitule_ue': str(row['intitule_ue']).strip(),
                        'credit': int(row['credit']),
                        'semestre': int(row['semestre']),
                        'seuil': int(row.get('seuil', 50)),
                        'categorie': categorie,
                        'classe': classe,
                    }
                )
                success_count += 1
            except Exception as e:
                error_count += 1
                errors.append(f"Ligne {index + 2}: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'total': total,
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors[:10]
        })
                
    except Exception as e:
        return JsonResponse({'error': f'Erreur lors de la lecture du fichier: {str(e)}'}, status=500)


@login_required
def import_ec(request):
    """Importer des EC depuis un fichier Excel"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            df = pd.read_excel(excel_file)
            success_count = 0
            error_count = 0
            errors = []
            
            # Debug: afficher les colonnes détectées
            print(f"DEBUG - Colonnes Excel: {list(df.columns)}")
            
            for index, row in df.iterrows():
                try:
                    # Récupérer l'UE
                    ue = UE.objects.get(code_ue=str(row['code_ue']).strip())
                    
                    # Récupérer la catégorie si spécifiée (valeur simple)
                    categorie = None
                    if 'code_categorie' in df.columns and pd.notna(row['code_categorie']):
                        categorie = str(row['code_categorie']).strip()
                        print(f"DEBUG - Ligne {index}: categorie = '{categorie}'")
                    
                    # Récupérer la classe si spécifiée
                    classe = None
                    if 'code_classe' in df.columns and pd.notna(row['code_classe']):
                        code_classe_val = str(row['code_classe']).strip()
                        classe = Classe.objects.filter(code_classe=code_classe_val).first()
                        print(f"DEBUG - Ligne {index}: code_classe = '{code_classe_val}', trouvé = {classe}")
                    
                    ec, created = EC.objects.update_or_create(
                        code_ec=str(row['code_ec']).strip(),
                        defaults={
                            'intitule_ue': str(row['intitule_ec']).strip(),
                            'credit': int(row['credit']),
                            'code_ue': ue,
                            'seuil': int(row.get('seuil', 8)),
                            'categorie': categorie,
                            'classe': classe,
                        }
                    )
                    print(f"DEBUG - EC créé/modifié: {ec.code_ec}, categorie={ec.categorie}, classe={ec.classe}")
                    success_count += 1
                except UE.DoesNotExist:
                    error_count += 1
                    errors.append(f"Ligne {index + 2}: UE '{row['code_ue']}' non trouvée")
                except Exception as e:
                    error_count += 1
                    errors.append(f"Ligne {index + 2}: {str(e)}")
            
            if success_count > 0:
                messages.success(request, f'{success_count} EC importé(s) avec succès!')
            if error_count > 0:
                messages.warning(request, f'{error_count} erreur(s) lors de l\'importation.')
                for error in errors[:5]:
                    messages.error(request, error)
                    
        except Exception as e:
            messages.error(request, f'Erreur lors de la lecture du fichier: {str(e)}')
        
        return redirect('gestion_ec')
    
    return render(request, 'gestion/import_excel_progress.html', {
        'titre': 'Éléments Constitutifs (EC)',
        'colonnes': ['code_ec', 'intitule_ec', 'credit', 'code_ue', 'seuil', 'code_categorie', 'code_classe'],
        'retour': 'gestion_ec'
    })


@login_required
def import_ec_ajax(request):
    """API AJAX pour importer des EC avec progression"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Accès non autorisé'}, status=403)
    
    if request.method != 'POST' or not request.FILES.get('excel_file'):
        return JsonResponse({'error': 'Fichier manquant'}, status=400)
    
    excel_file = request.FILES['excel_file']
    
    try:
        df = pd.read_excel(excel_file)
        total = len(df)
        success_count = 0
        error_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Récupérer l'UE parente
                ue = UE.objects.get(code_ue=str(row['code_ue']).strip())
                
                # Récupérer la classe si spécifiée
                classe = None
                if 'code_classe' in row and pd.notna(row['code_classe']):
                    classe = Classe.objects.filter(code_classe=str(row['code_classe']).strip()).first()
                
                # Récupérer la catégorie (A ou B)
                categorie = 'A'
                if 'code_categorie' in row and pd.notna(row['code_categorie']):
                    cat_value = str(row['code_categorie']).strip().upper()
                    if cat_value in ['A', 'B']:
                        categorie = cat_value
                
                EC.objects.update_or_create(
                    code_ec=str(row['code_ec']).strip(),
                    defaults={
                        'intitule_ue': str(row['intitule_ec']).strip(),
                        'credit': int(row['credit']),
                        'code_ue': ue,
                        'seuil': int(row.get('seuil', 8)),
                        'categorie': categorie,
                        'classe': classe,
                    }
                )
                success_count += 1
            except UE.DoesNotExist:
                error_count += 1
                errors.append(f"Ligne {index + 2}: UE '{row['code_ue']}' non trouvée")
            except Exception as e:
                error_count += 1
                errors.append(f"Ligne {index + 2}: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'total': total,
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors[:10]
        })
                
    except Exception as e:
        return JsonResponse({'error': f'Erreur lors de la lecture du fichier: {str(e)}'}, status=500)


@login_required
def import_inscriptions(request):
    """Importer des inscriptions depuis un fichier Excel"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            df = pd.read_excel(excel_file)
            success_count = 0
            error_count = 0
            errors = []
            
            from reglage.models import Classe as ReglageClasse
            from .models import Classe as CoreClasse, Cohorte
            
            for index, row in df.iterrows():
                try:
                    # Récupérer l'étudiant
                    etudiant = Etudiant.objects.get(matricule_et=str(row['matricule_etudiant']).strip())
                    
                    # Récupérer ou créer la classe dans core depuis reglage
                    code_cl = str(row['code_classe']).strip()
                    try:
                        classe = CoreClasse.objects.get(code_classe=code_cl)
                    except CoreClasse.DoesNotExist:
                        # Récupérer depuis reglage et créer dans core
                        reglage_classe = ReglageClasse.objects.get(code_classe=code_cl)
                        classe = CoreClasse.objects.create(
                            code_classe=reglage_classe.code_classe,
                            designation_cl=reglage_classe.designation_classe
                        )
                    
                    # Récupérer la cohorte si présente
                    cohorte = None
                    if 'cohorte' in row and pd.notna(row['cohorte']):
                        code_cohorte = str(row['cohorte']).strip()
                        try:
                            cohorte = Cohorte.objects.get(code_cohorte=code_cohorte)
                        except Cohorte.DoesNotExist:
                            pass  # Cohorte optionnelle, on continue sans
                    
                    Inscription.objects.update_or_create(
                        code_inscription=str(row['code_inscription']).strip(),
                        defaults={
                            'annee_academique': str(row['annee_academique']).strip(),
                            'matricule_etudiant': etudiant,
                            'code_classe': classe,
                            'cohorte': cohorte,
                        }
                    )
                    success_count += 1
                except Etudiant.DoesNotExist:
                    error_count += 1
                    errors.append(f"Ligne {index + 2}: Étudiant '{row['matricule_etudiant']}' non trouvé")
                except ReglageClasse.DoesNotExist:
                    error_count += 1
                    errors.append(f"Ligne {index + 2}: Classe '{row['code_classe']}' non trouvée dans reglage")
                except Exception as e:
                    error_count += 1
                    errors.append(f"Ligne {index + 2}: {str(e)}")
            
            if success_count > 0:
                messages.success(request, f'{success_count} inscription(s) importée(s) avec succès!')
            if error_count > 0:
                messages.warning(request, f'{error_count} erreur(s) lors de l\'importation.')
                for error in errors[:5]:
                    messages.error(request, error)
                    
        except Exception as e:
            messages.error(request, f'Erreur lors de la lecture du fichier: {str(e)}')
        
        return redirect('gestion_inscriptions')
    
    return render(request, 'gestion/import_excel_progress.html', {
        'titre': 'Inscriptions',
        'colonnes': ['code_inscription', 'annee_academique', 'matricule_etudiant', 'code_classe', 'cohorte (optionnel)'],
        'retour': 'gestion_inscriptions'
    })


@login_required
def import_inscriptions_ajax(request):
    """API AJAX pour importer des inscriptions avec progression"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Accès non autorisé'}, status=403)
    
    if request.method != 'POST' or not request.FILES.get('excel_file'):
        return JsonResponse({'error': 'Fichier manquant'}, status=400)
    
    excel_file = request.FILES['excel_file']
    
    try:
        df = pd.read_excel(excel_file)
        total = len(df)
        success_count = 0
        error_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Récupérer l'étudiant
                etudiant = Etudiant.objects.get(matricule_et=str(row['matricule_etudiant']).strip())
                
                # Récupérer la classe depuis reglage
                classe = None
                if 'code_classe' in row and pd.notna(row['code_classe']):
                    from reglage.models import Classe as ReglageClasse
                    classe = ReglageClasse.objects.get(code_classe=str(row['code_classe']).strip())
                
                # Récupérer la cohorte si spécifiée
                cohorte = None
                if 'cohorte' in row and pd.notna(row['cohorte']):
                    code_cohorte = str(row['cohorte']).strip()
                    try:
                        cohorte = Cohorte.objects.get(code_cohorte=code_cohorte)
                    except Cohorte.DoesNotExist:
                        pass  # Cohorte optionnelle, on continue sans
                
                Inscription.objects.update_or_create(
                    code_inscription=str(row['code_inscription']).strip(),
                    defaults={
                        'annee_academique': str(row['annee_academique']).strip(),
                        'matricule_etudiant': etudiant,
                        'code_classe': classe,
                        'cohorte': cohorte,
                    }
                )
                success_count += 1
            except Etudiant.DoesNotExist:
                error_count += 1
                errors.append(f"Ligne {index + 2}: Étudiant '{row['matricule_etudiant']}' non trouvé")
            except ReglageClasse.DoesNotExist:
                error_count += 1
                errors.append(f"Ligne {index + 2}: Classe '{row['code_classe']}' non trouvée dans reglage")
            except Exception as e:
                error_count += 1
                errors.append(f"Ligne {index + 2}: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'total': total,
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors[:10]
        })
                
    except Exception as e:
        return JsonResponse({'error': f'Erreur lors de la lecture du fichier: {str(e)}'}, status=500)


@login_required
def import_cohortes(request):
    """Importer des cohortes depuis un fichier Excel"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            df = pd.read_excel(excel_file)
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    Cohorte.objects.update_or_create(
                        code_cohorte=str(row['code_cohorte']).strip(),
                        defaults={
                            'lib_cohorte': str(row['lib_cohorte']).strip(),
                            'debut': pd.to_datetime(row['debut']).date() if pd.notna(row.get('debut')) else None,
                        }
                    )
                    success_count += 1
                except Exception as e:
                    error_count += 1
                    errors.append(f"Ligne {index + 2}: {str(e)}")
            
            if success_count > 0:
                messages.success(request, f'{success_count} cohorte(s) importée(s) avec succès!')
            if error_count > 0:
                messages.warning(request, f'{error_count} erreur(s) lors de l\'importation.')
                for error in errors[:5]:
                    messages.error(request, error)
                    
        except Exception as e:
            messages.error(request, f'Erreur lors de la lecture du fichier: {str(e)}')
        
        return redirect('gestion_cohortes')
    
    return render(request, 'gestion/import_excel_progress.html', {
        'titre': 'Cohortes',
        'colonnes': ['code_cohorte', 'lib_cohorte', 'debut'],
        'retour': 'gestion_cohortes'
    })


@login_required
def import_cohortes_ajax(request):
    """API AJAX pour importer des cohortes avec progression"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Accès non autorisé'}, status=403)
    
    if request.method != 'POST' or not request.FILES.get('excel_file'):
        return JsonResponse({'error': 'Fichier manquant'}, status=400)
    
    excel_file = request.FILES['excel_file']
    
    try:
        df = pd.read_excel(excel_file)
        total = len(df)
        success_count = 0
        error_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                Cohorte.objects.update_or_create(
                    code_cohorte=str(row['code_cohorte']).strip(),
                    defaults={
                        'lib_cohorte': str(row['lib_cohorte']).strip(),
                        'debut': int(row['debut']),
                    }
                )
                success_count += 1
            except Exception as e:
                error_count += 1
                errors.append(f"Ligne {index + 2}: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'total': total,
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors[:10]
        })
                
    except Exception as e:
        return JsonResponse({'error': f'Erreur lors de la lecture du fichier: {str(e)}'}, status=500)


@login_required
def telecharger_modele_excel(request, model_type):
    """Télécharger un modèle Excel pour l'importation"""
    if not request.user.is_staff:
        return redirect('home')
    
    modeles = {
        'etudiants': {
            'colonnes': ['matricule_etudiant', 'nom_complet', 'sexe', 'date_naiss', 'nationalite', 'telephone'],
            'exemple': ['ETU001', 'DUPONT Jean', 'M', '2000-01-15', 'Congolaise (RDC)', '+243123456789']
        },
        'enseignants': {
            'colonnes': ['matricule_en', 'nom_complet', 'code_fonction', 'code_grade', 'code_categorie', 'telephone', 'code_dpt', 'code_section'],
            'exemple': ['ENS001', 'MARTIN Pierre', 'FCT001', 'GRD001', 'CAT001', '+243123456789', 'DPT001', 'SEC001']
        },
        'attributions': {
            'colonnes': ['matricule_en', 'code_cours', 'type_charge', 'annee_academique'],
            'exemple': [
                ['ENS001', 'UE101', 'COURS', '2024-2025'],
                ['ENS002', 'EC101', 'TD', '2024-2025'],
                ['ENS001', 'UE102', '', '2024-2025'],
                ['ENS003', 'EC201', 'TP', '2024-2025']
            ],
    'instructions': [
        'matricule_en: Matricule de l\'enseignant (ex: ENS001)',
        'code_cours: Code UE ou EC (ex: UE101, EC101)',
        'type_charge: Type de charge (optionnel, ex: COURS, TD, TP)',
        'annee_academique: Année académique (ex: 2024-2025)'
    ]
        },
        'fonctions': {
            'colonnes': ['code_fonction', 'designation_fonction'],
            'exemple': ['FCT001', 'Enseignant']
        },
        'ue': {
            'colonnes': ['code_ue', 'intitule_ue', 'credit', 'semestre', 'seuil', 'categorie', 'code_classe'],
            'exemple': ['UE101', 'Mathématiques', 6, 1, 50, 'A', 'L1INFO']
        },
        'ec': {
            'colonnes': ['code_ec', 'intitule_ec', 'credit', 'code_ue', 'seuil', 'code_categorie', 'code_classe'],
            'exemple': ['EC101', 'Algèbre', 3, 'UE101', 50, 'CAT001', 'L1INFO']
        },
        'inscriptions': {
            'colonnes': ['code_inscription', 'annee_academique', 'matricule_etudiant', 'code_classe', 'cohorte'],
            'exemple': ['INS001', '2024-2025', 'ETU001', 'L1INFO', 'COH2024']
        },
        'cohortes': {
            'colonnes': ['code_cohorte', 'lib_cohorte', 'debut'],
            'exemple': ['COH2024', 'Cohorte 2024-2025', '2024-10-01']
        },
    }
    
    if model_type not in modeles:
        messages.error(request, 'Type de modèle inconnu.')
        return redirect('home')
    
    modele = modeles[model_type]
    df = pd.DataFrame([modele['exemple']], columns=modele['colonnes'])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=modele_{model_type}.xlsx'
    df.to_excel(response, index=False)
    
    return response


# ========== VUES DE SUPPRESSION EN MASSE ==========

@login_required
def supprimer_tout_ue(request):
    """Supprimer toutes les UE"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    if request.method == 'POST':
        count = UE.objects.count()
        UE.objects.all().delete()
        messages.success(request, f'{count} UE supprimée(s) avec succès!')
        return redirect('gestion_ue')
    
    return render(request, 'gestion/confirmer_suppression.html', {
        'titre': 'Supprimer toutes les UE',
        'message': f'Êtes-vous sûr de vouloir supprimer les {UE.objects.count()} UE ?',
        'retour': 'gestion_ue'
    })


@login_required
def supprimer_tout_ec(request):
    """Supprimer tous les EC"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    if request.method == 'POST':
        count = EC.objects.count()
        EC.objects.all().delete()
        messages.success(request, f'{count} EC supprimé(s) avec succès!')
        return redirect('gestion_ec')
    
    return render(request, 'gestion/confirmer_suppression.html', {
        'titre': 'Supprimer tous les EC',
        'message': f'Êtes-vous sûr de vouloir supprimer les {EC.objects.count()} EC ?',
        'retour': 'gestion_ec'
    })


@login_required
def supprimer_tout_etudiants(request):
    """Supprimer tous les étudiants"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    if request.method == 'POST':
        count = Etudiant.objects.count()
        Etudiant.objects.all().delete()
        messages.success(request, f'{count} étudiant(s) supprimé(s) avec succès!')
        return redirect('gestion_etudiants')
    
    return render(request, 'gestion/confirmer_suppression.html', {
        'titre': 'Supprimer tous les étudiants',
        'message': f'Êtes-vous sûr de vouloir supprimer les {Etudiant.objects.count()} étudiants ?',
        'retour': 'gestion_etudiants'
    })


@login_required
def reinitialiser_mdp_tous_etudiants(request):
    """Réinitialiser en lot les mots de passe des étudiants vers username+2025"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')

    users_qs = User.objects.filter(role='ETUDIANT', is_active=True)
    total = users_qs.count()

    if request.method == 'POST':
        count = 0
        for u in users_qs.only('id', 'username'):
            u.set_password(f"{u.username}2025")
            u.save(update_fields=['password'])
            count += 1
        messages.success(request, f'{count} mot(s) de passe étudiant(s) réinitialisé(s) avec succès!')
        return redirect('gestion_etudiants')

    return render(request, 'gestion/confirmer_reinitialisation_mdp_etudiants.html', {
        'titre': 'Réinitialiser les mots de passe des étudiants',
        'message': f'Êtes-vous sûr de vouloir réinitialiser les mots de passe de {total} étudiant(s) ?',
        'retour': 'gestion_etudiants'
    })


@login_required
def supprimer_tout_enseignants(request):
    """Supprimer tous les enseignants"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    if request.method == 'POST':
        count = Enseignant.objects.count()
        Enseignant.objects.all().delete()
        messages.success(request, f'{count} enseignant(s) supprimé(s) avec succès!')
        return redirect('gestion_enseignants')
    
    return render(request, 'gestion/confirmer_suppression.html', {
        'titre': 'Supprimer tous les enseignants',
        'message': f'Êtes-vous sûr de vouloir supprimer les {Enseignant.objects.count()} enseignants ?',
        'retour': 'gestion_enseignants'
    })


@login_required
def envoyer_recours(request):
    """Envoyer un recours au jury"""
    if request.user.role != 'ETUDIANT':
        messages.error(request, 'Accès réservé aux étudiants.')
        return redirect('home')
    
    if request.method == 'POST':
        from .models import Recours, FichierRecours
        
        objet = request.POST.get('objet')
        ue_ec = request.POST.get('ue_ec')
        description = request.POST.get('description')
        fichiers = request.FILES.getlist('fichiers')
        
        if not objet or not ue_ec or not description:
            messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
            return redirect('etudiant_notes')
        
        try:
            # Récupérer l'étudiant connecté
            etudiant = Etudiant.objects.get(id_lgn=request.user)
            
            # Créer le recours
            recours = Recours.objects.create(
                etudiant=etudiant,
                objet=objet,
                ue_ec_concerne=ue_ec,
                description=description
            )
            
            # Ajouter les fichiers joints
            for fichier in fichiers:
                if fichier.size > 10 * 1024 * 1024:  # 10MB
                    messages.warning(request, f'Le fichier {fichier.name} dépasse la taille maximale de 10MB.')
                    continue
                
                FichierRecours.objects.create(
                    recours=recours,
                    fichier=fichier
                )
            
            messages.success(request, f'Votre recours {recours.code_recours} a été envoyé avec succès ! Il sera traité par le jury.')
            return redirect('etudiant_notes')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'envoi du recours: {str(e)}')
            return redirect('etudiant_notes')
    
    return redirect('etudiant_notes')


@login_required
def supprimer_tout_inscriptions(request):
    """Supprimer toutes les inscriptions"""
    if not request.user.is_staff:
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')
    
    if request.method == 'POST':
        count = Inscription.objects.count()
        Inscription.objects.all().delete()
        messages.success(request, f'{count} inscription(s) supprimée(s) avec succès!')
        return redirect('gestion_inscriptions')
    
    return render(request, 'gestion/confirmer_suppression.html', {
        'titre': 'Supprimer toutes les inscriptions',
        'message': f'Êtes-vous sûr de vouloir supprimer les {Inscription.objects.count()} inscriptions ?',
        'retour': 'gestion_inscriptions'
    })


@login_required
@require_gestionnaire_or_admin
def statistiques(request):
    
    # Statistiques générales
    student_count = Etudiant.objects.count()
    teacher_count = Enseignant.objects.count()
    jury_count = Jury.objects.count()
    course_count = Attribution.objects.count()
    evaluation_count = EvaluationEnseignement.objects.count()
    
    # Évaluations par année académique
    evaluations_by_year = EvaluationEnseignement.objects.values('annee_academique').annotate(count=Count('id')).order_by('annee_academique')
    eval_labels = [e['annee_academique'] for e in evaluations_by_year]
    eval_data = [e['count'] for e in evaluations_by_year]
    
    # Répartition par rôle
    roles_data = {
        'labels': ['Étudiants', 'Enseignants', 'Jurys'],
        'data': [student_count, teacher_count, jury_count]
    }
    
    context = {
        'student_count': student_count,
        'teacher_count': teacher_count,
        'jury_count': jury_count,
        'course_count': course_count,
        'evaluation_count': evaluation_count,
        'roles_chart': json.dumps(roles_data),
        'eval_chart': json.dumps({'labels': eval_labels, 'data': eval_data}),
    }
    
    return render(request, 'core/statistiques.html', context)


@login_required
@require_admin
def historique_actions(request):
    """Vue pour afficher l'historique de toutes les actions - Admin seulement"""
    
    # Récupérer tous les historiques
    historiques = HistoriqueAction.objects.select_related('utilisateur').all()
    
    # Filtrer par utilisateur si demandé
    utilisateur_id = request.GET.get('utilisateur')
    if utilisateur_id:
        historiques = historiques.filter(utilisateur_id=utilisateur_id)
    
    # Filtrer par type d'action si demandé
    type_action = request.GET.get('type_action')
    if type_action:
        historiques = historiques.filter(type_action=type_action)
    
    # Filtrer par type d'objet si demandé
    type_objet = request.GET.get('type_objet')
    if type_objet:
        historiques = historiques.filter(type_objet=type_objet)
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(historiques, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Récupérer les listes pour les filtres
    utilisateurs = User.objects.filter(historiqueaction__isnull=False).distinct().order_by('username')
    
    context = {
        'page_obj': page_obj,
        'utilisateurs': utilisateurs,
        'type_actions': HistoriqueAction.TYPE_ACTIONS,
        'types_objets': HistoriqueAction.TYPES_OBJETS,
        'utilisateur_id': utilisateur_id,
        'type_action': type_action,
        'type_objet': type_objet,
    }
    
    return render(request, 'gestion/historique_actions.html', context)


# ========== VUES DÉDIÉES AU GESTIONNAIRE ==========

@login_required
@require_gestionnaire_or_admin
def gestionnaire_communiques_jury(request):
    """Vue pour que le gestionnaire voie tous les communiqués de jury"""
    try:
        from reglage.models import AnneeAcademique
        
        # Filtres
        classe_filter = request.GET.get('classe', '')
        annee_filter = request.GET.get('annee', '')
        
        # Listes pour les filtres
        classes = Classe.objects.all().order_by('code_classe')
        annees = AnneeAcademique.objects.all().order_by('-code_anac')
        
        # Récupérer tous les communiqués
        communiques = CommuniqueDeliberation.objects.select_related(
            'code_classe', 'cree_par'
        ).all()
        
        # Appliquer les filtres
        if classe_filter:
            communiques = communiques.filter(code_classe__code_classe=classe_filter)
        if annee_filter:
            communiques = communiques.filter(annee_academique=annee_filter)
        
        communiques = communiques.order_by('-date_deliberation', '-date_creation')
        
        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(communiques, 20)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'page_obj': page_obj,
            'classes': classes,
            'annees': annees,
            'classe_filter': classe_filter,
            'annee_filter': annee_filter,
            'total': communiques.count(),
        }
        return render(request, 'gestionnaire/communiques_jury.html', context)
        
    except Exception as e:
        messages.error(request, f'Erreur: {str(e)}')
        return redirect('home')


@login_required
@require_gestionnaire_or_admin
def gestionnaire_deliberations(request):
    """Vue pour que le gestionnaire voie les délibérations publiées par les jurys"""
    try:
        from reglage.models import AnneeAcademique
        
        # Récupérer tous les jurys
        jurys = Jury.objects.select_related('code_classe', 'id_lgn').order_by('code_jury')
        
        # Jury sélectionné (peut être None)
        jury_selected_id = request.GET.get('jury', '')
        jury_selected = None
        
        # Récupérer les délibérations
        if jury_selected_id:
            try:
                jury_selected = Jury.objects.get(code_jury=jury_selected_id)
                # Délibérations publiées par ce jury
                deliberations = Deliberation.objects.select_related(
                    'code_classe', 'matricule_etudiant', 'code_ue', 'code_ec', 'cree_par'
                ).filter(cree_par=jury_selected.id_lgn)
            except Jury.DoesNotExist:
                deliberations = Deliberation.objects.none()
        else:
            # Si pas de jury sélectionné, montrer aucun résultat au début
            deliberations = Deliberation.objects.none()
        
        # Filtres additionnels
        classe_filter = request.GET.get('classe', '')
        annee_filter = request.GET.get('annee', '')
        
        # Listes pour les filtres
        classes = Classe.objects.all().order_by('code_classe')
        annees = AnneeAcademique.objects.all().order_by('-code_anac')
        
        # Appliquer les filtres
        if classe_filter:
            deliberations = deliberations.filter(code_classe__code_classe=classe_filter)
        if annee_filter:
            deliberations = deliberations.filter(annee_academique=annee_filter)
        
        deliberations = deliberations.order_by('-date_creation')
        
        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(deliberations, 50)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'page_obj': page_obj,
            'jurys': jurys,
            'jury_selected': jury_selected,
            'jury_selected_id': jury_selected_id,
            'classes': classes,
            'annees': annees,
            'classe_filter': classe_filter,
            'annee_filter': annee_filter,
            'total': deliberations.count(),
        }
        return render(request, 'gestionnaire/deliberations.html', context)
        
    except Exception as e:
        messages.error(request, f'Erreur: {str(e)}')
        return redirect('home')


@login_required
@require_gestionnaire_or_admin
def gestionnaire_evaluations(request):
    """Vue pour que le gestionnaire voie toutes les évaluations des enseignants"""
    try:
        from reglage.models import AnneeAcademique
        
        # Filtres
        enseignant_filter = request.GET.get('enseignant', '')
        annee_filter = request.GET.get('annee', '')
        
        # Listes pour les filtres
        enseignants = Enseignant.objects.all().order_by('nom_complet')
        annees = AnneeAcademique.objects.all().order_by('-code_anac')
        
        # Récupérer toutes les évaluations
        evaluations = EvaluationEnseignement.objects.select_related(
            'etudiant', 'attribution__matricule_en'
        ).all()
        
        # Appliquer les filtres
        if enseignant_filter:
            evaluations = evaluations.filter(attribution__matricule_en__matricule_en=enseignant_filter)
        if annee_filter:
            evaluations = evaluations.filter(annee_academique=annee_filter)
        
        evaluations = evaluations.order_by('-date_creation')
        
        # Calculer statistiques
        total_evaluations = evaluations.count()
        if total_evaluations > 0:
            # Calculer la moyenne des notes de ponctualité (principale métrique)
            avg_note = evaluations.aggregate(
                avg=Avg('ponctualite')
            )['avg'] or 0
        else:
            avg_note = 0
        
        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(evaluations, 30)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'page_obj': page_obj,
            'enseignants': enseignants,
            'annees': annees,
            'enseignant_filter': enseignant_filter,
            'annee_filter': annee_filter,
            'total': total_evaluations,
            'avg_note': round(avg_note, 2),
        }
        return render(request, 'gestionnaire/evaluations.html', context)
        
    except Exception as e:
        messages.error(request, f'Erreur: {str(e)}')
        return redirect('home')


@login_required
@require_gestionnaire_or_admin
def gestionnaire_commentaires_etudiants(request):
    """Vue pour que le gestionnaire voie tous les commentaires des étudiants sur les enseignants"""
    try:
        from reglage.models import AnneeAcademique
        
        # Filtres
        classe_filter = request.GET.get('classe', '')
        annee_filter = request.GET.get('annee', '')
        
        # Listes pour les filtres
        classes = Classe.objects.all().order_by('code_classe')
        annees = AnneeAcademique.objects.all().order_by('-code_anac')
        
        # Récupérer tous les commentaires des étudiants
        commentaires = CommentaireCours.objects.select_related(
            'etudiant'
        ).all()
        
        # Appliquer les filtres - regarder les inscriptions de l'étudiant
        if classe_filter:
            # Récupérer les inscriptions de la classe
            inscriptions = Inscription.objects.filter(
                code_classe__code_classe=classe_filter
            ).values_list('matricule_etudiant', flat=True)
            commentaires = commentaires.filter(etudiant__matricule_et__in=inscriptions)
        
        if annee_filter:
            commentaires = commentaires.filter(annee_academique=annee_filter)
        
        commentaires = commentaires.order_by('-date_creation')
        
        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(commentaires, 30)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'page_obj': page_obj,
            'classes': classes,
            'annees': annees,
            'classe_filter': classe_filter,
            'annee_filter': annee_filter,
            'total': commentaires.count(),
        }
        return render(request, 'gestionnaire/commentaires_etudiants.html', context)
        
    except Exception as e:
        messages.error(request, f'Erreur: {str(e)}')
        return redirect('home')


@login_required
@require_gestionnaire_or_admin
def gestionnaire_commentaires_enseignants(request):
    """Vue pour que le gestionnaire voie tous les commentaires des enseignants sur les étudiants"""
    try:
        from reglage.models import AnneeAcademique
        
        # Filtres
        enseignant_filter = request.GET.get('enseignant', '')
        annee_filter = request.GET.get('annee', '')
        
        # Listes pour les filtres
        enseignants = Enseignant.objects.all().order_by('nom_complet')
        annees = AnneeAcademique.objects.all().order_by('-code_anac')
        
        # Récupérer tous les commentaires des enseignants sur les étudiants
        # D'abord, chercher si le modèle existe pour les commentaires des enseignants
        # Sinon, on affiche un message d'info
        
        # Pour maintenant, on va créer des données vides et montrer l'interface
        commentaires_enseignants = []
        
        # Appliquer les filtres simulés
        filtered_count = 0
        
        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(commentaires_enseignants, 30)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'page_obj': page_obj,
            'enseignants': enseignants,
            'annees': annees,
            'enseignant_filter': enseignant_filter,
            'annee_filter': annee_filter,
            'total': filtered_count,
            'message_info': 'Cette vue affichera les commentaires des enseignants sur les étudiants une fois le système complètement configuré.',
        }
        return render(request, 'gestionnaire/commentaires_enseignants.html', context)
        
    except Exception as e:
        messages.error(request, f'Erreur: {str(e)}')
        return redirect('home')


# === IMPORT EXCEL ===

@login_required
@require_gestionnaire_or_admin
def gestionnaire_import_attributions(request):
    """Vue pour importer des attributions depuis un fichier Excel"""
    try:
        import_result = None
        
        if request.method == 'POST' and request.FILES.get('file'):
            file = request.FILES['file']
            
            try:
                # Lire le fichier Excel
                df = pd.read_excel(file)
                
                # Normaliser les noms de colonnes
                df.columns = df.columns.str.strip().str.lower()
                
                # Détecter les colonnes (avec ou sans en-têtes)
                col_names = list(df.columns)
                
                # Déterminer les colonnes (peut être header ou positionnel)
                if 'matricule enseignant' in col_names or 'matricule_enseignant' in col_names:
                    matricule_col = 'matricule enseignant' if 'matricule enseignant' in col_names else 'matricule_enseignant'
                    code_cours_col = 'code cours' if 'code cours' in col_names else 'code_cours'
                    type_charge_col = 'type charge' if 'type charge' in col_names else 'type_charge'
                    annee_col = 'année académique' if 'année académique' in col_names else 'annee_academique'
                else:
                    # Format positionnel: col 0=matricule, col 1=code_cours, col 2=type_charge, col 3=annee
                    matricule_col = 0
                    code_cours_col = 1
                    type_charge_col = 2
                    annee_col = 3
                
                created_count = 0
                updated_count = 0
                errors = []
                
                updates_existing = request.POST.get('updates_existing') == '1'
                
                for idx, row in df.iterrows():
                    try:
                        # Extraire les données
                        matricule = str(row[matricule_col]).strip() if isinstance(row[matricule_col], (str, int, float)) else None
                        code_cours = str(row[code_cours_col]).strip() if isinstance(row[code_cours_col], (str, int, float)) else None
                        type_charge = str(row[type_charge_col]).strip() if isinstance(row[type_charge_col], (str, int, float)) and pd.notna(row[type_charge_col]) else None
                        annee = str(row[annee_col]).strip() if isinstance(row[annee_col], (str, int, float)) else None
                        
                        if not matricule or not code_cours or not annee:
                            errors.append(f"Ligne {idx+2}: Données manquantes (matricule/code_cours/année)")
                            continue
                        
                        # Vérifier que l'enseignant existe
                        try:
                            enseignant = Enseignant.objects.get(matricule_en=matricule)
                        except Enseignant.DoesNotExist:
                            errors.append(f"Ligne {idx+2}: Enseignant '{matricule}' non trouvé")
                            continue
                        
                        # Vérifier que le cours existe (UE ou EC)
                        cours_existe = UE.objects.filter(code_ue=code_cours).exists() or EC.objects.filter(code_ec=code_cours).exists()
                        if not cours_existe:
                            errors.append(f"Ligne {idx+2}: Cours '{code_cours}' non trouvé")
                            continue
                        
                        # Générer un code attribution unique
                        code_attribution = f"{matricule}_{code_cours}_{annee}"
                        
                        # Chercher ou créer l'attribution
                        type_charge_obj = None
                        if type_charge:
                            try:
                                type_charge_obj = TypeCharge.objects.get(code_typecharge=type_charge)
                            except TypeCharge.DoesNotExist:
                                pass
                        
                        attribution, created = Attribution.objects.update_or_create(
                            code_attribution=code_attribution,
                            defaults={
                                'matricule_en': enseignant,
                                'code_cours': code_cours,
                                'type_charge': type_charge_obj,
                                'annee_academique': annee,
                            }
                        )
                        
                        if created:
                            created_count += 1
                        elif updates_existing:
                            updated_count += 1
                        
                    except Exception as e:
                        errors.append(f"Ligne {idx+2}: {str(e)}")
                
                import_result = {
                    'created_count': created_count,
                    'updated_count': updated_count,
                    'error_count': len(errors),
                    'total_count': len(df),
                    'errors': errors[:10]  # Afficher max 10 erreurs
                }
                
                if created_count > 0 or updated_count > 0:
                    messages.success(request, f'{created_count} attribution(s) créée(s), {updated_count} mise(s) à jour')
                if errors:
                    messages.warning(request, f'{len(errors)} erreur(s) lors de l\'import')
                    
            except Exception as e:
                messages.error(request, f'Erreur lors de la lecture du fichier: {str(e)}')
        
        context = {
            'import_result': import_result,
        }
        return render(request, 'gestionnaire/import_attributions.html', context)
        
    except Exception as e:
        messages.error(request, f'Erreur: {str(e)}')
        return redirect('home')


@login_required
@require_gestionnaire_or_admin
def telecharger_template_attributions(request):
    """Télécharge un template Excel pour les attributions"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Attributions'
        
        # En-têtes
        headers = ['Matricule Enseignant', 'Code Cours', 'Type Charge', 'Année Académique']
        ws.append(headers)
        
        # Style en-têtes
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajouter quelques exemples
        examples = [
            ['ENS001', 'INF101', 'CM', '2024-2025'],
            ['ENS002', 'PROG', 'TD', '2024-2025'],
            ['ENS003', 'BDD', 'TP', '2024-2025'],
        ]
        
        for example in examples:
            ws.append(example)
        
        # Ajuster les largeurs
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        
        # Réponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="template_attributions.xlsx"'
        wb.save(response)
        
        return response
        
    except Exception as e:
        messages.error(request, f'Erreur: {str(e)}')
        return redirect('home')



from .views_jury_presence import (
    jury_presence_deliberation,
    jury_imprimable_releves_tous,
    jury_imprimable_profils_tous
)
from .views_releve_pdf import jury_imprimable_releve


@login_required
@require_staff_or_roles(['GESTIONNAIRE', 'AGENT'])
def suivi_cohorte(request):
    """Vue de suivi des UE par cohorte"""
    
    cohortes = Cohorte.objects.all().select_related('code_mention')
    ues = []
    cohorte_selectionnee = None
    total_credits = 0
    
    # Récupérer la cohorte sélectionnée
    cohorte_id = request.GET.get('cohorte', '')
    
    if cohorte_id:
        try:
            cohorte_selectionnee = Cohorte.objects.select_related('code_mention').get(code_cohorte=cohorte_id)
            
            # Récupérer toutes les UE liées à la mention de cette cohorte
            if cohorte_selectionnee.code_mention:
                # Récupérer les classes liées à cette mention
                from reglage.models import Classe
                classes = Classe.objects.filter(code_mention=cohorte_selectionnee.code_mention)
                
                # Récupérer les UE liées à ces classes
                ues = UE.objects.filter(classe__in=classes).select_related('classe').order_by('semestre', 'code_ue')
                
                # Calculer le total des crédits
                total_credits = sum(ue.credit for ue in ues)
        except Cohorte.DoesNotExist:
            messages.error(request, 'Cohorte non trouvée.')
    
    context = {
        'cohortes': cohortes,
        'ues': ues,
        'cohorte_selectionnee': cohorte_selectionnee,
        'total_credits': total_credits,
    }
    return render(request, 'suivi/cohorte.html', context)


# ========== FICHES DE COTATION ==========

@login_required
@require_gestionnaire_or_admin
def gestion_fiches_cotation(request):
    """Page de sélection de classe pour imprimer les fiches de cotation vides"""
    from reglage.models import AnneeAcademique
    classes = Classe.objects.all().order_by('code_classe')
    annees = AnneeAcademique.objects.all().order_by('-code_anac')
    annee_active = AnneeAcademique.get_annee_en_cours()
    annee_default = annee_active.code_anac if annee_active else ''
    # Semestres distincts issus des UE et EC
    semestres_ue = list(UE.objects.values_list('semestre', flat=True).distinct())
    semestres_ec = list(EC.objects.select_related('code_ue').values_list('code_ue__semestre', flat=True).distinct())
    semestres = sorted(set(s for s in semestres_ue + semestres_ec if s is not None))
    context = {
        'classes': classes,
        'annees': annees,
        'annee_default': annee_default,
        'semestres': semestres,
    }
    return render(request, 'gestion/fiches_cotation.html', context)


@login_required
@require_gestionnaire_or_admin
def gestion_fiches_cotation_pdf(request):
    """Génère un PDF avec toutes les fiches de cotation vides pour une classe donnée.
    Une fiche par cours (UE/EC), avec la liste des étudiants inscrits et colonnes CC/Examen vides.
    """
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reglage.models import AnneeAcademique

    classe_code = request.GET.get('classe', '')
    annee_code = request.GET.get('annee', '')
    semestre_filtre = request.GET.get('semestre', '')

    if not classe_code:
        messages.error(request, 'Veuillez sélectionner une classe.')
        return redirect('gestion_fiches_cotation')

    try:
        classe = Classe.objects.get(code_classe=classe_code)
    except Classe.DoesNotExist:
        messages.error(request, 'Classe introuvable.')
        return redirect('gestion_fiches_cotation')

    if not annee_code:
        annee_active = AnneeAcademique.get_annee_en_cours()
        annee_code = annee_active.code_anac if annee_active else ''

    # Récupérer les étudiants inscrits dans cette classe/année, triés par nom
    inscriptions = Inscription.objects.filter(
        code_classe=classe,
        annee_academique=annee_code,
    ).select_related('matricule_etudiant').order_by('matricule_etudiant__nom_complet')
    etudiants = [insc.matricule_etudiant for insc in inscriptions]

    # Construire la map des cours (UE + EC) appartenant à cette classe
    ues_qs = UE.objects.filter(classe_id=classe_code)
    ecs_qs = EC.objects.filter(classe_id=classe_code)
    if semestre_filtre:
        try:
            sem_int = int(semestre_filtre)
            ues_qs = ues_qs.filter(semestre=sem_int)
            ecs_qs = ecs_qs.filter(code_ue__semestre=sem_int)
        except ValueError:
            pass
    ues_classe = list(ues_qs.order_by('semestre', 'code_ue'))
    ecs_classe = list(ecs_qs.order_by('code_ec'))
    cours_map = {}
    for ue in ues_classe:
        cours_map[ue.code_ue] = ue
    for ec in ecs_classe:
        cours_map[ec.code_ec] = ec
    codes_cours_classe = set(cours_map.keys())

    # Récupérer les attributions pour cette classe/année filtrées directement
    attributions = Attribution.objects.filter(
        annee_academique=annee_code,
        code_cours__in=codes_cours_classe,
    ).select_related('matricule_en', 'type_charge').order_by('matricule_en__nom_complet', 'code_cours')

    fiches = []
    for attr in attributions:
        cours_obj = cours_map.get(attr.code_cours)
        if cours_obj is not None:
            fiches.append({'attribution': attr, 'cours': cours_obj})

    if not fiches:
        messages.warning(request, 'Aucun cours attribué trouvé pour cette classe et cette année.')
        return redirect('gestion_fiches_cotation')

    # Génération PDF
    import os
    from django.conf import settings
    from PIL import Image as PILImage
    from reportlab.platypus import Image as RLImage

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    page_w = A4[0] - 3 * cm  # largeur utile
    styles = getSampleStyleSheet()
    bureau_style = ParagraphStyle('bureau', parent=styles['Heading2'], fontSize=11, alignment=TA_CENTER, spaceAfter=2, textColor=colors.HexColor('#2c3e50'))
    title_style = ParagraphStyle('title', parent=styles['Heading1'], fontSize=13, alignment=TA_CENTER, spaceAfter=4, textColor=colors.HexColor('#2c3e50'))
    sub_style = ParagraphStyle('sub', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, spaceAfter=6)
    white_style = ParagraphStyle('white', parent=styles['Normal'], fontSize=8, alignment=TA_LEFT, textColor=colors.white)
    sig_style = ParagraphStyle('sig', parent=styles['Normal'], fontSize=8, alignment=TA_LEFT)

    from lmdmanagersystem.middleware import get_entete_path
    entete_path = get_entete_path()

    story = []

    for idx, fiche in enumerate(fiches):
        attr = fiche['attribution']
        cours = fiche['cours']
        enseignant = attr.matricule_en
        intitule = getattr(cours, 'intitule_ue', attr.code_cours)
        credit = getattr(cours, 'credit', '-')
        type_cours = attr.get_type_cours() or ''
        # Semestre et UE parente
        if type_cours == 'EC' and hasattr(cours, 'code_ue') and cours.code_ue:
            ue_parente = cours.code_ue  # FK vers UE
            semestre = getattr(ue_parente, 'semestre', '-')
            ue_label = ue_parente.intitule_ue
        else:
            semestre = getattr(cours, 'semestre', '-')
            ue_label = ''

        # Image d'en-tête
        if os.path.exists(entete_path):
            pil_img = PILImage.open(entete_path)
            iw, ih = pil_img.size
            desired_w = page_w
            desired_h = desired_w * (ih / iw)
            story.append(RLImage(entete_path, width=desired_w, height=desired_h))
            story.append(Spacer(1, 0.3 * cm))

        # Titres
        story.append(Paragraph(f"BUREAU DU JURY {classe.code_classe}", bureau_style))
        story.append(Paragraph("FICHE DE COTATION", title_style))
        story.append(Paragraph(
            f"Classe : <b>{classe.code_classe} – {classe.designation_classe}</b> &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Année académique : <b>{annee_code}</b>",
            sub_style
        ))
        story.append(Spacer(1, 0.2 * cm))

        # Infos cours/enseignant (texte blanc sur fond sombre)
        grade_code = enseignant.grade.code_grade if enseignant.grade else '-'
        if type_cours == 'EC':
            col1_label = Paragraph(f"<b>UE :</b> {ue_label}", white_style)
            col2_label = Paragraph(f"<b>EC :</b> {intitule}", white_style)
        else:
            col1_label = Paragraph(f"<b>UE :</b> {intitule}", white_style)
            col2_label = Paragraph("", white_style)
        row1 = [
            col1_label,
            col2_label,
            Paragraph(f"<b>Semestre :</b> S{semestre}", white_style),
            Paragraph(f"<b>Crédits :</b> {credit}", white_style),
        ]
        row2 = [
            Paragraph(f"<b>Enseignant :</b> {enseignant.nom_complet}", white_style),
            Paragraph(f"<b>Matricule :</b> {enseignant.matricule_en}", white_style),
            Paragraph(f"<b>Grade :</b> {grade_code}", white_style),
            Paragraph("", white_style),
        ]
        info_data = [row1, row2]
        col_info = [page_w * 0.38, page_w * 0.22, page_w * 0.20, page_w * 0.20]
        info_table = Table(info_data, colWidths=col_info)
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#4a6fa5')),
            ('PADDING', (0, 0), (-1, -1), 6),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.4 * cm))

        # Tableau des étudiants (vide, sans Matricule ni Observations)
        header = ['#', 'Nom complet', 'CC (/10)', 'Examen (/10)', 'Note Finale (/20)', 'Rattrapage (/20)']
        col_widths = [1*cm, page_w - 1*cm - 2.8*cm - 3.2*cm - 3.8*cm - 3*cm, 2.8*cm, 3.2*cm, 3.8*cm, 3*cm]
        table_data = [header]
        for i, etud in enumerate(etudiants, 1):
            table_data.append([str(i), etud.nom_complet, '', '', '', ''])

        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
            ('ROWHEIGHT', (0, 1), (-1, -1), 20),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.8 * cm))
        story.append(Paragraph(
            "<b>Signature de l'enseignant :</b> ____________________________"
            "&nbsp;&nbsp;&nbsp;&nbsp;<b>Date :</b> ____________________",
            sig_style
        ))

        # Saut de page entre fiches (sauf la dernière)
        if idx < len(fiches) - 1:
            story.append(PageBreak())

    doc.build(story)
    buffer.seek(0)
    filename = f"fiches_cotation_{classe_code}_{annee_code}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
def enseignant_fiche_cotation_pdf(request, code_cours, annee):
    """Génère la fiche de cotation PDF pour un enseignant (avec les notes saisies)."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    enseignant = get_simulated_enseignant(request)
    if not enseignant:
        messages.error(request, 'Profil enseignant non trouvé.')
        return redirect('home')

    # Vérifier l'attribution
    attribution = Attribution.objects.filter(
        matricule_en=enseignant,
        code_cours=code_cours,
        annee_academique=annee,
    ).first()
    if not attribution:
        messages.error(request, "Vous n'êtes pas autorisé à imprimer cette fiche.")
        return redirect('enseignant_mes_cours')

    # Infos cours
    cours_obj = attribution.get_cours_object()
    intitule = getattr(cours_obj, 'intitule_ue', code_cours)
    credit = getattr(cours_obj, 'credit', '-')
    type_cours = attribution.get_type_cours() or ''
    # classe_id est la PK string de Classe (FK brute)
    classe_code = getattr(cours_obj, 'classe_id', None)
    # Semestre et UE parente
    if type_cours == 'EC' and cours_obj and hasattr(cours_obj, 'code_ue') and cours_obj.code_ue:
        ue_parente = cours_obj.code_ue
        semestre = getattr(ue_parente, 'semestre', '-')
        ue_label = ue_parente.intitule_ue
    else:
        semestre = getattr(cours_obj, 'semestre', '-')
        ue_label = ''

    # Étudiants inscrits
    etudiants_data = []
    if classe_code:
        inscriptions = Inscription.objects.filter(
            code_classe_id=classe_code,
            annee_academique=annee,
        ).select_related('matricule_etudiant').order_by('matricule_etudiant__nom_complet')

        for i, insc in enumerate(inscriptions, 1):
            etud = insc.matricule_etudiant
            # Récupérer l'évaluation existante
            if type_cours == 'UE':
                ev = Evaluation.objects.filter(
                    matricule_etudiant=etud,
                    code_ue__code_ue=code_cours,
                    annee_academique=annee,
                ).first()
            else:
                ev = Evaluation.objects.filter(
                    matricule_etudiant=etud,
                    code_ec__code_ec=code_cours,
                    annee_academique=annee,
                ).first()

            cc = f"{ev.cc:.1f}" if ev and ev.cc is not None else ''
            examen = f"{ev.examen:.1f}" if ev and ev.examen is not None else ''
            note_finale = ''
            note_finale_val = None
            if ev and ev.cc is not None and ev.examen is not None:
                note_finale_val = round(ev.cc * 0.4 + ev.examen * 0.6, 2)
                note_finale = f"{note_finale_val:.2f}"
            rattrapage = f"{ev.rattrapage:.1f}" if ev and ev.rattrapage is not None else ''
            rachat = f"{ev.rachat:.1f}" if ev and ev.rachat is not None else ''

            etudiants_data.append({
                'num': i,
                'matricule': etud.matricule_et,
                'nom': etud.nom_complet,
                'cc': cc,
                'examen': examen,
                'note_finale': note_finale,
                'note_finale_val': note_finale_val,
                'rattrapage': rattrapage,
                'rachat': rachat,
            })

    # Génération PDF
    import os
    from django.conf import settings
    from PIL import Image as PILImage
    from reportlab.platypus import Image as RLImage

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    page_w = A4[0] - 3 * cm
    styles = getSampleStyleSheet()
    bureau_style = ParagraphStyle('bureau', parent=styles['Heading2'], fontSize=11, alignment=TA_CENTER, spaceAfter=2, textColor=colors.HexColor('#2c3e50'))
    title_style = ParagraphStyle('title', parent=styles['Heading1'], fontSize=13, alignment=TA_CENTER, spaceAfter=4, textColor=colors.HexColor('#2c3e50'))
    sub_style = ParagraphStyle('sub', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, spaceAfter=6)
    white_style = ParagraphStyle('white', parent=styles['Normal'], fontSize=8, alignment=TA_LEFT, textColor=colors.white)
    sig_style = ParagraphStyle('sig', parent=styles['Normal'], fontSize=8, alignment=TA_LEFT)

    from lmdmanagersystem.middleware import get_entete_path
    entete_path = get_entete_path()
    # Récupérer le code classe pour le titre
    classe_label = str(classe_code) if classe_code else code_cours

    story = []

    # Image d'en-tête
    if os.path.exists(entete_path):
        pil_img = PILImage.open(entete_path)
        iw, ih = pil_img.size
        desired_w = page_w
        desired_h = desired_w * (ih / iw)
        story.append(RLImage(entete_path, width=desired_w, height=desired_h))
        story.append(Spacer(1, 0.3 * cm))

    # Titres
    story.append(Paragraph(f"BUREAU DU JURY {classe_label}", bureau_style))
    story.append(Paragraph("FICHE DE COTATION", title_style))
    story.append(Paragraph(f"Année académique : <b>{annee}</b>", sub_style))
    story.append(Spacer(1, 0.2 * cm))

    # Infos cours/enseignant (texte blanc sur fond sombre)
    grade_code = enseignant.grade.code_grade if enseignant.grade else '-'
    if type_cours == 'EC':
        col1_label = Paragraph(f"<b>UE :</b> {ue_label}", white_style)
        col2_label = Paragraph(f"<b>EC :</b> {intitule}", white_style)
    else:
        col1_label = Paragraph(f"<b>UE :</b> {intitule}", white_style)
        col2_label = Paragraph("", white_style)
    col_info = [page_w * 0.38, page_w * 0.22, page_w * 0.20, page_w * 0.20]
    info_data = [
        [
            col1_label,
            col2_label,
            Paragraph(f"<b>Semestre :</b> S{semestre}", white_style),
            Paragraph(f"<b>Crédits :</b> {credit}", white_style),
        ],
        [
            Paragraph(f"<b>Enseignant :</b> {enseignant.nom_complet}", white_style),
            Paragraph(f"<b>Matricule :</b> {enseignant.matricule_en}", white_style),
            Paragraph(f"<b>Grade :</b> {grade_code}", white_style),
            Paragraph("", white_style),
        ],
    ]
    info_table = Table(info_data, colWidths=col_info)
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#4a6fa5')),
        ('PADDING', (0, 0), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.4 * cm))

    # Tableau des étudiants avec notes (sans Matricule, Observations ni Rachat)
    header = ['#', 'Nom complet', 'CC (/10)', 'Examen (/10)', 'Note Finale (/20)', 'Rattrapage (/20)']
    col_widths = [1*cm, page_w - 1*cm - 2.8*cm - 3.2*cm - 3.8*cm - 3*cm, 2.8*cm, 3.2*cm, 3.8*cm, 3*cm]
    table_data = [header]
    for etud in etudiants_data:
        table_data.append([
            str(etud['num']),
            etud['nom'],
            etud['cc'],
            etud['examen'],
            etud['note_finale'],
            etud['rattrapage'],
        ])

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
        ('ROWHEIGHT', (0, 1), (-1, -1), 20),
    ]
    # Coloriser en rouge les notes finales < 10
    for row_idx, etud in enumerate(etudiants_data, 1):
        val = etud.get('note_finale_val')
        if val is not None and val < 10:
            style_cmds.append(('TEXTCOLOR', (4, row_idx), (4, row_idx), colors.red))
            style_cmds.append(('FONTNAME', (4, row_idx), (4, row_idx), 'Helvetica-Bold'))
    t.setStyle(TableStyle(style_cmds))
    story.append(t)
    story.append(Spacer(1, 0.8 * cm))
    story.append(Paragraph(
        "<b>Signature de l'enseignant :</b> ____________________________"
        "&nbsp;&nbsp;&nbsp;&nbsp;<b>Date :</b> ____________________",
        sig_style
    ))

    doc.build(story)
    buffer.seek(0)
    filename = f"fiche_cotation_{code_cours}_{annee}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
def jury_fiche_cotation_excel(request, code_cours, annee):
    """Télécharger la fiche de cotation Excel d'un cours (accès jury).
    Réutilise la même logique que telecharger_grille_evaluation de l'enseignant."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')

    classe = jury.code_classe
    code_classe_str = classe.code_classe

    # Identifier le cours (UE ou EC)
    cours_info = {'code': code_cours, 'intitule': code_cours, 'type': None, 'classe': None}
    try:
        ue = UE.objects.get(code_ue=code_cours)
        cours_info['intitule'] = ue.intitule_ue
        cours_info['type'] = 'UE'
        cours_info['classe'] = ue.classe
    except UE.DoesNotExist:
        try:
            ec = EC.objects.get(code_ec=code_cours)
            cours_info['intitule'] = ec.intitule_ue
            cours_info['type'] = 'EC'
            cours_info['classe'] = ec.classe
        except EC.DoesNotExist:
            messages.error(request, 'Cours introuvable.')
            return redirect('jury_grille_cours')

    # Vérifier si rattrapage/rachat activé
    rattrapage_actif = False
    rachat_actif = False
    if cours_info.get('classe'):
        param_eval = ParametreEvaluation.objects.filter(
            code_classe__code_classe=cours_info['classe'],
            annee_academique=annee
        ).first()
        if param_eval:
            rattrapage_actif = param_eval.rattrapage_actif
            rachat_actif = param_eval.rachat_actif

    # Récupérer les étudiants et leurs évaluations
    data = []
    if cours_info.get('classe'):
        inscriptions = Inscription.objects.filter(
            code_classe__code_classe=cours_info['classe'],
            annee_academique=annee
        ).select_related('matricule_etudiant')

        for insc in inscriptions:
            eval_existante = Evaluation.objects.filter(
                matricule_etudiant=insc.matricule_etudiant,
                code_ue__code_ue=code_cours if cours_info['type'] == 'UE' else None,
                code_ec__code_ec=code_cours if cours_info['type'] == 'EC' else None,
            ).first()

            row = {
                'Matricule': insc.matricule_etudiant.matricule_et,
                'Nom Complet': insc.matricule_etudiant.nom_complet,
                'CC (0-10)': eval_existante.cc if eval_existante and eval_existante.cc else '',
                'Examen (0-10)': eval_existante.examen if eval_existante and eval_existante.examen else '',
            }
            if rattrapage_actif:
                row['Rattrapage (0-20)'] = eval_existante.rattrapage if eval_existante and eval_existante.rattrapage else ''
            if rachat_actif:
                row['Rachat (0-20)'] = eval_existante.rachat if eval_existante and eval_existante.rachat else ''
            data.append(row)

    # Créer le fichier Excel — même format que telecharger_grille_evaluation
    df = pd.DataFrame(data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Grille d'évaluation"

    # En-tête du document
    ws.merge_cells('A1:F1')
    ws['A1'] = f"GRILLE D'ÉVALUATION - {cours_info['intitule']}"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws['A2'] = f"Code: {code_cours}"
    ws['B2'] = f"Année: {annee}"
    ws['C2'] = f"Classe: {cours_info.get('classe', '-')}"
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 10

    headers = list(df.columns) if not df.empty else ['Matricule', 'Nom Complet', 'CC (0-10)', 'Examen (0-10)']
    header_row = 4

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 25

    data_alignment = Alignment(horizontal='center', vertical='center')
    data_font = Font(size=11)
    alt_fill = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")

    for row_idx, row_data in enumerate(df.values, header_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.font = data_font
            if col_idx <= 2:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = data_alignment
            if (row_idx - header_row) % 2 == 0:
                cell.fill = alt_fill

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14

    last_row = header_row + len(df) + 2
    ws.cell(row=last_row, column=1, value="Instructions:").font = Font(bold=True, italic=True)
    ws.cell(row=last_row + 1, column=1, value="• CC et Examen : notes sur 10").font = Font(italic=True, size=10)
    ws.cell(row=last_row + 2, column=1, value="• Rattrapage et Rachat : notes sur 20").font = Font(italic=True, size=10)
    ws.cell(row=last_row + 3, column=1, value="• Ne modifiez pas les colonnes Matricule et Nom").font = Font(italic=True, size=10, color="FF0000")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=grille_{code_cours}_{annee}.xlsx'
    return response


@login_required
def jury_importer_grille_evaluation(request, code_cours, annee):
    """Importer les notes depuis un fichier Excel - Vue Jury"""
    jury = get_jury_for_user(request)
    if not jury:
        messages.error(request, 'Profil jury non trouvé.')
        return redirect('home')

    try:
        classe = jury.code_classe

        # Vérifier que le cours appartient à la classe du jury
        cours_type = None
        cours_classe = None
        try:
            ue_obj = UE.objects.get(code_ue=code_cours)
            cours_type = 'UE'
            cours_classe = ue_obj.classe
        except UE.DoesNotExist:
            try:
                ec_obj = EC.objects.get(code_ec=code_cours)
                cours_type = 'EC'
                cours_classe = ec_obj.classe
            except EC.DoesNotExist:
                messages.error(request, 'Cours non trouvé.')
                return redirect('jury_grille_cours')

        cours_classe_code = cours_classe.code_classe if cours_classe else None
        if cours_classe_code != classe.code_classe:
            messages.error(request, "Ce cours n'appartient pas à votre classe.")
            return redirect('jury_grille_cours')

        if request.method == 'POST' and request.FILES.get('fichier_excel'):
            fichier = request.FILES['fichier_excel']

            try:
                # Les en-têtes sont à la ligne 4 (index 3) dans le fichier exporté
                df = pd.read_excel(fichier, header=3)
                if 'Matricule' not in df.columns:
                    fichier.seek(0)
                    df = pd.read_excel(fichier, header=0)
                if 'Matricule' not in df.columns:
                    raise ValueError("Le fichier Excel ne contient pas la colonne 'Matricule'. Utilisez le modèle téléchargé.")

                # Récupérer le nom du membre du jury connecté
                jury_nom = ''
                if hasattr(jury, 'president'):
                    try:
                        ens_jury = Enseignant.objects.get(matricule_en=jury.president)
                        jury_nom = ens_jury.nom_complet
                    except Enseignant.DoesNotExist:
                        jury_nom = jury.president
                if not jury_nom:
                    jury_nom = request.user.get_full_name() or request.user.username

                count = 0
                for _, row in df.iterrows():
                    matricule = str(row.get('Matricule', '')).strip()
                    if not matricule or matricule == 'nan':
                        continue

                    try:
                        etudiant = Etudiant.objects.get(matricule_et=matricule)

                        eval_data = {
                            'matricule_etudiant': etudiant,
                            'annee_academique': annee,
                        }

                        # CC
                        cc_val = row.get('CC (0-10)', '')
                        if pd.notna(cc_val) and cc_val != '':
                            eval_data['cc'] = float(cc_val)

                        # Examen
                        examen_val = row.get('Examen (0-10)', '')
                        if pd.notna(examen_val) and examen_val != '':
                            eval_data['examen'] = float(examen_val)

                        # Rattrapage
                        rattrapage_val = row.get('Rattrapage (0-20)', '')
                        if pd.notna(rattrapage_val) and rattrapage_val != '':
                            eval_data['rattrapage'] = float(rattrapage_val)

                        # Rachat
                        rachat_val = row.get('Rachat (0-20)', '')
                        if pd.notna(rachat_val) and rachat_val != '':
                            eval_data['rachat'] = float(rachat_val)

                        if cours_type == 'UE':
                            ue_obj_import = UE.objects.get(code_ue=code_cours)
                            eval_obj, created = Evaluation.objects.update_or_create(
                                matricule_etudiant=etudiant,
                                code_ue=ue_obj_import,
                                annee_academique=annee,
                                code_classe=cours_classe,
                                defaults=eval_data
                            )
                        else:
                            ec_obj_import = EC.objects.get(code_ec=code_cours)
                            eval_obj, created = Evaluation.objects.update_or_create(
                                matricule_etudiant=etudiant,
                                code_ec=ec_obj_import,
                                annee_academique=annee,
                                code_classe=cours_classe,
                                defaults=eval_data
                            )

                        # Marquer comme modifié par le jury si c'est une mise à jour
                        if not created:
                            from django.utils import timezone as tz
                            eval_obj.modifie_par_jury = True
                            eval_obj.jury_modificateur = jury_nom
                            eval_obj.date_modification_jury = tz.now()
                            eval_obj.save()

                        count += 1
                    except Etudiant.DoesNotExist:
                        continue

                messages.success(request, f'{count} notes importées avec succès!')
            except Exception as e:
                messages.error(request, f"Erreur lors de l'importation: {str(e)}")

        return redirect('jury_evaluer_cours', code_cours=code_cours, annee=annee)
    except Exception as e:
        messages.error(request, f"Erreur: {str(e)}")
        return redirect('jury_grille_cours')
